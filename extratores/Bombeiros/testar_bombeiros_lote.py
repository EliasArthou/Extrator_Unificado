"""
Extrator Bombeiros em Lote (Híbrido): busca inscrições CBM no banco e extrai
boletos usando Playwright + AntiCaptcha reCAPTCHA v2.

Uso:
    python testar_bombeiros_lote.py                              # usa banco padrão (Scai.WMB)
    python testar_bombeiros_lote.py -b "C:\\caminho\\Scai.WMB"  # banco específico
    python testar_bombeiros_lote.py -b dados.xlsx                # Excel
    python testar_bombeiros_lote.py -c 0954                      # a partir do cliente 0954
    python testar_bombeiros_lote.py -n 5                         # só os 5 primeiros
    python testar_bombeiros_lote.py -w 4                         # 4 workers em paralelo
    python testar_bombeiros_lote.py --visible                    # modo visível
"""

import os
import sys
import time
import queue
import threading
import argparse
import pandas as pd
from datetime import datetime
from types import SimpleNamespace
from dotenv import load_dotenv

load_dotenv()

# Suprime warnings de conexão (anticaptcha SDK)
import logging
import urllib3
import io
logging.getLogger("urllib3").setLevel(logging.CRITICAL)
logging.getLogger("requests").setLevel(logging.CRITICAL)
urllib3.disable_warnings()


class _FilteredStream:
    """Filtro de stdout/stderr que suprime linhas com palavras-chave indesejadas."""
    BLOCKED = ("ConnectionResetError", "Connection aborted", "RemoteDisconnected")

    def __init__(self, original):
        self._original = original

    def write(self, text):
        if any(kw in text for kw in self.BLOCKED):
            return len(text)
        return self._original.write(text)

    def flush(self):
        return self._original.flush()

    def __getattr__(self, name):
        return getattr(self._original, name)


from rich.console import Console
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    Progress, ProgressColumn, SpinnerColumn, TextColumn, BarColumn, Task,
)
from rich.text import Text

from auxiliares import utils as aux
from auxiliares import sensiveis as senha
from extratores.Bombeiros import bombeiros_hibrido as Bh

# ── Console Rich global ─────────────────────────────────────────────────────
console = Console(force_terminal=True, color_system="truecolor")


class _WorkerConsole:
    """
    Wrapper do Console com buffer por thread.
    Durante o processamento paralelo, cada worker acumula suas mensagens
    num buffer local. Ao chamar flush(), tudo é escrito de uma vez (com lock),
    garantindo saída sequencial e agrupada por cliente.
    """
    def __init__(self, base: Console, write_lock: threading.Lock):
        self._base = base
        self._write_lock = write_lock
        self._local = threading.local()

    def _buffer(self) -> list:
        if not hasattr(self._local, 'buf'):
            self._local.buf = []
        return self._local.buf

    def print(self, *args, **kwargs):
        name = threading.current_thread().name
        if name.startswith("worker-"):
            # Acumula no buffer local da thread
            self._buffer().append((args, kwargs))
        else:
            # Thread principal: escreve direto
            self._base.print(*args, **kwargs)

    def flush(self, header: str = ""):
        """Escreve todo o buffer acumulado no console, de forma sequencial."""
        buf = self._buffer()
        if not buf and not header:
            return
        wid = threading.current_thread().name.split("-")[-1]
        with self._write_lock:
            if header:
                self._base.print(header)
            for args, kwargs in buf:
                # Prefixa com [Wn] para identificar o worker
                if args:
                    args = (f"[dim]\\[W{wid}][/] {args[0]}",) + args[1:]
                self._base.print(*args, **kwargs)
        buf.clear()

    def __getattr__(self, name):
        return getattr(self._base, name)


class HMSElapsedColumn(ProgressColumn):
    """Mostra tempo decorrido no formato HH:MM:SS."""
    def render(self, task: Task) -> Text:
        elapsed = task.finished_time if task.finished else task.elapsed
        if elapsed is None:
            return Text("00:00:00", style="cyan")
        h, resto = divmod(int(elapsed), 3600)
        m, s = divmod(resto, 60)
        return Text(f"{h:02d}:{m:02d}:{s:02d}", style="cyan")


class HMSRemainingColumn(ProgressColumn):
    """Mostra tempo restante estimado no formato HH:MM:SS."""
    def render(self, task: Task) -> Text:
        remaining = task.time_remaining
        if remaining is None:
            return Text("--:--:--", style="cyan")
        h, resto = divmod(int(remaining), 3600)
        m, s = divmod(resto, 60)
        return Text(f"{h:02d}:{m:02d}:{s:02d}", style="cyan")


def _worker(worker_id: int, work_queue: queue.Queue, objeto,
            headless: bool, caminho_perfil_base: str,
            progress, task_id, lock: threading.Lock,
            log_resultados: list, counters: dict,
            worker_console: '_WorkerConsole'):
    """
    Função executada por cada thread worker.
    Cada worker tem seu próprio browser com perfil isolado.
    Mensagens são bufferizadas e impressas todas de uma vez ao final de cada cliente.
    """
    from playwright.sync_api import sync_playwright

    # Perfil isolado por worker (evita conflito entre browsers)
    perfil_worker = f"{caminho_perfil_base}_{worker_id}"

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=perfil_worker,
            headless=headless,
            accept_downloads=True,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"]
        )
        page = context.pages[0] if context.pages else context.new_page()

        try:
            while True:
                try:
                    i, linha, total = work_queue.get_nowait()
                except queue.Empty:
                    break

                cod = str(linha[Bh.Codigo])
                cbm_raw = str(linha[Bh.NrCBM]).strip().zfill(8)
                cbm_fmt = f"{cbm_raw[:7]}-{cbm_raw[7]}"
                iptu = str(linha[Bh.NrIPTU_CBM]).strip() if len(linha) > 2 else ''
                cidade = str(linha[Bh.NrCidade]).strip() if len(linha) > Bh.NrCidade else ''

                # Header do cliente (vai pro flush junto com os prints internos)
                header = (
                    f"\n[bold]{'='*60}[/]\n"
                    f"[bold cyan][W{worker_id}][/] [{i+1}/{total}] Cliente: {cod} | CBMERJ: {cbm_fmt}"
                    + (f" | IPTU: {iptu}" if iptu else "")
                    + (f" | {cidade}" if cidade else "")
                    + f"\n[bold]{'='*60}[/]"
                )

                try:
                    dados, df = Bh.extrairbombeiros_hibrido(page, objeto, linha, cota_unica=args.cota_unica)
                    status = dados[3] if len(dados) > 3 else 'ERRO'
                except Exception as e:
                    status = f"ERRO: {e}"

                # Flush: escreve header + todos os prints internos de uma vez
                worker_console.flush(header=header)

                with lock:
                    if 'OK' in status:
                        counters['ok'] += 1
                    elif 'Sem débito' in status or 'Sem débitos' in status:
                        counters['sem_debito'] += 1
                    else:
                        counters['erros'] += 1

                    log_resultados.append({
                        'Cod Cliente': cod,
                        'CBMERJ': cbm_fmt,
                        'IPTU': iptu,
                        'Cidade': cidade,
                        'Status': status,
                    })

                    progress.advance(task_id)

                work_queue.task_done()

        finally:
            context.close()


def main():
    # Ativa filtro de stdout/stderr pra suprimir ConnectionResetError do anticaptcha SDK
    sys.stdout = _FilteredStream(sys.stdout)
    sys.stderr = _FilteredStream(sys.stderr)

    parser = argparse.ArgumentParser(description="Extrator Bombeiros em Lote (Híbrido)")
    parser.add_argument("-b", "--banco", default="",
                        help="Caminho do banco de dados (.WMB ou .xlsx)")
    parser.add_argument("-c", "--cliente", default="0",
                        help="Código do cliente inicial (0 = todos)")
    parser.add_argument("-n", "--max", type=int, default=0,
                        help="Máximo de registros a processar (0 = todos)")
    parser.add_argument("-p", "--pasta", default="Downloads/Bombeiros",
                        help="Pasta para salvar os PDFs")
    parser.add_argument("-w", "--workers", type=int, default=1,
                        help="Número de workers em paralelo (padrão: 1)")
    parser.add_argument("--visible", action="store_true",
                        help="Abrir janela do navegador (padrão: headless)")
    parser.add_argument("--cota-unica", action="store_true", dest="cota_unica",
                        help="Filtrar pra 1 boleto por exercicio (cota unica). Default: parcelado.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Só lista as inscrições, sem extrair")
    args = parser.parse_args()

    num_workers = max(1, args.workers)

    # ── 1. Conecta ao banco ──────────────────────────────────
    caminho_banco = args.banco
    if not caminho_banco:
        caminho_projeto = aux.caminhoprojeto()
        candidato = os.path.join(caminho_projeto, 'Scai.WMB')
        if os.path.isfile(candidato):
            caminho_banco = candidato
        else:
            console.print("[red][ERRO][/] Banco não encontrado. Use -b para especificar o caminho.")
            console.print(f"  Procurado em: {candidato}")
            sys.exit(1)

    console.print(f"[cyan][BD][/] Abrindo banco: {caminho_banco}")
    bd = aux.Banco(caminho_banco)

    # ── 2. Executa consulta SQL ──────────────────────────────
    sql = senha.SQLCBM
    cliente_corte = str(args.cliente).zfill(4)
    if cliente_corte != '0000':
        sql = sql.replace(';', ' ') + f" WHERE Codigo >= '{cliente_corte}' ORDER BY Codigo;"

    console.print(f"[cyan][BD][/] Executando consulta CBM...")
    resultado = bd.consultar(sql)

    if not resultado:
        console.print("[yellow][BD][/] Nenhum registro encontrado.")
        sys.exit(0)

    resultado.sort(key=lambda x: x[0])

    total = len(resultado)
    if args.max > 0:
        resultado = resultado[:args.max]

    console.print(f"[cyan][BD][/] {total} inscrição(ões) encontrada(s)"
                  + (f", processando {len(resultado)}" if args.max > 0 else ""))

    # ── 3. Lista inscrições ──────────────────────────────────
    console.print(f"\n{'#':>4}  {'Código':<8}  {'CBMERJ':<12}  {'IPTU':<12}  {'Cidade'}")
    console.print(f"{'─'*4}  {'─'*8}  {'─'*12}  {'─'*12}  {'─'*20}")
    for i, linha in enumerate(resultado):
        cod = str(linha[Bh.Codigo])
        cbm_raw = str(linha[Bh.NrCBM]).strip().zfill(8)
        cbm_fmt = f"{cbm_raw[:7]}-{cbm_raw[7]}"
        iptu = str(linha[Bh.NrIPTU_CBM]) if len(linha) > 2 else ''
        cidade = str(linha[Bh.NrCidade]) if len(linha) > Bh.NrCidade else ''
        console.print(f"{i+1:>4}  {cod:<8}  {cbm_fmt:<12}  {iptu:<12}  {cidade}")

    if args.dry_run:
        console.print(f"\n[yellow][DRY-RUN][/] Nenhuma extração realizada.")
        return

    # ── 4. Prepara execução ──────────────────────────────────
    os.makedirs(args.pasta, exist_ok=True)

    caminho_perfil_base = os.path.join(
        os.path.dirname(__file__), "Profile",
        os.getenv('NOMEPROFILECBM', 'cbm_profile')
    )

    headless = not args.visible
    console.print(f"\n[bold cyan][PW][/] Iniciando {num_workers} worker(s) ({'visível' if args.visible else 'headless'})...")

    # Impede hibernação durante o processo (tela pode desligar normalmente)
    aux.prevent_sleep()

    tempo_inicio = time.time()
    log_resultados = []
    counters = {'ok': 0, 'sem_debito': 0, 'erros': 0}
    lock = threading.Lock()

    # Fila de trabalho
    work_queue = queue.Queue()
    for i, linha in enumerate(resultado):
        work_queue.put((i, linha, len(resultado)))

    # Objeto compartilhado (pastadownload é thread-safe pois só lê)
    objeto = SimpleNamespace(pastadownload=args.pasta)

    # Worker console: buffer por thread + escrita sequencial com lock
    _write_lock = threading.Lock()
    worker_console = _WorkerConsole(console, _write_lock)
    _console_original_bh = Bh.console
    Bh.console = worker_console

    # ── Barra de progresso Rich (única, compartilhada) ───────
    progress = Progress(
        SpinnerColumn(),
        TextColumn(
            f"[bold blue]{num_workers} worker(s)[/]" if num_workers > 1
            else "[bold blue]Bombeiros[/]",
            justify="left"
        ),
        BarColumn(bar_width=40),
        TextColumn("{task.completed:>4}/{task.total:<4}", style="bold cyan"),
        TextColumn("│ Restante:", style="dim"),
        HMSRemainingColumn(),
        TextColumn("│ Decorrido:", style="dim"),
        HMSElapsedColumn(),
        expand=False,
    )
    task_id = progress.add_task("", total=len(resultado))

    def _build_panel():
        return Panel(
            progress,
            title="[bold green]Progresso Bombeiros",
            border_style="green",
            padding=(0, 1),
        )

    live = Live(_build_panel(), console=console, refresh_per_second=1,
                transient=False, screen=False, vertical_overflow="visible")
    live.start()

    try:
        if num_workers == 1:
            # ── Modo sequencial (1 worker, sem threads) ──────
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=caminho_perfil_base,
                    headless=headless,
                    accept_downloads=True,
                    args=["--disable-blink-features=AutomationControlled", "--start-maximized"]
                )
                page = context.pages[0] if context.pages else context.new_page()

                try:
                    while not work_queue.empty():
                        try:
                            i, linha, total_itens = work_queue.get_nowait()
                        except queue.Empty:
                            break

                        cod = str(linha[Bh.Codigo])
                        cbm_raw = str(linha[Bh.NrCBM]).strip().zfill(8)
                        cbm_fmt = f"{cbm_raw[:7]}-{cbm_raw[7]}"
                        iptu = str(linha[Bh.NrIPTU_CBM]).strip() if len(linha) > 2 else ''
                        cidade = str(linha[Bh.NrCidade]).strip() if len(linha) > Bh.NrCidade else ''

                        console.print(
                            f"\n[bold]{'='*60}[/]\n"
                            f"[bold cyan][{i+1}/{total_itens}][/] Cliente: {cod} | CBMERJ: {cbm_fmt}"
                            + (f" | IPTU: {iptu}" if iptu else "")
                            + (f" | {cidade}" if cidade else "")
                            + f"\n[bold]{'='*60}[/]"
                        )

                        try:
                            dados, df = Bh.extrairbombeiros_hibrido(page, objeto, linha, cota_unica=args.cota_unica)
                            status = dados[3] if len(dados) > 3 else 'ERRO'
                        except Exception as e:
                            status = f"ERRO: {e}"

                        if 'OK' in status:
                            counters['ok'] += 1
                        elif 'Sem débito' in status or 'Sem débitos' in status:
                            counters['sem_debito'] += 1
                        else:
                            counters['erros'] += 1

                        log_resultados.append({
                            'Cod Cliente': cod,
                            'CBMERJ': cbm_fmt,
                            'IPTU': iptu,
                            'Cidade': cidade,
                            'Status': status,
                        })

                        progress.advance(task_id)
                        work_queue.task_done()

                finally:
                    context.close()

        else:
            # ── Modo paralelo (N workers em threads) ─────────
            threads = []
            for wid in range(num_workers):
                t = threading.Thread(
                    target=_worker,
                    args=(wid, work_queue, objeto, headless,
                          caminho_perfil_base, progress, task_id,
                          lock, log_resultados, counters, worker_console),
                    daemon=True,
                    name=f"worker-{wid}",
                )
                threads.append(t)

            for t in threads:
                t.start()

            for t in threads:
                t.join()

    finally:
        live.stop()
        Bh.console = _console_original_bh
        # Libera bloqueio de hibernação
        aux.allow_sleep()

    # ── 5. Gera log Excel ────────────────────────────────────
    tempo_total = time.time() - tempo_inicio
    horas, resto = divmod(tempo_total, 3600)
    minutos, segundos = divmod(resto, 60)
    tempo_fmt = f"{int(horas):02d}:{int(minutos):02d}:{int(segundos):02d}"

    if log_resultados:
        timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        nome_log = f"Log_Bombeiros_{timestamp}.xlsx"
        pasta_log = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logs")
        os.makedirs(pasta_log, exist_ok=True)
        caminho_log = os.path.join(pasta_log, nome_log)

        df_log = pd.DataFrame(log_resultados)

        try:
            df_log.to_excel(caminho_log, index=False, engine='openpyxl')

            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = load_workbook(caminho_log)
            ws = wb.active

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                status_cell = row[4]
                if status_cell.value:
                    if 'OK' in str(status_cell.value):
                        status_cell.font = Font(color='008000', bold=True)
                    elif 'Sem débito' in str(status_cell.value):
                        status_cell.font = Font(color='0070C0')
                    else:
                        status_cell.font = Font(color='FF0000', bold=True)

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 30

            wb.save(caminho_log)
        except Exception as e:
            console.print(f"[yellow][AVISO][/] Erro ao formatar Excel: {e}")

        console.print(f"\n[green][LOG][/] Resultado salvo em: {caminho_log}")

    # ── 6. Resumo ────────────────────────────────────────────
    console.print(f"\n[bold]{'='*60}[/]")
    console.print(f"[bold green]RESUMO[/]")
    console.print(f"[bold]{'='*60}[/]")
    console.print(f"  Total processados: {len(resultado)}")
    console.print(f"  Workers:           {num_workers}")
    console.print(f"  [green]OK:                {counters['ok']}[/]")
    console.print(f"  [cyan]Sem débito:        {counters['sem_debito']}[/]")
    console.print(f"  [red]Erros:             {counters['erros']}[/]")
    console.print(f"  Tempo total:       {tempo_fmt}")
    console.print(f"  Pasta:             {os.path.abspath(args.pasta)}")

    if counters['erros'] > 0:
        console.print(f"\n[bold red]Registros com erro:[/]")
        for reg in log_resultados:
            if 'OK' not in reg['Status'] and 'Sem débito' not in reg['Status']:
                console.print(f"  {reg['Cod Cliente']} | {reg['CBMERJ']} | {reg['Status']}")


if __name__ == "__main__":
    main()
