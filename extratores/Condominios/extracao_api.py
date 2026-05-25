"""
extracao_api.py — Extrator de boletos via HTTP (sem Playwright).

Famílias suportadas:
  • Superlógica      — *.superlogica.net                     (API JSON)
  • LiveFacilities   — *.livefacilities.com.br,
                       live.bap.com.br, app.bcfadm.com.br,
                       fffacilities.com.br, zirtaweb.com      (ASP.NET WebForms)
  • Condomob         — app.condomob.net                       (Firebase + REST JSON)
  • Webware          — webware.com.br                        (ASP legado + Vue.js)
  • Protel           — condominio.protel.com.br               (jQuery)
  • Nacional         — sistemas.admnacional.com.br            (ASP clássico)
  • CIPA             — cipafacil.digital                     (API JSON REST)

Uso standalone:
    py extracao_api.py                           # todas as admins suportadas
    py extracao_api.py ADIPLANTEC                # só uma
    py extracao_api.py ESTASA "LUMEN ADMINISTRADORA"   # várias
"""

import re
import json
import sys
import os
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse, urljoin

import threading

import requests
import pandas as pd
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from rich.console import Console
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, TextColumn, BarColumn,
    TaskProgressColumn, TimeRemainingColumn, TimeElapsedColumn,
)

load_dotenv()

# ── Console Rich global ────────────────────────────────────────────────────────
console = Console(force_terminal=True, color_system="truecolor")


# ═══════════════════════════════════════════════════════════════════════════════
#  Barra de progresso fixa (idêntica à de fluxo-pw-new.py)
# ═══════════════════════════════════════════════════════════════════════════════

class FixedProgressLogger:
    """Barra de progresso na parte inferior; logs rolam normalmente acima."""

    def __init__(self, total: int, description: str = "Processando..."):
        self.total = total
        self.description = description
        self.current = 0
        self.lock = threading.Lock()
        self.live = None

        self.progress = Progress(
            SpinnerColumn(),
            TextColumn("[bold blue]{task.description}", justify="left"),
            BarColumn(bar_width=40),
            TextColumn("{task.completed:>4}/{task.total:<4}", style="bold cyan"),
            TaskProgressColumn(),
            TextColumn("│ Restante:", style="dim"),
            TimeRemainingColumn(),
            TextColumn("│ Decorrido:", style="dim"),
            TimeElapsedColumn(),
            expand=False,
        )
        self.task_id = self.progress.add_task(description, total=total)

    def _build_display(self):
        return Panel(
            self.progress,
            title="[bold green]⏳ Progresso ",
            border_style="green",
            padding=(0, 1),
        )

    def start(self):
        self.live = Live(
            self._build_display(),
            console=console,
            auto_refresh=False,   # sem auto-refresh; atualiza só em advance()/update_description()
            transient=False,
            screen=False,
            vertical_overflow="visible",
        )
        self.live.start()

    def stop(self):
        if self.live:
            self.live.stop()
            self.live = None

    def log(self, msg: str):
        with self.lock:
            if self.live:
                self.live.console.print(msg)
            else:
                console.print(msg)

    def advance(self, amount: int = 1):
        with self.lock:
            self.current += amount
            self.progress.advance(self.task_id, amount)
            if self.live:
                self.live.update(self._build_display(), refresh=True)

    def update_description(self, desc: str):
        with self.lock:
            self.description = desc
            self.progress.update(self.task_id, description=desc)
            if self.live:
                self.live.update(self._build_display(), refresh=True)


# Instância global do logger de progresso
_progress_logger: "FixedProgressLogger | None" = None

# Lock global para flush atômico de backlog (prints agrupados por administradora)
_print_lock = threading.Lock()

# Listas globais para geração dos logs Excel (análogas às do fluxo-pw-new.py)
_boletos_processados: list = []   # dicts retornados por processar_boleto_baixado
_boletos_duplicados:  list = []   # dicts {Cliente, Cod_Barras, Caminho}

# Pasta do ciclo (ex: Downloads/Condomínios/2026_05_Mai). Setada no __main__ ao
# parsear o argumento de ciclo da CLI. Quando setada, força todos os PDFs do
# ciclo para essa pasta — sobrepõe o vencimento extraído do barcode (que pode
# estar mal lido). Permanece None quando o módulo é usado como biblioteca.
_PASTA_CICLO_ATUAL: "Path | None" = None

# ── Compatibilidade com o restante do projeto ─────────────────────────────────
try:
    from extratores.Condominios.fluxo_pw_new import (
        DOWNLOADS_DIR,
        verificar_boleto_duplicado,
        registrar_codigo_barras,
        registrar_boleto_duplicado,
    )
    _usa_log_global = True
except ImportError:
    DOWNLOADS_DIR = Path("./Downloads")
    _usa_log_global = False


def set_downloads_dir(pasta: "str | Path"):
    """Permite ao chamador externo definir a pasta de downloads."""
    global DOWNLOADS_DIR
    DOWNLOADS_DIR = Path(pasta)


def get_boletos_processados() -> list:
    """Retorna e limpa a lista de boletos processados (dados do barcodereader)."""
    global _boletos_processados
    copia = list(_boletos_processados)
    _boletos_processados.clear()
    return copia


def get_boletos_duplicados() -> list:
    """Retorna e limpa a lista de boletos duplicados."""
    global _boletos_duplicados
    copia = list(_boletos_duplicados)
    _boletos_duplicados.clear()
    return copia


def _log_fallback(msg: str):
    """Log padrão quando _progress_logger não está ativo."""
    if _progress_logger is not None:
        _progress_logger.log(msg)
    else:
        console.print(msg)


def _flush_log(log_buffer: list):
    """Imprime todas as mensagens acumuladas de uma vez (thread-safe)."""
    with _print_lock:
        for m in log_buffer:
            console.print(m)
    log_buffer.clear()

try:
    from auxiliares import utils as aux
    from auxiliares import aux_patches  # registra adicionarcabecalhopdf_topo_adaptativo em aux
    from core import boletos as _boletos_mod
    _tem_aux = True
except ImportError:
    _tem_aux = False


# ═══════════════════════════════════════════════════════════════════════════════
#  Configuração de sites (lido do .env)
# ═══════════════════════════════════════════════════════════════════════════════

def _carregar_mapa_sites() -> dict:
    """Lê LISTAADMINISTRADORA_API (ou fallback LISTAADMINISTRADORA) do .env."""
    raw = os.getenv("LISTAADMINISTRADORA_API") or os.getenv("LISTAADMINISTRADORA", "[]")
    try:
        return {item["nomereal"]: item["site"] for item in json.loads(raw)}
    except Exception:
        return {}

MAPA_SITES: dict = _carregar_mapa_sites()


def _familia_from_url(url: str) -> str:
    """Detecta a família do portal a partir da URL."""
    u = (url or "").lower()
    if "superlogica" in u:
        return "superlogica"
    if "livefacilities" in u or "bap.com.br" in u or "app.bcfadm.com.br" in u or "fffacilities.com.br" in u or "zirtaweb.com" in u or "prosiga.protest.com.br" in u:
        return "livefacilities"
    if "webware" in u:
        return "webware"
    if "immobileweb.com.br" in u:
        return "immobileweb"
    if "protel.com.br" in u:
        return "protel"
    if "admnacional.com.br" in u:
        return "nacional"
    if "cipafacil.digital" in u:
        return "cipa"
    if "condomob.net" in u:
        return "condomob"
    return "desconhecido"

FAMILIAS_SUPORTADAS = {"superlogica", "livefacilities", "webware", "immobileweb", "protel", "nacional", "cipa", "condomob"}

ADMINS_SUPORTADAS: set = {
    nome for nome, site in MAPA_SITES.items()
    if _familia_from_url(site) in FAMILIAS_SUPORTADAS
}


# ═══════════════════════════════════════════════════════════════════════════════
#  Helpers de texto
# ═══════════════════════════════════════════════════════════════════════════════

def _normalizar(texto: str) -> str:
    t = re.sub(r"[^\w\s]", " ", texto or "")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t

def _similaridade(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalizar(a), _normalizar(b)).ratio()

def _melhor_match(alvo: str, opcoes: list, cutoff: float = 0.70) -> "int | None":
    if not opcoes:
        return None
    scores = [(_similaridade(alvo, op), i) for i, op in enumerate(opcoes)]
    melhor_score, melhor_idx = max(scores)
    return melhor_idx if melhor_score >= cutoff else None


def _numeros(texto: str) -> list:
    """Extrai sequências numéricas de um texto, ignorando zeros à esquerda."""
    return [str(int(n)) for n in re.findall(r'\d+', texto)]


def _melhor_match_score(alvo: str, opcoes: list, cutoff: float = 0.70,
                        validar_numero: bool = False) -> "tuple[int, float] | tuple[None, float]":
    """
    Como _melhor_match, mas retorna (idx, score).

    validar_numero=True: descarta candidatos cujo maior número difere do alvo
    (evita falsos positivos como 'APT. 604' → 'AP 603').
    """
    if not opcoes:
        return None, 0.0

    nums_alvo = _numeros(alvo)
    scores = []
    for i, op in enumerate(opcoes):
        score = _similaridade(alvo, op)
        if validar_numero and nums_alvo:
            nums_op = _numeros(op)
            # O maior número costuma ser o número do apartamento
            principal_alvo = max(nums_alvo, key=len)
            principal_op   = max(nums_op,  key=len) if nums_op else ""
            if principal_op and principal_op != principal_alvo:
                score = 0.0          # penalidade total: número diferente
        scores.append((score, i))

    melhor_score, melhor_idx = max(scores)
    if melhor_score >= cutoff:
        return melhor_idx, melhor_score
    return None, melhor_score


def _limpar_nome_condo_site(nome: str) -> str:
    """Remove prefixo de código numérico ('0548 - ', '0742 - ') do nome do
    condomínio retornado pelo site. Esse prefixo é ruído puro pra fuzzy match."""
    return re.sub(r"^\s*\d+\s*[-–—]\s*", "", nome or "").strip()


def _melhor_match_estruturado(alvo_condo: str, alvo_apto: str,
                              unidades: list,
                              cutoff: float = 0.60) -> "tuple[int, float] | tuple[None, float]":
    """
    Match estruturado em 2 campos (Opção B):
      - Compara `alvo_condo` contra `u["condo"]` (sem prefixo de código).
        Usa substring matching: se alvo está contido no nome do site (≥4 chars),
        score = 1.0. Resolve 'WATERWAYS' vs '0742 - WATERWAYS SOUTHWEST' onde
        SequenceMatcher dava 0.545 pela diferença de tamanho.
      - Compara `alvo_apto` contra `u["bloco"] + " " + u["unidade"]`.
      - Score final = média dos dois (50/50). Quando `alvo_apto` está vazio,
        score final = score_condo (100% no condomínio).

    Retorna (idx_melhor, score) ou (None, melhor_score_obtido).
    """
    if not unidades:
        return None, 0.0

    apto_vazio = _campo_vazio(alvo_apto)
    alvo_condo_norm = _normalizar(alvo_condo)
    alvo_apto_norm  = _normalizar(alvo_apto)

    scores = []
    for i, u in enumerate(unidades):
        condo_site_limpo = _limpar_nome_condo_site(u.get("condo", ""))
        condo_site_norm  = _normalizar(condo_site_limpo)

        # 1) Substring match (reforça matches parciais legítimos)
        if (alvo_condo_norm and condo_site_norm
                and len(alvo_condo_norm) >= 4
                and (alvo_condo_norm in condo_site_norm
                     or condo_site_norm in alvo_condo_norm)):
            sc_condo = 1.0
        else:
            sc_condo = _similaridade(alvo_condo, condo_site_limpo)

        if apto_vazio:
            score = sc_condo
        else:
            bloco_apto_site = f"{u.get('bloco', '')} {u.get('unidade', '')}".strip()
            bloco_apto_norm = _normalizar(bloco_apto_site)
            if (alvo_apto_norm and bloco_apto_norm
                    and len(alvo_apto_norm) >= 3
                    and alvo_apto_norm in bloco_apto_norm):
                sc_apto = 1.0
            else:
                sc_apto = _similaridade(alvo_apto, bloco_apto_site)
            score = (sc_condo + sc_apto) / 2
        scores.append((score, i))

    melhor_score, melhor_idx = max(scores)
    if melhor_score >= cutoff:
        return melhor_idx, melhor_score
    return None, melhor_score


def _log_fuzzy_match(tipo: str, alvo_banco: str, encontrado_site: str, score: float, log):
    """Loga resultado de fuzzy match no mesmo formato do fluxo-pw-new.py."""
    if score >= 0.999:
        log(f"[FUZZY {tipo}] ✅ Match exato (score={score:.2f}):")
        log(f"              📋 Buscado (Base):    '{alvo_banco}'")
        log(f"              🌐 Encontrado (Site): '{encontrado_site}'")
        log(f"              🎯 Método: Comparação direta")
    else:
        log(f"[FUZZY {tipo}] 🔍 Match parcial (score={score:.2f}):")
        log(f"              📋 Buscado (Base):    '{alvo_banco}'")
        log(f"              🌐 Encontrado (Site): '{encontrado_site}'")


def _log_sem_match(tipo: str, alvo: str, opcoes: list, log,
                   label_alvo: str = "Base", label_opcoes: str = "Site"):
    """Loga falha de fuzzy match.

    label_alvo / label_opcoes permitem indicar a origem real dos valores
    (ex: 'Site' / 'Base' quando a busca é do site contra o banco).
    """
    log(f"[FUZZY {tipo}] ❌ Nenhum match (cutoff não atingido):")
    log(f"              📋 Buscado ({label_alvo}):    '{alvo}'")
    top = opcoes[:5] if len(opcoes) > 5 else opcoes
    log(f"              🌐 Opções na {label_opcoes}:   {top}")

def _campo_vazio(valor: str) -> bool:
    """True se o campo do banco está vazio ou zerado ('', '0')."""
    return not valor or valor == "0"

def _fmt_opcoes(opcoes: list, n: int = 8) -> str:
    """Formata lista de opções disponíveis como string compacta para o Excel."""
    top = opcoes[:n] if len(opcoes) > n else opcoes
    return " | ".join(str(o) for o in top)

def _base_url(site_url: str) -> str:
    p = urlparse(site_url)
    return f"{p.scheme}://{p.netloc}"

def _scrape_hidden(soup, form=None) -> dict:
    """Extrai todos os inputs hidden de um formulário."""
    container = form or soup
    return {
        i["name"]: i.get("value", "")
        for i in container.find_all("input", {"type": "hidden"})
        if i.get("name")
    }

def _headers_base() -> dict:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        )
    }


def _nova_sessao() -> requests.Session:
    """Cria uma Session com verify=False (proxy corporativo) e warnings suprimidos."""
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    s = requests.Session()
    s.verify = False
    s.headers.update(_headers_base())
    return s


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Superlógica
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoSuperlogica:

    def __init__(self, base_url: str):
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(base_url.rstrip("/"))
        qs = parse_qs(parsed.query)
        self.licenca = qs.get("licenca", [None])[0]   # ex: "qualityhouse"
        # Novo formato: superlogica.net?licenca=xxx  →  base = superlogica.net
        # Antigo formato: qualityhouse.superlogica.net →  licenca = None
        self.base_url = f"{parsed.scheme}://{parsed.netloc}"
        self.session = _nova_sessao()

    def _lq(self) -> str:
        """Query string de licença para URLs que precisam dela."""
        return f"?licenca={self.licenca}" if self.licenca else ""

    def login(self, email: str, senha: str):
        email, senha = email.strip(), senha.strip()
        self.session.headers["X-Requested-With"] = "XMLHttpRequest"
        self.session.get(f"{self.base_url}/clients/areadocondomino{self._lq()}", timeout=15)
        payload = {
            "email": email, "senha": senha,
            "url": "", "CHAVE": self.licenca or "", "idCondominio": "",
            "FL_LOGIN_WEB": "1", "hashemail": "", "salvar": "Entrar",
        }
        r = self.session.post(
            f"{self.base_url}/areadocondomino/atual/publico/auth",
            data=payload, timeout=15,
        )
        data = r.json()
        if str(data.get("status")) != "202":
            raise ValueError(f"Login falhou ({self.base_url}{self._lq()}): {data.get('msg', 'sem mensagem')}")

    def listar_condominios(self) -> dict:
        self.session.headers.pop("X-Requested-With", None)
        r = self.session.get(
            f"{self.base_url}/clients/areadocondomino/index/index{self._lq()}",
            timeout=15, allow_redirects=True,
        )
        m = re.search(r"var condominios\s*=\s*(\{[^}]+\})", r.text)
        return json.loads(m.group(1)) if m else {}

    def trocar_condominio(self, id_condo: str):
        self.session.headers["X-Requested-With"] = "XMLHttpRequest"
        self.session.post(
            f"{self.base_url}/areadocondomino/atual/condominio/alterarcondominioatual",
            data={"ID_CONDOMINIO_COND": id_condo}, timeout=15,
        )

    def listar_cobrancas_em_aberto(self) -> list:
        self.session.headers["X-Requested-With"] = "XMLHttpRequest"
        r = self.session.get(
            f"{self.base_url}/areadocondomino/atual/cobranca/index", timeout=15,
        )
        return [
            item for item in r.json().get("data", [])
            if item.get("fl_status_recb") == "0"
            and not item.get("boleto_indisponivel")
            and item.get("link2viaboleto")
        ]

    def baixar_pdf(self, link2viaboleto: str) -> bytes:
        self.session.headers.pop("X-Requested-With", None)
        url = link2viaboleto.replace("FaturaHtml", "FaturaPdf")
        r = self.session.get(url, timeout=30, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Boleto indisponível — Content-Type: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão LiveFacilities  (BAP, BCF, HFLEX, VÓRTEX)
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoLiveFacilities:
    """
    Automação via requests para portais LiveFacilities (ASP.NET WebForms).

    Campos do formulário de login:
      ucLoginSistema$tbNomeEntrar  → usuário
      ucLoginSistema$tbSenhaEntrar → senha
      ucLoginSistema$btEntrar      → submit
      + __VIEWSTATE, __EVENTVALIDATION, __VIEWSTATEGENERATOR,
        _TSM_HiddenField_, hf_controleconteudo, ucLoginSistema$hfLoginApp
    """

    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.session = _nova_sessao()
        self._login_html: str = ""

    # ── Login ──────────────────────────────────────────────────────────────────

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()

        # 1. GET página de login → scrape tokens ASP.NET
        r = self.session.get(f"{self.base_url}/Index.aspx", timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        form = soup.find("form")
        if not form:
            raise ValueError("Formulário de login não encontrado")

        payload = _scrape_hidden(soup, form)
        payload["ucLoginSistema$tbNomeEntrar"] = usuario
        payload["ucLoginSistema$tbSenhaEntrar"] = senha
        payload["ucLoginSistema$btEntrar"]      = "Entrar"

        # 2. POST credenciais
        action = form.get("action") or f"{self.base_url}/"
        if not action.startswith("http"):
            action = urljoin(self.base_url, action)

        r2 = self.session.post(action, data=payload, timeout=15,
                               allow_redirects=True)
        self._login_html = r2.text

        # Verifica falha: se ainda há campo de senha visível, login falhou
        if "ucLoginSistema_tbSenhaEntrar" in r2.text:
            raise ValueError(f"Login falhou para {usuario} em {self.base_url}")

        # ── Primeiro Acesso (BAP, possivelmente outros) ───────────────────────
        # Algumas contas LiveFacilities, no primeiro login, são redirecionadas
        # pra /Portal/PessoaAltera.aspx ("ESTE É SEU PRIMEIRO ACESSO — confirme
        # seus dados"). O portal BLOQUEIA qualquer outra página até o cadastro
        # ser submetido. Detecta e re-submete o form com os mesmos valores
        # (sem alterar nada).
        url_final = r2.url.lower() if hasattr(r2, "url") else ""
        if "pessoaaltera" in url_final or "primeiro acesso" in r2.text.lower():
            self._confirmar_primeiro_acesso(r2)

    def _confirmar_primeiro_acesso(self, r_redirect):
        """
        Fluxo de 'Primeiro Acesso' BAP / LiveFacilities (3 etapas):

          1) Re-submete o form de cadastro (PessoaAltera.aspx) com os MESMOS
             valores que o servidor mandou — sem alterar nada.
          2) Após sucesso, aparece popup "Dados alterados com sucesso" + redirect
             pra página de termos LGPD.
          3) Aceita os termos clicando no botão CONCORDO (senão o portal faz
             logout automático).

        Levanta ValueError com detalhes se qualquer etapa falhar.
        """
        # ── ETAPA 1: cadastro ──────────────────────────────────────────────────
        r_post = self._submeter_form_inalterado(
            r_redirect,
            etapa="cadastro (PessoaAltera)",
            submit_pred=lambda name: "confirm" in name.lower(),
        )

        url_post = (r_post.url or "").lower()
        text_post_low = r_post.text.lower()

        # Sucesso da etapa 1 = NÃO voltou pra PessoaAltera com mensagem de erro
        # de validação visível. Popup "alterado com sucesso" também é sucesso.
        ainda_em_cadastro = "pessoaaltera" in url_post
        tem_sucesso_msg = ("alterado com sucesso" in text_post_low
                           or "dados alterados" in text_post_low)

        if ainda_em_cadastro and not tem_sucesso_msg:
            # Procura mensagem de validação útil
            msg_soup = BeautifulSoup(r_post.text, "html.parser")
            err_txt = ""
            for sel in [
                {"class_": re.compile(r"erro|error", re.I)},
                {"class_": re.compile(r"valid|alert", re.I)},
                {"id": re.compile(r"erro|error|valid", re.I)},
            ]:
                el = msg_soup.find(**sel)
                if el:
                    txt = el.get_text(" ", strip=True)
                    if txt:
                        err_txt = txt[:200]
                        break
            # Salva HTML pra debug
            try:
                import tempfile
                fd, dump = tempfile.mkstemp(prefix="bap_primeiro_acesso_",
                                            suffix=".html", text=True)
                with os.fdopen(fd, "w", encoding="utf-8", errors="ignore") as f:
                    f.write(r_post.text)
            except Exception:
                dump = "(falhou salvar dump)"
            raise ValueError(
                f"Primeiro Acesso BAP — cadastro: confirmação falhou "
                f"({err_txt or 'sem mensagem de erro identificável'}). "
                f"HTML salvo em: {dump}. URL final: {r_post.url}"
            )

        # ── ETAPA 2 e 3: aceitar termos LGPD ───────────────────────────────────
        self._aceitar_termos_lgpd(r_post)

        # Atualiza HTML de login pra refletir nova sessão liberada
        self._login_html = r_post.text

    def _aceitar_termos_lgpd(self, r_apos_cadastro):
        """
        Após o cadastro, BAP redireciona pra página de termos LGPD/privacidade
        com botões CONCORDO / NÃO CONCORDO. Sem clicar CONCORDO o portal faz
        logout automático.

        Procura o form de termos na resposta atual; se não achar, tenta URLs
        comuns. Submete o botão que contenha 'concord' no nome/value.
        """
        # Possíveis fontes da página de termos:
        candidatos = []
        url_atual = (r_apos_cadastro.url or "").lower()
        text_atual_low = r_apos_cadastro.text.lower()

        # 1) Talvez já estamos na página de termos
        if any(k in text_atual_low for k in ("concordo", "termos de uso",
                                             "política de privacidade",
                                             "politica de privacidade",
                                             "lgpd")):
            candidatos.append(r_apos_cadastro)

        # 2) URLs prováveis (fallback)
        urls_provaveis = [
            "/Portal/TermoLGPD.aspx",
            "/Portal/Termos.aspx",
            "/Portal/TermosUso.aspx",
            "/Portal/PoliticaPrivacidade.aspx",
            "/Portal/PessoaTermos.aspx",
            "/WebForms/Condomino/TermoLGPD.aspx",
            "/WebForms/Condomino/Termos.aspx",
        ]
        for path in urls_provaveis:
            if not candidatos:  # só tenta se ainda não achou
                try:
                    r_try = self.session.get(self.base_url + path, timeout=15)
                except requests.RequestException:
                    continue
                low = r_try.text.lower()
                if (r_try.status_code == 200
                        and any(k in low for k in ("concordo", "termos",
                                                    "privacidade", "lgpd"))):
                    candidatos.append(r_try)
                    break

        if not candidatos:
            # Sem página de termos detectada — assume que não é necessário aceitar
            return

        r_termos = candidatos[0]

        # Procura o form de termos com botão CONCORDO
        def _submit_pred(name: str) -> bool:
            n = (name or "").lower()
            return ("concord" in n) and ("nao" not in n) and ("não" not in n)

        # Procura também por value=CONCORDO
        def _submit_pred_value(value: str) -> bool:
            v = (value or "").lower()
            return ("concord" in v) and ("nao" not in v) and ("não" not in v)

        r_concord = self._submeter_form_inalterado(
            r_termos,
            etapa="termos LGPD (CONCORDO)",
            submit_pred=_submit_pred,
            submit_pred_value=_submit_pred_value,
        )

        # Se ainda há "concordo" como botão visível, tentar de novo
        # falhou silenciosamente — caller decide
        url_final = (r_concord.url or "").lower()
        if "termo" in url_final and "concordo" in r_concord.text.lower():
            raise ValueError(
                "Primeiro Acesso BAP — termos LGPD: botão CONCORDO foi submetido "
                "mas página de termos persiste. Acesse manualmente e aceite os termos."
            )

    def _submeter_form_inalterado(self, r_pagina, etapa: str,
                                  submit_pred=None,
                                  submit_pred_value=None):
        """
        Helper: extrai o primeiro <form> da página, monta payload com TODOS os
        inputs/textareas/selects atuais (sem alterar valores) e submete via POST.

        Args:
            r_pagina: Response da página atual (deve ter o form).
            etapa: descrição da etapa pro erro.
            submit_pred: callback(name) → bool. Adiciona apenas o submit cujo
                NAME satisfaça o predicado.
            submit_pred_value: callback(value) → bool (opcional). Se nenhum
                botão match por name, tenta por value.

        Returns:
            Response do POST.
        """
        soup = BeautifulSoup(r_pagina.text, "html.parser")
        form = soup.find("form")
        if not form:
            raise ValueError(
                f"Primeiro Acesso BAP — {etapa}: <form> não encontrado em "
                f"{r_pagina.url}. Acesse manualmente."
            )

        payload = {}
        for inp in form.find_all("input"):
            name = inp.get("name")
            if not name:
                continue
            tipo = (inp.get("type") or "text").lower()
            if tipo in ("submit", "button", "image"):
                continue
            if tipo in ("checkbox", "radio"):
                if inp.has_attr("checked"):
                    payload[name] = inp.get("value", "on")
            else:
                payload[name] = inp.get("value", "")

        for ta in form.find_all("textarea"):
            name = ta.get("name")
            if name:
                payload[name] = ta.get_text() or ""

        for sel in form.find_all("select"):
            name = sel.get("name")
            if not name:
                continue
            opt_sel = sel.find("option", selected=True)
            if opt_sel is None:
                opt_sel = sel.find("option")
            payload[name] = opt_sel.get("value", "") if opt_sel else ""

        # Adiciona o botão certo (ASP.NET espera o name do submit)
        submit_adicionado = False
        if submit_pred:
            for btn in form.find_all(["input", "button"]):
                tipo = (btn.get("type") or "").lower()
                if tipo not in ("submit", "image", "button"):
                    continue
                name = btn.get("name") or ""
                if name and submit_pred(name):
                    payload[name] = btn.get("value", "")
                    submit_adicionado = True
                    break

        # Fallback: tenta por value
        if not submit_adicionado and submit_pred_value:
            for btn in form.find_all(["input", "button"]):
                tipo = (btn.get("type") or "").lower()
                if tipo not in ("submit", "image", "button"):
                    continue
                value = btn.get("value") or btn.get_text(strip=True) or ""
                if value and submit_pred_value(value):
                    name = btn.get("name") or ""
                    if name:
                        payload[name] = value
                        submit_adicionado = True
                        break

        action = form.get("action") or r_pagina.url
        if not action.startswith("http"):
            action = urljoin(r_pagina.url, action)

        return self.session.post(action, data=payload, timeout=15,
                                 allow_redirects=True)

    # ── Unidades ───────────────────────────────────────────────────────────────

    def listar_unidades(self) -> list[dict]:
        """
        Navega para 'MEUS DADOS > MINHAS UNIDADES' e retorna lista de unidades.
        Cada item: {condo, bloco, unidade, postback_target, soup_pagina, url_pagina}
        """
        # Tenta URL direta da página de unidades.
        # Instâncias LiveFacilities usam prefixos diferentes:
        #   /WebForms/ → HFLEX, BCF, VORTEX, FFF (instalação padrão)
        #   /Portal/   → BAP (live.bap.com.br) e possivelmente outros
        #
        # OBS: BAP usa /Portal/Unidade.aspx (sem /Condomino/). Os primeiros
        # paths do array são pra outras instâncias; o /Portal/Unidade.aspx
        # cobre o BAP especificamente.
        r = None
        for path in [
            "/WebForms/Condomino/MinhasUnidades.aspx",
            "/WebForms/Condomino/Unidades.aspx",
            "/Portal/Condomino/MinhasUnidades.aspx",
            "/Portal/Condomino/Unidades.aspx",
            "/Portal/Unidade.aspx",       # BAP
            "/WebForms/Unidade.aspx",     # variantes possíveis
        ]:
            r_try = self.session.get(self.base_url + path, timeout=15)
            if r_try.status_code == 200 and "table" in r_try.text.lower() \
                    and "documento" in r_try.text.lower():
                r = r_try
                break
        if r is None:
            # Fallback: tenta parsear HTML pós-login para encontrar link de unidades
            soup = BeautifulSoup(self._login_html, "html.parser")
            link = soup.find("a", string=re.compile("MINHAS UNIDADES", re.I))
            if not link:
                raise ValueError("Página de unidades não encontrada")
            href = urljoin(self.base_url, link["href"])
            r = self.session.get(href, timeout=15)

        soup = BeautifulSoup(r.text, "html.parser")
        url_pagina = r.url  # URL real da página (depois de redirects) — usada
                            # como base para resolver form.action relativo
        tabelas = soup.find_all("table", class_="documento")
        if not tabelas:
            raise ValueError("Nenhuma unidade encontrada na página")

        unidades = []
        for tab in tabelas:
            def _span(pattern):
                el = tab.find("span", id=re.compile(pattern, re.I))
                return el.get_text(strip=True) if el else ""

            condo   = _span(r"lbListaEmpreendimentoNome")
            bloco   = _span(r"lbListaBlocoNome")
            unidade = _span(r"lbListaUnidadeCodigo")
            link    = tab.find("a", title="Alterar unidade")

            if not link:
                continue

            # Extrai __doPostBack target do href
            href = link.get("href", "")
            m = re.search(r"__doPostBack\('([^']+)'", href)
            postback_target = m.group(1) if m else ""

            unidades.append({
                "condo":   condo,
                "bloco":   bloco,
                "unidade": unidade,
                "texto":   f"{condo} {bloco} {unidade}",
                "postback_target": postback_target,
                "soup_pagina": soup,
                "url_pagina": url_pagina,
            })

        return unidades

    def selecionar_unidade(self, unidade_dict: dict, soup_pagina):
        """Executa o __doPostBack para selecionar a unidade.

        IMPORTANTE: o form action no BAP é './Unidade.aspx' (relativo).
        Resolver contra self.base_url (raiz do site) gera 'https://.../Unidade.aspx'
        (sem /Portal/), que o servidor responde 404 e perde o link de boletos.
        Resolução correta: contra a URL real da página (url_pagina), que já
        contém /Portal/.
        """
        target = unidade_dict["postback_target"]
        payload = _scrape_hidden(soup_pagina)
        payload["__EVENTTARGET"]   = target
        payload["__EVENTARGUMENT"] = ""

        # Determina action do form — resolver contra a URL real da página
        form = soup_pagina.find("form")
        action = form.get("action") if form else ""
        base_para_resolver = unidade_dict.get("url_pagina") or self.base_url
        if not action or not action.startswith("http"):
            action = urljoin(base_para_resolver, action or "")

        # Headers que o servidor BAP exige (Referer, Origin) — sem isso, parte
        # das instâncias rejeita o postback silenciosamente
        headers = {
            "Referer": base_para_resolver,
            "Origin": "{0.scheme}://{0.netloc}".format(urlparse(base_para_resolver)),
        }

        r = self.session.post(action, data=payload, timeout=15,
                              headers=headers, allow_redirects=True)
        return BeautifulSoup(r.text, "html.parser")

    # ── Boletos ────────────────────────────────────────────────────────────────

    def obter_url_boletos(self, soup_pos_selecao) -> "str | None":
        """Encontra o link 'boletoLista' após selecionar unidade."""
        link = soup_pos_selecao.find("a", attrs={"data-url": re.compile("boletoLista", re.I)})
        if link:
            return link["data-url"]
        link = soup_pos_selecao.find("a", href=re.compile("boletoLista", re.I))
        if link:
            return urljoin(self.base_url, link["href"])
        return None

    def listar_boletos(self, url_boletos: str) -> list[dict]:
        """
        Acessa a lista de boletos e extrai a URL real do PDF de cada um.

        Três formatos suportados (variam por instância LiveFacilities e
        possivelmente por perfil da conta no BAP):

          1) **href direto pra pCli_BoletoNovo** — usado por HFLEX/BCF/etc.
             Pega `href` e resolve com urljoin.

          2) **cordova_iab JS wrapper** — usado pelo BAP em algumas contas:
             `href="javascript:webkit.messageHandlers.cordova_iab.postMessage(
              ...'webauthlink||URL||' + $('#hfIdChaveAlternativa').val())"`
             Extrai URL entre `webauthlink||` e `||` via regex.

          3) **__doPostBack('...lkSeleciona','')** — usado pelo BAP em outras
             contas: o postback retorna HTML com iframe `src=/Operacional/
             PopUp/pCli_BoletoNovo.aspx?i=...&portal=...`. Dispara o postback
             e extrai a URL do iframe.
        """
        r = self.session.get(url_boletos, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")

        # Verifica erro
        erro_div = soup.find(id="ContentBody_ucBoletoLista_pListaErro")
        if erro_div and erro_div.get_text(strip=True):
            raise ValueError(erro_div.get_text(strip=True))

        # Coleta TODOS os <a title="Abrir Boleto"> (formatos diferentes do href)
        botoes = soup.find_all("a", title="Abrir Boleto")
        if not botoes:
            # Fallback: qualquer link com pCli_BoletoNovo
            botoes = soup.find_all("a", href=re.compile(r"pCli_BoletoNovo", re.I))

        boletos = []
        for btn in botoes:
            href = btn.get("href", "") or ""

            # Formato 2: cordova_iab JS — URL entre webauthlink|| e ||
            m_cordova = re.search(r"webauthlink\s*\|\|\s*([^|]+?)\s*\|\|", href)
            if m_cordova:
                boletos.append({"pdf_url": m_cordova.group(1).strip()})
                continue

            # Formato 3: __doPostBack — dispara postback e extrai URL do iframe
            m_postback = re.search(r"__doPostBack\(\s*['\"]([^'\"]+)['\"]", href)
            if m_postback:
                target = m_postback.group(1)
                url_pdf = self._postback_e_extrair_pdf(url_boletos, soup, target)
                if url_pdf:
                    boletos.append({"pdf_url": url_pdf})
                continue

            # Formato 1: href direto pra pCli_BoletoNovo
            if "pCli_BoletoNovo" in href:
                boletos.append({"pdf_url": urljoin(url_boletos, href)})
                continue

        return boletos

    def _postback_e_extrair_pdf(self, url_boletos: str,
                                soup_lista, target: str) -> "str | None":
        """
        Dispara __doPostBack do botão 'Abrir Boleto' e extrai a URL do PDF
        da resposta. BAP responde com HTML contendo iframe ou JS com a URL
        real do PDF (path: /Operacional/PopUp/pCli_BoletoNovo.aspx).
        """
        payload = _scrape_hidden(soup_lista)
        payload["__EVENTTARGET"]   = target
        payload["__EVENTARGUMENT"] = ""

        form = soup_lista.find("form")
        action = form.get("action") if form else ""
        if not action or not action.startswith("http"):
            action = urljoin(url_boletos, action or "")

        headers = {
            "Referer": url_boletos,
            "Origin": "{0.scheme}://{0.netloc}".format(urlparse(url_boletos)),
        }
        try:
            r = self.session.post(action, data=payload, timeout=15,
                                  headers=headers, allow_redirects=True)
        except requests.RequestException:
            return None

        # Procura URL do PDF na resposta — várias possibilidades:
        # 1) iframe com src apontando pra pCli_BoletoNovo
        soup = BeautifulSoup(r.text, "html.parser")
        iframe = soup.find("iframe", src=re.compile(r"pCli_BoletoNovo", re.I))
        if iframe:
            return urljoin(url_boletos, iframe.get("src"))

        # 2) JS cordova_iab no corpo da resposta
        m_cordova = re.search(r"webauthlink\s*\|\|\s*([^|]+?)\s*\|\|", r.text)
        if m_cordova:
            return m_cordova.group(1).strip()

        # 3) URL bruta no JS
        m_raw = re.search(
            r"https?://[^\"'\s]*pCli_BoletoNovo\.aspx\?[^\"'\s]+",
            r.text,
        )
        if m_raw:
            return m_raw.group(0)

        # 4) Path relativo bruto (sem domínio)
        m_path = re.search(
            r"/[^\"'\s]*pCli_BoletoNovo\.aspx\?[^\"'\s]+",
            r.text,
        )
        if m_path:
            return urljoin(url_boletos, m_path.group(0))

        return None

    def baixar_pdf(self, pdf_url: str) -> bytes:
        r = self.session.get(pdf_url, timeout=30, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Boleto indisponível — Content-Type: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Protel
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoProtel:
    """
    Automação via requests para condominio.protel.com.br (jQuery, sem CAPTCHA).

    Login:  POST /condominio/login  {usuario[login], usuario[senha]}
    Boleto: GET  /condominio/segunda_via  → links  a > img[src*='icone_pdf.gif']
    """

    BASE = "https://condominio.protel.com.br"

    def __init__(self):
        self.session = _nova_sessao()

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()
        # Visita login page para pegar cookies iniciais
        self.session.get(f"{self.BASE}/login", timeout=15)
        r = self.session.post(
            f"{self.BASE}/condominio/login",
            data={"usuario[login]": usuario, "usuario[senha]": senha},
            timeout=15, allow_redirects=True,
        )
        # Se ainda tiver o form de login, credenciais inválidas
        if 'name="usuario[senha]"' in r.text or 'id="usuario_senha"' in r.text:
            raise ValueError(f"Login Protel falhou para {usuario}")
        self._pos_login_html = r.text
        self._pos_login_url  = r.url

    def listar_boletos(self) -> list[dict]:
        """
        Navega para /condominio/segunda_via e retorna links de PDF.
        Padrão: <a href="..."><img src="...icone_pdf.gif"></a>
        """
        r = self.session.get(f"{self.BASE}/condominio/segunda_via",
                             timeout=15, allow_redirects=True)
        soup = BeautifulSoup(r.text, "html.parser")

        # Verifica se não há boletos
        conteudo = soup.find(id="conteudo")
        if conteudo:
            texto = conteudo.get_text(strip=True).lower()
            sem_boleto = any(t in texto for t in [
                "nenhum boleto", "não há boleto", "sem boleto", "não encontrado"
            ])
            if sem_boleto:
                return []

        # Links que contêm imagem icone_pdf.gif
        links = soup.find_all("a", href=True)
        boletos = []
        for a in links:
            img = a.find("img", src=re.compile(r"icone_pdf\.gif", re.I))
            if img:
                href = a["href"]
                if not href.startswith("http"):
                    href = urljoin(self.BASE, href)
                boletos.append({"pdf_url": href})

        # Fallback: qualquer link .pdf direto
        if not boletos:
            boletos = [
                {"pdf_url": urljoin(self.BASE, a["href"])}
                for a in soup.find_all("a", href=re.compile(r"\.pdf", re.I))
            ]

        return boletos

    def baixar_pdf(self, pdf_url: str) -> bytes:
        r = self.session.get(pdf_url, timeout=30, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Protel: Content-Type inesperado: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Administradora Nacional  (ASP puro)
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoNacional:
    """
    Automação via requests para sistemas.admnacional.com.br (ASP clássico).

    Login:  POST /condnet/login.asp  {_CONSULTA=S, login, senha, submit1}
    Boleto: navega por links que contenham '2ª Via', 'Boleto', 'Segunda Via'
            e busca hrefs que retornem PDF.
    """

    BASE = "https://sistemas.admnacional.com.br"
    LOGIN_URL = "https://sistemas.admnacional.com.br/condnet/login.asp"

    def __init__(self):
        self.session = _nova_sessao()
        self._pos_login_soup: "BeautifulSoup | None" = None
        self._pos_login_url: str = ""

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()
        self.session.get(self.LOGIN_URL, timeout=15)
        r = self.session.post(
            self.LOGIN_URL,
            data={"_CONSULTA": "S", "login": usuario,
                  "senha": senha, "submit1": "Entrar"},
            timeout=15, allow_redirects=True,
        )
        # Se voltar para o form de login, credenciais inválidas
        if 'name="login"' in r.text and 'name="senha"' in r.text:
            raise ValueError(f"Login Nacional falhou para {usuario}")
        self._pos_login_soup = BeautifulSoup(r.text, "html.parser")
        self._pos_login_url  = r.url

    def _navegar_boleto(self) -> "BeautifulSoup | None":
        """Tenta localizar a página de 2ª via de boleto."""
        soup = self._pos_login_soup
        if soup is None:
            return None

        # Busca link de 2ª via / boleto no menu
        termos = re.compile(r"2.*via|boleto|segunda.via", re.I)
        link = soup.find("a", string=termos)
        if not link:
            link = soup.find("a", href=termos)
        if not link:
            return soup  # Retorna página atual como fallback

        href = link.get("href", "")
        if not href.startswith("http"):
            href = urljoin(self._pos_login_url, href)

        r = self.session.get(href, timeout=15, allow_redirects=True)
        return BeautifulSoup(r.text, "html.parser")

    def listar_boletos(self) -> list[dict]:
        soup = self._navegar_boleto()
        if not soup:
            return []

        boletos = []
        # Tenta links diretos para PDF
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if re.search(r"\.pdf|boleto|segunda.via|2via", href, re.I):
                if not href.startswith("http"):
                    href = urljoin(self.BASE, href)
                boletos.append({"pdf_url": href, "texto": a.get_text(strip=True)})

        return boletos

    def baixar_pdf(self, pdf_url: str) -> bytes:
        r = self.session.get(pdf_url, timeout=30, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Nacional: Content-Type inesperado: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão CIPA  (cipafacil.digital)
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoCipa:
    """
    Automação via requests para cipafacil.digital (API REST JSON pura).

    Fluxo:
      1. POST back.cipafacil.digital/sign-in  {username, password}
         → access_token (JWT HS256, sem exp — duração longa)
      2. GET  back.cipafacil.digital/users/me
         → document (CNPJ da gestora), units[] (id, condominium_id, code, name)
      3. GET  back.cipafacil.digital/condominiums
         → items[] (id, code, name, units[])
      4. POST api.cipa.com.br/cipa_app_doc/boleto/historico
         {codigoCondominio, unidades[], status:["Emitido","Baixado"],
          limite:0, perfil:"CONDOMINO", documento, anoMesInicial:null, anoMesFinal:null}
         → historicos[] — filtra status=="Emitido" para boletos em aberto
      5. POST api.cipa.com.br/cipa_app_doc/boleto/informacoes/arquivo
         {numeroRecibo}  → PDF binário
    """

    BACK  = "https://back.cipafacil.digital"
    API   = "https://api.cipa.com.br"

    def __init__(self):
        self.session = _nova_sessao()
        self._token: str = ""
        self._documento: str = ""          # CNPJ da gestora
        self._units: list = []             # [{id, condominium_id, code, name}]
        self._condo_code:  dict = {}       # {condominium_id → code}
        self._condo_names: dict = {}       # {condominium_id → name}

    # ── Autenticação ────────────────────────────────────────────────────────────

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()
        r = self.session.post(
            f"{self.BACK}/sign-in",
            json={"username": usuario, "password": senha},
            timeout=20,
        )
        if r.status_code not in (200, 201):
            raise ValueError(f"Login CIPA falhou ({r.status_code}): {r.text[:200]}")
        data = r.json()
        self._token = data.get("access_token", "")
        if not self._token:
            raise ValueError("Login CIPA: access_token não retornado")
        self.session.headers["Authorization"] = f"Bearer {self._token}"

    # ── Perfil e condomínios ────────────────────────────────────────────────────

    def carregar_perfil(self):
        """Carrega documento (CNPJ) e lista de unidades do usuário."""
        r = self.session.get(f"{self.BACK}/users/me", timeout=15)
        r.raise_for_status()
        data = r.json()
        self._documento = data.get("document", "")
        self._units = data.get("units", [])

    def carregar_condominios(self):
        """Carrega o mapeamento condominium_id → code e nome."""
        r = self.session.get(f"{self.BACK}/condominiums", timeout=15)
        r.raise_for_status()
        items = r.json().get("items", [])
        self._condo_code  = {item["id"]: item["code"]                          for item in items}
        self._condo_names = {item["id"]: item.get("name", str(item["code"]))   for item in items}

    # ── Boletos ─────────────────────────────────────────────────────────────────

    def listar_boletos_em_aberto(self, codigo_condo: int,
                                  unidades: list) -> list:
        """
        Retorna boletos com status 'Emitido' (= Em Aberto) para as unidades dadas.
        """
        payload = {
            "codigoCondominio": codigo_condo,
            "unidades": unidades,
            "status": ["Emitido", "Baixado"],
            "limite": 0,
            "perfil": "CONDOMINO",
            "documento": self._documento,
            "anoMesInicial": None,
            "anoMesFinal": None,
        }
        r = self.session.post(
            f"{self.API}/cipa_app_doc/boleto/historico",
            json=payload, timeout=20,
        )
        r.raise_for_status()
        historicos = r.json().get("historicos", [])
        return [h for h in historicos if h.get("status") == "Emitido"]

    def baixar_pdf(self, numero_recibo: int) -> bytes:
        r = self.session.post(
            f"{self.API}/cipa_app_doc/boleto/informacoes/arquivo",
            json={"numeroRecibo": numero_recibo},
            timeout=30,
        )
        ct = r.headers.get("Content-Type", "")

        # PDF binário direto
        if "application/pdf" in ct:
            return r.content

        # JSON com campo "boleto" em base64 (formato real da API CIPA)
        if "application/json" in ct or "text/" in ct:
            try:
                import base64 as _b64
                dados = r.json()
                b64 = dados.get("boleto") or dados.get("pdf") or dados.get("arquivo")
                if b64:
                    return _b64.b64decode(b64)
            except Exception:
                pass

        raise ValueError(f"CIPA: Content-Type inesperado: {ct} | {r.text[:120]}")


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Webware  (CONAC, LOWNDES, ACIR)
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoWebware:
    """
    Automação via requests para portais Webware (ASP legado).

    Login:
      GET  {site}?adm=XXXXX   → scrape campos hidden
      POST https://www.webware.com.br/bin/login.asp
           mem=USUARIO, pass=SENHA + campos hidden

    Após login → página com cards de condomínio e iframe .prestacao-interativa
    O iframe carrega Vue.js; a URL do PDF está em:
      //a[contains(@href,'/relatorioboleto/BuscarBoletoPorRecibo')]
    """

    LOGIN_URL = "https://www.webware.com.br/bin/login.asp"

    def __init__(self, site_url: str):
        self.site_url = site_url.rstrip("/")
        # Extrai adm da query string (ex: adm=19834090)
        m = re.search(r"adm=(\d+)", self.site_url)
        self.adm_id = m.group(1) if m else ""
        self.session = _nova_sessao()
        self._pos_login_html: str = ""

    # ── Login ──────────────────────────────────────────────────────────────────

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()

        # 1. GET página de login → scrape campos hidden
        r = self.session.get(self.site_url, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        form = soup.find("form")
        if not form:
            raise ValueError("Formulário de login Webware não encontrado")

        payload = _scrape_hidden(soup, form)
        payload["mem"]  = usuario
        payload["pass"] = senha

        # 2. POST para /bin/login.asp
        r2 = self.session.post(
            self.LOGIN_URL, data=payload, timeout=15, allow_redirects=True,
        )
        self._pos_login_html = r2.text

        # Verifica falha: formulário de login ainda presente
        if 'name="mem"' in r2.text or 'name="pass"' in r2.text:
            raise ValueError(f"Login Webware falhou para {usuario}")

    # ── Condomínios / Unidades ─────────────────────────────────────────────────

    def listar_cards_condominios(self) -> list[dict]:
        """Retorna lista de {titulo, href} dos cards de condomínio."""
        soup = BeautifulSoup(self._pos_login_html, "html.parser")
        cards = []
        for card in soup.find_all(class_="card"):
            titulo_el = card.find(class_="txt-titulo")
            link_el   = card.find("a", class_="bt")
            if titulo_el and link_el:
                cards.append({
                    "titulo": titulo_el.get_text(strip=True),
                    "href":   urljoin("https://www.webware.com.br", link_el["href"]),
                })
        return cards

    def obter_iframe_boleto(self, card_href: str) -> "str | None":
        """
        Navega para o card de 2ª via de boleto e retorna a src do iframe
        .prestacao-interativa (Vue.js).
        """
        r = self.session.get(card_href, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")

        iframe = soup.find("iframe", class_="prestacao-interativa")
        if not iframe:
            iframe = soup.find("iframe", src=re.compile("prestacao|boleto", re.I))
        if iframe:
            return urljoin("https://www.webware.com.br", iframe.get("src", ""))
        return None

    def listar_boletos_via_iframe(self, iframe_src: str) -> list[dict]:
        """
        Tenta extrair links de boleto da página do iframe.
        O Vue.js renderiza client-side, então esta abordagem pode não funcionar
        se os dados forem carregados apenas por JavaScript.
        Retorna lista de {pdf_url}.
        """
        r = self.session.get(iframe_src, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")

        links = soup.find_all("a", href=re.compile(r"BuscarBoletoPorRecibo", re.I))
        return [{"pdf_url": urljoin("https://www.webware.com.br", a["href"])} for a in links]

    def baixar_pdf(self, pdf_url: str) -> bytes:
        r = self.session.get(pdf_url, timeout=30, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Boleto indisponível — Content-Type: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Immobileweb  (Alterdata — usado por PRED RIO, IRIGON, etc.)
# ═══════════════════════════════════════════════════════════════════════════════

class SessaoImmobileweb:
    """
    Automação via requests para portais Immobileweb (Alterdata Software).

    Fluxo:
      1. GET  /login/<nome_curto>                    → seta cookie NOME_CURTO + ASP.NET_SessionId
      2. POST /Login.aspx/Acessar                    → {usuario, senha, nomecurto} JSON
         response: {"d":{"Success":bool, "Response":"<msg ou URL com autenticacao=UUID>", "Model":null}}
         - Sucesso: Response é URL contendo "autenticacao=<UUID>"
         - Falha:   Response é mensagem de erro (não é URL)
      3. POST /Documentos_condomino/Cobrancas.aspx/GetCobrancas → lista paginada
         body: {"pagina":"1","filtroQuantidade":N}
         response: {"d":{"ListaPaginada":[...], "TotalPaginas":N, "TotalCount":N}}
      4. GET  /VisualizarBoleto.aspx?ID=<Id>&Tipo=V&Ex=.PDF    → bytes do PDF

    O campo `Situacao` do BoletoViewModel diferencia "Baixado" (pago) de
    "Baixa pendente" (em aberto). Só os pendentes são interessantes pra extração.
    """

    # Endpoint comum pós-login (independente de WebForms vs MVC).
    # Funciona pra ambos porque o cookie 'jwt' é setado no parent domain
    # '.immobileweb.com.br' e vale tanto em www.* quanto locacao.*.
    BASE_PORTAL = "https://www.immobileweb.com.br"

    def __init__(self, site_url: str):
        # site_url ex:
        #   - "https://www.immobileweb.com.br/login/744862"           (PRED RIO — WebForms)
        #   - "https://locacao.immobileweb.com.br/irigon"             (IRIGON — MVC)
        from urllib.parse import urlparse
        parsed = urlparse(site_url.rstrip("/"))
        self.BASE_LOGIN = f"{parsed.scheme}://{parsed.netloc}"
        # Detecta fluxo pelo subdomínio
        self.is_mvc = "locacao." in parsed.netloc.lower()
        # nome_curto: /login/<x> (WebForms) ou /<x> (MVC)
        m = re.search(r"/login/([^/?]+)", parsed.path)
        if m:
            self.nome_curto = m.group(1)
            self.path_login = f"/login/{self.nome_curto}"
        else:
            self.nome_curto = parsed.path.strip("/").split("/")[0]
            self.path_login = parsed.path or f"/{self.nome_curto}"
        self.autenticacao = ""
        self.session = _nova_sessao()

    def login(self, usuario: str, senha: str):
        usuario, senha = usuario.strip(), senha.strip()
        if self.is_mvc:
            self._login_mvc(usuario, senha)
        else:
            self._login_webforms(usuario, senha)

    def _login_webforms(self, usuario: str, senha: str):
        """Login na versão WebForms v3.20 (PRED RIO) — AJAX JSON."""
        # 1. GET página de login pra setar cookie NOME_CURTO + ASP.NET_SessionId
        self.session.get(f"{self.BASE_LOGIN}{self.path_login}", timeout=15)

        # 2. POST AJAX JSON
        payload = {"usuario": usuario, "senha": senha, "nomecurto": self.nome_curto}
        r = self.session.post(
            f"{self.BASE_LOGIN}/Login.aspx/Acessar",
            json=payload,
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=15,
        )

        try:
            data = r.json().get("d", {})
        except Exception:
            raise ValueError(f"Resposta inesperada do login Immobileweb (HTTP {r.status_code})")

        response_str = str(data.get("Response", ""))

        # Sucesso: Response é URL contendo "autenticacao=<UUID>"
        m = re.search(r"autenticacao=([0-9a-f-]{36})", response_str)
        if m:
            self.autenticacao = m.group(1)
            return

        raise ValueError(f"Login Immobileweb falhou para {usuario}: {response_str or 'sem mensagem'}")

    def _login_mvc(self, usuario: str, senha: str):
        """Login na versão MVC v3.13 (IRIGON) — form POST tradicional.

        Sucesso: o servidor redireciona pra outra URL (PrincipalCondomino,
        Dashboard, www.immobileweb.com.br, etc.) — URL final diferente da
        página de login. Cookie 'jwt' geralmente também é setado mas nem
        sempre o requests.Session captura (depende de Domain/SameSite).
        Falha: volta pra mesma página de login, com 'validation-msg'
        preenchido (ex: 'Login Inexistente').
        """
        login_url = f"{self.BASE_LOGIN}{self.path_login}"

        # 1. GET da pagina de login (estabelece sessao/cookies iniciais)
        self.session.get(login_url, timeout=15)

        # 2. POST tradicional com Login/Senha (CamelCase) — headers de browser real
        r = self.session.post(
            login_url,
            data={"Login": usuario, "Senha": senha},
            headers={
                "Referer": login_url,
                "Origin": self.BASE_LOGIN,
                "Content-Type": "application/x-www-form-urlencoded",
            },
            timeout=15,
            allow_redirects=True,
        )

        # Sucesso primário: redirecionado pra outra URL (login OK)
        if r.url.rstrip("/") != login_url.rstrip("/"):
            return

        # Caiu de volta na mesma URL — provavel erro. Tenta extrair msg.
        msg_match = re.search(
            r'id="validation-msg"[^>]*value="([^"]*)"', r.text
        )
        msg = msg_match.group(1).strip() if msg_match else ""
        if msg:
            raise ValueError(f"Login Immobileweb (MVC) falhou para {usuario}: {msg}")

        # Sem mensagem explícita — fallback: checa cookie 'jwt' como sinal de sessao
        if any(c.name == "jwt" for c in self.session.cookies):
            return

        raise ValueError(
            f"Login Immobileweb (MVC) falhou para {usuario}: "
            f"URL final {r.url}, sem mensagem de erro nem cookie jwt"
        )

    def listar_boletos_pendentes(self) -> list[dict]:
        """
        Lista boletos com Situacao == 'Baixa pendente'.
        O backend não suporta filtro por situação — filtra no client.
        Endpoint vive sempre em www.immobileweb.com.br (mesmo pra IRIGON, que
        loga em locacao.* — o cookie jwt é parent-domain).
        """
        # filtroQuantidade alto pra pegar tudo em uma chamada
        r = self.session.post(
            f"{self.BASE_PORTAL}/Documentos_condomino/Cobrancas.aspx/GetCobrancas",
            json={"pagina": "1", "filtroQuantidade": 500},
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=20,
        )
        try:
            d = r.json().get("d") or {}
        except Exception:
            raise ValueError(f"Resposta inesperada de GetCobrancas (HTTP {r.status_code})")

        # d pode ser None se a sessão não tem acesso (ex: síndico tentando ver
        # cobranças no portal de condômino). Tratar como "sem boletos".
        if not isinstance(d, dict):
            return []
        lista = d.get("ListaPaginada") or []
        return [b for b in lista if b.get("Situacao") == "Baixa pendente"]

    def baixar_pdf(self, id_boleto: str) -> bytes:
        """
        Baixa PDF. id_boleto é o campo `Id` do BoletoViewModel (não IdBoleto).
        Endpoint vive sempre em www.immobileweb.com.br.
        """
        r = self.session.get(
            f"{self.BASE_PORTAL}/VisualizarBoleto.aspx",
            params={"ID": id_boleto, "Tipo": "V", "Ex": ".PDF"},
            timeout=30,
            allow_redirects=True,
        )
        ct = r.headers.get("Content-Type", "")
        if not ct.startswith("application/pdf"):
            raise ValueError(f"Boleto indisponível — Content-Type: {ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Sessão Condomob  (app.condomob.net)
# ═══════════════════════════════════════════════════════════════════════════════

_CONDOMOB_API      = "https://condomob.appspot.com/app"
_CONDOMOB_FB_KEY   = "AIzaSyAB4ndLltP1nj-5whOLCL2KZoMzWa48Vh4"
_CONDOMOB_APP_INFO = json.dumps({
    "appVersion": "0.25.491", "platform": "WEB",
    "platformVersion": "Chrome 135",
    "extra": {"os": "Windows 10", "isMobile": False,
              "isBrowser": True, "isDesktop": True},
}, separators=(",", ":"))


class SessaoCondomob:
    """Automação via requests para portais Condomob (app.condomob.net).

    Fluxo:
      1. Firebase signInWithPassword  → idToken
      2. /cidade/list + /condominio/list → catálogo de condomínios da administradora
      3. /condominio/unidades          → identificadores de unidades
      4. /acesso/social/login          → tokenCondomob (por unidade)
      5. /financeiro/list              → boletos em aberto
      6. GET link (público)            → PDF
    """

    def __init__(self, administradora_id: str):
        self.administradora_id = administradora_id
        self.session = _nova_sessao()
        self._fb_token: str = ""
        self._token_cond: str = ""
        self._cache_unidades: dict = {}   # condo_id → [identificadores]

    # ── Cabeçalhos ─────────────────────────────────────────────────────────────

    def _h_firebase(self) -> dict:
        return {"Authorization": self._fb_token, "AppInfo": _CONDOMOB_APP_INFO}

    def _h_cond(self) -> dict:
        return {"Authorization": self._token_cond, "AppInfo": _CONDOMOB_APP_INFO}

    # ── Login Firebase ──────────────────────────────────────────────────────────

    def login(self, email: str, senha: str):
        r = self.session.post(
            f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword"
            f"?key={_CONDOMOB_FB_KEY}",
            json={"email": email.strip(), "password": senha.strip(),
                  "returnSecureToken": True},
            timeout=15,
        )
        data = r.json()
        if "idToken" not in data:
            msg = data.get("error", {}).get("message", str(data))
            raise ValueError(f"Firebase login falhou: {msg}")
        self._fb_token = data["idToken"]

    # ── Catálogo de condomínios ─────────────────────────────────────────────────

    def listar_condominios(self) -> list:
        """Retorna [{id, nome, ...}] para todos os condomínios da administradora."""
        r = self.session.get(
            f"{_CONDOMOB_API}/cidade/list",
            params={"administradora": self.administradora_id},
            headers=self._h_firebase(), timeout=15,
        )
        r.raise_for_status()
        try:
            _data = r.json()
            cidades = _data if isinstance(_data, list) else []
        except Exception:
            cidades = []

        condos = []
        for cidade in cidades:
            r2 = self.session.get(
                f"{_CONDOMOB_API}/condominio/list",
                params={"cidade": cidade["id"],
                        "administradora": self.administradora_id},
                headers=self._h_firebase(), timeout=15,
            )
            try:
                _data2 = r2.json() if r2.ok else []
                if isinstance(_data2, list):
                    condos.extend(_data2)
            except Exception:
                pass
        return condos

    # ── Unidades registradas de um condomínio ──────────────────────────────────

    def listar_unidades(self, condo_id) -> list:
        """Retorna lista de dicts {id: Long, identificador: str} para o condomínio.

        Usa /acesso/social/list (requer Firebase token) que retorna objetos completos
        com unidade.id numérico — necessário para /acesso/social/login.
        /condominio/unidades retorna apenas strings e NÃO pode ser usado para login.
        """
        if condo_id in self._cache_unidades:
            return self._cache_unidades[condo_id]
        r = self.session.get(
            f"{_CONDOMOB_API}/acesso/social/list",
            params={"condominio": condo_id, "administradora": self.administradora_id},
            headers=self._h_firebase(),
            timeout=15,
        )
        r.raise_for_status()
        try:
            _data = r.json()
            raw = _data if isinstance(_data, list) else []
        except Exception:
            raw = []
        # Normaliza para [{id, identificador}]
        unidades = [
            {"id": item["unidade"]["id"],
             "identificador": item["unidade"].get("identificador", "")}
            for item in raw if "unidade" in item
        ]
        self._cache_unidades[condo_id] = unidades
        return unidades

    # ── Login de unidade (obtém tokenCondomob) ──────────────────────────────────

    def login_unidade(self, condo_id, unidade_id):
        """unidade_id deve ser o ID numérico Long (de /acesso/social/list), não o identificador string."""
        r = self.session.post(
            f"{_CONDOMOB_API}/acesso/social/login",
            params={"condominio": condo_id, "unidade": unidade_id},
            headers={**self._h_firebase(),
                     "Content-Type": "application/x-www-form-urlencoded"},
            timeout=15,
        )
        if not r.ok:
            try:
                corpo = r.json()
            except Exception:
                corpo = r.text[:300]
            raise ValueError(f"HTTP {r.status_code} em login_unidade "
                             f"(condo={condo_id}, unidade_id={unidade_id}): {corpo}")
        try:
            data = r.json()
        except Exception:
            raise ValueError(f"login_unidade: resposta não-JSON (condo={condo_id}, unidade={unidade_id}): {r.text[:200]}")
        # O token pode vir em data["token"] ou diretamente como string
        if isinstance(data, dict):
            token = data.get("token") or data.get("tokenCondomob", "")
        elif isinstance(data, str):
            token = data
        else:
            token = ""
        if not token:
            raise ValueError(f"tokenCondomob não encontrado: {data}")
        self._token_cond = token

    # ── Boletos e PDF ───────────────────────────────────────────────────────────

    def listar_boletos(self) -> list:
        r = self.session.get(
            f"{_CONDOMOB_API}/financeiro/list",
            headers=self._h_cond(), timeout=15,
        )
        r.raise_for_status()
        try:
            _data = r.json()
            return _data if isinstance(_data, list) else []
        except Exception:
            return []  # corpo vazio = sem boletos

    def baixar_pdf(self, link: str) -> bytes:
        r = self.session.get(link, timeout=30)
        ct = r.headers.get("Content-Type", "")
        ok_status = 200 <= r.status_code < 300
        if not ok_status or "pdf" not in ct.lower():
            raise ValueError(f"PDF não obtido: status={r.status_code} ct={ct}")
        return r.content


# ═══════════════════════════════════════════════════════════════════════════════
#  Processador principal
# ═══════════════════════════════════════════════════════════════════════════════

def _processar_grupo(nome_adm, email, senha_adm, grupo, progress=None, progress_cb=None, idx_offset=0, total_geral=0):
    """
    Processa um único grupo (administradora) com backlog de prints.
    Acumula todas as mensagens em _log e faz flush atômico no final.
    Retorna lista de dicts {identificador, status, mensagem, arquivo}.
    """
    _log: list[str] = []
    _msg = lambda m: _log.append(m)

    site_url = MAPA_SITES.get(nome_adm, "")
    familia  = _familia_from_url(site_url)

    _msg(f"\n[bold]{'━'*70}[/bold]")
    _msg(f"[bold]  {nome_adm}  \\[{familia}]  |  {email}  |  {len(grupo)} registro(s)[/bold]")
    _msg(f"[bold]{'━'*70}[/bold]")

    resultados = []

    # ── Credenciais vazias → não tenta login ──────────────────────────────────
    _cred_vazia = not email and not senha_adm
    _login_vazio = not email
    _senha_vazia = not senha_adm
    if _login_vazio or _senha_vazia:
        if _cred_vazia:
            _aviso = "E-mail e senha não preenchidos no banco"
        elif _login_vazio:
            _aviso = "E-mail (LogAdm) não preenchido no banco"
        else:
            _aviso = "Senha (SenAdm) não preenchida no banco"
        _msg(f"[yellow]  ⚠ SEM CREDENCIAL — {_aviso}. Extração ignorada.[/yellow]")
        for l in grupo:
            resultados.append(_resultado(
                str(l[0]).strip(),
                "sem_credencial",
                _aviso,
                None,
            ))
        if progress:
            progress.advance(len(grupo))
        _flush_log(_log)
        return resultados

    try:
        if familia == "superlogica":
            resultados += _processar_superlogica(
                _base_url(site_url), email, senha_adm, grupo, _msg
            )
        elif familia == "livefacilities":
            resultados += _processar_livefacilities(
                _base_url(site_url), email, senha_adm, grupo, _msg
            )
        elif familia == "webware":
            resultados += _processar_webware(
                site_url, email, senha_adm, grupo, _msg
            )
        elif familia == "immobileweb":
            resultados += _processar_immobileweb(
                site_url, email, senha_adm, grupo, _msg
            )
        elif familia == "protel":
            resultados += _processar_protel(
                email, senha_adm, grupo, _msg
            )
        elif familia == "nacional":
            resultados += _processar_nacional(
                email, senha_adm, grupo, _msg
            )
        elif familia == "cipa":
            resultados += _processar_cipa(
                email, senha_adm, grupo, _msg
            )
        elif familia == "condomob":
            resultados += _processar_condomob(
                site_url, email, senha_adm, grupo, _msg
            )
    except Exception as exc:
        _msg(f"[red]  \\[ERRO INESPERADO] {nome_adm} | {type(exc).__name__}: {exc}[/red]")
        for l in grupo:
            resultados.append(_resultado(str(l[0]).strip(), "erro_inesperado", str(exc), None))

    # ── Garante que todos os registros do grupo tenham resultado ──────────────
    ids_com_resultado = {r["identificador"] for r in resultados}
    for l in grupo:
        cod = str(l[0]).strip()
        if cod not in ids_com_resultado:
            condo_banco = str(l[4]).strip() if len(l) > 4 else ''
            unid_banco  = str(l[5]).strip() if len(l) > 5 else ''
            # Fallback genérico: handler não devolveu resultado pra esse ID.
            # Pode ser: handler retornou cedo (ex: 0 cards no site), exceção
            # engolida, ou bug. Não é necessariamente "match falhou".
            detalhe = (
                f"Condo='{condo_banco}' Unid='{unid_banco}' — "
                f"sem resultado do handler {familia}"
            )
            _msg(f"[yellow]    \\[{cod}] {detalhe}[/yellow]")
            resultados.append(_resultado(cod, "sem_resultado_handler", detalhe, None))

    # ── Contabiliza resultados no backlog ─────────────────────────────────────
    ok  = sum(1 for r in resultados if r["status"] == "ok")
    err = sum(1 for r in resultados if r["status"] not in ("ok", "duplicado", "sem_credencial"))
    dup = sum(1 for r in resultados if r["status"] == "duplicado")
    if ok > 0:
        _msg(f"[green]  ✔ {ok} boleto(s) baixado(s)[/green]")
    if dup > 0:
        _msg(f"[yellow]  ⚠ {dup} duplicado(s)[/yellow]")
    if err > 0:
        _msg(f"[red]  ✘ {err} erro(s)[/red]")

    if progress:
        progress.advance(len(grupo))

    _flush_log(_log)
    return resultados


def processar(
    linhas: list,
    administradoras: "list | None" = None,
    log=None,
    progress: "FixedProgressLogger | None" = None,
    skip_ids: "set | None" = None,
    num_workers: int = 1,
    progress_cb=None,
) -> list:
    """
    Processa registros via API/requests (Superlógica, LiveFacilities, Webware,
    Protel, Nacional, CIPA, Condomob).

    Parâmetros
    ----------
    linhas          : registros do banco [CODIGO, LogAdm, SenAdm, NomeAdm,
                                          NomeCond, Unidade, ...]
    administradoras : lista de nomes para filtrar (None = todas suportadas)
    log             : função de log (padrão: usa _progress_logger ou console)
    progress        : instância de FixedProgressLogger para avançar a barra
    skip_ids        : conjunto de IDs já extraídos a pular
    num_workers     : número de workers paralelos (default 1 = sequencial)
    progress_cb     : callback(idx, total, cod, admin) para atualizar GUI

    Retorna lista de dicts {identificador, status, mensagem, arquivo}.
    """
    if log is None:
        log = _log_fallback

    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    resultados = []
    _results_lock = threading.Lock()

    # ── Filtra por administradora (se informado) ───────────────────────────────
    if administradoras:
        admins_alvo = {a.strip().upper() for a in administradoras}
        linhas = [l for l in linhas if str(l[3]).strip().upper() in admins_alvo]

    # ── Pula IDs já extraídos neste ciclo ──────────────────────────────────────
    if skip_ids:
        pulados = [l for l in linhas if str(l[0]).strip()[:4] in skip_ids]
        linhas  = [l for l in linhas if str(l[0]).strip()[:4] not in skip_ids]
        if pulados:
            log(f"[API] Pulando {len(pulados)} registro(s) já extraídos neste mês.")
        if progress:
            progress.advance(len(pulados))

    # ── Separa suportadas de não-suportadas ────────────────────────────────────
    linhas_ok, linhas_skip = [], []
    for l in linhas:
        nome = str(l[3]).strip()
        site = MAPA_SITES.get(nome, "")
        if _familia_from_url(site) in FAMILIAS_SUPORTADAS:
            linhas_ok.append(l)
        else:
            linhas_skip.append(l)

    if linhas_skip:
        adms = {str(l[3]).strip() for l in linhas_skip}
        log(f"[API] Sem suporte via requests: {', '.join(sorted(adms))}")

    if not linhas_ok:
        log("[API] Nenhum registro para processar.")
        return resultados

    # ── Agrupa por (NomeAdm, LogAdm, SenAdm) ──────────────────────────────────
    grupos: dict = defaultdict(list)
    for linha in linhas_ok:
        chave = (
            str(linha[3]).strip(),
            str(linha[1]).strip(),
            str(linha[2]).strip(),
        )
        grupos[chave].append(linha)

    total_geral = len(linhas_ok)
    items = list(grupos.items())
    _counter = [0]  # linhas processadas (para progress_cb)

    def _worker(item):
        (nome_adm, email, senha_adm), grupo = item
        res = _processar_grupo(
            nome_adm, email, senha_adm, grupo,
            progress=progress,
            progress_cb=progress_cb,
            total_geral=total_geral,
        )
        # Callback GUI por grupo
        if progress_cb:
            with _results_lock:
                _counter[0] += len(grupo)
                cnt = _counter[0]
            progress_cb(cnt, total_geral, "", nome_adm)
        with _results_lock:
            resultados.extend(res)

    if num_workers <= 1:
        for item in items:
            _worker(item)
    else:
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=num_workers) as pool:
            # list() força consumo do iterador — propaga exceções
            list(pool.map(_worker, items))

    return resultados


# ═══════════════════════════════════════════════════════════════════════════════
#  Processadores por família
# ═══════════════════════════════════════════════════════════════════════════════

def _processar_superlogica(base_url, email, senha, grupo, log) -> list:
    resultados = []
    sessao = SessaoSuperlogica(base_url)

    try:
        sessao.login(email, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    condominios = sessao.listar_condominios()
    if not condominios:
        log("  Nenhum condomínio encontrado")
        return [_resultado(l[0], "sem_condominio", "Sem condomínios", None) for l in grupo]

    log(f"  Condominios: {list(condominios.values())}")

    login_unico = len(grupo) == 1 and not grupo[0][6]
    if login_unico:
        id_unico = str(grupo[0][0]).strip()
        log(f"  [{id_unico}] Login único — baixando todas as cobranças sem matching")

    i_global_unico = 0  # contador global de boletos no caminho login_unico (entre todos os condos)
    for id_condo, nome_condo_api in condominios.items():
        sessao.trocar_condominio(id_condo)
        cobrancas = sessao.listar_cobrancas_em_aberto()

        if not cobrancas:
            log(f"  [{nome_condo_api}] Sem cobranças em aberto")
            continue

        log(f"\n  [{nome_condo_api}] {len(cobrancas)} cobrança(s)")

        if login_unico:
            # Login dedicado — atribui todas as cobranças ao identificador único
            for cobranca in cobrancas:
                unidade_api = cobranca["st_unidade_uni"].strip()
                log(f"    [{id_unico}] {nome_condo_api} / {unidade_api or '(sem nome)'}")
                try:
                    pdf_bytes = sessao.baixar_pdf(cobranca["link2viaboleto"])
                except ValueError as e:
                    log(f"    Erro download: {e}")
                    resultados.append(_resultado(id_unico, "erro_download", str(e), None))
                    continue
                sufixo = ""  # Era f"_{i_global_unico}"; agora deixa _salvar_pdf decidir via colisao
                resultados.append(_salvar_pdf(
                    pdf_bytes, id_unico,
                    f"OK | {nome_condo_api} / {unidade_api or '(sem nome)'}",
                    log, sufixo,
                ))
                i_global_unico += 1
            continue

        nomes_banco = [str(l[4]).strip() for l in grupo]
        for i, (l, nome) in enumerate(zip(grupo, nomes_banco)):
            if _campo_vazio(nome):
                log(f"    ⚠ [{str(l[0]).strip()}] NomeCond vazio no banco — registro ignorado no match de condomínio")
        linhas_condo = [
            grupo[i] for i, nome in enumerate(nomes_banco)
            if not _campo_vazio(nome) and _similaridade(nome, nome_condo_api) >= 0.60
        ]
        if not linhas_condo:
            disponiveis = [n for n in nomes_banco if not _campo_vazio(n)]
            _log_sem_match("CONDOMÍNIO", nome_condo_api, disponiveis, log,
                           label_alvo="Site", label_opcoes="Base")
            continue

        for cobranca in cobrancas:
            unidade_api = cobranca["st_unidade_uni"].strip()

            if not unidade_api:
                # Unidade sem nome no site → salva em "Em avaliação" com id do primeiro registro
                identificador = str(linhas_condo[0][0]).strip() if linhas_condo else "0000"
                log(f"    Unidade sem nome no site — salvando em 'Em avaliação' [{identificador}]")
                try:
                    pdf_bytes = sessao.baixar_pdf(cobranca["link2viaboleto"])
                except ValueError as e:
                    log(f"    Erro download: {e}")
                    resultados.append(_resultado(identificador, "erro_download", str(e), None))
                    continue
                resultados.append(_salvar_pdf_em_avaliacao(pdf_bytes, identificador, log))
                continue

            identificador, _opcoes_disp = _match_identificador(unidade_api, linhas_condo, log)
            if not identificador:
                log(f"    Unidade '{unidade_api}' não encontrada no banco")
                continue
            log(f"    [{identificador}] {nome_condo_api} / {unidade_api}")
            try:
                pdf_bytes = sessao.baixar_pdf(cobranca["link2viaboleto"])
            except ValueError as e:
                log(f"    Erro download: {e}")
                resultados.append(_resultado(identificador, "erro_download", str(e), None))
                continue
            resultados.append(_salvar_pdf(
                pdf_bytes, identificador,
                f"OK | {nome_condo_api} / {unidade_api}",
                log,
            ))

    # ── Registros do grupo sem resultado → sem_match com opções do site ───────
    ids_ok = {r["identificador"] for r in resultados}
    condos_site = list(condominios.values()) if condominios else []
    for l in grupo:
        cod = str(l[0]).strip()
        if cod not in ids_ok:
            condo_banco = str(l[4]).strip() if len(l) > 4 else ''
            unid_banco  = str(l[5]).strip() if len(l) > 5 else ''
            resultados.append(_resultado(
                cod, "sem_match",
                f"Condo='{condo_banco}' Unid='{unid_banco}' — não encontrado",
                None, opcoes=_fmt_opcoes(condos_site)))

    return resultados


def _processar_livefacilities(base_url, usuario, senha, grupo, log) -> list:
    resultados = []
    sessao = SessaoLiveFacilities(base_url)

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        unidades = sessao.listar_unidades()
    except ValueError as e:
        log(f"  Erro ao listar unidades: {e}")
        return [_resultado(l[0], "erro_unidades", str(e), None) for l in grupo]

    log(f"  {len(unidades)} unidade(s) encontrada(s)")

    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)  # LOGINMULTIPLO=0 + 1 unica linha no grupo

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            # Login dedicado a esta unidade — usa todas as unidades disponíveis sem fuzzy match
            unidades_alvo = unidades
            log(f"  [{identificador}] Login único — sem matching, usando {len(unidades)} unidade(s)")
        else:
            alvo_condo = str(linha[4]).strip()
            alvo_apto  = str(linha[5]).strip()
            if _campo_vazio(alvo_condo):
                log(f"  [{identificador}] ⚠ NomeCond vazio no banco — match de unidade ficará impreciso")
            if _campo_vazio(alvo_apto):
                log(f"  [{identificador}] ⚠ Unidade vazia no banco — match de unidade ficará impreciso")
            alvo_texto = f"{alvo_condo} {alvo_apto}"

            # Match estruturado em 2 campos (condo + apto separadamente)
            # — evita o ruído de "código + bloco + apto" da string do site
            # competindo com o nome curto do banco. Ver _melhor_match_estruturado.
            idx, score = _melhor_match_estruturado(
                alvo_condo, alvo_apto, unidades, cutoff=0.60,
            )
            if idx is None:
                textos = [u["texto"] for u in unidades]
                _log_sem_match("APARTAMENTO", alvo_texto, textos, log)
                resultados.append(_resultado(identificador, "sem_unidade",
                                             f"Unidade não encontrada: {alvo_texto}", None,
                                             opcoes=_fmt_opcoes(textos)))
                continue

            _log_fuzzy_match("APARTAMENTO", alvo_texto, unidades[idx]["texto"], score, log)
            unidades_alvo = [unidades[idx]]

        _termos_sem_boleto = [
            "sem boleto", "nenhum boleto", "não há boleto",
            "não existem boleto", "não possui boleto", "no momento não",
        ]
        baixou_algum = False
        i_global     = 0  # contador global entre todas as unidades do cliente
        for unidade_dict in unidades_alvo:
            try:
                soup_pos = sessao.selecionar_unidade(unidade_dict,
                                                      unidade_dict["soup_pagina"])
                url_boletos = sessao.obter_url_boletos(soup_pos)
                if not url_boletos:
                    if not login_unico:
                        texto_pag = soup_pos.get_text(separator=" ", strip=True).lower()
                        if any(t in texto_pag for t in _termos_sem_boleto):
                            log(f"  [{identificador}] Sem boletos em aberto")
                            resultados.append(_resultado(identificador, "sem_boleto",
                                                         "Sem boletos em aberto", None))
                        else:
                            log(f"  [{identificador}] Sem boletos em aberto (link de boletos ausente na página)")
                            resultados.append(_resultado(identificador, "sem_boleto",
                                                         "Sem boletos (link ausente)", None))
                    continue

                boletos = sessao.listar_boletos(url_boletos)
                if not boletos:
                    if not login_unico:
                        log(f"  [{identificador}] Sem boletos em aberto")
                        resultados.append(_resultado(identificador, "sem_boleto",
                                                     "Sem boletos em aberto", None))
                    continue

                for boleto in boletos:
                    try:
                        pdf_bytes = sessao.baixar_pdf(boleto["pdf_url"])
                    except ValueError as e:
                        log(f"  [{identificador}] Erro download: {e}")
                        resultados.append(_resultado(identificador, "erro_download", str(e), None))
                        continue
                    sufixo = ""  # Era f"_{i_global}"; agora deixa _salvar_pdf decidir via colisao
                    resultados.append(_salvar_pdf(
                        pdf_bytes, identificador,
                        f"OK | {unidade_dict['texto']}",
                        log, sufixo,
                    ))
                    baixou_algum = True
                    i_global += 1

            except ValueError as e:
                log(f"  [{identificador}] Erro: {e}")
                resultados.append(_resultado(identificador, "erro", str(e), None))

        if login_unico and not baixou_algum:
            log(
                f"  [{identificador}] Sem boletos em aberto "
                f"({len(unidades_alvo)} unidade(s) checada(s), nenhuma com cobranças)"
            )
            resultados.append(_resultado(identificador, "sem_boleto", "Sem boletos em aberto", None))

    return resultados


def _processar_webware(site_url, usuario, senha, grupo, log) -> list:
    resultados = []
    sessao = SessaoWebware(site_url)

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    cards = sessao.listar_cards_condominios()
    log(f"  {len(cards)} card(s) de condomínio encontrado(s)")

    # ── Diagnóstico/Erro: site retornou 0 cards ─────────────────────────────────
    # Distingue dos casos "encontrou cards mas não bateu match". As 3 hipóteses
    # mais comuns pra 0 cards no webware:
    #   (a) Cards renderizados via Vue.js — requests não enxerga (precisa Playwright)
    #   (b) Login sem condomínios associados (credencial válida mas vazia)
    #   (c) CSS class do portal mudou — seletor desatualizado
    if not cards:
        html = sessao._pos_login_html or ""
        tem_vue   = "vue" in html.lower() or "v-app" in html.lower()
        tem_logout = "logout" in html.lower() or "sair" in html.lower()
        diag = (
            f"html={len(html)}B, vue={tem_vue}, logout={tem_logout}"
        )
        log(f"  ⚠ Site retornou 0 cards de condomínio ({diag})")
        if tem_vue:
            motivo = "Site usa Vue.js — cards renderizados client-side. Use Playwright."
        elif not tem_logout:
            motivo = "Pós-login sem indício de sessão autenticada — possível login falso."
        else:
            motivo = "Login OK mas sem condomínios associados (ou seletor de card desatualizado)."
        return [
            _resultado(str(l[0]).strip(), "sem_cards_site", motivo, None)
            for l in grupo
        ]

    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_condo    = str(linha[4]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)  # LOGINMULTIPLO=0 + 1 unica linha no grupo

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            cards_alvo = cards
            log(f"  [{identificador}] Login único — sem matching, iterando {len(cards)} card(s)")
        else:
            if _campo_vazio(alvo_condo):
                log(f"  [{identificador}] ⚠ NomeCond vazio no banco — match de condomínio ficará impreciso")
            titulos = [c["titulo"] for c in cards]
            idx, score = _melhor_match_score(alvo_condo, titulos, cutoff=0.60)
            if idx is None:
                _log_sem_match("CONDOMÍNIO", alvo_condo, titulos, log)
                resultados.append(_resultado(identificador, "sem_condo",
                                             f"Condomínio não encontrado: {alvo_condo}", None,
                                             opcoes=_fmt_opcoes(titulos)))
                continue
            _log_fuzzy_match("CONDOMÍNIO", alvo_condo, cards[idx]["titulo"], score, log)
            cards_alvo = [cards[idx]]

        i_global = 0  # contador global entre todos os cards do cliente
        for card in cards_alvo:
            try:
                iframe_src = sessao.obter_iframe_boleto(card["href"])
                if not iframe_src:
                    raise ValueError("iframe de boleto não encontrado (Vue.js pode requerer Playwright)")

                boletos = sessao.listar_boletos_via_iframe(iframe_src)
                if not boletos:
                    log(f"  [{identificador}] Sem boletos (Vue.js renderiza client-side — use Playwright)")
                    resultados.append(_resultado(identificador, "sem_boleto_api",
                                                 "Vue.js: use Playwright para este portal", None))
                    continue

                for boleto in boletos:
                    try:
                        pdf_bytes = sessao.baixar_pdf(boleto["pdf_url"])
                    except ValueError as e:
                        log(f"  [{identificador}] Erro download: {e}")
                        resultados.append(_resultado(identificador, "erro_download", str(e), None))
                        continue
                    sufixo = ""  # Era f"_{i_global}"; agora deixa _salvar_pdf decidir via colisao
                    resultados.append(_salvar_pdf(
                        pdf_bytes, identificador,
                        f"OK | {card['titulo']}",
                        log, sufixo,
                    ))
                    i_global += 1

            except ValueError as e:
                log(f"  [{identificador}] Erro: {e}")
                resultados.append(_resultado(identificador, "erro", str(e), None))

    return resultados


def _processar_immobileweb(site_url, usuario, senha, grupo, log) -> list:
    """
    Immobileweb (Alterdata): cada credencial vê apenas os boletos da própria
    unidade do morador. Não há seleção de condomínio/unidade no protocolo —
    o servidor já filtra pela credencial.

    Lista os boletos com Situacao 'Baixa pendente' (em aberto) e baixa o PDF
    de cada um. Se houver mais de um registro no grupo (login_unico), todos
    recebem o mesmo conjunto de boletos.

    Suporta dois fluxos de login:
      - WebForms v3.20 (PRED RIO em www.immobileweb.com.br/login/<num>):
        AJAX JSON POST pra /Login.aspx/Acessar
      - MVC v3.13 (IRIGON em locacao.immobileweb.com.br/<slug>):
        form POST tradicional pra /<slug> com campos Login/Senha

    Após o login, ambos usam os MESMOS endpoints em www.immobileweb.com.br
    para listar boletos e baixar PDFs — o cookie 'jwt' é setado no parent
    domain '.immobileweb.com.br' e funciona nos dois subdomínios.
    """
    resultados = []
    sessao = SessaoImmobileweb(site_url)

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        boletos = sessao.listar_boletos_pendentes()
    except Exception as e:
        log(f"  Erro ao listar boletos: {e}")
        return [_resultado(l[0], "erro_boletos", str(e), None) for l in grupo]

    if not boletos:
        log("  Sem boletos em aberto (todos com Situacao=Baixado)")
        return [_resultado(l[0], "sem_boleto", "Sem boletos em aberto", None) for l in grupo]

    log(f"  {len(boletos)} boleto(s) em aberto encontrado(s)")

    # ── Associa boletos aos registros do grupo ────────────────────────────────
    # No Immobileweb cada credencial é dedicada a uma unidade (login_unico
    # geralmente True). Se houver matching de unidade pra fazer, usa o campo
    # `Unidade` do boleto contra `linha[5]` da base.
    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_apto     = str(linha[5]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)

        if login_unico:
            if login_unico_feito:
                continue  # ja' baixou tudo na primeira iteracao
            login_unico_feito = True
            boletos_alvo = boletos
            log(f"  [{identificador}] Login único — baixando todos os {len(boletos)} boleto(s)")
        elif len(boletos) == 1:
            boletos_alvo = boletos
            log(f"  [{identificador}] Boleto único — associado diretamente")
        else:
            if _campo_vazio(alvo_apto):
                log(f"  [{identificador}] ⚠ Unidade vazia no banco — fallback para todos os boletos")
                boletos_alvo = boletos
            else:
                unidades = [b.get("Unidade", "") for b in boletos]
                idx, score = _melhor_match_score(alvo_apto, unidades, cutoff=0.50, validar_numero=True)
                if idx is not None:
                    _log_fuzzy_match("APARTAMENTO", alvo_apto, unidades[idx], score, log)
                    boletos_alvo = [boletos[idx]]
                else:
                    _log_sem_match("APARTAMENTO", alvo_apto, unidades, log)
                    log(f"  [{identificador}] Sem match — pulando")
                    resultados.append(_resultado(identificador, "sem_unidade",
                                                 f"Unidade não encontrada: {alvo_apto}", None,
                                                 opcoes=_fmt_opcoes(unidades)))
                    continue

        i_global = 0
        for boleto in boletos_alvo:
            id_pdf = boleto.get("Id")
            if not id_pdf:
                log(f"  [{identificador}] Boleto sem Id — pulando")
                continue
            try:
                pdf_bytes = sessao.baixar_pdf(id_pdf)
            except ValueError as e:
                log(f"  [{identificador}] Erro download: {e}")
                resultados.append(_resultado(identificador, "erro_download", str(e), None))
                continue

            unidade = boleto.get("Unidade", "")
            referencia = boleto.get("Referencia", "")
            sufixo = ""  # Era f"_{i_global}"; agora deixa _salvar_pdf decidir via colisao
            resultados.append(_salvar_pdf(
                pdf_bytes, identificador,
                f"OK | {unidade} / {referencia}",
                log, sufixo,
            ))
            i_global += 1

    return resultados


def _processar_protel(usuario, senha, grupo, log) -> list:
    """
    Protel: cada usuário tem seu próprio login e vê apenas seus boletos.
    Não há seleção de condomínio/unidade — o portal filtra pelo login.
    """
    resultados = []
    sessao = SessaoProtel()

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        boletos = sessao.listar_boletos()
    except Exception as e:
        log(f"  Erro ao listar boletos: {e}")
        return [_resultado(l[0], "erro_boletos", str(e), None) for l in grupo]

    if not boletos:
        log("  Sem boletos em aberto")
        return [_resultado(l[0], "sem_boleto", "Sem boletos em aberto", None) for l in grupo]

    log(f"  {len(boletos)} boleto(s) encontrado(s)")

    # Associa cada boleto ao registro do banco via fuzzy match de unidade
    # (se houver mais de um registro no grupo, tenta emparelar pela unidade)
    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_apto     = str(linha[5]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            # Login dedicado — baixa todos os boletos sem matching
            boletos_alvo = boletos
            log(f"  [{identificador}] Login único — baixando todos os {len(boletos)} boleto(s)")
        elif len(boletos) == 1:
            boletos_alvo = boletos
            log(f"  [{identificador}] Boleto único — associado diretamente")
        else:
            if _campo_vazio(alvo_apto):
                log(f"  [{identificador}] ⚠ Unidade vazia no banco — fallback para primeiro boleto")
                boletos_alvo = [boletos[0]]
            else:
                textos = [b.get("texto", b["pdf_url"]) for b in boletos]
                idx, score = _melhor_match_score(alvo_apto, textos, cutoff=0.50, validar_numero=True)
                if idx is not None:
                    _log_fuzzy_match("APARTAMENTO", alvo_apto, textos[idx], score, log)
                    boletos_alvo = [boletos[idx]]
                else:
                    _log_sem_match("APARTAMENTO", alvo_apto, textos, log)
                    log(f"  [{identificador}] Usando primeiro boleto como fallback")
                    boletos_alvo = [boletos[0]]

        for i, boleto_alvo in enumerate(boletos_alvo):
            try:
                pdf_bytes = sessao.baixar_pdf(boleto_alvo["pdf_url"])
            except ValueError as e:
                log(f"  [{identificador}] Erro download: {e}")
                resultados.append(_resultado(identificador, "erro_download", str(e), None))
                continue
            sufixo = ""  # Era f"_{i}"; agora deixa _salvar_pdf decidir via colisao
            resultados.append(_salvar_pdf(
                pdf_bytes, identificador,
                f"OK | Protel / {alvo_apto}",
                log, sufixo,
            ))

    return resultados


def _processar_nacional(usuario, senha, grupo, log) -> list:
    """
    Administradora Nacional: login individual por usuário,
    boletos listados via navegação de menu.
    """
    resultados = []
    sessao = SessaoNacional()

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        boletos = sessao.listar_boletos()
    except Exception as e:
        log(f"  Erro ao listar boletos: {e}")
        return [_resultado(l[0], "erro_boletos", str(e), None) for l in grupo]

    if not boletos:
        log("  Sem boletos encontrados")
        return [_resultado(l[0], "sem_boleto", "Sem boletos encontrados", None) for l in grupo]

    log(f"  {len(boletos)} boleto(s) encontrado(s)")

    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_apto     = str(linha[5]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            boletos_alvo = boletos
            log(f"  [{identificador}] Login único — baixando todos os {len(boletos)} boleto(s)")
        elif len(boletos) == 1:
            boletos_alvo = boletos
            log(f"  [{identificador}] Boleto único — associado diretamente")
        else:
            if _campo_vazio(alvo_apto):
                log(f"  [{identificador}] ⚠ Unidade vazia no banco — fallback para primeiro boleto")
                boletos_alvo = [boletos[0]]
            else:
                textos = [b.get("texto", b["pdf_url"]) for b in boletos]
                idx, score = _melhor_match_score(alvo_apto, textos, cutoff=0.50, validar_numero=True)
                if idx is not None:
                    _log_fuzzy_match("APARTAMENTO", alvo_apto, textos[idx], score, log)
                    boletos_alvo = [boletos[idx]]
                else:
                    _log_sem_match("APARTAMENTO", alvo_apto, textos, log)
                    log(f"  [{identificador}] Usando primeiro boleto como fallback")
                    boletos_alvo = [boletos[0]]

        for i, boleto_alvo in enumerate(boletos_alvo):
            try:
                pdf_bytes = sessao.baixar_pdf(boleto_alvo["pdf_url"])
            except ValueError as e:
                log(f"  [{identificador}] Erro download: {e}")
                resultados.append(_resultado(identificador, "erro_download", str(e), None))
                continue
            sufixo = ""  # Era f"_{i}"; agora deixa _salvar_pdf decidir via colisao
            resultados.append(_salvar_pdf(
                pdf_bytes, identificador,
                f"OK | Nacional / {alvo_apto}",
                log, sufixo,
            ))

    return resultados


def _processar_cipa(usuario, senha, grupo, log) -> list:
    """
    CIPA (cipafacil.digital) — API JSON pura, sem browser.

    Cada credencial (username/password) dá acesso a N unidades em N condomínios.
    O mapeamento entre unidade da API e registro do banco é feito por fuzzy match
    de (NomeCond + Unidade) contra (name do condomínio + code da unidade).
    """
    resultados = []
    sessao = SessaoCipa()

    try:
        sessao.login(usuario, senha)
        log("  Login OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        sessao.carregar_perfil()
        sessao.carregar_condominios()
    except Exception as e:
        log(f"  Erro ao carregar perfil/condominios: {e}")
        return [_resultado(l[0], "erro_perfil", str(e), None) for l in grupo]

    log(f"  Perfil: doc={sessao._documento} | {len(sessao._units)} unidade(s)")

    # Monta mapa: condominium_id → (code, [unit_codes])
    from collections import defaultdict as _dd
    condo_units: dict = _dd(list)
    for u in sessao._units:
        condo_units[u["condominium_id"]].append(u)

    # Pré-computa listas de condos para lookup
    nomes_condo_api_glob = []
    condo_ids_lista_glob = []
    for cid, units in condo_units.items():
        nome_api = sessao._condo_names.get(cid, str(sessao._condo_code.get(cid, cid)))
        nomes_condo_api_glob.append(nome_api)
        condo_ids_lista_glob.append(cid)

    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_condo    = str(linha[4]).strip()
        alvo_apto     = str(linha[5]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            # Login dedicado — itera todos os condos e todas as unidades sem matching
            condos_iter = list(zip(condo_ids_lista_glob, nomes_condo_api_glob))
            log(f"  [{identificador}] Login único — sem matching, iterando {len(condos_iter)} condo(s)")
        else:
            if _campo_vazio(alvo_condo):
                log(f"  [{identificador}] ⚠ NomeCond vazio no banco — match de condomínio ficará impreciso")

            idx_condo, score_condo = _melhor_match_score(alvo_condo, nomes_condo_api_glob, cutoff=0.50)
            if idx_condo is None:
                nomes_por_code = [str(sessao._condo_code.get(cid, cid)) for cid in condo_ids_lista_glob]
                idx_condo, score_condo = _melhor_match_score(alvo_condo, nomes_por_code, cutoff=0.50)

            if idx_condo is None:
                _log_sem_match("CONDOMÍNIO", alvo_condo, nomes_condo_api_glob, log)
                resultados.append(_resultado(identificador, "sem_condo",
                                             f"Condomínio não encontrado: {alvo_condo}", None,
                                             opcoes=_fmt_opcoes(nomes_condo_api_glob)))
                continue

            _log_fuzzy_match("CONDOMÍNIO", alvo_condo, nomes_condo_api_glob[idx_condo], score_condo, log)
            condos_iter = [(condo_ids_lista_glob[idx_condo], nomes_condo_api_glob[idx_condo])]

        i_global = 0  # contador de boletos salvo entre todos os condos do cliente (login_unico)
        for condo_id, condo_nome_site in condos_iter:
            condo_code  = sessao._condo_code.get(condo_id)
            units_condo = condo_units[condo_id]

            if condo_code is None:
                log(f"  [{identificador}] Code do condomínio não encontrado (id={condo_id})")
                resultados.append(_resultado(identificador, "sem_condo",
                                             f"Code não encontrado para condo_id={condo_id}", None))
                continue

            if login_unico:
                units_alvo = [u["code"] for u in units_condo]
                log(f"  [{identificador}] {condo_nome_site} — {len(units_alvo)} unidade(s)")
            else:
                # Fuzzy match da unidade
                # validar_numero=False: unit codes CIPA são concatenados bloco+apto (ex: '1402' = BL1 AP402)
                if not alvo_apto:
                    log(f"  [{identificador}] Unidade vazia no banco — sem match possível")
                    resultados.append(_resultado(identificador, "sem_unidade",
                                                 "Unidade vazia no banco", None))
                    continue
                unit_codes = [u["code"] for u in units_condo]
                idx_unit, score_unit = _melhor_match_score(alvo_apto, unit_codes, cutoff=0.50,
                                                           validar_numero=False)
                if idx_unit is None:
                    _log_sem_match("APARTAMENTO", alvo_apto, unit_codes, log)
                    resultados.append(_resultado(identificador, "sem_unidade",
                                                 f"Unidade não encontrada: {alvo_apto}", None,
                                                 opcoes=_fmt_opcoes(unit_codes)))
                    continue
                _log_fuzzy_match("APARTAMENTO", alvo_apto, unit_codes[idx_unit], score_unit, log)
                units_alvo = [unit_codes[idx_unit]]

            try:
                boletos = sessao.listar_boletos_em_aberto(condo_code, units_alvo)
            except Exception as e:
                log(f"  [{identificador}] Erro ao listar boletos: {e}")
                resultados.append(_resultado(identificador, "erro_boletos", str(e), None))
                continue

            if not boletos:
                log(f"  [{identificador}] Sem boletos em aberto")
                resultados.append(_resultado(identificador, "sem_boleto",
                                             "Sem boletos em aberto", None))
                continue

            log(f"  [{identificador}] {len(boletos)} boleto(s) em aberto")

            for boleto in boletos:
                recibo = boleto["numeroRecibo"]
                venc   = boleto.get("dataVencimento", "")
                unit_code = boleto.get("unidade", units_alvo[0] if len(units_alvo) == 1 else "?")
                try:
                    pdf_bytes = sessao.baixar_pdf(recibo)
                except ValueError as e:
                    log(f"  [{identificador}] Erro download recibo {recibo}: {e}")
                    resultados.append(_resultado(identificador, "erro_download", str(e), None))
                    continue

                sufixo = ""  # Era f"_{i_global}"; agora deixa _salvar_pdf decidir via colisao
                log(f"  [{identificador}] recibo={recibo} venc={venc}")
                resultados.append(_salvar_pdf(
                    pdf_bytes, identificador,
                    f"OK | CIPA / {condo_code} / {unit_code} venc={venc}",
                    log, sufixo,
                ))
                i_global += 1

    return resultados


def _processar_condomob(site_url: str, email: str, senha: str, grupo: list, log) -> list:
    """
    Condomob (app.condomob.net) — Firebase auth + REST API JSON.

    Fluxo por registro:
      1. Firebase login (1× por grupo de credenciais)
      2. Listar condomínios da administradora (cidades → condos)
      3. Fuzzy match NomeCond → condomínio
      4. Listar unidades do condomínio (cacheado por condo_id)
      5. Fuzzy match Unidade → identificador da unidade
      6. Login de unidade → tokenCondomob
      7. Listar boletos em aberto
      8. Download do PDF (link público)
    """
    from urllib.parse import urlparse, parse_qs
    resultados = []

    # Extrai administradora_id da URL (ex: ?administradora=4566022178996224)
    qs = parse_qs(urlparse(site_url).query)
    adm_id = qs.get("administradora", [None])[0]
    if not adm_id:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  [{ids}] URL sem parâmetro 'administradora': {site_url}")
        return [_resultado(l[0], "erro", "URL sem administradora_id", None) for l in grupo]

    sessao = SessaoCondomob(adm_id)

    try:
        sessao.login(email, senha)
        log("  Login Firebase OK")
    except ValueError as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro login: [{ids}] {e}")
        return [_resultado(l[0], "erro_login", str(e), None) for l in grupo]

    try:
        condos = sessao.listar_condominios()
    except Exception as e:
        ids = ", ".join(str(l[0]).strip() for l in grupo)
        log(f"  Erro ao listar condomínios: [{ids}] {e}")
        return [_resultado(l[0], "erro", f"Erro ao listar condomínios: {e}", None) for l in grupo]

    log(f"  {len(condos)} condomínio(s) encontrado(s) na administradora")
    nomes_condo_api = [str(c.get("nome", c.get("id", ""))).strip() for c in condos]

    login_unico_feito = False
    for linha in grupo:
        identificador = str(linha[0]).strip()
        alvo_condo    = str(linha[4]).strip()
        alvo_apto     = str(linha[5]).strip()
        login_unico   = (not linha[6]) and (len(grupo) == 1)  # LOGINMULTIPLO=0 + 1 unica linha no grupo

        if login_unico:
            if login_unico_feito:
                continue  # já baixou tudo na primeira iteração
            login_unico_feito = True
            # Login dedicado — usa todos os condomínios/unidades registrados sem fuzzy match
            condos_alvo = condos
            log(f"  [{identificador}] Login único — sem matching, iterando {len(condos)} condomínio(s)")
        else:
            if _campo_vazio(alvo_condo):
                log(f"  [{identificador}] ⚠ NomeCond vazio no banco — match de condomínio ficará impreciso")

            idx_condo, score_condo = _melhor_match_score(alvo_condo, nomes_condo_api, cutoff=0.50)
            if idx_condo is None:
                _log_sem_match("CONDOMÍNIO", alvo_condo, nomes_condo_api, log)
                resultados.append(_resultado(identificador, "sem_condo",
                                             f"Condomínio não encontrado: {alvo_condo}", None,
                                             opcoes=_fmt_opcoes(nomes_condo_api)))
                continue

            _log_fuzzy_match("CONDOMÍNIO", alvo_condo, nomes_condo_api[idx_condo], score_condo, log)
            condos_alvo = [condos[idx_condo]]

        i_global           = 0     # contador global de boletos entre todos os condos do cliente
        unidades_visitadas = 0     # total de unidades iteradas (para log de sem_boleto)
        baixou_algum       = False

        for condo in condos_alvo:
            condo_id   = condo["id"]
            condo_nome = str(condo.get("nome", condo_id)).strip()

            try:
                unidades = sessao.listar_unidades(condo_id)
            except Exception as e:
                log(f"  [{identificador}] Erro ao listar unidades ({condo_nome}): {e}")
                resultados.append(_resultado(identificador, "erro", f"Erro unidades: {e}", None))
                continue

            if not unidades:
                log(f"  [{identificador}] Nenhuma unidade registrada em {condo_nome}")
                resultados.append(_resultado(identificador, "sem_unidade",
                                             "Nenhuma unidade registrada no Condomob para esta conta", None))
                continue

            if login_unico:
                unidades_alvo = unidades
                log(f"  [{identificador}] {condo_nome} — {len(unidades)} unidade(s), baixando sem matching")
            else:
                if _campo_vazio(alvo_apto):
                    log(f"  [{identificador}] ⚠ Unidade vazia no banco — sem match possível")
                    resultados.append(_resultado(identificador, "sem_unidade",
                                                 "Unidade vazia no banco", None))
                    continue

                # Fuzzy match no identificador (string) — validar_numero=False pois
                # identificadores Condomob têm prefixos/separadores (ex: "4A-0402")
                identificadores = [u["identificador"] for u in unidades]
                idx_unit, score_unit = _melhor_match_score(alvo_apto, identificadores, cutoff=0.50,
                                                           validar_numero=False)
                if idx_unit is None:
                    _log_sem_match("APARTAMENTO", alvo_apto, identificadores, log)
                    resultados.append(_resultado(identificador, "sem_unidade",
                                                 f"Unidade não encontrada: {alvo_apto}", None,
                                                 opcoes=_fmt_opcoes(identificadores)))
                    continue

                _log_fuzzy_match("APARTAMENTO", alvo_apto, unidades[idx_unit]["identificador"],
                                 score_unit, log)
                unidades_alvo = [unidades[idx_unit]]

            for unidade_match in unidades_alvo:
                unidades_visitadas += 1
                unit_id_numerico   = unidade_match["id"]
                unit_identificador = unidade_match["identificador"]

                try:
                    sessao.login_unidade(condo_id, unit_id_numerico)
                except Exception as e:
                    log(f"  [{identificador}] Erro login unidade ({unit_identificador}): {e}")
                    resultados.append(_resultado(identificador, "erro_login",
                                                 f"Login unidade: {e}", None))
                    continue

                try:
                    boletos = sessao.listar_boletos()
                except Exception as e:
                    log(f"  [{identificador}] Erro ao listar boletos: {e}")
                    resultados.append(_resultado(identificador, "erro_boletos", str(e), None))
                    continue

                if not boletos:
                    if not login_unico:
                        log(f"  [{identificador}] Sem boletos em aberto")
                        resultados.append(_resultado(identificador, "sem_boleto",
                                                     "Sem boletos em aberto", None))
                    continue

                log(f"  [{identificador}] {len(boletos)} boleto(s) em aberto")

                for boleto in boletos:
                    link = boleto.get("link", "")
                    venc = boleto.get("dataVencimento", "")
                    if not link:
                        continue
                    try:
                        pdf_bytes = sessao.baixar_pdf(link)
                    except ValueError as e:
                        log(f"  [{identificador}] Erro download: {e}")
                        resultados.append(_resultado(identificador, "erro_download", str(e), None))
                        continue
                    sufixo = ""  # Era f"_{i_global}"; agora deixa _salvar_pdf decidir via colisao
                    resultados.append(_salvar_pdf(
                        pdf_bytes, identificador,
                        f"OK | Condomob / {condo_nome} / {unit_identificador} venc={venc}",
                        log, sufixo,
                    ))
                    baixou_algum = True
                    i_global += 1

        if login_unico and not baixou_algum:
            log(
                f"  [{identificador}] Sem boletos em aberto "
                f"({unidades_visitadas} unidade(s) checada(s), nenhuma com cobranças)"
            )
            resultados.append(_resultado(identificador, "sem_boleto", "Sem boletos em aberto", None))

    return resultados


# ═══════════════════════════════════════════════════════════════════════════════
#  Helpers internos
# ═══════════════════════════════════════════════════════════════════════════════

def _match_identificador(unidade_api: str, linhas_condo: list, log) -> "tuple[str | None, list]":
    """Retorna (identificador, opcoes_disponíveis).
    opcoes_disponíveis é a lista de unidades do banco (para log de não-encontrado).
    """
    # Filtra entradas do banco sem nome de unidade (Unidade = '') para não poluir o match
    linhas_validas = [(i, l) for i, l in enumerate(linhas_condo) if str(l[5]).strip()]
    if not linhas_validas:
        # Só há registros sem unidade no banco — não há como identificar
        _log_sem_match("APARTAMENTO", unidade_api, [], log, label_alvo="Site", label_opcoes="Base")
        return None, []

    apartamentos_validos = [str(l[5]).strip() for _, l in linhas_validas]
    idx, score = _melhor_match_score(unidade_api, apartamentos_validos, cutoff=0.65, validar_numero=True)

    if idx is None:
        # Exibe código do cliente junto ao nome da unidade para facilitar diagnóstico
        todos_aptos = [f"{str(l[0]).strip()}: {str(l[5]).strip()}" for l in linhas_condo]
        _log_sem_match("APARTAMENTO", unidade_api, todos_aptos, log,
                       label_alvo="Site", label_opcoes="Base")
        return None, [str(l[5]).strip() for l in linhas_condo]

    linha_original = linhas_validas[idx][1]
    _log_fuzzy_match("APARTAMENTO", apartamentos_validos[idx], unidade_api, score, log)
    return str(linha_original[0]).strip(), []



def _salvar_pdf_em_avaliacao(pdf_bytes: bytes, identificador: str, log) -> dict:
    """
    Salva PDF em DOWNLOADS_DIR/Condomínios/Em avaliação/{identificador}.pdf
    quando a unidade não tem nome no site (não foi possível fazer o match).
    Usa sufixo _n se o arquivo já existir.
    """
    pasta = DOWNLOADS_DIR / "Condomínios" / "Em avaliação"
    pasta.mkdir(parents=True, exist_ok=True)
    destino = pasta / f"{identificador}.pdf"
    k = 0
    while destino.exists():
        k += 1
        destino = pasta / f"{identificador}_{k}.pdf"
    destino.write_bytes(pdf_bytes)
    log(f"    Salvo em avaliação: {destino.name}")
    return _resultado(identificador, "em_avaliacao",
                      "Unidade sem nome — salvo em 'Em avaliação'", str(destino))


def _salvar_pdf(pdf_bytes: bytes, identificador: str, mensagem_ok: str,
                log, sufixo: str = "") -> dict:
    """
    Salva pdf_bytes em disco, adiciona cabeçalho e verifica duplicata.
    Usado por todos os processadores.

    Parâmetros
    ----------
    pdf_bytes     : conteúdo binário do PDF
    identificador : código do cliente (4+ chars)
    mensagem_ok   : texto da mensagem em caso de sucesso
    log           : função de log
    sufixo        : sufixo extra antes de .pdf (ex: "_1") para múltiplos boletos
    """
    nome_base = f"{identificador}{sufixo}"
    pasta_condominios = DOWNLOADS_DIR / "Condomínios"
    destino = pasta_condominios / f"{nome_base}.pdf"
    k = 0
    while destino.exists():
        k += 1
        destino = pasta_condominios / f"{nome_base}_{k}.pdf"

    pasta_condominios.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(pdf_bytes)
    log(f"    Salvo: {destino.name}")

    # ── 1. Lê barcode ANTES do cabeçalho ──────────────────────────────────────
    # O pypdf reescreve o PDF (decripta + recompacta) de uma forma que faz
    # o pyzbar perder a capacidade de decodificar o I25. Por isso lê o barcode
    # no PDF cru e passa o resultado pra processar_boleto_baixado.
    #
    # Caminho ABSOLUTO obrigatório: barcodereader → calculate_pixels faz
    # os.path.join(path, path) que com path relativo duplica o caminho.
    info_barcode = None
    if _tem_aux:
        try:
            caminho_abs = str(destino.resolve())
            info_barcode = _boletos_mod.barcodereader(caminho_abs, renomear=False)
            # Diagnóstico: mostra o que o barcodereader retornou
            _msg_err = "Impossível ler código de barras"
            _cb = (info_barcode.get('Código de Barras') or '') if info_barcode else ''
            _vc = (info_barcode.get('Data Vencimento') or '') if info_barcode else ''
            if not _cb or _cb == _msg_err:
                log(f"    🔍 Barcode: NÃO LIDO no PDF cru")
            else:
                log(f"    🔍 Barcode lido: venc={_vc} (linha={_cb[:20]}...)")
        except Exception as e:
            log(f"    Barcode (exceção): {e}")

    # ── 2. Cabeçalho ──────────────────────────────────────────────────────────
    if _tem_aux:
        try:
            codigo = identificador[:4] if len(identificador) >= 4 else identificador
            destino_abs = str(destino.resolve())
            aux.adicionarcabecalhopdf_topo_adaptativo(destino_abs, destino_abs, codigo)
        except Exception as e:
            log(f"    Cabeçalho: {e}")

    # ── 3. Move pra subpasta do mês usando info_barcode já lida ──────────────
    caminho_final = destino
    cod_barras = ""
    dados = None
    if _tem_aux:
        try:
            import io as _io
            _null = _io.StringIO()
            _stdout_orig = sys.stdout
            sys.stdout = _null
            try:
                dados = _boletos_mod.processar_boleto_baixado(
                    caminho_pdf=destino.resolve(),
                    codigo_cliente=identificador,
                    pasta_downloads=DOWNLOADS_DIR,
                    pasta_destino_forcada=_PASTA_CICLO_ATUAL,
                    info_barcode=info_barcode,
                )
            finally:
                sys.stdout = _stdout_orig
            if dados:
                caminho_final = Path(dados.get("Caminho_Completo", str(destino)))
                cod_barras    = dados.get("Cod_Barras", "")
                # Log explícito do que aconteceu com o arquivo
                pasta_mes_arq = dados.get("Pasta_Mes", "")
                venc_arq      = dados.get("Vencimento", "")
                if pasta_mes_arq:
                    log(f"    → {pasta_mes_arq}/ (venc {venc_arq})")
                else:
                    if not cod_barras:
                        log(f"    ⚠ Sem barcode legível — mantido na raiz")
                    elif not venc_arq:
                        log(f"    ⚠ Sem vencimento — mantido na raiz")
                    else:
                        log(f"    ⚠ Vencimento suspeito ({venc_arq}) — mantido na raiz")
        except Exception as e:
            log(f"    processar_boleto: {e}")

    # ── Verifica duplicata (só quando as funções de cache estão disponíveis) ───
    if cod_barras and _usa_log_global:
        try:
            if verificar_boleto_duplicado(cod_barras):
                log(f"    [DUPLICADO]")
                registrar_boleto_duplicado(identificador, cod_barras, str(caminho_final))
                _boletos_duplicados.append({
                    "Cliente":    identificador,
                    "Cod_Barras": cod_barras,
                    "Caminho":    str(caminho_final),
                })
                if caminho_final.exists():
                    caminho_final.unlink()
                return _resultado(identificador, "duplicado",
                                  "Código de barras já existia", str(caminho_final))
            else:
                registrar_codigo_barras(cod_barras)
        except Exception as e:
            log(f"    Verificação duplicata: {e}")

    # ── Registra no log de boletos processados ────────────────────────────────
    if dados:
        _boletos_processados.append(dados)

    return _resultado(identificador, "ok", mensagem_ok, str(caminho_final))


def _resultado(identificador, status, mensagem, arquivo, opcoes=None):
    return {"identificador": identificador, "status": status,
            "mensagem": mensagem, "arquivo": arquivo,
            "opcoes": opcoes or ""}


# ═══════════════════════════════════════════════════════════════════════════════
#  Execução standalone
# ═══════════════════════════════════════════════════════════════════════════════

_MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

def _pasta_ciclo(ano_mes: "str | None") -> Path:
    """
    Retorna o Path da subpasta do ciclo em DOWNLOADS_DIR.

    ano_mes : 'YYYY_MM'  (ex: '2026_04') — obrigatório via argumento CLI.
              Se None, usa o mês atual como fallback de segurança.

    Exemplos de resultado:
      '2026_04'  →  downloads/2026_04_Abr
      '2026_12'  →  downloads/2026_12_Dez
    """
    base = DOWNLOADS_DIR / "Condomínios"
    if ano_mes:
        partes = ano_mes.split("_")
        if len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
            ano, mes = int(partes[0]), int(partes[1])
            if 1 <= mes <= 12:
                return base / f"{ano}_{mes:02d}_{_MESES_PT[mes - 1]}"
        raise ValueError(f"Formato inválido para ciclo: '{ano_mes}' — use YYYY_MM (ex: 2026_04)")
    # fallback: mês atual
    from datetime import datetime as _dt
    hoje = _dt.now()
    return base / f"{hoje.year}_{hoje.month:02d}_{_MESES_PT[hoje.month - 1]}"


def _ids_ja_baixados(pasta_mes: Path) -> set:
    """
    Retorna o conjunto de identificadores (4 chars) que já têm PDF na pasta do mês.

    Trata todos os formatos possíveis de nome:
      2938.pdf          → '2938'
      2938_1.pdf        → '2938'
      2938_2.pdf        → '2938'
      2938_n.pdf        → '2938'
      202604_2938.pdf   → '2938'   (prefixo de data do boletos.py)
      202604_2938_1.pdf → '2938'
    """
    if not pasta_mes.exists():
        return set()
    # Padrão: prefixo de data opcional (6 dígitos + _), depois o identificador
    # (4+ alfanum), depois sufixo _N opcional
    _pat = re.compile(r'^(?:\d{6}_)?([A-Za-z0-9]{4}[A-Za-z0-9]*)(?:_\d+)?$')
    ids = set()
    for pdf in pasta_mes.glob("*.pdf"):
        m = _pat.match(pdf.stem)
        if m:
            ids.add(m.group(1)[:4])
    return ids


def _bloquear_suspensao():
    """Impede que o Windows suspenda/desligue o PC durante o processamento."""
    try:
        import ctypes
        ES_CONTINUOUS       = 0x80000000
        ES_SYSTEM_REQUIRED  = 0x00000001
        ES_DISPLAY_REQUIRED = 0x00000002
        ctypes.windll.kernel32.SetThreadExecutionState(
            ES_CONTINUOUS | ES_SYSTEM_REQUIRED
        )
    except Exception:
        pass


def _liberar_suspensao():
    """Restaura o comportamento normal de suspensão do Windows."""
    try:
        import ctypes
        ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)  # ES_CONTINUOUS
    except Exception:
        pass


if __name__ == "__main__":
    from auxiliares import utils as aux

    # ── Parsing de argumentos ──────────────────────────────────────────────────
    # Uso:
    #   py extracao_api.py 2026_04                        → ciclo normal (pula IDs já em 2026_04_Abr)
    #   py extracao_api.py 2026_04 --reset                → apaga TODOS os PDFs de downloads/ e baixa tudo
    #   py extracao_api.py 2026_04 ADIPLANTEC             → ciclo normal + filtro de admin
    #   py extracao_api.py 2026_04 --reset ADIPLANTEC     → reset total + filtro de admin
    #   py extracao_api.py 2026_04 -b "C:\caminho\Scai.WMB"  → banco em outro caminho (default: <raiz>/Scai.WMB)
    _args = sys.argv[1:]

    # Extrai --banco/-b <path> se presente (opcional)
    # Sem ele, usa o padrao: <raiz>/Scai.WMB
    _caminho_banco_arg = None
    for _i in range(len(_args)):
        if _args[_i] in ("--banco", "-b") and _i + 1 < len(_args):
            _caminho_banco_arg = _args[_i + 1]
            _args = _args[:_i] + _args[_i + 2:]
            break

    primeiro_ciclo = "--reset" in _args
    _args = [a for a in _args if a != "--reset"]

    # Detecta YYYY_MM (ex: "2026_04") como primeiro argumento posicional
    _ciclo_pat = re.compile(r'^\d{4}_\d{2}$')
    ciclo_arg = next((a for a in _args if _ciclo_pat.match(a)), None)
    if ciclo_arg is None:
        console.print("[bold red]Erro:[/bold red] informe o ciclo como primeiro argumento (ex: 2026_04)")
        sys.exit(1)
    _args = [a for a in _args if a != ciclo_arg]

    filtro = _args if _args else None

    if filtro:
        admins_alvo = [a.strip() for a in filtro]
    else:
        admins_alvo = sorted(ADMINS_SUPORTADAS)

    adms_formatados = ", ".join(f"'{a}'" for a in admins_alvo)

    SQL = f"""
        SELECT LEFT(TRIM(Tabela.CODIGO),4)                                    AS Codigo,
               TRIM(Tabela.LogAdm)                                             AS LogAdm,
               TRIM(Tabela.SenAdm)                                             AS SenAdm,
               TRIM(Tabela.NomeAdm)                                            AS NomeAdm,
               TRIM(Tabela.NomeCond)                                           AS Condominio,
               TRIM(Tabela.Unidade)                                            AS Unidade,
               Consulta.LoginMultiplo,
               '' , '' , '' , '' , '',
               REPLACE(REPLACE(REPLACE(TRIM(Prop.CGCCPF),'.',''),'-',''),'/','') AS CPF
          FROM ((Inquilin_New AS Tabela
          LEFT JOIN Prop
            ON LEFT(TRIM(Tabela.CODIGO),4) = Prop.CODIGO)
          LEFT JOIN (
               SELECT T.NomeAdm, T.LogAdm, T.SenAdm,
                      Count(T.Cod) > 1 AS LoginMultiplo
                 FROM (SELECT DISTINCT NomeAdm, LogAdm, SenAdm,
                              LEFT(TRIM(Codigo),4) AS Cod
                         FROM Inquilin_New) AS T
                GROUP BY T.NomeAdm, T.LogAdm, T.SenAdm
               HAVING (((T.NomeAdm)<>''))
          ) AS Consulta
            ON  Tabela.NomeAdm = Consulta.NomeAdm
            AND Tabela.LogAdm  = Consulta.LogAdm
            AND Tabela.SenAdm  = Consulta.SenAdm)
         WHERE Tabela.Situa NOT IN ('E','F','K','V')
               AND TRIM(Tabela.NomeAdm) IN ({adms_formatados})
         ORDER BY Tabela.NomeAdm, Tabela.LogAdm, Tabela.NomeCond, Tabela.Unidade DESC
    """

    # Usa caminho do --banco se passado, senao default <raiz>/Scai.WMB
    caminho_banco = _caminho_banco_arg or os.path.join(aux.caminhoprojeto(), 'Scai.WMB')
    if not os.path.isfile(caminho_banco):
        console.print(f"[bold red]Erro:[/bold red] arquivo de banco nao encontrado: {caminho_banco}")
        console.print(f"[dim]Use --banco <caminho> pra especificar outro arquivo[/dim]")
        sys.exit(1)
    console.print(f"[cyan][BD][/] Usando banco: {caminho_banco}")
    banco = aux.Banco(caminho_banco)
    linhas = banco.consultar(SQL)
    total = len(linhas)

    _bloquear_suspensao()

    # ── Lógica de ciclo ────────────────────────────────────────────────────────
    pasta_mes = _pasta_ciclo(ciclo_arg)
    skip_ids: set = set()

    if primeiro_ciclo:
        # Apaga TODOS os PDFs de TODAS as subpastas de DOWNLOADS_DIR
        total_apagados = 0
        erros_apagar = 0
        if DOWNLOADS_DIR.exists():
            for pdf in DOWNLOADS_DIR.rglob("*.pdf"):
                try:
                    pdf.unlink()
                    total_apagados += 1
                except Exception as e:
                    console.print(f"[yellow]Não foi possível apagar {pdf.name}: {e}[/yellow]")
                    erros_apagar += 1
        console.print(
            f"[bold yellow]⚠ --reset:[/bold yellow] {total_apagados} PDF(s) apagados de "
            f"[cyan]{DOWNLOADS_DIR}[/cyan]"
            + (f" ({erros_apagar} erro(s))" if erros_apagar else "")
        )
    else:
        # Ciclo subsequente: pula IDs que já têm PDF na pasta do ciclo informado
        skip_ids = _ids_ja_baixados(pasta_mes)
        if skip_ids:
            console.print(
                f"[cyan]ℹ Ciclo {ciclo_arg}:[/cyan] {len(skip_ids)} identificador(es) já extraídos "
                f"em [cyan]{pasta_mes.name}[/cyan] serão pulados"
            )

    console.print(f"\n[bold cyan]ExtratorAPI[/bold cyan] — [green]{total}[/green] registros carregados")

    _progress_logger = FixedProgressLogger(
        total=total,
        description="[green]Processando condomínios...",
    )
    _progress_logger.start()

    try:
        resultados = processar(linhas, administradoras=filtro, progress=_progress_logger,
                               skip_ids=skip_ids if not primeiro_ciclo else None)
    finally:
        _progress_logger.stop()
        _liberar_suspensao()

    # ── Resumo final ──────────────────────────────────────────────────────────
    ok  = sum(1 for r in resultados if r["status"] == "ok")
    dup = sum(1 for r in resultados if r["status"] == "duplicado")
    err = len(resultados) - ok - dup

    console.print(f"\n[bold]{'='*70}[/bold]")
    console.print(
        f"[bold green]RESUMO[/bold green]  "
        f"OK=[bold green]{ok}[/bold green]  "
        f"Duplicados=[bold yellow]{dup}[/bold yellow]  "
        f"Erros=[bold red]{err}[/bold red]  "
        f"Total=[bold]{len(resultados)}[/bold]"
    )
    console.print(f"[bold]{'='*70}[/bold]\n")

    for r in resultados:
        icone = (
            "[green]OK [/green]"  if r["status"] == "ok"
            else "[yellow]DUP[/yellow]" if r["status"] == "duplicado"
            else "[red]ERR[/red]"
        )
        console.print(f"  {icone} [{r['identificador']}] {r['mensagem']}")

    # ── Geração dos logs Excel (igual ao fluxo-pw-new.py) ─────────────────────
    from datetime import datetime as _dt
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    timestamp_comum = _dt.now().strftime("%Y%m%d_%H%M%S")
    # Logs ficam dentro de Condomínios/ (junto com os PDFs do ciclo) — não na
    # raiz do projeto, pra manter tudo organizado por administradora.
    PASTA_LOGS = DOWNLOADS_DIR / "Condomínios" / "logs"
    PASTA_LOGS.mkdir(parents=True, exist_ok=True)

    # ── 1. Boletos_Processados ────────────────────────────────────────────────
    if _boletos_processados:
        df_bol = pd.DataFrame(_boletos_processados)
        try:
            df_bol["Vencimento_dt"] = pd.to_datetime(df_bol["Vencimento"], format="%d/%m/%Y")
            df_bol = df_bol.sort_values("Vencimento_dt").drop("Vencimento_dt", axis=1)
        except Exception:
            pass
        cols_excluir = [c for c in ("Pasta_Mes", "Caminho_Completo") if c in df_bol.columns]
        df_excel = df_bol.drop(columns=cols_excluir)

        arq_bol = str(PASTA_LOGS / f"Boletos_Processados_{timestamp_comum}.xlsx")
        df_excel.to_excel(arq_bol, index=False, engine="openpyxl")

        try:
            wb = load_workbook(arq_bol)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            col_w = {"A": 12, "B": 50, "C": 12, "D": 25, "E": 50, "F": 15, "G": 15}
            for col, w in col_w.items():
                ws.column_dimensions[col].width = w
            wb.save(arq_bol)
        except Exception as e:
            console.print(f"[yellow]Formatação Excel boletos: {e}[/yellow]")

        # Resumo por mês
        if "Pasta_Mes" in df_bol.columns and "Cliente" in df_bol.columns:
            resumo = df_bol.groupby("Pasta_Mes").agg(
                Qtd=("Cliente", "count"), Valor=("Valor", "sum")
            )
            resumo_fmt = resumo.copy()
            resumo_fmt["Qtd"]   = resumo_fmt["Qtd"].apply(lambda x: f"{x:,}".replace(",", "."))
            resumo_fmt["Valor"] = resumo_fmt["Valor"].apply(lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            console.print(f"\n[bold]Resumo por mês:[/bold]\n{resumo_fmt.to_string()}")

        console.print(f"\n[green]✅ {len(df_bol)} boleto(s) salvos em: {arq_bol}[/green]")
    else:
        console.print("\n[yellow]⚠️ Nenhum boleto processado com sucesso.[/yellow]")

    # ── 2. Boletos_Duplicados ─────────────────────────────────────────────────
    if _boletos_duplicados:
        arq_dup = str(PASTA_LOGS / f"Boletos_Duplicados_{timestamp_comum}.xlsx")
        pd.DataFrame(_boletos_duplicados).to_excel(arq_dup, index=False, engine="openpyxl")
        console.print(f"[yellow]⚠️ {len(_boletos_duplicados)} duplicado(s) em: {arq_dup}[/yellow]")

    # ── 3. Resultado_Boletos (input + status por identificador) ───────────────
    res_por_id: dict = {}
    for r in resultados:
        res_por_id.setdefault(r["identificador"], []).append(
            f"[{r['status']}] {r['mensagem']}"
        )

    # Mapa resultado por identificador (pode haver múltiplos por id, ex: múltiplos boletos)
    mapa_res: dict = {}
    for r in resultados:
        mapa_res.setdefault(str(r["identificador"]).strip(), []).append(r)

    # Inclui TODOS os registros do banco — os sem resultado aparecem como "nao_processado"
    linhas_res = []
    for l in linhas:
        id_ = str(l[0]).strip()
        usuario      = str(l[1]).strip() if len(l) > 1 else ""
        administradora = str(l[3]).strip() if len(l) > 3 else ""
        condominio   = str(l[4]).strip() if len(l) > 4 else ""
        apartamento  = str(l[5]).strip() if len(l) > 5 else ""

        res_lista = mapa_res.get(id_)
        if res_lista:
            for r in res_lista:
                linhas_res.append({
                    "IDENTIFICADOR":     id_,
                    "USUARIO":           usuario,
                    "ADMINISTRADORA":    administradora,
                    "CONDOMINIO":        condominio,
                    "APARTAMENTO":       apartamento,
                    "STATUS":            r["status"],
                    "MENSAGEM":          r["mensagem"],
                    "ARQUIVO":           r.get("arquivo", "") or "",
                    "OPCOES_DISPONIVEIS": r.get("opcoes", "") or "",
                })
        else:
            linhas_res.append({
                "IDENTIFICADOR":     id_,
                "USUARIO":           usuario,
                "ADMINISTRADORA":    administradora,
                "CONDOMINIO":        condominio,
                "APARTAMENTO":       apartamento,
                "STATUS":            "nao_processado",
                "MENSAGEM":          "",
                "ARQUIVO":           "",
                "OPCOES_DISPONIVEIS": "",
            })

    arq_res = str(PASTA_LOGS / f"Resultado_API_{timestamp_comum}.xlsx")
    df_res = pd.DataFrame(linhas_res, columns=[
        "IDENTIFICADOR","USUARIO","ADMINISTRADORA","CONDOMINIO",
        "APARTAMENTO","STATUS","MENSAGEM","ARQUIVO","OPCOES_DISPONIVEIS",
    ])
    df_res.to_excel(arq_res, index=False, engine="openpyxl")

    try:
        wb = load_workbook(arq_res)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        col_w_res = {
            "A": 12,  # IDENTIFICADOR
            "B": 30,  # USUARIO
            "C": 35,  # ADMINISTRADORA
            "D": 35,  # CONDOMINIO
            "E": 15,  # APARTAMENTO
            "F": 18,  # STATUS
            "G": 60,  # MENSAGEM
            "H": 60,  # ARQUIVO
            "I": 80,  # OPCOES_DISPONIVEIS
        }
        for col, w in col_w_res.items():
            ws.column_dimensions[col].width = w
        wb.save(arq_res)
    except Exception as e:
        console.print(f"[yellow]Formatação Excel resultado: {e}[/yellow]")

    console.print(f"[green]✅ Resultado salvo em: {arq_res}[/green]\n")
