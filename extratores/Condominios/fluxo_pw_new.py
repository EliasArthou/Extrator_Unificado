# -*- coding: utf-8 -*-

import os
import re
import json
import unicodedata
from urllib.parse import urlparse
from pathlib import Path
from dotenv import load_dotenv
from rapidfuzz import fuzz
from playwright.sync_api import sync_playwright, TimeoutError
import pandas as pd
from datetime import datetime
import time
import threading
from collections import deque
from auxiliares import utils as aux
from auxiliares import aux_patches  # registra adicionarcabecalhopdf_topo_adaptativo em aux

# --- Rich imports ---
from rich.console import Console, Group
from rich.live import Live
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeRemainingColumn, \
    TimeElapsedColumn
from rich.text import Text

# ====== módulo que devolve o site por administradora ======
from auxiliares import sensiveis as info

# ====== Perfil do Chrome para baixar PDF direto ======

# Inicializa o console global do Rich
console = Console(force_terminal=True, color_system="truecolor")


# ===================== PROGRESS LOGGER COM BARRA FIXA =====================
class FixedProgressLogger:
    """
    Barra de progresso na parte inferior, logs rolam normalmente acima.
    """

    def __init__(self, total: int, description: str = "Processando..."):
        self.total = total
        self.description = description
        self.current = 0
        self.lock = threading.Lock()
        self.live = None

        # Cria o Progress com alinhamento fixo
        self.progress = Progress(
            SpinnerColumn(),
            TextColumn("[bold blue]{task.description}", justify="left"),
            BarColumn(bar_width=40),
            TextColumn("{task.completed:>4}/{task.total:<4}", style="bold cyan"),
            TaskProgressColumn(),
            TextColumn("│ Restante:", style="dim"),
            TimeRemainingColumn(),
            TextColumn("│ Decorrido:", style="dim"),
            TimeElapsedColumn(),
            expand=False,  # Não expande para largura do terminal
        )
        self.task_id = self.progress.add_task(description, total=total)

    def _build_display(self):
        """Constrói apenas a barra de progresso."""
        return Panel(
            self.progress,
            title="[bold green]⏳ Progresso ",
            border_style="green",
            padding=(0, 1),
        )

    def start(self):
        """Inicia o display."""
        self.live = Live(
            self._build_display(),
            console=console,
            refresh_per_second=4,
            transient=False,
            screen=False,
            vertical_overflow="visible",
        )
        self.live.start()

    def stop(self):
        """Para o display."""
        if self.live:
            self.live.stop()
            self.live = None

    def log(self, msg: str):
        """Imprime log acima da barra (scroll normal do terminal)."""
        with self.lock:
            if self.live:
                self.live.console.print(msg)
            else:
                console.print(msg)

    def advance(self, amount: int = 1):
        """Avança a barra de progresso."""
        with self.lock:
            self.current += amount
            self.progress.advance(self.task_id, amount)
            if self.live:
                self.live.update(self._build_display())

    def update_description(self, desc: str):
        """Atualiza a descrição da tarefa."""
        with self.lock:
            self.description = desc
            self.progress.update(self.task_id, description=desc)
            if self.live:
                self.live.update(self._build_display())


# Instância global do logger
_progress_logger: FixedProgressLogger = None


def log(msg: str):
    """Função global de log - usa o FixedProgressLogger se disponível."""
    global _progress_logger
    if _progress_logger is not None:
        _progress_logger.log(msg)
    else:
        console.print(msg)


USER_DATA_DIR = Path("./chrome-profile")
DOWNLOADS_DIR = Path("./downloads")

# ===== CACHE DE CÓDIGOS DE BARRAS PARA DETECTAR DUPLICATAS ===== ↓
CODIGOS_BARRAS_EXISTENTES = set()  # Set global com códigos já baixados
BOLETOS_DUPLICADOS = []  # Lista de boletos duplicados para log


def carregar_codigos_barras_existentes():
    """
    Lê todos os PDFs na pasta de downloads (incluindo subpastas) e extrai os códigos de barras.
    Retorna um set com todos os códigos encontrados.
    """
    global CODIGOS_BARRAS_EXISTENTES
    from core import boletos
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn

    log(f"📂 [CACHE] Carregando códigos de barras existentes de {DOWNLOADS_DIR}...")

    codigos = set()
    arquivos_lidos = 0
    arquivos_erro = 0

    # Lista todos os PDFs primeiro para saber o total
    lista_pdfs = list(DOWNLOADS_DIR.rglob("*.pdf"))
    total_pdfs = len(lista_pdfs)

    if total_pdfs == 0:
        log(f"ℹ️ [CACHE] Nenhum PDF encontrado na pasta de downloads")
        CODIGOS_BARRAS_EXISTENTES = codigos
        return codigos

    # Barra de progresso com Rich
    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(bar_width=40),
        TaskProgressColumn(),
        TextColumn("({task.completed}/{task.total})"),
        TimeElapsedColumn(),
        console=console,
        transient=True  # Remove a barra ao finalizar
    ) as progress:
        task = progress.add_task("Lendo códigos de barras...", total=total_pdfs)

        for pdf_path in lista_pdfs:
            try:
                # Usa a função existente do módulo boletos
                resultado = boletos.barcodereader(str(pdf_path), qualidade=200, renomear=False)

                if resultado and resultado.get('Código de Barras') != "Impossível ler código de barras":
                    codigo = resultado.get('Código de Barras', '')
                    if codigo:
                        codigos.add(codigo)
                        arquivos_lidos += 1
            except Exception as e:
                arquivos_erro += 1
                # Não loga cada erro individual para não poluir

            progress.update(task, advance=1)

    CODIGOS_BARRAS_EXISTENTES = codigos
    log(f"✅ [CACHE] {len(codigos)} códigos de barras carregados de {arquivos_lidos}/{total_pdfs} arquivos")
    if arquivos_erro > 0:
        log(f"⚠️ [CACHE] {arquivos_erro} arquivos não puderam ser lidos")

    return codigos


def verificar_boleto_duplicado(codigo_barras: str) -> bool:
    """
    Verifica se um código de barras já existe no cache.
    Retorna True se for duplicado, False se for novo.
    """
    return codigo_barras in CODIGOS_BARRAS_EXISTENTES


def registrar_codigo_barras(codigo_barras: str):
    """
    Adiciona um novo código de barras ao cache (para evitar duplicatas na mesma sessão).
    """
    global CODIGOS_BARRAS_EXISTENTES
    CODIGOS_BARRAS_EXISTENTES.add(codigo_barras)


def registrar_boleto_duplicado(codigo_cliente: str, codigo_barras: str, arquivo_descartado: str):
    """
    Registra um boleto duplicado para log posterior.
    """
    global BOLETOS_DUPLICADOS
    BOLETOS_DUPLICADOS.append({
        "Cliente": codigo_cliente,
        "Cod_Barras": codigo_barras,
        "Arquivo_Descartado": arquivo_descartado,
        "Motivo": "Código de barras já existe na pasta de downloads"
    })
# ===== FIM CACHE ===== ↑

# ===== CONSTANTES DO .ENV (ÍNDICES DE COLUNAS) ===== ↓
IDENTIFICADOR = 0
USUARIO = 1
SENHA = 2
ADMINISTRADORA = 3
CONDOMINIO = 4
APARTAMENTO = 5
LOGINMULTIPLO = 6
RESPOSTA = 7
CHECKARQUIVO = 8
CHECKERRO = 9
NOMEFUNCAO = 10
PROBLEMALOGIN = 11
CPF = 12
# ===== FIM CONSTANTES ===== ↑

# ===== LISTAS DE DICIONÁRIOS ===== ↓
# Dicionário DE-PARA para normalização de apartamentos
# Chave: termo original (minúsculo), Valor: abreviação padronizada
DE_PARA_APTO = {
    "apartamento": "ap",
    "apto": "ap",
    "unidade": "und",
    "loja": "lj",
    "sala": "sl",
    "salas": "sl",
    "bloco": "bl",
    "box": "bx",
    "casa": "cs",
    "pavimento": "pav"
}

# Dicionário DE-PARA para normalização de condomínios
# Chave: termo original (minúsculo), Valor: abreviação padronizada
DE_PARA_CONDO = {
    "condominio": "cond",
    "edificio": "ed",
    "residencial": "res",
    "grupamento": "grup",
    "agrupamento": "grup",
    "associacao": "assoc",
    "proprietarios": "prop",
    "moradores": "mor",
    "gleba": "glb",
    "setor": "set",
    "resort": "",
    "deck": ""
}


# ===== FIM ===== ↑


def preparar_perfil():
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
    prefs_path = USER_DATA_DIR / "Default" / "Preferences"
    prefs_path.parent.mkdir(parents=True, exist_ok=True)
    prefs = {
        "download": {"default_directory": str(DOWNLOADS_DIR.resolve()), "prompt_for_download": False},
        "plugins": {"always_open_pdf_externally": True},
    }
    if prefs_path.exists():
        try:
            cur = json.load(open(prefs_path, "r", encoding="utf-8"))
        except Exception:
            cur = {}
        cur.setdefault("download", {}).update(prefs["download"])
        cur.setdefault("plugins", {}).update(prefs["plugins"])
        data = cur
    else:
        data = prefs
    json.dump(data, open(prefs_path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)


# ===================== Helpers =====================

def _norm_apto(txt: str) -> str:
    """Normaliza apartamentos usando De-Para e limpando zeros à esquerda."""
    t = _norm(txt)  # Converte para minúsculo/limpo
    tokens = t.split()
    t_padronizado = " ".join([DE_PARA_APTO.get(token, token) for token in tokens])
    # Remove zeros: 'lj 0110' -> 'lj 110'
    return re.sub(r'\b0+(\d+)', r'\1', t_padronizado).strip()


def _norm_condo_v2(txt: str) -> str:
    """Normaliza condomínios removendo códigos iniciais e aplicando De-Para."""
    t = _norm(txt)
    # Remove código inicial: '4109 - ALOHA' -> 'aloha'
    t = re.sub(r'^\d+\s*[-–—]?\s*', '', t)
    tokens = t.split()
    return " ".join([DE_PARA_CONDO.get(token, token) for token in tokens]).strip()

    return t_padronizado.strip()


def _extrair_numero_apto(texto: str) -> str:
    """
    Extrai o número do apartamento/loja.
    Padrão comum: número VEM ANTES da palavra-chave (ex: '0119 LJ', '101 APTO')

    Usa as CHAVES do dicionário DE_PARA_APTO para construir os padrões dinamicamente.

    Ex: '0119 LJ' -> '119'
    Ex: '101 APTO' -> '101'
    Ex: 'Bloco 1 0203 SALA' -> '203' (ignora bloco, pega número antes de SALA)
    """
    texto = str(texto).upper().strip()

    # Extrai palavras-chave das CHAVES do dicionário DE_PARA_APTO
    palavras = [k.upper() for k in DE_PARA_APTO.keys()]

    # Constrói o padrão regex dinamicamente baseado no dicionário
    # Adiciona variações com ponto opcional (ex: APT, APT., APTO, APTO.)
    palavras_regex = []
    for p in palavras:
        # Adiciona a palavra base
        palavras_regex.append(re.escape(p))
        # Se a palavra não termina com 'S', adiciona versão com 'S' opcional (ex: SALA -> SALAS?)
        if not p.endswith('S'):
            palavras_regex.append(re.escape(p) + 'S?')

    # Remove duplicatas e junta com | (ordena por tamanho para evitar matches parciais)
    palavras_pattern = '|'.join(sorted(set(palavras_regex), key=len, reverse=True))

    # 1ª tentativa: Número ANTES de palavra-chave (padrão mais comum)
    # Captura: "0119 LJ", "101 APTO", "203 SALA"
    padrao_antes = rf'(\d+)\s*(?:{palavras_pattern})\.?\b'
    match = re.search(padrao_antes, texto, re.IGNORECASE)
    if match:
        return str(int(match.group(1)))  # Remove zeros à esquerda

    # 2ª tentativa: Número DEPOIS de palavra-chave (casos raros)
    # Captura: "APTO 101", "LOJA 0119"
    padrao_depois = rf'(?:{palavras_pattern})\.?\s*(\d+)'
    match = re.search(padrao_depois, texto, re.IGNORECASE)
    if match:
        return str(int(match.group(1)))

    # 3ª tentativa: Pega o primeiro número (fallback para casos sem palavra-chave)
    numeros = re.findall(r'(\d+)', texto)
    if numeros:
        return str(int(numeros[0]))

    return ""


def _apto_from_item_text(raw: str) -> str:
    s = (raw or "").strip()
    s = re.split(r"\bVencimento\b", s, flags=re.I)[0]
    s = s.split(" - ")[0]
    s = s.split(" — ")[0]
    return s.strip()


def _idx_to_excel_col(idx: int) -> str:
    """Converte índice numérico (0-based) para letra de coluna Excel (A, B, C, ...)"""
    return chr(65 + idx)  # 65 = 'A' em ASCII


def _norm(s: str) -> str:
    s = (s or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def _ensure_env_index(name: str) -> int:
    v = os.getenv(name)
    if v is None or not str(v).isdigit():
        raise RuntimeError(f"ENV '{name}' ausente/inválida (precisa ser índice int).")
    return int(v)


def _make_context_from_row(linha):
    """
    Constrói um dicionário com nomes -> valores da linha, usando os índices
    vindos de variáveis de ambiente. Ex.: ADMINISTRADORA, USUARIO, SENHA, CONDOMINIO, APARTAMENTO, IDENTIFICADOR, CPF.
    """
    names = ["ADMINISTRADORA", "USUARIO", "SENHA", "CONDOMINIO", "APARTAMENTO", "IDENTIFICADOR", "CPF"]
    ctx = {}
    for n in names:
        try:
            idx = _ensure_env_index(n)
            ctx[n] = linha[idx]
        except Exception:
            pass

    # SITE específico por administradora (URL)
    # Prioriza LISTAADMINISTRADORA_API (URLs migradas pro Superlógica)
    # Fallback pra LISTAADMINISTRADORA se não achar na API
    if "ADMINISTRADORA" in ctx:
        ctx["SITE"] = None
        try:
            from auxiliares import utils as _aux
            lista_api = _aux.carregar_lista_administradoras_api()
            match = next((x for x in lista_api if x.get("nomereal") == ctx["ADMINISTRADORA"]), None)
            if match and match.get("site"):
                ctx["SITE"] = match["site"]
        except Exception:
            pass
        if not ctx["SITE"]:
            try:
                ctx["SITE"] = info.retornaradministradora("nomereal", ctx["ADMINISTRADORA"], "site")
            except Exception:
                ctx["SITE"] = None
    # derivado
    if "APARTAMENTO" in ctx:
        ctx["APTO_SAFE"] = _norm(str(ctx["APARTAMENTO"])).replace(" ", "_")
    return ctx


def _resolve_placeholders(val, context: dict):
    if not isinstance(val, str):
        return val
    if val.startswith("${") and val.endswith("}"):
        key = val[2:-1]
        return context.get(key, val)
    return val


def creds_norm(ctx):
    adm = (ctx.get("ADMINISTRADORA") or "").strip()
    usr = (ctx.get("USUARIO") or "").strip().lower()
    sen = (ctx.get("SENHA") or "")
    return (adm, usr, sen)


def mask_creds(creds):
    """Mascara a senha em uma tupla de credenciais para exibição em logs."""
    if creds is None:
        return None
    if isinstance(creds, tuple) and len(creds) >= 3:
        adm, usr, sen = creds[0], creds[1], creds[2]
        masked_sen = "***" if sen else ""
        return (adm, usr, masked_sen)
    return creds


def extrair_primeiro_locator_login():
    """
    Extrai o primeiro locator de cada plataforma de LOGINS.
    Usa para detectar tela de login automaticamente.
    """
    selectors = {}

    for platform, steps in LOGINS.items():
        for step in steps:
            # Procura primeiro passo "fill" (geralmente é o campo de usuário/email)
            if step.get("action") == "fill":
                selectors[platform] = step["locator"]
                break

    return selectors


def tela_login_visivel(pagina, timeout_ms=800):
    """
    Detecta tela de login automaticamente.
    Reutiliza configuração de LOGINS.
    """
    selectors = extrair_primeiro_locator_login()

    for platform, locator_spec in selectors.items():
        try:
            by = locator_spec.get("by", "css")

            if by == "css":
                loc = pagina.locator(locator_spec["selector"])
            elif by == "role":
                role = locator_spec["role"]
                name = locator_spec.get("name")
                exact = locator_spec.get("exact", False)
                loc = pagina.get_by_role(role, name=name, exact=exact) if name else pagina.get_by_role(role)
            elif by == "xpath":
                loc = pagina.locator(f"xpath={locator_spec['selector']}")
            else:
                continue

            loc.first.wait_for(state="visible", timeout=timeout_ms)
            return True

        except Exception:
            continue

    return False


def esta_logado(page, site_family: str, timeout_ms=700) -> bool:
    """Heurística por família de site após login."""
    fam = (site_family or "").strip().lower()
    try:
        if fam == "superlogica":
            # Elemento do menu/condomínios da área logada
            page.locator("ul#lista").first.wait_for(state="visible", timeout=timeout_ms)
            return True
        # ===== LIVEFACILITIES (NOVO) =====
        elif fam == "livefacilities":
            # Verifica se o menu "MEUS DADOS" ou "SAIR" está visível
            # O link "MEUS DADOS" é padrão no menu lateral
            try:
                page.get_by_role("link", name="MEUS DADOS").first.wait_for(state="visible", timeout=timeout_ms)
                return True
            except:
                # Fallback: tenta achar o botão de Sair
                page.get_by_text("Sair", exact=True).first.wait_for(state="visible", timeout=timeout_ms)
                return True
        # ===== CIPA =====
        elif "cipa" in fam:
            # CIPA: Verifica se está logado pelo texto "Logado como" ou menu de condomínios
            try:
                # Tenta encontrar "Logado como" que aparece quando está logado
                if page.locator("text=/Logado como/i").first.is_visible(timeout=timeout_ms):
                    return True
            except:
                pass
            try:
                # Fallback: verifica se está na área de condomínios ou dentro de um
                if page.locator("div.filter-article").first.is_visible(timeout=timeout_ms):
                    return True
            except:
                pass
            try:
                # Verifica se está dentro de um condomínio (menu lateral visível)
                if page.locator("a[href*='/boletos']").first.is_visible(timeout=timeout_ms):
                    return True
            except:
                pass
            return False
    except Exception:
        pass
    return False


def fuzzy_match_items_core(items_data: list[dict], alvo: str, cutoff: float = 0.75,
                           validar_numero: bool = False, is_condo: bool = False, tie_tol: float = 1e-6) -> dict:
    if not items_data:
        raise ValueError("Lista de items vazia")

    # 1. Define qual função de limpeza usar
    norm_fn = _norm_condo_v2 if is_condo else _norm_apto
    tipo_busca = "CONDOMÍNIO" if is_condo else "APARTAMENTO"

    # 2. Match Exato (Original vs Original)
    alvo_norm_simples = _norm(alvo)
    for item in items_data:
        if _norm(item['text']) == alvo_norm_simples:
            log_lines = [
                f"[FUZZY {tipo_busca}] ✅ Match exato (score=1.00):",
                f"              📋 Buscado (Base): '{alvo}'",
                f"              🌐 Encontrado (Site): '{item['text']}'",
                f"              🎯 Método: Comparação direta"
            ]
            return {'winners': [item], 'best_score': 1.0, 'best_txt': item['text'], 'log_lines': log_lines}

    # 3. Cálculo Fuzzy sobre as versões normalizadas
    for item in items_data:
        tn_site = norm_fn(item['text'])
        tn_alvo = norm_fn(alvo)

        w_score = fuzz.WRatio(tn_site, tn_alvo) / 100.0
        p_score = fuzz.partial_ratio(tn_site, tn_alvo) / 100.0

        item['w_score'] = w_score
        item['p_score'] = p_score
        # 🚀 A MUDANÇA: Sempre pega o maior score para não 'atropelar' matches parciais perfeitos
        item['score'] = max(w_score, p_score)

    best_score = max(item['score'] for item in items_data)
    empatados = [it for it in items_data if abs(it['score'] - best_score) <= tie_tol]

    # 4. Validação Rígida de Número (Apenas para Aptos)
    if validar_numero and not is_condo and empatados:
        numero_alvo = _extrair_numero_apto(alvo)
        if numero_alvo:
            validados = [it for it in empatados if _extrair_numero_apto(it['text']) == numero_alvo]
            if validados:
                empatados = validados
            else:
                raise ValueError(f"Número '{numero_alvo}' não encontrado nos candidatos.")

    if best_score < cutoff:
        raise ValueError(f"Melhor score {best_score:.2f} < cutoff {cutoff:.2f}. Buscado: '{alvo}'")

    vencedor = empatados[0]

    # 📝 MONTAGEM DO LOG COM TEXTOS ORIGINAIS (Conforme solicitado)
    log_lines = [
        f"[FUZZY {tipo_busca}] 🔍 Match parcial (score={best_score:.2f}):",
        f"              📋 Buscado (Base): '{alvo}'",
        f"              🌐 Encontrado (Site): '{vencedor['text']}'",
        f"              📊 WRatio: {vencedor['w_score']:.2f} | partial: {vencedor['p_score']:.2f}"
    ]

    return {
        'winners': empatados, 'best_score': best_score, 'best_txt': vencedor['text'],
        'log_lines': log_lines
    }


def _gravar_resposta(linha: list, texto: str):
    """Grava texto no índice RESPOSTA (do .env) dentro da lista 'linha'."""
    try:
        idx = _ensure_env_index("RESPOSTA")
    except Exception:
        return
    if idx >= len(linha):
        linha.extend([""] * (idx - len(linha) + 1))
    linha[idx] = texto


def _host_from_url(u: str) -> str:
    try:
        return (urlparse(u).netloc or "").lower()
    except Exception:
        return ""


# ---- LOGOUT helpers ----

def _fechar_overlay_cipa(page, timeout_ms=2000) -> bool:
    """
    Fecha overlay/modal aberto na CIPA antes de tentar clicar em outros elementos.
    Retorna True se fechou algo, False se não havia overlay.
    """
    try:
        # Verifica se há overlay visível
        overlay = page.locator("div.cdk-overlay-backdrop").first
        if overlay.is_visible(timeout=500):
            # Método 1: Pressiona ESC para fechar
            page.keyboard.press("Escape")
            page.wait_for_timeout(500)

            # Verifica se fechou
            if not overlay.is_visible(timeout=300):
                return True

            # Método 2: Clica no backdrop
            try:
                overlay.click(force=True, timeout=1000)
                page.wait_for_timeout(500)
                return True
            except:
                pass

            # Método 3: Remove via JS
            try:
                page.evaluate("""() => {
                    const overlays = document.querySelectorAll('.cdk-overlay-backdrop, .cdk-overlay-pane');
                    overlays.forEach(o => o.remove());
                }""")
                page.wait_for_timeout(300)
                return True
            except:
                pass
    except:
        pass
    return False


def _try_click_logout(page, timeout_ms=2000) -> bool:
    # Variações comuns na Superlógica
    try:
        page.get_by_role("link", name=re.compile(r"^Sair$", re.I)).first.click(timeout=800)
        page.locator("input[name='email']").first.wait_for(state="visible", timeout=timeout_ms)
        return True
    except Exception:
        pass
    try:
        page.locator("a.item-sair.hasPopover").first.click(timeout=800)
        page.locator("input[name='email']").first.wait_for(state="visible", timeout=timeout_ms)
        return True
    except Exception:
        pass
    try:
        page.locator("span.texto-menu", has_text=re.compile(r"^Sair$", re.I)).first.click(timeout=800)
        page.locator("input[name='email']").first.wait_for(state="visible", timeout=timeout_ms)
        return True
    except Exception:
        pass
    return False


def _force_logout_by_url(page, site_url: str, timeout_ms=3000) -> bool:
    # Superlógica costuma responder a estes endpoints
    try:
        base = (site_url or "").rstrip("/")
        for path in ("/publico/login/out", "/acesso/morador/publico/logout"):
            try:
                page.goto(base + path)
                page.locator("input[name='email']").first.wait_for(state="visible", timeout=timeout_ms)
                return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def efetuar_logout(page, site_url: str) -> None:
    """
    Logout robusto:
    1. Tenta clicar no botão 'Sair'.
    2. Se houver botão 'Novo Login', clica nele OU executa o JS dele.
    3. Navega para a URL inicial.
    """
    # Se já está na tela de login, não faz nada
    if tela_login_visivel(page, timeout_ms=600):
        return

    log("🔄 [LOGOUT] Iniciando procedimento...")
    url_lower = (site_url or "").lower()

    # ===== 1. LIVEFACILITIES =====
    if "livefacilities" in url_lower or "servicos.condominio" in url_lower or "bap.com.br" in url_lower:
        try:
            seletor_sair = "a[id$='lbSair'], a[id$='lkSair'], a[title='Sair']"
            botao_sair = page.locator(seletor_sair).first
            if botao_sair.is_visible(timeout=3000):
                botao_sair.click()
                page.wait_for_timeout(1000)
            _forcar_home(page, site_url)
            return
        except:
            pass

    # ===== 2. PROTEL =====
    if "protel" in url_lower:
        try:
            botao_sair = page.locator("#botao_logout, img[alt='Botao_sair']").first
            if botao_sair.is_visible(timeout=3000):
                botao_sair.click()
                page.wait_for_load_state("networkidle")
            _forcar_home(page, site_url)
            return
        except:
            pass

    # ===== 3. FERNANDO E FERNANDES =====
    if "fernandoefernandes" in url_lower:
        try:
            menu = page.locator("a.dropdown-toggle").first
            btn = page.locator("a[href*='inform=off'], a:has-text('Logout')").first
            if not btn.is_visible() and menu.is_visible():
                menu.click()
                page.wait_for_timeout(500)
            if btn.is_visible():
                btn.click()
                page.wait_for_load_state("networkidle")
                _forcar_home(page, site_url)
                return
        except:
            pass

    # ===== 4. IMODATA (FINAL) =====
    if "imodata" in url_lower:
        try:
            # 1. Clicar em SAIR
            botao_sair = page.locator("a[href='/logout/0']").first

            if not botao_sair.is_visible():
                menu_avatar = page.locator("a.dropdown-user-link, a.dropdown-toggle .avatar").first
                if menu_avatar.is_visible():
                    log("   🖱️ Abrindo menu suspenso (Imodata)...")
                    menu_avatar.click()
                    page.wait_for_timeout(800)

            if botao_sair.is_visible(timeout=3000):
                log("   🖱️ Clicando em 'Sair' (Imodata)...")
                botao_sair.click()
                page.wait_for_timeout(1500)

                # 2. CLICAR EM NOVO LOGIN
            log("   🔍 Procurando botão 'Novo Login'...")
            btn_novo = page.locator("input[value='Novo Login'], input[onclick*='show_form']").first

            sucesso_reset = False

            # Tenta clicar
            if btn_novo.is_visible(timeout=3000):
                log("   🖱️ Botão visível, clicando...")
                btn_novo.click(force=True)
                page.wait_for_timeout(1000)
                sucesso_reset = True
            else:
                # Se não clicou, injeta JS
                log("   💉 Injetando JS para restaurar formulário...")
                page.evaluate("""() => {
                    try {
                        document.getElementById('show_form').style.display='';
                        document.getElementById('show_relogin').style.display='none';
                    } catch(e) {}
                }""")
                page.wait_for_timeout(1000)
                # Verifica se o form apareceu
                if page.locator("#show_form").is_visible():
                    sucesso_reset = True

            # --- AQUI ESTÁ O TRUQUE ---
            # Se conseguimos resetar o formulário, NÃO navegamos para a home.
            # Deixamos a página parada aqui para o próximo login aproveitar o form aberto.
            if sucesso_reset:
                log("   ✅ Formulário de login restaurado. Mantendo página atual.")
                return

        except Exception as e:
            log(f"   ⚠️ Erro Imodata Logout: {e}")

    # ===== 5. CIPA - Logout INTELIGENTE ===== ↓
    if "cipafacil" in url_lower:
        try:
            # ===== 1. VERIFICA USUÁRIO LOGADO ===== ↓
            usuario_esperado = os.getenv("USUARIO", "").strip().lower()
            usuario_logado = None

            try:
                # Procura texto "Logado como [email]"
                texto_logado = page.locator("text=/Logado como/i").first
                if texto_logado.is_visible(timeout=2000):
                    texto_completo = texto_logado.inner_text()
                    # Extrai email (ex: "Logado como\ncondominios@wmartins.com.br")
                    linhas = texto_completo.split('\n')
                    if len(linhas) > 1:
                        usuario_logado = linhas[1].strip().lower()
                        log(f"   🔍 Usuário logado: {usuario_logado}")
            except Exception as e:
                log(f"   ⚠️ Não conseguiu verificar usuário logado: {e}")
            # ===== FIM VERIFICAÇÃO ===== ↑

            # ===== 2. DECIDE ESTRATÉGIA ===== ↓
            usuario_correto = (usuario_logado and usuario_esperado and
                               usuario_logado == usuario_esperado)

            if usuario_correto:
                log(f"   ✅ Usuário CORRETO ({usuario_logado})")

                # ===== ESTRATÉGIA A: SÓ VOLTA PARA ÁREA DE CONDOMÍNIOS ===== ↓
                try:
                    # Fecha overlay/modal se estiver aberto
                    _fechar_overlay_cipa(page)

                    # Verifica se está dentro de um condomínio (menu com nome)
                    menu_condo = page.locator("div[class*='cursor-pointer']").filter(
                        has_text=re.compile(r"Gestao|Imobiliaria|Condominio|Wmartins", re.I)
                    ).first

                    if menu_condo.is_visible(timeout=2000):
                        log(f"   🏢 Dentro de condomínio - voltando para área de seleção...")

                        # Fecha overlay novamente antes de clicar
                        _fechar_overlay_cipa(page)

                        # Clica no menu
                        menu_condo.click(timeout=5000)
                        page.wait_for_timeout(1000)

                        # Clica em "Sair" (volta para lista de condomínios)
                        btn_sair_condo = page.get_by_role("menuitem", name="Sair").first
                        if btn_sair_condo.is_visible(timeout=3000):
                            btn_sair_condo.click(timeout=5000)
                            page.wait_for_timeout(1500)
                            log(f"   ✅ Voltou para área de condomínios")
                            return  # ← PARA AQUI! Não faz logout do usuário

                    else:
                        log(f"   ℹ️  Já está na área de condomínios")
                        return  # ← JÁ ESTÁ NO LUGAR CERTO!

                except Exception as e:
                    log(f"   ⚠️ Erro ao voltar para área de condomínios: {e}")
                # ===== FIM ESTRATÉGIA A ===== ↑

            else:
                log(
                    f"   ❌ Usuário INCORRETO (esperado: {usuario_esperado}, logado: {usuario_logado or 'desconhecido'})")

                # ===== ESTRATÉGIA B: LOGOUT COMPLETO ===== ↓
                # Primeiro: Fecha qualquer overlay/modal aberto
                _fechar_overlay_cipa(page)

                # Nível 1: Sai do condomínio (se estiver dentro)
                try:
                    menu_condo = page.locator("div[class*='cursor-pointer']").filter(
                        has_text=re.compile(r"Gestao|Imobiliaria|Condominio|Wmartins", re.I)
                    ).first

                    if menu_condo.is_visible(timeout=2000):
                        log(f"   🏢 Saindo do condomínio...")
                        _fechar_overlay_cipa(page)  # Garante overlay fechado
                        menu_condo.click(timeout=5000)
                        page.wait_for_timeout(1000)

                        page.get_by_role("menuitem", name="Sair").click(timeout=5000)
                        page.wait_for_timeout(1500)
                        log(f"   ✅ Saiu do condomínio")
                except Exception as e:
                    log(f"   ℹ️  Não estava em condomínio: {e}")

                # Nível 2: Logout do usuário
                try:
                    log(f"   👤 Fazendo logout do usuário...")
                    _fechar_overlay_cipa(page)  # Garante overlay fechado

                    # Procura ícone de perfil
                    icone_perfil = None
                    seletores_perfil = [
                        ".mat-icon.notranslate > svg",
                        ".mat-icon.icon-size-4 > svg",
                        "button[class*='avatar']",
                        "[class*='user-menu']"
                    ]

                    for sel in seletores_perfil:
                        try:
                            icone = page.locator(sel).first
                            if icone.is_visible(timeout=2000):
                                icone_perfil = icone
                                break
                        except:
                            continue

                    if icone_perfil:
                        icone_perfil.click(timeout=5000)
                        page.wait_for_timeout(1000)

                        # Clica em "Sair" do usuário
                        page.get_by_role("menuitem", name="Sair").click(timeout=5000)
                        page.wait_for_load_state("networkidle", timeout=10000)
                        log(f"   ✅ Logout do usuário OK")

                        # Navega para login
                        _navegar_e_esperar(page, "https://cipafacil.digital/sign-in")
                        return

                except Exception as e:
                    log(f"   ⚠️ Erro no logout do usuário: {e}")
                # ===== FIM ESTRATÉGIA B ===== ↑

        except Exception as e:
            log(f"   ⚠️ Erro CIPA Logout: {e}")
    # ===== FIM CIPA ===== ↑

    # ===== 6. MÉTODO GENÉRICO =====

    if _try_click_logout(page, timeout_ms=2500):
        return
    if _force_logout_by_url(page, site_url, timeout_ms=2500):
        return

    try:
        page.context.clear_cookies()
    except:
        pass
    _forcar_home(page, site_url)


def _forcar_home(page, url: str) -> None:
    """Navega para página inicial usando helper genérico."""
    log(f"   🏠 Voltando para página inicial...")
    _navegar_e_esperar(page, url)


def _navegar_e_esperar(page, url: str, timeout_ms: int = 30000) -> None:
    """
    Navega para URL e aguarda página carregar completamente.
    Centraliza lógica de navegação usada em vários lugares.
    Lança exceção se a navegação falhar.
    """
    if not url:
        return

    try:
        log(f"   🌐 Navegando para: {url}")
        page.goto(url, timeout=timeout_ms)
        page.wait_for_load_state("networkidle", timeout=min(timeout_ms, 15000))
        page.wait_for_timeout(1000)  # Espera extra para JS carregar

        # Verifica se caiu em página de erro comum ou about:blank
        from urllib.parse import urlparse
        url_atual = page.url.lower()

        # Lista de páginas problemáticas que indicam falha
        paginas_erro = [
            "about:blank",
            "chrome-error:",
            "edge-error:",
            "data:text/html,chromewebdata",
            "chrome://",
            "edge://",
        ]

        # Se está em alguma página de erro, falha
        if any(url_atual.startswith(p) for p in paginas_erro):
            raise ValueError(f"Navegação falhou: está em página de erro ({url_atual})")

        # Extrai domínio base (sem www, sem subdomínios)
        def extrair_dominio_base(url_str):
            """Extrai domínio principal (ex: 'protel.com.br' de 'www.condominio.protel.com.br')"""
            netloc = urlparse(url_str).netloc.lower()
            # Remove www
            netloc = netloc.replace('www.', '')
            # Pega os 2 últimos níveis (empresa.com.br ou empresa.com)
            partes = netloc.split('.')
            if len(partes) >= 2:
                return '.'.join(partes[-2:]) if len(partes[-1]) <= 3 else '.'.join(partes[-3:])
            return netloc

        dominio_esperado = extrair_dominio_base(url)
        dominio_atual = extrair_dominio_base(page.url)

        # Aceita se o domínio base é o mesmo (permite redirecionamentos)
        if dominio_esperado and dominio_atual and dominio_esperado != dominio_atual:
            raise ValueError(f"Navegação falhou: esperava domínio '{dominio_esperado}' mas está em '{dominio_atual}'")

        log(f"   ✅ Página carregada")
    except Exception as e:
        log(f"   ❌ Erro ao navegar: {e}")
        # Lança exceção para interromper o fluxo
        raise


# ===================== Classe Executor =====================
class ExecutorFluxoPW:
    """
    Ações:
      - goto | fill | click | click_expect_popup
      - verificar_elemento | esperar_selector | esperar_url | parar_se_visivel
      - expandir_lista
      - fuzzy_lista_fechada
      - loop_apartamentos | loop_apartamentos_fuzzy_best
      - download_imprimir
    """

    def __init__(self, page, console_rich=None):
        self.page = page
        self.console = console_rich or Console()
        self.resposta = ""
        self.mensagemerro = ""
        self.passo_interrompido = None
        self.baixados = 0
        self.duplicados = 0  # Contador de boletos duplicados descartados
        self.match_log = []
        self.login_ok = None
        self.erros_log = []
        self.codigo_cliente = ""
        self.condo_selecionado_site = None
        self.boletos_processados = []
        self.cpf_senha = None  # CPF para desbloquear PDFs

    # ---------- core loop ----------
    def executar(self, passos: list[dict], linha=None):
        self.baixados = 0
        linha = linha or {}
        skip_mode = None  # "SKIP_LOGIN" ou "SKIP_TO_BOLETOS"

        # Armazena linha como variável de instância para uso nas ações
        self.context_linha = linha

        # Extrai CPF do contexto para usar em PDFs protegidos
        self.cpf_senha = linha.get("CPF", None)

        for i, passo in enumerate(passos):
            try:
                scope = passo.get("scope", "flow")
                action = passo.get("action")

                # --- LÓGICA DE SINALIZADOR (FLAG) ---
                if skip_mode == "SKIP_FLOW":
                    # Se estamos pulando o fluxo, só executamos o que for "always"
                    # E que NÃO seja ação de interação com boletos (ex: logout)
                    if scope != "always":
                        continue
                    else:
                        # Resetamos o sinalizador se chegamos em um passo de "limpeza" final
                        if action in ["goto", "verificar_erro_login"]:
                            skip_mode = None

                # Se estiver em modo skip, pula passos apropriados
                if skip_mode == "SKIP_LOGIN" and scope == "login":
                    log(f"[SKIP] Passo {i + 1} (login) - sessão reutilizada")
                    continue

                if skip_mode == "SKIP_TO_BOLETOS":
                    # Pula passos de login e seleção de condomínio
                    if scope in ["login", "condo"]:
                        log(f"[SKIP] Passo {i + 1} ({scope}) - indo direto para boletos")
                        continue
                    else:
                        skip_mode = None  # Para de pular (chegou no scope=flow)

                log(f"[PASSO {i + 1}/{len(passos)}] action={action} scope={scope}")
                self._run_step(passo, linha)

            except StopIteration as e:
                # Tratamento especial para reuso de sessão
                msg = str(e)
                if "SKIP_LOGIN" in msg:
                    log(f"♻️  [SESSÃO] Reutilizando login - pulando passos de autenticação")
                    skip_mode = "SKIP_LOGIN"
                    continue
                elif "SKIP_TO_BOLETOS" in msg:
                    log(f"♻️  [SESSÃO] Reutilizando tudo - pulando para boletos")
                    skip_mode = "SKIP_TO_BOLETOS"
                    continue

                elif "SKIP_FLOW" in msg:
                    log(f"⚠️ [SINALIZADOR] Passo opcional falhou. Pulando restante do fluxo desta linha...")
                    skip_mode = "SKIP_FLOW"
                    # Gravamos a resposta para o Excel saber o que houve
                    self.resposta = "Boleto não disponível (botão de gerar ausente)"
                    continue

                else:
                    raise

            except Exception as e:
                self.resposta = str(e)
                self.passo_interrompido = passo
                break

    def get_status(self):
        return {
            "ok": not bool(self.resposta),
            "resposta": self.resposta,
            "passo_interrompido": self.passo_interrompido,
            "mensagemerro": self.mensagemerro,
            "baixados": self.baixados,
            "duplicados": self.duplicados,  # Boletos duplicados descartados
            "match_log": "\n".join(self.match_log),
            "login_ok": self.login_ok,
            "erros": "\n".join(self.erros_log),
            "condo_site": self.condo_selecionado_site,
        }

    def reset_status(self):
        self.resposta = ""
        self.mensagemerro = ""
        self.passo_interrompido = None
        self.baixados = 0
        self.duplicados = 0  # Reset duplicados também
        self.match_log = []
        self.login_ok = None
        self.erros_log = []
        self.codigo_cliente = ""
        self.condo_selecionado_site = None
        # NÃO reseta boletos_processados (acumula durante toda execução)

    def _verificar_usuario_ativo(self, step, contexto):
        """
        Verifica sessão ativa. Se usuário incorreto, FORÇA o logout limpando cookies.
        """
        selector_email = step["locator_email"]
        selector_logout = step["locator_logout"]
        selector_menu = step.get("locator_logout_menu")

        usuario_esperado = contexto.get("USUARIO", "").strip().lower()

        log(f"   🕵️ [SMART LOGIN] Verificando sessão ativa...")

        try:
            # 1. Tenta achar o e-mail na tela
            loc_email = self.page.locator(self._locator_from_json(selector_email)).first

            if loc_email.is_visible(timeout=3000):
                email_tela = (loc_email.text_content() or "").strip().lower()
                log(f"       👤 Usuário logado detectado: {email_tela}")

                if usuario_esperado in email_tela:
                    log(f"       ✅ Usuário CORRETO! Mantendo sessão e pulando login.")
                    return "BREAK_FLOW"
                else:
                    log(f"       ⛔ Usuário INCORRETO (Esperado: {usuario_esperado})...")

                    # 2. TENTA LOGOUT VIA UI (Visual)
                    try:
                        if selector_menu:
                            log("       🖱️ Abrindo menu de perfil...")
                            # Force=True ajuda a clicar mesmo se o ícone estiver meio sobreposto
                            self.page.locator(self._locator_from_json(selector_menu)).first.click(force=True)
                            self.page.wait_for_timeout(1000)  # Espera menu animar

                        log("       🖱️ Clicando em Sair...")
                        self.page.locator(self._locator_from_json(selector_logout)).first.click(force=True)
                        self.page.wait_for_timeout(1000)
                    except Exception as e:
                        log(f"       ⚠️ Falha no clique de logout visual: {e}")

                    # 3. LOGOUT NUCLEAR (Limpa Cookies)
                    # Isso garante que a sessão morra, mesmo se o clique falhar
                    log("       🛡️ Forçando limpeza de cookies para garantir logout...")
                    self.page.context.clear_cookies()

                    return None  # Segue para o login
            else:
                log("       ⚪ Nenhuma sessão ativa visível. Seguindo para login...")
                return None

        except Exception as e:
            log(f"       ⚠️ Erro na verificação de sessão: {e}")
            return None

    # ---------- locator resolvers ----------
    def _locator_from_json(self, spec: dict):
        # Resolve o contexto (Página ou Frame)
        context = self.page
        if "frame" in spec:
            # Se tiver frame, usa frame_locator
            f_spec = spec["frame"]
            # Suporta frame por nome, url ou seletor
            if "name" in f_spec:
                context = self.page.frame_locator(f"iframe[name='{f_spec['name']}']")
            elif "selector" in f_spec:
                context = self.page.frame_locator(f_spec["selector"])

        by = spec.get("by", "css")
        if by == "css":
            return context.locator(spec["selector"])

        elif by == "xpath":
            return context.locator(f"xpath={spec['selector']}")

        elif by == "role":
            role = spec["role"]
            name = spec.get("name")
            exact = bool(spec.get("exact", False))
            if name is None:
                return context.get_by_role(role)
            return context.get_by_role(role, name=name, exact=exact)
        else:
            raise ValueError(f"by='{by}' não suportado")

    # ---------- NOVA AÇÃO: Remover Elemento Bloqueante (Nuclear) ----------
    def _remover_elemento_bloqueante(self, selector):
        """Executa JS para remover um elemento do DOM e limpar o backdrop do modal"""
        log(f"   💣 [NUCLEAR] Removendo elemento bloqueante: {selector}")
        try:
            self.page.evaluate(f"""() => {{
                // 1. Remove o elemento alvo
                const el = document.querySelector('{selector}');
                if (el) {{ el.remove(); }}

                // 2. Remove o fundo escuro (backdrop) que sobra as vezes
                const bds = document.querySelectorAll('.modal-backdrop');
                bds.forEach(bd => bd.remove());

                // 3. Destrava o scroll do body
                document.body.classList.remove('modal-open');
                document.body.style.overflow = 'auto';
            }}""")
            self.page.wait_for_timeout(500)
        except Exception as e:
            log(f"   ⚠️ Erro ao remover elemento: {e}")

    def _locator_from_page(self, page_like, spec: dict):
        by = spec.get("by", "css")
        if by == "css":
            return page_like.locator(spec["selector"])

        elif by == "xpath":
            return page_like.locator(f"xpath={spec['selector']}")

        elif by == "role":
            role = spec["role"]
            name = spec.get("name")
            exact = bool(spec.get("exact", False))
            if name is None:
                return page_like.get_by_role(role)
            return page_like.get_by_role(role, name=name, exact=exact)
        raise ValueError(f"by='{by}' não suportado")

    # ---------- atomic: click + expect_popup ----------
    def _click_expect_popup_atomic(self, locator_spec: dict, popup_timeout: int = 15000):
        loc = self._locator_from_json(locator_spec).first
        loc.wait_for(state="visible", timeout=popup_timeout)
        with self.page.expect_popup(timeout=popup_timeout) as pop_info:
            loc.click()
        popup = pop_info.value
        popup.wait_for_load_state("domcontentloaded")
        return popup

    # ---------- baixar com prioridade ----------
    def _baixar_com_prioridade(self, page_like, specs_em_ordem: list, destino: Path, timeout_ms: int = 20000):
        """
        Tenta baixar o boleto tratando erros de popup e múltiplos botões.
        """
        log(f"   📥 [DOWNLOAD] Iniciando _baixar_com_prioridade para {destino.name}")

        # 1. INICIALIZAÇÃO (Evita o erro 'name texto_erro is not defined')
        texto_erro = "Boleto indisponível ou erro no carregamento"

        # 2. VERIFICAÇÃO DE ERRO NA TELA (Cálculo mantido e protegido)
        try:
            # Detecta elementos de erro comuns no Superlógica
            erro_el = page_like.locator(".alert-danger, .avisoErro").first

            # Se o elemento de erro estiver visível em até 2s, processamos a mensagem
            if erro_el.is_visible(timeout=2000):
                # Tenta capturar a mensagem específica dentro da div .alerta
                try:
                    msg_especifica = page_like.locator(".alerta").first
                    if msg_especifica.is_visible(timeout=500):
                        texto = msg_especifica.inner_text().strip()
                        # Substitui quebras de linha por espaço
                        texto = texto.replace('\n', ' ').replace('\r', ' ')
                        texto = re.sub(r' +', ' ', texto).strip()
                        # Filtra textos muito longos que podem ser lixo de HTML
                        if 0 < len(texto) < 250:
                            texto_erro = texto
                except:
                    pass  # Mantém o valor padrão se falhar ao ler o texto

                # Registra o erro e interrompe para não tentar clicar em botões inexistentes
                self.mensagemerro = texto_erro
                self.resposta = texto_erro
                log(f"⚠️ [ERRO NO POPUP] {texto_erro}")
                self.erros_log.append(f"[BOLETO] {texto_erro}")

                # Tenta fechar o alerta para limpar a tela (opcional)
                try:
                    page_like.locator(".close, a[data-dismiss='alert']").first.click(timeout=1000)
                except:
                    pass

                raise ValueError(f"Erro: {texto_erro}")

        except TimeoutError:
            # Se não achou div de erro em 2s, assume que a página carregou o boleto
            pass
        except ValueError:
            # Repassa o erro de negócio para o executor parar
            raise

        # 3. LOOP DE DOWNLOAD (Tenta cada botão da lista de prioridades)
        for spec in specs_em_ordem or []:
            if not spec:
                continue
            try:
                # Localiza o botão (Imprimir, Pagar, etc)
                btn = self._locator_from_page(page_like, spec).first

                # ESPERA ROBUSTA: Aguarda até 5s para o botão aparecer (essencial para sites lentos)
                btn.wait_for(state="visible", timeout=5000)

                # Verifica se é o botão "Pagar" - nesse caso, clica e depois seleciona boleto
                btn_name = spec.get("name", "").lower() if isinstance(spec, dict) else ""
                if btn_name == "pagar":
                    # Clica no botão Pagar para abrir o modal de pagamento
                    btn.click(force=True, timeout=3000)
                    log(f"   🔄 Clicou em 'Pagar' - verificando opções de pagamento")

                    # Aguarda os radio buttons de pagamento aparecerem (carregados via AJAX)
                    try:
                        radio_pix = page_like.locator("#parcela-pix")
                        radio_pix.wait_for(state="attached", timeout=3000)

                        # Se PIX existe, seleciona boleto bancário
                        radio_boleto = page_like.locator("#parcela-0")
                        if radio_boleto.count() > 0 and not radio_boleto.is_checked():
                            radio_boleto.click(force=True, timeout=2000)
                            log(f"   🔄 Selecionado 'Boleto Bancário'")
                            page_like.wait_for_timeout(500)
                    except Exception:
                        # Se não encontrar radios, continua normalmente (página pode não ter PIX)
                        pass

                    # Agora procura o botão Imprimir dentro do modal
                    try:
                        btn_imprimir = page_like.locator("#btnSubmitParcelamentoCartao")
                        btn_imprimir.wait_for(state="visible", timeout=3000)

                        # Dispara a captura do download
                        with page_like.expect_download(timeout=timeout_ms) as d_info:
                            btn_imprimir.click(force=True, timeout=3000)

                        # Salva o arquivo no destino final
                        d = d_info.value
                        d.save_as(str(destino))

                        self.baixados += 1
                        log(f"   ✅ Download realizado com sucesso: {destino.name}")
                        return destino
                    except Exception as e:
                        # Se o botão Imprimir não existir, tenta o próximo spec
                        continue

                else:
                    # Para outros botões (Imprimir direto, etc), fluxo normal
                    # Dispara a captura do download
                    with page_like.expect_download(timeout=timeout_ms) as d_info:
                        # CLIQUE FORÇADO: Ignora overlays ou elementos transparentes na frente
                        btn.click(force=True, timeout=3000)

                    # Salva o arquivo no destino final
                    d = d_info.value
                    d.save_as(str(destino))

                    self.baixados += 1
                    log(f"   ✅ Download realizado com sucesso: {destino.name}")
                    return destino

            except Exception as e:
                # Se este botão falhou, tenta o próximo da lista sem alarde
                continue

        # 4. FINALIZAÇÃO: Se percorreu todos os botões e nenhum funcionou
        self.erros_log.append(f"[DOWNLOAD] Nenhum botão de download respondeu no popup")
        raise TimeoutError("Nenhum dos botões fornecidos disparou download.")

    # ---------- expandir_lista (Mais Itens) ----------
    def _expandir_lista(self, locator_spec: dict, timeout=1500, max_clicks=999, settle_ms=800):
        """
        Clica em 'Mais itens' enquanto estiver visível.
        Tenta primeiro o locator informado; se não rolar, tenta fallback por CSS id^='btn_mais_itens_cobrancas'.
        """

        def _mk(spec):
            by = spec.get("by", "css")
            if by == "role":
                role = spec["role"]
                name = spec.get("name")
                exact = bool(spec.get("exact", False))
                return (self.page.get_by_role(role) if name is None
                        else self.page.get_by_role(role, name=name, exact=exact)).first
            elif by == "css":
                return self.page.locator(spec["selector"]).first
            else:
                raise ValueError(f"by='{by}' não suportado em expandir_lista")

        primary = locator_spec
        fallback = {"by": "css", "selector": "a[id^='btn_mais_itens_cobrancas']"}

        clicks = 0
        used = None

        for tentativa in ("primary", "fallback"):
            spec = primary if tentativa == "primary" else fallback
            used = tentativa

            # Tenta esperar o botão ficar visível antes de começar o loop de cliques
            try:
                _mk(spec).wait_for(state="visible", timeout=timeout)
            except:
                pass  # Se não aparecer, tenta o fallback ou segue

            while clicks < max_clicks:
                loc = _mk(spec)
                try:
                    if not loc.is_visible():
                        break
                    try:
                        loc.scroll_into_view_if_needed(timeout=300)
                    except Exception:
                        pass
                    loc.click(timeout=timeout)
                    clicks += 1
                    self.page.wait_for_timeout(settle_ms)
                except Exception:
                    # tenta dar um scroll no fim e rechecar uma vez
                    try:
                        self.page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
                        self.page.wait_for_timeout(250)
                        if not _mk(spec).is_visible():
                            break
                    except Exception:
                        break
            if clicks:
                break

        log(f"[expandir_lista] cliques={clicks} via={used}")

    def _verificar_erro_login(self, locator_spec: dict, timeout: int = 3000,
                              deve_parar: bool = True):
        """
        Verifica se existe elemento de erro de login.

        Se deve_parar=True e erro encontrado:
        - Lança ValueError imediatamente
        - Para execução do fluxo
        - FAZ LOGOUT para limpar estado (NOVO!)
        """
        try:
            loc = self._locator_from_json(locator_spec).first
            loc.wait_for(state="visible", timeout=timeout)

            # Elemento de ERRO encontrado!
            try:
                texto_erro = loc.inner_text().strip()
                # Substitui quebras de linha por espaço para evitar textos grudados
                texto_erro = texto_erro.replace('\n', ' ').replace('\r', ' ')
                # Remove espaços múltiplos
                texto_erro = re.sub(r' +', ' ', texto_erro).strip()
            except:
                texto_erro = "Erro de login detectado"

            # Registra erro
            self.mensagemerro = texto_erro
            self.resposta = texto_erro

            log(f"❌ [ERRO LOGIN DETECTADO] {texto_erro}")

            # Se deve_parar=True, lança exceção imediatamente
            if deve_parar:
                self.login_ok = False
                self.erros_log.append(f"[LOGIN] {texto_erro}")

                # ===== FORÇA LOGOUT PARA LIMPAR ESTADO ===== ↓
                log(f"🔄 [LIMPANDO ESTADO] Fazendo logout após erro de login...")
                try:
                    # Tenta limpar cookies e storage
                    self.page.context.clear_cookies()
                    self.page.evaluate("() => { localStorage.clear(); sessionStorage.clear(); }")

                    # Aguarda um pouco
                    self.page.wait_for_timeout(500)

                    log(f"✅ [ESTADO LIMPO] Logout realizado")
                except Exception as e:
                    log(f"⚠️ [LOGOUT] Erro ao limpar estado: {e}")
                # ===== FIM ===== ↑

                raise ValueError(f"Erro de login: {texto_erro}")

            # Quando OK:
            if self.login_ok is None:
                self.login_ok = True
            return True, texto_erro

        except ValueError:
            # Re-lança ValueError (é o erro de login que queremos propagar)
            raise

        except Exception:
            # Elemento NÃO encontrado (TimeoutError ou outro)
            # Isso é NORMAL - significa login OK
            if self.login_ok is None:
                self.login_ok = True
            log(f"✅ [LOGIN OK] Nenhum erro detectado")
            return False, None

    def _fuzzy_lista_fechada(self, alvo: str, container_spec: dict, items_css: str = "a",
                             toggle_spec: dict | None = None, cutoff: float = 0.80, timeout=10000):
        cont = self._locator_from_json(container_spec).first
        cont.wait_for(state="attached", timeout=timeout)

        # Espera o primeiro item aparecer
        cont.locator(items_css).first.wait_for(state="visible", timeout=10000)
        self.page.wait_for_timeout(1000)

        # Extrai textos via JavaScript
        textos = cont.evaluate(
            f"root => Array.from(root.querySelectorAll('{items_css}')).map(a => (a.textContent||'').trim())"
        )
        if not textos:
            raise RuntimeError(f"Container não possui itens com '{items_css}'.")

        # ===== PREPARA DADOS PARA O MATCHER ===== ↓
        items_data = [
            {'idx': i, 'text': t, 'element': None}  # element será resolvido depois
            for i, t in enumerate(textos)
        ]
        # ===== FIM ===== ↑

        # ===== USA FUNÇÃO REUTILIZÁVEL ===== ↓
        try:
            result = self._fuzzy_match_items(
                items_data=items_data,
                alvo=alvo,
                cutoff=cutoff,
                validar_numero=False,
                is_condo=True
            )

            best_i = result['winners'][0]['idx']
            best_txt = result['best_txt']

        except ValueError as e:
            raise ValueError(f"'{alvo}' não atingiu cutoff {cutoff:.2f}: {e}")
        # ===== FIM ===== ↑

        # ===== CLICA NO ELEMENTO (se tem toggle) ===== ↓
        if toggle_spec:
            tog = self._locator_from_json(toggle_spec).first
            tog.wait_for(state="visible", timeout=timeout)
            tog.click()

            # Aguarda menu abrir
            cont.wait_for(state="visible", timeout=timeout)
            self.page.wait_for_load_state("domcontentloaded", timeout=3000)

            # Pega elemento real pelo índice
            elemento = cont.locator(items_css).nth(best_i)

            # Aguarda ser visível
            try:
                elemento.wait_for(state="visible", timeout=5000)
            except:
                pass

            # Scroll
            try:
                elemento.scroll_into_view_if_needed(timeout=2000)
            except:
                pass

            self.page.wait_for_timeout(500)

            # ===== CLIQUE ROBUSTO (3 ESTRATÉGIAS) ===== ↓
            try:
                elemento.click(timeout=5000)
                log(f"  ✅ Clicou em '{best_txt}' (clique normal)")
            except Exception as e1:
                log(f"  ⚠️  Clique normal falhou: {e1}")
                try:
                    elemento.click(force=True, timeout=5000)
                    log(f"  ✅ Clicou em '{best_txt}' (force click)")
                except Exception as e2:
                    log(f"  ⚠️  Force click falhou: {e2}")
                    try:
                        elemento.evaluate("el => el.click()")
                        self.page.wait_for_timeout(800)
                        log(f"  ✅ Clicou em '{best_txt}' (JavaScript)")
                    except Exception as e3:
                        log(f"  ❌ Todas as 3 estratégias falharam!")
                        raise ValueError(f"Não conseguiu clicar em '{best_txt}': {e3}")
            # ===== FIM CLIQUE ROBUSTO ===== ↑
        # ===== FIM ===== ↑

        return best_txt  # Retorna texto vencedor

    # ---------- loop (fuzzy best com empate) ----------
    def _loop_apartamentos_fuzzy_best(
            self,
            alvo_apartamento: str,
            container_spec: dict,
            item_css: str,
            visualizar_spec: dict,
            pagar_spec: dict | None,
            imprimir_spec: dict,
            identificador_base: str,
            cutoff: float = 0.75,
            tie_tol: float = 1e-6,
            timeout: int = 20000,
            apto_parts: list[str] | None = None,
            click_target: str | None = None,
            download_priority: list | None = None,
    ):
        cont = self._locator_from_json(container_spec).first
        cont.wait_for(state="attached", timeout=timeout)

        items = cont.evaluate(
            """(root, opts) => {
                const { itemSel, parts } = opts;
                const result = [];
                const list = root.querySelectorAll(itemSel);
                for (let i = 0; i < list.length; i++) {
                  const el = list[i];
                  const texts = (Array.isArray(parts) && parts.length)
                    ? parts.map(sel => {
                        const n = el.querySelector(sel);
                        return (n && (n.textContent || '').trim()) || '';
                      }).filter(Boolean)
                    : [ (el.textContent || '').trim() ];
                  result.push({
                    idx: i,
                    id: el.id || '',
                    parts: texts,
                    combined: texts.join(' ').trim(),
                    full: (el.textContent || '').trim()
                  });
                }
                return result;
            }""",
            {"itemSel": item_css, "parts": (apto_parts or [])},
        )

        # Dedup por id/combined
        uniq = {}
        for it in items:
            key = it["id"] or it["combined"] or it["full"]
            if key not in uniq:
                uniq[key] = it
        items = list(uniq.values())
        if not items:
            raise TimeoutError("Nenhum item encontrado para o apartamento.")

        alvo = (_apto_from_item_text(alvo_apartamento) or "").strip()

        # ===== PREPARA DADOS PARA O MATCHER ===== ↓
        items_data = [
            {
                'idx': it['idx'],
                'text': (_apto_from_item_text(it['combined']) or "").strip(),
                'element': it  # Guarda o dict completo
            }
            for it in items
        ]
        # ===== FIM ===== ↑

        # ===== USA FUNÇÃO REUTILIZÁVEL ===== ↓
        try:
            result = self._fuzzy_match_items(
                items_data=items_data,
                alvo=alvo,
                cutoff=cutoff,
                validar_numero=True,  # Apartamentos PRECISAM validar
                tie_tol=tie_tol
            )

            # Pega winners (contêm o dict original do item)
            winners = [w['element'] for w in result['winners']]

        except ValueError as e:
            raise TimeoutError(f"Apartamento não encontrado: {e}")
        # ===== FIM ===== ↑

        # Dedup winners
        seen = set()
        wuniq = []
        for w in winners:
            key = w["id"] or w["combined"] or w["full"]
            if key in seen:
                continue
            seen.add(key)
            wuniq.append(w)
        winners = wuniq

        if download_priority is None:
            download_priority = [imprimir_spec, pagar_spec]

        def _click_stable(el):
            try:
                el.scroll_into_view_if_needed()
            except Exception:
                pass
            try:
                (el.locator(click_target).first if click_target else el).click(timeout=1500)
                return True
            except Exception:
                pass
            try:
                (el.locator(click_target).first if click_target else el).click(force=True, timeout=1000)
                return True
            except Exception:
                pass
            try:
                (el.locator(click_target).first if click_target else el).evaluate("e => e.click()")
                return True
            except Exception:
                pass
            try:
                box = (el.locator(click_target).first if click_target else el).bounding_box()
                if box:
                    self.page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
                    return True
            except Exception:
                pass
            return False

        for k, w in enumerate(winners):
            popup = None

            try:
                if w["id"]:
                    el = cont.locator(f"{item_css}[id='{w['id']}']").first
                else:
                    el = cont.locator(item_css).nth(w["idx"])

                try:
                    cont.evaluate(
                        """(root, sel, id, idx) => {
                            const list = root.querySelectorAll(sel);
                            const t = id ? root.querySelector(sel+'#'+CSS.escape(id)) : list[idx];
                            if (t) t.scrollIntoView({block:'center'});
                        }""",
                        item_css,
                        w["id"],
                        w["idx"],
                    )
                except Exception:
                    pass

                if not _click_stable(el):
                    raise TimeoutError(f"Não consegui clicar no item {w['id'] or w['idx']}.")

                popup = self._click_expect_popup_atomic(visualizar_spec, popup_timeout=timeout)

                nome = f"{identificador_base}.pdf" if k == 0 else f"{identificador_base}_{k}.pdf"
                destino = DOWNLOADS_DIR / nome
                self._baixar_com_prioridade(popup, download_priority, destino=destino, timeout_ms=timeout)
                log(f"[OK] Baixado: {destino}")

                # ===== ADICIONA CABEÇALHO (já remove senha internamente!) ===== ↓
                try:
                    from auxiliares import utils as aux
                    codigo = identificador_base[:4] if len(identificador_base) >= 4 else identificador_base
                    aux.adicionarcabecalhopdf_topo_adaptativo(str(destino), str(destino), codigo,
                                                              senha_base=self.cpf_senha)
                    log(f"✅ [CABEÇALHO] Código {codigo} adicionado")
                except Exception as e:
                    log(f"⚠️ [CABEÇALHO] Erro: {e}")
                # ===== FIM ===== ↑

                # ===== PROCESSA BOLETO E MOVE PARA PASTA DO MÊS ===== ↓
                try:
                    from core import boletos  # ← IMPORTA O MÓDULO CORRETO
                    dados_boleto = boletos.processar_boleto_baixado(
                        caminho_pdf=destino,
                        codigo_cliente=self.codigo_cliente,
                        pasta_downloads=DOWNLOADS_DIR
                    )
                    if dados_boleto:
                        # ===== VERIFICA DUPLICATA ===== ↓
                        codigo_barras = dados_boleto.get('Cod_Barras', '')
                        if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                            # Boleto duplicado - conta como baixado mas registra como descartado
                            self.duplicados += 1
                            log(f"⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                            registrar_boleto_duplicado(
                                self.codigo_cliente,
                                codigo_barras,
                                dados_boleto.get('Caminho_Completo', str(destino))
                            )
                            # Remove o arquivo duplicado
                            try:
                                caminho_duplicado = Path(dados_boleto.get('Caminho_Completo', str(destino)))
                                if caminho_duplicado.exists():
                                    caminho_duplicado.unlink()
                                    log(f"🗑️ [DUPLICADO] Arquivo removido: {caminho_duplicado.name}")
                            except Exception as del_err:
                                log(f"⚠️ [DUPLICADO] Erro ao remover arquivo: {del_err}")
                        else:
                            # Boleto novo - adiciona ao cache e ao DataFrame
                            if codigo_barras:
                                registrar_codigo_barras(codigo_barras)
                            self.boletos_processados.append(dados_boleto)
                            log(f"✅ [DATAFRAME] Boleto adicionado à tabela")
                        # ===== FIM VERIFICA DUPLICATA ===== ↑
                except Exception as e:
                    log(f"⚠️ [DATAFRAME] Erro ao processar: {e}")

            except ValueError as e:  # ← CAPTURA ERROS DE NEGÓCIO
                log(f"⚠️ [ERRO CAPTURADO NO LOOP] {e}")
                # Não re-lança, continua próximo apartamento

            finally:
                if popup:
                    try:
                        popup.close()
                        log(f"🔒 [POPUP/ABA FECHADO]")
                    except Exception:
                        pass

    # ---------- loop simples (por name exato) ----------
    def _loop_apartamentos(self, apto_texto: str,
                           item_spec: dict,
                           popup_spec: dict | None,
                           pagar_spec: dict | None,
                           imprimir_spec: dict,
                           identificador_base: str):
        loc = self._locator_from_json(item_spec)
        loc.first.wait_for(state="visible", timeout=15000)
        total = loc.count()
        log(f"[INFO] Encontrados {total} itens para '{apto_texto}'.")
        for i in range(total):
            loc = self._locator_from_json(item_spec)
            if loc.count() <= i:
                break
            item = loc.nth(i)
            try:
                item.scroll_into_view_if_needed()
            except Exception:
                pass
            item.click()
            if popup_spec:
                popup = self._click_expect_popup_atomic(popup_spec, popup_timeout=20000)
                destino = DOWNLOADS_DIR / (f"{identificador_base}.pdf" if i == 0 else f"{identificador_base}_{i}.pdf")
                self._baixar_com_prioridade(popup, [imprimir_spec, pagar_spec], destino=destino, timeout_ms=20000)
                try:
                    popup.close()
                except Exception:
                    pass

    def _loop_apartamentos_dinamico(self, item_spec_links: dict, item_spec_buttons: dict, identificador_base: str,
                                    stop_if_not_found: bool = False):
        """
        Loop dinâmico: tenta baixar via LINKS primeiro, se não encontrar tenta via BOTÕES.
        Se falhar em ambos, pode sinalizar SKIP_FLOW para evitar timeouts seguintes.
        """
        # ===== 0. VERIFICA SE HÁ MENSAGEM DE ERRO PRIMEIRO ===== ↓
        try:
            # Tenta verificar se existe mensagem de erro/indisponibilidade
            if "frame" in item_spec_links:
                frame_loc = self.page.frame_locator(item_spec_links["frame"]["selector"])
                msg_erro = frame_loc.locator("#mensagem-indisponivel, .v-alert__text, .mensagem-erro").first
            else:
                msg_erro = self.page.locator("#mensagem-indisponivel, .v-alert__text, .mensagem-erro").first

            # Espera até 2 segundos para ver se a mensagem aparece
            msg_erro.wait_for(state="attached", timeout=2000)

            # Se chegou aqui, a mensagem existe - captura o texto
            texto_erro = ""
            try:
                texto_erro = (msg_erro.text_content() or "").strip()
                if not texto_erro:
                    texto_erro = (msg_erro.inner_text() or "").strip()

                # Trata quebras de linha para evitar textos grudados
                texto_erro = texto_erro.replace('\n', ' ').replace('\r', ' ')
                # Remove espaços duplos
                while '  ' in texto_erro:
                    texto_erro = texto_erro.replace('  ', ' ')
                texto_erro = texto_erro.strip()
            except:
                texto_erro = "Mensagem de indisponibilidade detectada"

            log(f"⚠️  [MENSAGEM DE ERRO DETECTADA] {texto_erro}")

            # Se tem mensagem de erro, não tenta baixar
            if stop_if_not_found:
                self.resposta = texto_erro
                raise StopIteration("SKIP_FLOW")
            else:
                raise ValueError(texto_erro)

        except TimeoutError:
            # Mensagem de erro não apareceu - prossegue normalmente
            pass
        except (StopIteration, ValueError):
            # Re-lança nossas próprias exceções
            raise
        except Exception as e:
            # Outros erros - apenas loga e continua
            log(f"⚠️  Erro ao verificar mensagem: {type(e).__name__}")
            pass

        # ===== 1. TENTA LINKS PRIMEIRO ===== ↓
        try:
            log(f"🔍 [BUSCANDO] Procurando por links de boletos...")
            loc_links = self._locator_from_json(item_spec_links)
            loc_links.first.wait_for(state="visible", timeout=3000)
            total_links = loc_links.count()

            if total_links > 0:
                log(f"✅ [RESULTADO] Encontrados {total_links} link(s)")
                for i in range(total_links):
                    try:
                        loc = self._locator_from_json(item_spec_links).nth(i)
                        destino = DOWNLOADS_DIR / (
                            f"{identificador_base}.pdf" if i == 0 else f"{identificador_base}_{i}.pdf")

                        with self.page.expect_download(timeout=20000) as d_info:
                            loc.click()

                        d = d_info.value
                        d.save_as(str(destino))
                        self.baixados += 1
                        log(f"[OK] Baixado: {destino.name}")

                        # Processamento (Cabeçalho e Dataframe)
                        try:
                            from auxiliares import utils as aux
                            codigo = identificador_base[:4] if len(identificador_base) >= 4 else identificador_base
                            aux.adicionarcabecalhopdf_topo_adaptativo(
                                pdf_entrada=str(destino),
                                pdf_saida=str(destino),
                                codigo=codigo,
                                senha_base=self.cpf_senha,
                                margem=3
                            )
                        except:
                            pass

                        try:
                            from core import boletos
                            dados = boletos.processar_boleto_baixado(destino, self.codigo_cliente, DOWNLOADS_DIR)
                            if dados:
                                # Verifica duplicata
                                codigo_barras = dados.get('Cod_Barras', '')
                                if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                                    self.duplicados += 1
                                    log(f"⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                                    registrar_boleto_duplicado(self.codigo_cliente, codigo_barras, dados.get('Caminho_Completo', ''))
                                    try:
                                        Path(dados.get('Caminho_Completo', '')).unlink(missing_ok=True)
                                    except: pass
                                else:
                                    if codigo_barras: registrar_codigo_barras(codigo_barras)
                                    self.boletos_processados.append(dados)
                        except:
                            pass

                    except Exception as e:
                        log(f"⚠️ Erro ao baixar link {i + 1}: {e}")
                return  # Sucesso - sai da função

        except Exception:
            log(f"ℹ️  [INFO] Links não encontrados - tentando botões...")

        # ===== 2. TENTA BOTÕES SE NÃO ACHOU LINKS ===== ↓
        try:
            log(f"🔍 [BUSCANDO] Procurando por botões 'GERAR BOLETO'...")
            loc_buttons = self._locator_from_json(item_spec_buttons)
            loc_buttons.first.wait_for(state="visible", timeout=5000)
            total_buttons = loc_buttons.count()

            if total_buttons > 0:
                log(f"✅ [RESULTADO] Encontrados {total_buttons} botão(ões)")
                for i in range(total_buttons):
                    try:
                        loc = self._locator_from_json(item_spec_buttons).nth(i)
                        destino = DOWNLOADS_DIR / (
                            f"{identificador_base}.pdf" if i == 0 else f"{identificador_base}_{i}.pdf")

                        with self.page.expect_download(timeout=20000) as d_info:
                            loc.click()

                        d = d_info.value
                        d.save_as(str(destino))
                        self.baixados += 1
                        log(f"[OK] Baixado: {destino.name}")

                        # Processamento igual ao bloco de links...
                        try:
                            from auxiliares import utils as aux
                            codigo = identificador_base[:4] if len(identificador_base) >= 4 else identificador_base
                            aux.adicionarcabecalhopdf_topo_adaptativo(str(destino), str(destino), codigo,
                                                                      senha_base=self.cpf_senha)
                        except:
                            pass

                        try:
                            from core import boletos
                            dados = boletos.processar_boleto_baixado(destino, self.codigo_cliente, DOWNLOADS_DIR)
                            if dados:
                                # Verifica duplicata
                                codigo_barras = dados.get('Cod_Barras', '')
                                if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                                    self.duplicados += 1
                                    log(f"⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                                    registrar_boleto_duplicado(self.codigo_cliente, codigo_barras, dados.get('Caminho_Completo', ''))
                                    try:
                                        Path(dados.get('Caminho_Completo', '')).unlink(missing_ok=True)
                                    except: pass
                                else:
                                    if codigo_barras: registrar_codigo_barras(codigo_barras)
                                    self.boletos_processados.append(dados)
                        except:
                            pass

                    except Exception as e:
                        log(f"⚠️ Erro ao baixar botão {i + 1}: {e}")
                return  # Sucesso - sai da função

        except Exception:
            pass  # Segue para o tratamento de falha final

        # ===== 3. TRATAMENTO DE FALHA FINAL (SINALIZADOR) ===== ↓
        # Primeiro tenta capturar mensagem de erro da página (se houver)
        msg = None
        try:
            # Procura por elementos comuns de erro/aviso
            erro_locators = [
                ".alert-danger", ".alert-warning", ".mensagem-erro", ".erro",
                "#mensagem-indisponivel", ".v-alert__text", ".alerta",
                "[class*='erro']", "[class*='alert']", "[class*='aviso']"
            ]

            for selector in erro_locators:
                try:
                    elem_erro = self.page.locator(selector).first
                    elem_erro.wait_for(state="visible", timeout=1000)
                    texto = (elem_erro.inner_text() or "").strip()
                    if texto:
                        # Trata quebras de linha
                        texto = texto.replace('\n', ' ').replace('\r', ' ')
                        texto = re.sub(r' +', ' ', texto).strip()
                        if len(texto) > 0 and len(texto) < 500:  # Ignora textos muito longos
                            msg = texto
                            break
                except:
                    continue
        except:
            pass

        # Se não encontrou mensagem de erro, verifica se existe tabela de boletos
        if not msg:
            try:
                tabela_boletos = self.page.locator("table.table, table.table-striped, tbody tr").first
                tabela_boletos.wait_for(state="attached", timeout=2000)
                # Se chegou aqui, existe tabela mas sem botões de download
                msg = "Boletos existem mas download não está disponível (podem estar vencidos ou bloqueados pelo sistema)"
            except:
                # Não tem tabela - realmente não há boletos
                msg = "Nenhum boleto encontrado (links e botões ausentes)"

        log(f"⚠️  [AVISO] {msg}")

        if stop_if_not_found:
            self.resposta = msg
            # DISPARA O SINALIZADOR para o método executar() pular o restante do fluxo
            raise StopIteration("SKIP_FLOW")
        else:
            raise ValueError(msg)

    def _download_recibos_icondo(self, partial_text: str, identificador: str,
                                 max_boletos: int = 50, timeout: int = 120000):
        """
        Baixa todos os boletos com texto parcial "Recibo No." (padrão iCondo).

        Baseado na função ICondo() do condominios.py (linhas 912-933):
        - boletos = site.verificarobjetoexiste('PARTIAL_LINK_TEXT', 'Recibo No.', itemunico=False)
        - Clica em cada link e aguarda download
        - Renomeia com identificador + contador

        Args:
            partial_text: Texto parcial do link (ex: "Recibo No.")
            identificador: Base para nome do arquivo (ex: "0001")
            max_boletos: Limite máximo de boletos
            timeout: Timeout total em ms
        """

        try:
            log(f"[download_recibos_icondo] Procurando links '{partial_text}'...")

            # Procurar todos os links com texto parcial
            # Equivalente ao Selenium: verificarobjetoexiste('PARTIAL_LINK_TEXT', 'Recibo No.', itemunico=False)
            links = self.page.get_by_role("link", name=re.compile(re.escape(partial_text), re.I)).all()

            if not links:
                # Tenta com locator genérico
                links = self.page.locator(f"a:has-text('{partial_text}')").all()

            if not links:
                raise ValueError(f"Nenhum link com texto '{partial_text}' encontrado")

            total_encontrados = len(links)
            log(f"[INFO] {total_encontrados} link(s) '{partial_text}' encontrado(s)")

            if total_encontrados > max_boletos:
                log(f"⚠️ Limitando a {max_boletos} boletos")
                links = links[:max_boletos]

            baixados = 0
            falhas = 0

            # Loop equivalente ao Selenium (condominios.py linhas 914-933)
            for i in range(len(links)):
                try:
                    # Re-obter os links a cada iteração (como no Selenium linha 932)
                    # Isso evita StaleElementReferenceError
                    current_links = self.page.get_by_role("link", name=re.compile(re.escape(partial_text), re.I)).all()
                    if not current_links:
                        current_links = self.page.locator(f"a:has-text('{partial_text}')").all()

                    if i >= len(current_links):
                        log(f"⚠️ Link {i + 1} não existe mais, parando")
                        break

                    link = current_links[i]

                    log(f"Baixando boleto {i + 1}/{len(links)}...")

                    # Define nome do arquivo (condominios.py linhas 919-923)
                    if baixados > 0:
                        nome_arquivo = f"{identificador}_{baixados}.pdf"
                    else:
                        nome_arquivo = f"{identificador}.pdf"

                    caminho_destino = DOWNLOADS_DIR / nome_arquivo

                    # Aguarda download (Playwright equivalente ao monitorar_downloads_sem_href)
                    with self.page.expect_download(timeout=timeout) as download_info:
                        link.click()

                    download = download_info.value

                    # Salva arquivo
                    download.save_as(str(caminho_destino))

                    # Verifica se salvou
                    if caminho_destino.exists():
                        baixados += 1
                        log(f"✓ Boleto {i + 1} salvo: {nome_arquivo}")

                        # Adiciona cabeçalho PDF se header_fn estiver disponível
                        # (condominios.py linhas 924-929)
                        if hasattr(self, 'header_fn') and self.header_fn:
                            try:
                                codigo = identificador[:4] if len(identificador) >= 4 else identificador
                                self.header_fn(str(caminho_destino), str(caminho_destino), codigo)
                            except Exception as e:
                                log(f"⚠️ Erro ao adicionar cabeçalho PDF: {e}")
                    else:
                        falhas += 1
                        log(f"✗ Falha ao salvar boleto {i + 1}")

                    # Pequena pausa entre downloads (como no Selenium)
                    time.sleep(0.5)

                except Exception as e:
                    falhas += 1
                    log(f"✗ Erro ao baixar boleto {i + 1}: {e}")
                    continue

            # Atualiza contador de baixados (usado no get_status)
            self.baixados = baixados

            # Define resposta
            if baixados > 0:
                msg = f"{baixados} boleto{'s' if baixados > 1 else ''} baixado{'s' if baixados > 1 else ''}"

                if falhas > 0:
                    msg += f" ({falhas} falha(s))"
                log(f"[INFO] {msg}")
            else:
                msg = f"Nenhum boleto baixado ({falhas} falha{'s' if falhas > 1 else ''})"
                log(f"[ERRO] {msg}")
                raise ValueError(msg)

        except Exception as e:
            error_msg = f"Erro ao baixar recibos iCondo: {str(e)}"
            log(f"✗ {error_msg}")
            raise ValueError(error_msg)

    # ---------- NOVA AÇÃO: Loop para Boletos em Popup ----------
    def _loop_apartamentos_popup(self, item_spec, identificador_base):
        """
        Clica em elementos que abrem boletos em Nova Janela (window.open).
        Salva o conteúdo da janela como PDF.
        """
        log(f"🔍 [POPUP MODE] Procurando boletos...")

        # 1. Tenta localizar os elementos
        try:
            loc = self._locator_from_json(item_spec)
            # Espera um pouco para garantir que a lista carregou
            loc.first.wait_for(state="attached", timeout=5000)
            total = loc.count()
        except:
            total = 0

        if total == 0:
            log("⚠️ Nenhum boleto encontrado para clicar.")
            return

        log(f"✅ Encontrados {total} boleto(s).")

        for i in range(total):
            try:
                # Recarrega o locator a cada volta para evitar erro de elemento stale
                loc = self._locator_from_json(item_spec)
                elemento = loc.nth(i)

                # Scroll para garantir visibilidade
                elemento.scroll_into_view_if_needed()

                # Prepara o nome do arquivo
                nome_arq = f"{identificador_base}.pdf" if i == 0 else f"{identificador_base}_{i}.pdf"
                destino = DOWNLOADS_DIR / nome_arq

                log(f"   🔄 Processando boleto {i + 1}...")

                # CLICA E ESPERA A NOVA PÁGINA (POPUP)
                # Aumentamos o timeout para 30s pois o popup pode demorar
                with self.page.expect_popup(timeout=30000) as p_info:
                    # Força o clique se necessário
                    elemento.click(force=True)

                popup = p_info.value
                popup.wait_for_load_state("networkidle")  # Espera carregar o boleto visual

                # SALVA COMO PDF
                popup.pdf(path=str(destino))
                log(f"   📥 Salvo como PDF: {nome_arq}")

                # Fecha a janelinha do boleto
                popup.close()

                self.baixados += 1

                # Processamento extra (Cabeçalho/Dataframe)
                try:
                    from auxiliares import utils as aux
                    from core import boletos
                    codigo = self.codigo_cliente[:4]

                    # Adiciona cabeçalho
                    aux.adicionarcabecalhopdf_topo_adaptativo(str(destino), str(destino), codigo,
                                                              senha_base=self.cpf_senha)

                    # Adiciona ao relatório
                    dados = boletos.processar_boleto_baixado(destino, self.codigo_cliente, DOWNLOADS_DIR)
                    if dados:
                        # Verifica duplicata
                        codigo_barras = dados.get('Cod_Barras', '')
                        if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                            self.duplicados += 1
                            log(f"⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                            registrar_boleto_duplicado(self.codigo_cliente, codigo_barras, dados.get('Caminho_Completo', ''))
                            try:
                                Path(dados.get('Caminho_Completo', '')).unlink(missing_ok=True)
                            except: pass
                        else:
                            if codigo_barras: registrar_codigo_barras(codigo_barras)
                            self.boletos_processados.append(dados)

                except Exception as e_proc:
                    log(f"   ⚠️ Erro processamento pós-download: {e_proc}")

            except Exception as e:
                log(f"   ❌ Erro ao processar boleto {i + 1}: {e}")

    def _fuzzy_match_items(self, items_data: list[dict], alvo: str,
                           cutoff: float = 0.75, validar_numero: bool = False,
                           is_condo: bool = False, tie_tol: float = 1e-6) -> dict:
        """
        Wrapper que chama a função pura E adiciona logs ao self.match_log.
        """
        result = fuzzy_match_items_core(
            items_data=items_data,
            alvo=alvo,
            cutoff=cutoff,
            validar_numero=validar_numero,
            is_condo=is_condo,
            tie_tol=tie_tol
        )

        # Adiciona ao log da classe para aparecer no Excel/Console
        for line in result['log_lines']:
            log(line)
        self.match_log.append("\n".join(result['log_lines']))

        return result

    # ---------- dispatcher ----------
    def _run_step(self, step: dict, context: dict):
        action = step.get("action")
        step = {k: (_resolve_placeholders(v, context) if isinstance(v, str) else v) for k, v in step.items()}

        if action == "goto":
            url = step["value"]  # ← ERRO! ${SITE} não é resolvido!
            _navegar_e_esperar(self.page, url)

        elif action == "fill":
            loc = self._locator_from_json(step["locator"]).first
            loc.wait_for(state="visible", timeout=int(step.get("timeout", 10000)))
            loc.fill(str(step["value"]))

        elif action == "click":
            if bool(step.get("js", False)):
                # Clica via JavaScript - bypassa menus, overlays e visibilidade CSS
                # Se o elemento tiver data-url, navega direto; senão chama .click()
                # Passa o seletor como argumento para evitar problemas de aspas no f-string
                locator_spec = step["locator"]
                selector = locator_spec.get("selector", "")
                self.page.evaluate(
                    """(sel) => {
                        const el = document.querySelector(sel);
                        if (!el) throw new Error('Elemento não encontrado: ' + sel);
                        const dataUrl = el.getAttribute('data-url');
                        if (dataUrl) window.location.href = dataUrl;
                        else el.click();
                    }""",
                    selector
                )
                self.page.wait_for_load_state("networkidle", timeout=int(step.get("timeout", 15000)))
            else:
                loc = self._locator_from_json(step["locator"]).first
                loc.wait_for(state="visible", timeout=int(step.get("timeout", 10000)))
                loc.click(force=bool(step.get("force", False)))

        elif action == "hover":
            loc = self._locator_from_json(step["locator"]).first
            loc.wait_for(state="visible", timeout=int(step.get("timeout", 10000)))
            loc.hover()
            self.page.wait_for_timeout(int(step.get("wait_after", 500)))

        elif action == "click_expect_popup":
            popup = self._click_expect_popup_atomic(step["locator"], popup_timeout=int(step.get("timeout", 15000)))
            if step.get("switch_to_popup", True):
                self.page = popup

        elif action == "verificar_elemento":
            loc = self._locator_from_json(step["locator"]).first
            try:
                loc.wait_for(state="visible", timeout=int(step.get("timeout", 5000)))
                texto = (loc.inner_text() or "").strip()
                # Substitui quebras de linha por espaço
                texto = texto.replace('\n', ' ').replace('\r', ' ')
                self.mensagemerro = re.sub(r' +', ' ', texto).strip()
                self.resposta = f"[verificar_elemento] '{self.mensagemerro}'"
                if step.get("stop_if_found", False):
                    raise ValueError(self.resposta)
            except TimeoutError:
                if step.get("stop_if_missing", False):
                    raise ValueError(f"[verificar_elemento] não encontrado.")

        elif action == "esperar_selector":
            self._locator_from_json(step["locator"]).first.wait_for(
                state=step.get("state", "visible"), timeout=int(step.get("timeout", 20000))
            )

        elif action == "remover_elemento":
            self._remover_elemento_bloqueante(step["selector"])

        elif action == "esperar_url":
            padrao = step["pattern"]
            self.page.wait_for_url(re.compile(padrao, re.I), timeout=int(step.get("timeout", 20000)))

        elif action == "expandir_lista":
            self._expandir_lista(
                locator_spec=step["locator"],
                timeout=int(step.get("timeout", 3000)),
                max_clicks=int(step.get("max_clicks", 500)),
                settle_ms=int(step.get("settle_ms", 400)),
            )

        elif action == "fuzzy_lista_fechada":
            alvo = step["value"]
            texto_selecionado = self._fuzzy_lista_fechada(
                alvo=alvo,
                container_spec=step["container"],
                items_css=step.get("items_css", "a"),
                toggle_spec=step.get("toggle"),
                cutoff=float(step.get("cutoff", 0.80)),
                timeout=int(step.get("timeout", 10000)),
                # debug=bool(step.get("debug", False)),
            )
            # Guardar para tracking
            self.condo_selecionado_site = texto_selecionado



        elif action == "parar_se_visivel":
            try:
                loc = self._locator_from_json(step["locator"]).first
                timeout_ms = int(step.get("timeout", 5000))

                try:
                    loc.wait_for(state="attached", timeout=timeout_ms)
                except:
                    loc.wait_for(state="visible", timeout=timeout_ms)

                # 🔍 CAPTURA DINÂMICA (MUDANÇA: Prioriza inner_text para manter <br>)
                texto_do_site = ""
                try:
                    # inner_text() no Playwright preserva as quebras de linha do layout
                    texto_do_site = (loc.inner_text() or "").strip()
                    if not texto_do_site:
                        texto_do_site = (loc.text_content() or "").strip()

                    # Substitui quebras de linha por espaço para evitar textos grudados
                    texto_do_site = texto_do_site.replace('\n', ' ').replace('\r', ' ')
                    # Normaliza espaços múltiplos
                    texto_do_site = re.sub(r' +', ' ', texto_do_site).strip()

                except:
                    pass

                mensagem_final = texto_do_site if texto_do_site else step.get("message", "Aviso detectado.")
                log(f"⚠️  [PARAR_SE_VISIVEL] Mensagem capturada:\n{mensagem_final}")
                raise ValueError(mensagem_final)

            except TimeoutError:
                pass

            except ValueError as ve:
                raise ve

            except Exception as e:
                log(f"⚠️  [PARAR_SE_VISIVEL] Exceção ignorada: {type(e).__name__}: {e}")
                pass

        elif action == "parar_se_nao_visivel":
            try:
                loc = self._locator_from_json(step["locator"]).first
                timeout_ms = int(step.get("timeout", 5000))

                try:
                    # Tenta encontrar o elemento
                    loc.wait_for(state="attached", timeout=timeout_ms)
                    # Se encontrou, não faz nada (elemento existe, tudo ok)
                    log(f"✅ [PARAR_SE_NAO_VISIVEL] Elemento encontrado, continuando...")

                except TimeoutError:
                    # Não encontrou o elemento - isso é um erro!
                    # Captura o conteúdo da página para mostrar no erro
                    texto_pagina = ""
                    try:
                        conteudo = self.page.locator("#conteudo").first
                        texto_pagina = (conteudo.inner_text() or "").strip()
                        texto_pagina = texto_pagina.replace('\n', ' ').replace('\r', ' ')
                        texto_pagina = re.sub(r' +', ' ', texto_pagina).strip()
                    except:
                        texto_pagina = step.get("message", "Elemento não encontrado na página.")

                    mensagem_final = texto_pagina if texto_pagina else "Elemento esperado não foi encontrado."
                    log(f"⚠️  [PARAR_SE_NAO_VISIVEL] Elemento não encontrado:\n{mensagem_final}")
                    raise ValueError(mensagem_final)

            except ValueError as ve:
                raise ve

            except Exception as e:
                log(f"⚠️  [PARAR_SE_NAO_VISIVEL] Exceção: {type(e).__name__}: {e}")
                raise

        elif action == "loop_apartamentos":
            item_spec = step["item_spec"]
            if item_spec.get("by") == "role" and isinstance(item_spec.get("name"), str):
                item_spec["name"] = _resolve_placeholders(item_spec["name"], context)
            self._loop_apartamentos(
                apto_texto=context.get("APARTAMENTO", ""),
                item_spec=item_spec,
                popup_spec=step["visualizar_spec"],
                pagar_spec=step.get("pagar_spec"),
                imprimir_spec=step["imprimir_spec"],
                identificador_base=str(step.get("identificador", f"boleto_{context.get('APTO_SAFE', 'apto')}")),
            )

        elif action == "loop_apartamentos_fuzzy_best":
            self._loop_apartamentos_fuzzy_best(
                alvo_apartamento=step.get("alvo_apartamento", ""),
                container_spec=step["container"],
                item_css=step.get("items_css", "a"),
                visualizar_spec=step["visualizar_spec"],
                pagar_spec=step.get("pagar_spec"),
                imprimir_spec=step["imprimir_spec"],
                identificador_base=str(step.get("identificador", f"boleto_{context.get('APTO_SAFE', 'apto')}")),
                cutoff=float(step.get("cutoff", 0.75)),
                tie_tol=float(step.get("tie_tol", 0.0)),
                timeout=int(step.get("timeout", 20000)),
                apto_parts=step.get("apto_parts"),
                click_target=step.get("click_target"),
                download_priority=step.get("download_priority"),
            )

        elif action == "download_imprimir":
            destino = DOWNLOADS_DIR / str(step.get("arquivo", "boleto.pdf"))
            self._baixar_com_prioridade(self.page, [step["locator"]], destino=destino)

        elif action == "loop_apartamentos_dinamico":  # ← ADICIONE AQUI!
            self._loop_apartamentos_dinamico(
                item_spec_links=step["item_spec_links"],
                item_spec_buttons=step["item_spec_buttons"],
                identificador_base=str(step.get("identificador", f"boleto_{context.get('APTO_SAFE', 'apto')}")),
                stop_if_not_found=step.get("stop_flow_if_fail", False)
            )

        elif action == "loop_apartamentos_popup":
            self._loop_apartamentos_popup(
                item_spec=step["locator"],
                identificador_base=str(step.get("identificador", context.get("IDENTIFICADOR")))
            )

        elif action == "try_click_multi":
            """
            Tenta clicar testando múltiplos seletores.
            Para quando conseguir clicar no primeiro que encontrar.
            """
            locators = step["locators"]

            sucesso = False
            for idx, loc_spec in enumerate(locators):
                try:
                    loc = self._locator_from_json(loc_spec).first
                    loc.wait_for(state="visible", timeout=3000)
                    loc.click()
                    log(f"✅ [CLICK MULTI] Clicou usando seletor {idx + 1}/{len(locators)}")
                    sucesso = True
                    break  # ← Para no primeiro que funcionar
                except Exception as e:
                    log(f"⚠️ [CLICK MULTI] Seletor {idx + 1} falhou: {e}")
                    continue

            if not sucesso:
                raise ValueError("Nenhum dos seletores para clique funcionou")


        elif action == "try_click":

            """
            TENTA clicar em elemento. 
            MUDANÇA: Espera 'attached' para clicar mesmo se o menu estiver recolhido.
            """

            try:
                loc = self._locator_from_json(step["locator"]).first
                # Mudamos 'visible' para 'attached' para não travar no menu lateral da CIPA
                loc.wait_for(state="attached", timeout=int(step.get("timeout", 7000)))
                # O force=True (que já está no seu código) faz o clique ignorar se está escondido
                loc.click(force=True)
                log(f"✅ [CLIQUE] Elemento clicado com force=True")

            except Exception as e:
                log(f"ℹ️  [CLIQUE] Elemento não disponível - pulando...")
                # SE o sinalizador de parada estiver ativo no JSON, avisamos o executor
                if step.get("stop_flow_if_fail"):
                    raise StopIteration("SKIP_FLOW")

        elif action == "try_fuzzy_lista_fechada":
            """
            TENTA fazer fuzzy_lista_fechada, mas NÃO falha se não encontrar.
            """
            try:
                alvo = step["value"]
                alvo_apto = step.get("alvo_apartamento")
                container_spec = step["container"]
                items_css = step.get("items_css", "a")
                cutoff = float(step.get("cutoff", 0.95))
                timeout = int(step.get("timeout", 3000))
                text_selector = step.get("text_selector")
                click_selector = step.get("click_selector")

                log(f"🔍 [TRY FUZZY] Procurando items_css='{items_css}'")
                if text_selector:
                    log(f"   text_selector='{text_selector}'")
                if click_selector:
                    log(f"   click_selector='{click_selector}'")
                if alvo_apto:
                    log(f"   alvo_apartamento='{alvo_apto}'")

                cont = self._locator_from_json(container_spec).first
                cont.wait_for(state="attached", timeout=timeout)

                # ===== EXTRAI TÍTULO, APARTAMENTO E TEXTO COMPLETO ===== ↓
                if text_selector:
                    log(f"🔍 [TRY FUZZY] Executando JavaScript para extrair textos...")

                    result = cont.evaluate(
                        f"""root => {{
                            const items = root.querySelectorAll('{items_css}');
                            return Array.from(items).map(item => {{
                                const fullText = item.textContent.trim();
                                const titleEl = item.querySelector('{text_selector}');
                                const title = titleEl ? titleEl.textContent.trim() : '';

                                // Pega segundo <p> (apartamento)
                                const paras = item.querySelectorAll('p');
                                const apto = paras.length > 1 ? paras[1].textContent.trim() : '';

                                return {{
                                    fullText: fullText,
                                    title: title,
                                    apto: apto
                                }};
                            }});
                        }}"""
                    )

                    textos = [r['fullText'] for r in result]
                    titulos = [r['title'] for r in result]
                    aptos = [r['apto'] for r in result]
                else:
                    textos = cont.evaluate(
                        f"root => Array.from(root.querySelectorAll('{items_css}')).map(a => (a.textContent||'').trim())"
                    )
                    titulos = textos
                    aptos = [""] * len(textos)

                if not textos:
                    raise RuntimeError(f"Container não possui itens com '{items_css}'.")

                # ===== FUZZY MATCHING COM SCORES DETALHADOS ===== ↓
                alvo_norm = _norm(alvo)
                alvo_apto_norm = _norm(alvo_apto) if alvo_apto else None

                best_i = -1
                best_txt = ""
                best_titulo = ""
                best_apto = ""
                best_score = 0.0
                best_w_score = 0.0
                best_p_score = 0.0
                is_exact = False

                # Match exato (usando normalização robusta)
                alvo_norm_v2 = _norm_condo_v2(alvo)
                for i, t in enumerate(textos):
                    t_norm_v2 = _norm_condo_v2(t)

                    # PULA se o item estiver vazio após normalização (evita match falso de itens sem texto)
                    if not t_norm_v2:
                        continue

                    if alvo_apto_norm:
                        # Se tem apto, checa se o condomínio normalizado bate E o apto está no texto original
                        if alvo_norm_v2 == t_norm_v2 and alvo_apto_norm in _norm(t):
                            best_i = i
                            best_txt = t
                            best_titulo = titulos[i]
                            best_apto = aptos[i]
                            best_score = 1.0
                            best_w_score = 1.0
                            best_p_score = 1.0
                            is_exact = True
                            break
                    else:
                        if alvo_norm_v2 == t_norm_v2:
                            best_i = i
                            best_txt = t
                            best_titulo = titulos[i]
                            best_apto = aptos[i]
                            best_score = 1.0
                            best_w_score = 1.0
                            best_p_score = 1.0
                            is_exact = True
                            break

                # Fuzzy se não achou exato
                if best_i < 0:
                    # alvo_norm_v2 já definido acima
                    for i, t in enumerate(textos):
                        tn_v2 = _norm_condo_v2(t)
                        if not tn_v2: continue

                        w_score = fuzz.WRatio(tn_v2, alvo_norm_v2) / 100.0
                        p_score = fuzz.partial_ratio(tn_v2, alvo_norm_v2) / 100.0
                        score = max(w_score, p_score)

                        # Bonus se achar apartamento (opcional no condo)
                        if alvo_apto_norm and alvo_apto_norm in _norm(t):
                            score = min(score + 0.1, 1.0)

                        if score > best_score:
                            best_i = i
                            best_txt = t
                            best_titulo = titulos[i]  # Garante que pega o título correto
                            best_apto = aptos[i]
                            best_score = score
                            best_w_score = w_score
                            best_p_score = p_score

                # ===== LOG FORMATADO (IGUAL AO ORIGINAL!) ===== ↓
                status = "✅ Match exato" if is_exact else "🔍 Match parcial"
                lines = []
                lines.append(f"[FUZZY CONDOMÍNIO] {status} (score={best_score:.2f}):")
                lines.append(f"                   📋 Buscado na BASE:   '{alvo}' + '{alvo_apto or 'qualquer'}'")
                lines.append(f"                   🌐 Encontrado no SITE: '{best_titulo}' - {best_apto}")

                if not is_exact:
                    lines.append(f"                   📊 WRatio: {best_w_score:.2f} | partial_ratio: {best_p_score:.2f}")
                    metodo = "WRatio" if best_w_score >= cutoff else "max(WRatio, partial)"
                    lines.append(f"                   🎯 Método usado: {metodo}")
                else:
                    lines.append(f"                   🎯 Método usado: Comparação direta")

                # Imprime
                for line in lines:
                    log(line)

                # Tracking
                self.match_log.append("\n".join(lines))
                # ===== FIM LOG ===== ↑

                if best_i < 0 or best_score < cutoff:
                    raise ValueError(f"Score {best_score:.2f} < cutoff {cutoff:.2f}")

                # Segurança adicional: Se o título for vazio, algo deu errado na extração
                if not best_titulo and not best_txt:
                    raise ValueError("Item encontrado está vazio (erro de extração).")

                # Clica no elemento
                elemento = cont.locator(items_css).nth(best_i)

                if click_selector:
                    botao = elemento.locator(click_selector).first
                    botao.click(timeout=5000)
                    log(f"  ✅ Clicou no card correto (via {click_selector})")
                else:
                    elemento.click(timeout=5000)
                    log(f"  ✅ Clicou no card correto")

                self.condo_selecionado_site = f"{best_titulo} - {best_apto}"
                log(f"✅ [CARD SELECIONADO]")

            except Exception as e:
                log(f"⚠️ [TRY FUZZY] ERRO: {e}")
                log(f"ℹ️  [SELEÇÃO OPCIONAL] Não encontrou lista de condomínios")
                log(f"   → Continuando sem selecionar...")


            except Exception as e:
                log(f"⚠️ [TRY FUZZY] ERRO: {e}")
                log(f"ℹ️  [SELEÇÃO OPCIONAL] Não encontrou lista")
                log(f"   → Continuando...")

        elif action == "verificar_e_deslogar_livefacilities":
            """
            Verifica se já está logado no LiveFacilities.
            - Se estiver logado com o MESMO usuário: Define skip_mode="SKIP_LOGIN" para pular login
            - Se estiver logado com usuário DIFERENTE: Faz logout
            - Se não estiver logado: Continua normal
            """
            try:
                # Pega o usuário esperado do contexto (self.context_linha)
                if not self.context_linha:
                    log("⚠️  [LIVEFACILITIES] Contexto não disponível, pulando verificação")
                else:
                    usuario_esperado = (self.context_linha.get("USUARIO", "") or "").strip().lower()

                    # Verifica se está na página logada (se existe o link "SAIR")
                    link_sair = self.page.get_by_role("link", name="SAIR").first

                    try:
                        link_sair.wait_for(state="visible", timeout=3000)
                        # Está logado! Precisa verificar se é o usuário correto
                        log("ℹ️  [LIVEFACILITIES] Detectado login ativo, verificando usuário...")

                        # Tenta capturar o nome do usuário logado
                        # O LiveFacilities geralmente mostra o login no campo de busca ou no menu
                        usuario_logado = None
                        try:
                            # Tenta pegar do input de login (se estiver preenchido)
                            input_login = self.page.locator(
                                'input[name*="login" i], input[placeholder*="login" i]').first
                            usuario_logado = input_login.input_value(timeout=1000).strip().lower()
                        except:
                            pass

                        # Se não conseguiu pelo input, tenta por texto visível na página
                        if not usuario_logado:
                            try:
                                # Procura por elemento que mostre o usuário logado
                                texto_pagina = self.page.locator("body").inner_text(timeout=2000)
                                # Verifica se o usuário esperado aparece no texto
                                if usuario_esperado and usuario_esperado in texto_pagina.lower():
                                    usuario_logado = usuario_esperado
                            except:
                                pass

                        # Se conseguiu identificar o usuário E é o mesmo
                        if usuario_logado and usuario_esperado and usuario_esperado == usuario_logado:
                            log(f"✅ [LIVEFACILITIES] Já logado com usuário correto: {usuario_logado}")
                            log("⏭️  [LIVEFACILITIES] Pulando etapa de login...")
                            # Sinaliza para pular os passos de login
                            skip_mode = "SKIP_LOGIN"
                            return  # Retorna sem fazer logout

                        # Se não conseguiu identificar OU é usuário diferente → Faz logout
                        log(f"🚪 [LIVEFACILITIES] Fazendo logout (usuário: {usuario_logado or 'não identificado'} ≠ {usuario_esperado})...")
                        link_sair.click()

                        # Aguarda voltar para tela de login
                        self.page.wait_for_selector(
                            'input[type="text"][name*="login"], input[type="text"][placeholder*="login" i]',
                            timeout=5000)
                        log("✅ [LIVEFACILITIES] Logout realizado com sucesso")

                    except:
                        # Não está logado, tudo ok
                        log("ℹ️  [LIVEFACILITIES] Não há login ativo, prosseguindo...")
                        pass

            except Exception as e:
                log(f"⚠️  [LIVEFACILITIES] Erro ao verificar login: {e}")
                # Não interrompe o fluxo, continua tentando fazer login
                pass

        elif action == "check_already_logged_superlogica":
            """
            Verifica se já está logado no Superlógica.
            Se já estiver logado, pula os passos de login.
            """
            try:
                # Tenta detectar se já está na página de condomínios (já logado)
                # Superlógica mostra "Meus condomínios" ou elementos específicos quando logado
                logged_in_selector = self.page.locator(
                    "a[href*='condominio'], .navbar, #menu-principal, .condominio-card")

                # Aguarda um pouco para ver se está logado
                logged_in_selector.first.wait_for(state="visible", timeout=3000)

                # Se chegou aqui, está logado
                log(f"   ✅ [SUPERLOGICA] Já está logado - pulando login")

                # Usa o sistema de skip existente
                raise StopIteration("SKIP_LOGIN")

            except TimeoutError:
                # Não está logado, continua normalmente
                log(f"   🔓 [SUPERLOGICA] Não está logado - fazendo login")
                pass

        elif action == "verificar_erro_login":
            """
            Verifica erro de login.

            Parâmetros do step:
            - locator: Seletor do elemento de erro
            - timeout: Timeout em ms (padrão 3000)
            - stop_if_error: Se True (padrão), para quando encontrar erro

            Se erro encontrado e stop_if_error=True:
            - Lança ValueError imediatamente
            - Para execução do login
            - Grava mensagem em self.mensagemerro

            Se erro não encontrado:
            - Continua normalmente
            """

            deve_parar = step.get("stop_if_error", True)  # Padrão: PARAR se erro
            tem_erro, mensagem = self._verificar_erro_login(
                locator_spec=step["locator"],
                timeout=int(step.get("timeout", 3000)),
                deve_parar=deve_parar

            )

            # Se chegou aqui e tem erro, é porque deve_parar=False
            # Nesse caso, só registra mas não interrompe

            if tem_erro:
                log(f"⚠️ Erro detectado mas continuando (stop_if_error=False)")

        elif action == "try_fuzzy_livefacilities_unidade":
            """
            Faz fuzzy match de CONDOMÍNIO + BLOCO + UNIDADE no LiveFacilities.
            Precisa incluir condomínio porque pode ter múltiplas "LOJA 110" em condomínios diferentes!
            """
            try:
                # Conta quantas unidades existem
                tabelas = self.page.locator("table.documento").all()

                if len(tabelas) == 0:
                    raise ValueError("Nenhuma unidade encontrada")

                if len(tabelas) == 1:
                    log(f"ℹ️  [LIVEFACILITIES] Apenas 1 unidade - selecionando automaticamente")
                    try:
                        link = tabelas[0].locator("a[title='Alterar unidade']")
                        link.click()
                        self.page.wait_for_load_state("networkidle", timeout=10000)
                        self.page.wait_for_timeout(2000)
                        log(f"✅ [LIVEFACILITIES] Unidade única selecionada")
                    except Exception as e:
                        log(f"⚠️  [LIVEFACILITIES] Erro ao selecionar unidade única: {e}")
                    return

                # ===== MÚLTIPLAS UNIDADES ===== ↓
                alvo_condo = step["value"]  # ${CONDOMINIO} = "CENTRO COMERCIAL BARRA PLAZA"
                alvo_apto = step.get("apartamento", "")  # ${APARTAMENTO} = "LOJA  110"
                cutoff = float(step.get("cutoff", 0.60))  # Reduzido de 0.75 para 0.60

                # log(f"🔍 [LIVEFACILITIES] {len(tabelas)} unidades encontradas")
                #
                # Extrai info de cada tabela
                opcoes = []
                for i, tabela in enumerate(tabelas):
                    try:
                        # Nome do condomínio (ex: "0395 - CENTRO COMERCIAL BARRA PLAZA")
                        condo_completo = tabela.locator("span[id*='lbListaEmpreendimentoNome']").inner_text().strip()
                        # Bloco (ex: "LOJA")
                        bloco = tabela.locator("span[id*='lbListaBlocoNome']").inner_text().strip()
                        # Unidade (ex: "110")
                        unidade = tabela.locator("span[id*='lbListaUnidadeCodigo']").inner_text().strip()

                        # Link de alterar
                        link = tabela.locator("a[title='Alterar unidade']")

                        opcoes.append({
                            "idx": i,
                            "condo_completo": condo_completo,
                            "bloco": bloco,
                            "unidade": unidade,
                            "link": link,
                            # ===== TEXTO PARA BUSCA (COMPLETO!) ===== ↓
                            "texto_busca": f"{condo_completo} {bloco} {unidade}"
                        })

                        # log(f"   [{i}] {condo_completo} - {bloco} {unidade}")

                    except Exception as e:
                        log(f"⚠️ Erro ao extrair tabela {i}: {e}")
                        continue

                if not opcoes:
                    raise ValueError("Não conseguiu extrair informações das unidades")

                # ===== FUZZY MATCHING (CONDOMÍNIO + APARTAMENTO) ===== ↓
                alvo_norm = _norm(f"{alvo_condo} {alvo_apto}")

                log(f"   📋 Buscando: '{alvo_condo}' + '{alvo_apto}'")
                log(f"   🔍 Normalizado: '{alvo_norm}'\n")

                best_i = -1
                best_opt = None
                best_score = 0.0
                best_w_score = 0.0
                best_p_score = 0.0
                is_exact = False

                # Match exato
                for opt in opcoes:
                    if _norm(opt["texto_busca"]) == alvo_norm:
                        best_i = opt["idx"]
                        best_opt = opt
                        best_score = 1.0
                        best_w_score = 1.0
                        best_p_score = 1.0
                        is_exact = True
                        break

                # Fuzzy
                if best_i < 0:
                    log(f"   🔍 Calculando scores fuzzy...")
                    for opt in opcoes:
                        tn = _norm(opt["texto_busca"])
                        w_score = fuzz.WRatio(tn, alvo_norm) / 100.0
                        p_score = fuzz.partial_ratio(tn, alvo_norm) / 100.0
                        score = max(w_score, p_score)

                        # Log apenas se score >= 0.50 para não poluir
                        if score >= 0.50:
                            log(f"      [{opt['idx']}] score={score:.2f} (W:{w_score:.2f}, P:{p_score:.2f}) - {opt['condo_completo'][:50]} - {opt['bloco']} {opt['unidade']}")

                        if score > best_score:
                            best_i = opt["idx"]
                            best_opt = opt
                            best_score = score
                            best_w_score = w_score
                            best_p_score = p_score

                # ===== LOG (SEM ASPAS CONFUSAS!) ===== ↓
                status = "✅ Match exato" if is_exact else "🔍 Match parcial"
                lines = []
                lines.append(f"[FUZZY LIVEFACILITIES] {status} (score={best_score:.2f}):")
                lines.append(f"                       📋 Buscado na BASE:   {alvo_condo} + {alvo_apto}")
                lines.append(
                    f"                       🌐 Encontrado no SITE: {best_opt['condo_completo']} - {best_opt['bloco']} {best_opt['unidade']}")

                if not is_exact:
                    lines.append(
                        f"                       📊 WRatio: {best_w_score:.2f} | partial_ratio: {best_p_score:.2f}")
                    metodo = "WRatio" if best_w_score >= cutoff else "max(WRatio, partial)"
                    lines.append(f"                       🎯 Método usado: {metodo}")
                else:
                    lines.append(f"                       🎯 Método usado: Comparação direta")

                for line in lines:
                    log(line)

                self.match_log.append("\n".join(lines))
                # ===== FIM LOG ===== ↑

                if best_score < cutoff:
                    # ===== MOSTRA OPÇÕES DISPONÍVEIS ===== ↓
                    log(f"\n❌ [LIVEFACILITIES] Nenhuma opção atingiu o cutoff {cutoff:.2f}!")
                    log(f"   Melhor score encontrado: {best_score:.2f}")
                    log(f"   Opções disponíveis:")
                    for opt in opcoes[:10]:
                        log(f"   - {opt['condo_completo']} - {opt['bloco']} {opt['unidade']}")
                    if len(opcoes) > 10:
                        log(f"   ... e mais {len(opcoes) - 10}")
                    # ===== FIM ===== ↑
                    raise ValueError(f"❌ Unidade não encontrada! Melhor score: {best_score:.2f} < cutoff {cutoff:.2f}")

                # Clica no link
                best_opt["link"].click()
                log(f"✅ [LIVEFACILITIES] Unidade selecionada")

                # === CORREÇÃO AQUI ===
                # Força esperar o site processar a troca de unidade (refresh)
                try:
                    self.page.wait_for_load_state("networkidle", timeout=10000)
                    self.page.wait_for_timeout(2000)  # Pausa extra de segurança para scripts lentos
                except:
                    pass
                # =====================

                # Salva tracking
                self.condo_selecionado_site = f"{best_opt['condo_completo']} - {best_opt['bloco']} {best_opt['unidade']}"

            except Exception as e:
                log(f"⚠️ [LIVEFACILITIES UNIDADE] ERRO: {e}")
                raise


        elif action == "download_livefacilities_boletos":

            """
            Fluxo LIVEFACILITIES (Apenas Download)
            Assumes-se que a navegação (Financeiro -> 2ª Via) foi feita pelo JSON.
            1. Verifica: Erros na tela
            2. Loop: Abre Modal -> Entra no Iframe -> Baixa PDF
            3. Processamento: Aplica cabeçalho
            """

            try:
                # Imports locais
                import os
                from auxiliares import utils as aux
                from core import boletos

                identificador = step.get("identificador", "boleto")
                timeout = 60000
                log("🚀 [LIVEFACILITIES] Iniciando verificação e download de boletos...")

                # ===== 1. VERIFICAÇÃO DE ERRO =====
                # Verifica se há mensagem de erro na lista (ex: boleto indisponível)
                loc_erro = self.page.locator("#ContentBody_ucBoletoLista_pListaErro")
                # Só considera erro se estiver visível E tiver texto (ignora div vazia)
                if loc_erro.is_visible() and loc_erro.inner_text().strip() != "":
                    texto_erro = loc_erro.inner_text().strip()
                    # Substitui quebras de linha por espaço
                    texto_erro = texto_erro.replace('\n', ' ').replace('\r', ' ')
                    texto_erro = re.sub(r' +', ' ', texto_erro).strip()
                    log(f"⚠️ Mensagem de erro detectada na página: {texto_erro}")
                    # MUDANÇA: Agora ele levanta um erro para gravar o ❌ no Excel
                    raise ValueError(texto_erro)

                # ===== 2. LOOP DE DOWNLOAD (MODAL + IFRAME) =====
                # IMPORTANTE: Ignora boletos da seção "Outros boletos deste pagador"
                # Pega apenas os boletos FORA da div #ContentBody_ucBoletoLista_dvOutrosBoletosPagador
                seletor_abrir = '#ContentBody_ucBoletoLista_upListaBoleto > div.content:not(#ContentBody_ucBoletoLista_dvOutrosBoletosPagador) a[title="Abrir Boleto"]'

                # Aguarda um pouco para garantir que a lista carregou
                try:
                    self.page.wait_for_selector(seletor_abrir, timeout=5000)
                except:
                    pass  # Se não aparecer em 5s, checa count() abaixo

                qtd_boletos = self.page.locator(seletor_abrir).count()
                if qtd_boletos == 0:
                    msg_vazio = "Nenhum boleto encontrado nesta unidade (lista vazia)."
                    log(f"ℹ️ {msg_vazio}")
                    # MUDANÇA: Grita o erro para o Excel gravar o ❌
                    raise ValueError(msg_vazio)

                log(f"🔍 Encontrados {qtd_boletos} boletos.")
                baixados = 0

                # Itera sobre os boletos encontrados
                for i in range(qtd_boletos):
                    log(f"\n🔄 Processando boleto {i + 1}/{qtd_boletos}...")
                    try:
                        # Re-localiza o elemento a cada volta (necessário pós-refresh)
                        botao_abrir = self.page.locator(seletor_abrir).nth(i)
                        botao_abrir.scroll_into_view_if_needed()
                        botao_abrir.click()
                        # --- LIDA COM O FRAME DO MODAL ---
                        # Localiza o iframe do modal (classe padrão do plugin iziModal)
                        iframe_modal = self.page.frame_locator(".iziModal-iframe")
                        # Botão de imprimir DENTRO do iframe
                        botao_imprimir = iframe_modal.locator('a[href*="/pCli_BoletoNovo.aspx"]').first
                        log("   ⏳ Aguardando botão de download no modal...")
                        # Dispara download
                        with self.page.expect_download(timeout=timeout) as download_info:
                            # force=True ajuda se o modal ainda estiver animando
                            botao_imprimir.click(force=True)
                        download = download_info.value
                        # --- SALVAMENTO E NOMEAÇÃO ---
                        codigo_cliente = identificador[:4] if len(identificador) >= 4 else identificador
                        sufixo = f"_{i}.pdf" if i > 0 else ".pdf"
                        nome_arquivo = f"{codigo_cliente}{sufixo}"
                        destino = os.path.join(DOWNLOADS_DIR, nome_arquivo)

                        download.save_as(destino)
                        log(f"   📥 Download salvo: {nome_arquivo}")
                        # ===== SEU BLOCO CORRIGIDO AQUI ===== ↓

                        # 1. Adiciona Cabeçalho
                        try:
                            # aux.adicionarcabecalhopdf_topo_adaptativo espera string
                            aux.adicionarcabecalhopdf_topo_adaptativo(str(destino), str(destino), codigo_cliente,
                                                                      senha_base=self.cpf_senha)
                            log(f"   ✅ [CABEÇALHO] Código {codigo_cliente} adicionado")
                        except Exception as e:
                            log(f"   ⚠️ [CABEÇALHO] Erro: {e}")

                        # 2. Processa e Move para Pasta do Mês
                        try:
                            log(f"   📂 Classificando...")
                            dados_boleto = boletos.processar_boleto_baixado(
                                caminho_pdf=destino,  # Passa o objeto Path
                                codigo_cliente=self.codigo_cliente,  # Usa o código da classe
                                pasta_downloads=DOWNLOADS_DIR
                            )

                            if dados_boleto:
                                # Verifica duplicata
                                codigo_barras = dados_boleto.get('Cod_Barras', '')
                                if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                                    self.duplicados += 1
                                    baixados += 1  # Conta como baixado (download foi OK)
                                    log(f"   ⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                                    registrar_boleto_duplicado(self.codigo_cliente, codigo_barras, dados_boleto.get('Caminho_Completo', ''))
                                    try:
                                        Path(dados_boleto.get('Caminho_Completo', '')).unlink(missing_ok=True)
                                    except: pass
                                else:
                                    if codigo_barras: registrar_codigo_barras(codigo_barras)
                                    self.boletos_processados.append(dados_boleto)
                                    log(f"   ✅ [SUCESSO] Classificado e movido!")
                                    baixados += 1
                            else:
                                log(f"   ⚠️ Arquivo baixado, mas não classificado (mantido na raiz).")
                                baixados += 1  # Conta como baixado pois o arquivo existe

                        except Exception as e_class:
                            log(f"   ❌ Erro ao processar/mover boleto: {e_class}")
                            baixados += 1

                        # ===== FIM DO BLOCO ===== ↑

                    except Exception as e_inner:
                        log(f"   ❌ Erro ao baixar boleto {i + 1}: {e_inner}")

                    finally:
                        # --- RESET: REFRESH NA PÁGINA ---
                        # Essencial para fechar o modal e limpar o DOM
                        log("   🔄 Atualizando página para fechar modal...")
                        self.page.reload()
                        self.page.wait_for_load_state("networkidle")

                # =========================
                # Atualiza o contador da classe para o relatório final pegar o valor correto
                self.baixados = baixados
                # =========================
                if baixados == 0:
                    raise ValueError("Boletos listados mas nenhum download concluído.")

            except Exception as e:
                log(f"❌ [ERRO CRÍTICO] Falha no fluxo LiveFacilities: {e}")
                raise

        elif action == "try_fill_multi":
            """
            Tenta preencher campo usando múltiplos seletores (fallback).
            Útil quando o seletor exato é incerto.
            """
            try:
                locators_list = step["locators"]
                valor = step["value"]

                sucesso = False
                for i, loc_spec in enumerate(locators_list):
                    try:
                        loc = self._locator_from_json(loc_spec).first
                        loc.wait_for(state="visible", timeout=2000)
                        loc.fill(valor)
                        log(f"✅ [FILL MULTI] Preencheu usando seletor {i + 1}/{len(locators_list)}")
                        sucesso = True
                        break
                    except Exception as e:
                        log(f"⚠️ [FILL MULTI] Seletor {i + 1} falhou: {e}")
                        continue

                if not sucesso:
                    raise ValueError("Nenhum dos seletores funcionou para preencher o campo")

            except Exception as e:
                log(f"❌ [FILL MULTI] ERRO: {e}")
                raise

        elif action == "loop_cipa_download":
            """
            Loop especial para CIPA - clica nos ícones de download.
            """
            identificador = step.get("identificador", "boleto")
            timeout = int(step.get("timeout", 120000))

            log(f"🔍 [CIPA] Procurando boletos para download...")

            try:
                # Localiza todos os ícones de download
                loc = self._locator_from_json(step["locator"])

                # Aguarda pelo menos um aparecer
                try:
                    loc.first.wait_for(state="visible", timeout=10000)
                except:
                    log(f"⚠️ [CIPA] Nenhum ícone de download encontrado")
                    return

                total = loc.count()
                log(f"✅ [CIPA] Encontrados {total} boleto(s) para download")

                if total == 0:
                    return

                for i in range(total):
                    try:
                        # Re-localiza a cada iteração (DOM pode mudar)
                        loc = self._locator_from_json(step["locator"])

                        if loc.count() <= i:
                            log(f"⚠️ [CIPA] Boleto {i + 1} não existe mais")
                            break

                        elemento = loc.nth(i)

                        # Define nome do arquivo
                        codigo = identificador[:4] if len(identificador) >= 4 else identificador
                        nome_arquivo = f"{codigo}.pdf" if i == 0 else f"{codigo}_{i}.pdf"
                        destino = DOWNLOADS_DIR / nome_arquivo

                        log(f"   🔄 Baixando boleto {i + 1}/{total}...")

                        # Clica e aguarda download
                        with self.page.expect_download(timeout=timeout) as d_info:
                            elemento.click(force=True)

                        download = d_info.value
                        download.save_as(str(destino))
                        self.baixados += 1
                        log(f"   📥 Salvo: {nome_arquivo}")

                        # Adiciona cabeçalho
                        try:
                            from auxiliares import utils as aux
                            aux.adicionarcabecalhopdf_topo_adaptativo(str(destino), str(destino), codigo,
                                                                      senha_base=self.cpf_senha)
                            log(f"   ✅ [CABEÇALHO] Código {codigo} adicionado")
                        except Exception as e:
                            log(f"   ⚠️ [CABEÇALHO] Erro: {e}")

                        # Processa boleto
                        try:
                            from core import boletos
                            dados_boleto = boletos.processar_boleto_baixado(
                                caminho_pdf=destino,
                                codigo_cliente=self.codigo_cliente,
                                pasta_downloads=DOWNLOADS_DIR
                            )
                            if dados_boleto:
                                # Verifica duplicata
                                codigo_barras = dados_boleto.get('Cod_Barras', '')
                                if codigo_barras and verificar_boleto_duplicado(codigo_barras):
                                    self.duplicados += 1
                                    log(f"   ⚠️ [DUPLICADO] Código de barras já existe! Boleto baixado mas descartado.")
                                    registrar_boleto_duplicado(self.codigo_cliente, codigo_barras, dados_boleto.get('Caminho_Completo', ''))
                                    try:
                                        Path(dados_boleto.get('Caminho_Completo', '')).unlink(missing_ok=True)
                                    except: pass
                                else:
                                    if codigo_barras: registrar_codigo_barras(codigo_barras)
                                    self.boletos_processados.append(dados_boleto)
                                    log(f"   ✅ [DATAFRAME] Boleto adicionado")
                        except Exception as e:
                            log(f"   ⚠️ [DATAFRAME] Erro: {e}")

                        # Pausa entre downloads
                        self.page.wait_for_timeout(1000)

                    except Exception as e:
                        log(f"   ❌ Erro ao baixar boleto {i + 1}: {e}")
                        continue

                if self.baixados == 0:
                    raise ValueError("Nenhum boleto baixado com sucesso")

            except Exception as e:
                log(f"❌ [CIPA] Erro no loop de download: {e}")
                raise


        elif action == "verificar_sessao_cipa":
            """
            Verifica sessão ativa no CIPA e decide ação.
            """

            usuario_esperado = step.get("usuario", "").strip().lower()
            condo_esperado = step.get("condominio", "").strip()
            log(f"🔍 [CIPA SESSÃO] Verificando sessão ativa...")
            log(f"   📋 Esperado: {usuario_esperado} | {condo_esperado}")
            try:
                # 1. Verifica se está na área logada

                try:
                    self.page.locator("a:has-text('Dashboard')").first.wait_for(state="visible", timeout=3000)
                except:
                    log(f"   ⚪ Não está logado - seguindo para login")
                    return

                # 2. Extrai usuário logado

                usuario_logado = None
                try:
                    el_email = self.page.locator("div.mat-menu-trigger div.opacity-50").first
                    if el_email.is_visible(timeout=2000):
                        usuario_logado = el_email.inner_text().strip().lower()
                        log(f"   👤 Usuário logado: {usuario_logado}")
                except:
                    pass

                # 3. Extrai condomínio atual

                condo_logado = None
                try:
                    el_condo = self.page.locator("span.text-sm.font-semibold.uppercase.text-primary-900").first
                    if el_condo.is_visible(timeout=2000):
                        condo_logado = el_condo.inner_text().strip()
                        log(f"   🏢 Condomínio logado: {condo_logado}")
                except:
                    pass

                # 4. Compara

                usuario_igual = False
                if usuario_logado and usuario_esperado:
                    usuario_igual = (usuario_esperado in usuario_logado or
                                     usuario_logado in usuario_esperado)
                condo_igual = False
                if condo_logado and condo_esperado:
                    condo_logado_norm = _norm(condo_logado)
                    condo_esperado_norm = _norm(condo_esperado)
                    if condo_esperado_norm in condo_logado_norm or condo_logado_norm in condo_esperado_norm:
                        condo_igual = True
                    else:
                        w_score = fuzz.WRatio(condo_logado_norm, condo_esperado_norm) / 100.0
                        condo_igual = w_score >= 0.75
                log(f"   🔍 Usuário={'✅' if usuario_igual else '❌'} | Condomínio={'✅' if condo_igual else '❌'}")

                # 5. AÇÃO DIRETA (sem funções extras!)

                if not usuario_igual:
                    log(f"   ❌ Usuário DIFERENTE - logout completo")
                    self.page.context.clear_cookies()
                    _navegar_e_esperar(self.page, "https://cipafacil.digital/sign-in")
                    return  # Continua login normal

                elif not condo_igual:
                    log(f"   🔄 Condomínio DIFERENTE - voltando para seleção")
                    # Limpa cache de unidades do condomínio anterior
                    self.cipa_unidades_cache = []
                    _navegar_e_esperar(self.page, "https://cipafacil.digital/condominios")
                    raise StopIteration("SKIP_LOGIN")

                else:
                    # Mesmo usuário e condomínio - vai direto para boletos
                    # O filtro_cipa vai capturar as unidades se necessário
                    log(f"   ✅ Sessão REUTILIZADA - indo direto para boletos")
                    _navegar_e_esperar(self.page, "https://cipafacil.digital/boletos")
                    raise StopIteration("SKIP_TO_BOLETOS")

            except StopIteration:
                raise

            except Exception as e:
                log(f"   ⚠️ Erro: {e} - continuando normalmente")
                return

        elif action == "capturar_unidades_cipa":
            """
            Captura lista completa de unidades da página "Unidades".
            Faz scroll até o fim para carregar todas (lazy loading).
            Usa cache por condomínio - se já tiver as unidades do condo atual, não recaptura.
            """
            # Inicializa dict de cache por condomínio se não existir
            if not hasattr(self, 'cipa_unidades_por_condo'):
                self.cipa_unidades_por_condo = {}

            # Pega nome do condomínio atual do contexto
            condo_atual = (self.context_linha.get("CONDOMINIO", "") if self.context_linha else "").strip().upper()

            # Verifica se já temos cache para este condomínio
            if condo_atual and condo_atual in self.cipa_unidades_por_condo:
                unidades_cached = self.cipa_unidades_por_condo[condo_atual]
                log(f"📋 [CIPA] Usando cache de {len(unidades_cached)} unidades do condomínio '{condo_atual}'")
                self.cipa_unidades_cache = unidades_cached
                for u in unidades_cached:
                    log(f"      • {u}")
                return

            log(f"📋 [CIPA] Capturando lista de unidades para '{condo_atual}'...")

            try:
                # Clica em "Unidades" no menu lateral
                link_unidades = self.page.locator("a[href*='/unidades']").first
                if not link_unidades.is_visible(timeout=3000):
                    # Tenta pelo texto
                    link_unidades = self.page.get_by_role("link", name=re.compile(r"Unidades", re.I)).first

                link_unidades.wait_for(state="visible", timeout=10000)
                link_unidades.click()

                self.page.wait_for_load_state("networkidle")
                self.page.wait_for_timeout(2000)

                # ===== SCROLL ATÉ O FIM PARA CARREGAR TODAS AS UNIDADES (LAZY LOADING) =====
                log("   📜 Fazendo scroll para carregar todas as unidades...")

                # Faz scroll até o fim para carregar lazy loading
                prev_count = 0
                max_scrolls = 50

                for scroll_i in range(max_scrolls):
                    # Conta unidades atuais
                    current_count = self.page.locator("span.text-default.font-bold.text-md").count()

                    if current_count == prev_count and scroll_i > 3:
                        # Sem novas unidades após alguns scrolls - provavelmente carregou tudo
                        break

                    prev_count = current_count

                    # Faz scroll via JavaScript no container correto
                    # O container é o div com overflow-y-auto que tem max-height (área de conteúdo principal)
                    self.page.evaluate("""
                        () => {
                            // Procura o container com overflow-y-auto que contém as unidades
                            const container = document.querySelector('div.overflow-y-auto[style*="max-height"]') ||
                                              document.querySelector('div.flex.flex-col.flex-auto.overflow-y-auto') ||
                                              document.querySelector('app-layout div.overflow-y-auto');
                            if (container) {
                                container.scrollTop = container.scrollHeight;
                            }
                        }
                    """)
                    self.page.wait_for_timeout(300)

                log(f"   📜 Scroll completo após {scroll_i + 1} iterações ({current_count} elementos)")
                # ===== FIM SCROLL =====

                # Captura todas as unidades após scroll
                unidades = []

                # Seletor principal baseado no HTML fornecido
                seletores = [
                    "span.text-default.font-bold.text-md",
                    "div.w-full span.text-default.font-bold",
                    "td.mat-cell span",
                    ".unidade-nome",
                    "table tbody tr td:first-child"
                ]

                for seletor in seletores:
                    elementos = self.page.locator(seletor).all()
                    if elementos:
                        for el in elementos:
                            try:
                                texto = el.inner_text().strip()
                                if texto and len(texto) > 1:
                                    unidades.append(texto)
                            except:
                                pass
                        if unidades:
                            break

                # Remove duplicatas mantendo ordem
                seen = set()
                unidades_unicas = []
                for u in unidades:
                    u_norm = u.upper().strip()
                    if u_norm not in seen and u_norm:
                        seen.add(u_norm)
                        unidades_unicas.append(u)

                # Salva no cache global e no cache por condomínio
                self.cipa_unidades_cache = unidades_unicas
                if condo_atual:
                    self.cipa_unidades_por_condo[condo_atual] = unidades_unicas
                    log(f"   💾 Cache salvo para '{condo_atual}'")

                if unidades_unicas:
                    log(f"   ✅ {len(unidades_unicas)} unidades capturadas:")
                    for u in unidades_unicas:
                        log(f"      • {u}")
                else:
                    log(f"   ⚠️ Nenhuma unidade encontrada na página")

            except Exception as e:
                log(f"   ⚠️ Erro ao capturar unidades: {e}")
                self.cipa_unidades_cache = []

        elif action == "filtro_cipa":
            """
            Filtra boletos por unidade e status "Em Aberto".
            Usa cache de unidades POR CONDOMÍNIO + fuzzy match para encontrar a melhor correspondência.
            Se não tiver cache para o condomínio atual, captura as unidades primeiro.
            """
            apartamento_alvo = step.get("apartamento", "")
            cutoff = float(step.get("cutoff", 0.60))
            log(f"🔍 [CIPA FILTRO] Configurando filtros para: {apartamento_alvo}")

            try:
                # Inicializa dict de cache por condomínio se não existir
                if not hasattr(self, 'cipa_unidades_por_condo'):
                    self.cipa_unidades_por_condo = {}

                # Pega nome do condomínio atual do contexto
                condo_atual = (self.context_linha.get("CONDOMINIO", "") if self.context_linha else "").strip().upper()

                # ===== 0. VERIFICA CACHE POR CONDOMÍNIO ===== ↓
                # Primeiro verifica se temos cache específico para este condomínio
                if condo_atual and condo_atual in self.cipa_unidades_por_condo:
                    self.cipa_unidades_cache = self.cipa_unidades_por_condo[condo_atual]
                    log(f"   💾 Usando cache de {len(self.cipa_unidades_cache)} unidades para '{condo_atual}'")

                # Se não tem cache (nem geral nem por condo), captura
                elif not hasattr(self, 'cipa_unidades_cache') or not self.cipa_unidades_cache:
                    log(f"   ⚠️ Sem cache de unidades para '{condo_atual}' - capturando agora...")
                    # Navega para Unidades
                    link_unidades = self.page.locator("a[href*='/unidades']").first
                    if link_unidades.is_visible(timeout=3000):
                        link_unidades.click()
                        self.page.wait_for_load_state("networkidle")
                        self.page.wait_for_timeout(2000)

                        # ===== SCROLL ATÉ O FIM PARA CARREGAR TODAS AS UNIDADES (LAZY LOADING) =====
                        log(f"   📜 Fazendo scroll para carregar todas as unidades...")

                        # Faz scroll até o fim para carregar lazy loading
                        prev_count = 0
                        max_scrolls = 50

                        for scroll_i in range(max_scrolls):
                            # Conta unidades atuais
                            current_count = self.page.locator("span.text-default.font-bold.text-md").count()

                            if current_count == prev_count and scroll_i > 3:
                                # Sem novas unidades após alguns scrolls - provavelmente carregou tudo
                                break

                            prev_count = current_count

                            # Faz scroll via JavaScript no container correto
                            # O container é o div com overflow-y-auto que tem max-height (área de conteúdo principal)
                            self.page.evaluate("""
                                () => {
                                    // Procura o container com overflow-y-auto que contém as unidades
                                    const container = document.querySelector('div.overflow-y-auto[style*="max-height"]') ||
                                                      document.querySelector('div.flex.flex-col.flex-auto.overflow-y-auto') ||
                                                      document.querySelector('app-layout div.overflow-y-auto');
                                    if (container) {
                                        container.scrollTop = container.scrollHeight;
                                    }
                                }
                            """)
                            self.page.wait_for_timeout(300)

                        log(f"   📜 Scroll completo após {scroll_i + 1} iterações ({current_count} elementos)")

                        # Captura as unidades
                        unidades = []
                        elementos = self.page.locator("span.text-default.font-bold.text-md").all()
                        for el in elementos:
                            try:
                                texto = el.inner_text().strip()
                                if texto and len(texto) > 1:
                                    unidades.append(texto)
                            except:
                                pass

                        # Remove duplicatas
                        seen = set()
                        self.cipa_unidades_cache = []
                        for u in unidades:
                            u_norm = u.upper().strip()
                            if u_norm not in seen:
                                seen.add(u_norm)
                                self.cipa_unidades_cache.append(u)

                        # Salva no cache por condomínio
                        if condo_atual:
                            self.cipa_unidades_por_condo[condo_atual] = self.cipa_unidades_cache
                            log(f"   💾 Cache salvo para '{condo_atual}'")

                        log(f"   ✅ {len(self.cipa_unidades_cache)} unidades capturadas")
                        for u in self.cipa_unidades_cache:
                            log(f"      • {u}")

                        # Volta para Boletos
                        log(f"   🔙 Voltando para página de boletos...")
                        link_boletos = self.page.locator("a[href*='/boletos']").first
                        link_boletos.click()
                        self.page.wait_for_load_state("networkidle")
                        self.page.wait_for_timeout(2000)

                        # Aguarda o botão de filtros aparecer (indica que a página carregou)
                        try:
                            self.page.get_by_role("button", name=re.compile(r"Filtros", re.I)).wait_for(state="visible",
                                                                                                        timeout=10000)
                            log(f"   ✅ Página de boletos carregada")
                        except:
                            log(f"   ⚠️ Botão Filtros não encontrado após voltar")
                # ===== FIM CAPTURA ===== ↑

                # ===== 1. FUZZY MATCH COM CACHE DE UNIDADES (COM VALIDAÇÃO DE NÚMERO!) ===== ↓
                unidade_selecionada = None

                if hasattr(self, 'cipa_unidades_cache') and self.cipa_unidades_cache:
                    log(f"   📋 Buscando melhor match em {len(self.cipa_unidades_cache)} unidades (cutoff={cutoff})...")

                    # Usa a função fuzzy_match_items_core com validar_numero=True
                    items_data = [
                        {'idx': i, 'text': u, 'element': None}
                        for i, u in enumerate(self.cipa_unidades_cache)
                    ]

                    try:
                        result = fuzzy_match_items_core(
                            items_data=items_data,
                            alvo=apartamento_alvo,
                            cutoff=cutoff,
                            validar_numero=True,  # IMPORTANTE: Valida números!
                            is_condo=False
                        )

                        unidade_selecionada = result['best_txt']
                        best_score = result['best_score']

                        # Log do resultado
                        for line in result.get('log_lines', []):
                            log(f"   {line}")

                    except ValueError as e:
                        # Não encontrou match válido - NÃO baixa boleto
                        raise ValueError(f"❌ Unidade '{apartamento_alvo}' não encontrada!\n   {e}")
                else:
                    # Sem cache - erro porque não pode validar
                    raise ValueError(f"❌ Sem cache de unidades - não é possível validar '{apartamento_alvo}'")
                # ===== FIM FUZZY MATCH ===== ↑

                # ===== 2. VERIFICA SE UNIDADE JÁ ESTÁ SELECIONADA ===== ↓
                # O badge da unidade atual está em: span.leading-relaxed.whitespace-nowrap
                # Formato: "20403 - APT 403 SNI BL 2" ou "307 - Apto 307"
                # IMPORTANTE: O badge só existe na página de BOLETOS
                unidade_ja_selecionada = False
                try:
                    log(f"   🔍 Verificando badge da unidade atual...")
                    log(f"   📍 URL atual: {self.page.url}")

                    # Garante que está na página de boletos
                    if "/boletos" not in self.page.url:
                        log(f"   ⚠️ Não está na página de boletos - navegando...")
                        self.page.locator("a[href*='/boletos']").first.click()
                        self.page.wait_for_load_state("networkidle")

                    # Aguarda a página carregar completamente e o botão Filtros aparecer
                    self.page.wait_for_timeout(1500)
                    try:
                        self.page.get_by_role("button", name=re.compile(r"Filtros", re.I)).wait_for(state="visible",
                                                                                                    timeout=5000)
                    except:
                        pass

                    # Tenta múltiplos seletores para o badge
                    # O badge fica dentro de: div.inline-flex.rounded-full > span.leading-relaxed.whitespace-nowrap
                    badge_unidade = None
                    badge_seletores = [
                        "div.inline-flex.rounded-full span.leading-relaxed.whitespace-nowrap",
                        "div.rounded-full span.leading-relaxed.whitespace-nowrap",
                        "div.bg-gray-200 span.leading-relaxed.whitespace-nowrap",
                        "span.leading-relaxed.whitespace-nowrap",
                    ]

                    # Log: quantos elementos existem para cada seletor
                    for sel in badge_seletores:
                        count = self.page.locator(sel).count()
                        if count > 0:
                            log(f"   📊 Seletor '{sel}': {count} elemento(s)")

                    for sel in badge_seletores:
                        loc = self.page.locator(sel).first
                        if loc.count() > 0:
                            try:
                                if loc.is_visible(timeout=2000):
                                    badge_unidade = loc
                                    log(f"   📍 Badge encontrado com seletor: {sel}")
                                    break
                            except Exception as e:
                                log(f"   ⚠️ Seletor {sel} falhou: {e}")
                                continue

                    if badge_unidade:
                        try:
                            badge_texto = badge_unidade.inner_text().strip().upper()
                            unidade_sel_norm = unidade_selecionada.upper().strip()
                            log(f"   📋 Badge encontrado: '{badge_texto}' | Buscando: '{unidade_sel_norm}'")

                            # Verifica se a unidade selecionada está no badge
                            # O badge pode ter formato "CODIGO - UNIDADE" então verifica se contém
                            if unidade_sel_norm in badge_texto or badge_texto.endswith(unidade_sel_norm):
                                log(f"   ✅ Unidade já selecionada no badge!")
                                unidade_ja_selecionada = True
                            else:
                                log(f"   🔄 Unidade diferente - precisa trocar")
                        except Exception as e:
                            log(f"   ⚠️ Erro ao ler texto do badge: {e}")
                    else:
                        log(f"   ⚠️ Badge não encontrado ou não visível")
                except Exception as e:
                    log(f"   ⚠️ Erro ao verificar badge: {e}")

                # ===== 2b. ABRE MODAL DE FILTROS (SEMPRE) ===== ↓
                log("   🖱️ Abrindo modal de filtros...")
                btn_filtros = self.page.get_by_role("button", name=re.compile(r"Filtros", re.I))
                btn_filtros.wait_for(state="visible", timeout=10000)
                btn_filtros.click()

                # Aguarda o modal abrir E ter conteúdo
                self.page.locator("mat-dialog-container").wait_for(state="visible", timeout=5000)

                # Aguarda os campos do modal carregarem
                try:
                    self.page.locator("mat-dialog-container mat-form-field").first.wait_for(state="visible",
                                                                                            timeout=5000)
                except:
                    pass
                self.page.wait_for_timeout(500)

                # Verifica se o modal realmente abriu com campos
                modal_campos = self.page.locator("mat-dialog-container mat-form-field").count()
                log(f"   📊 Modal: {modal_campos} campos encontrados")

                if modal_campos == 0:
                    log(f"   ⚠️ Modal sem campos - tentando novamente...")
                    # Fecha e tenta abrir de novo
                    self.page.keyboard.press("Escape")
                    self.page.wait_for_timeout(1000)
                    btn_filtros.click()
                    self.page.locator("mat-dialog-container").wait_for(state="visible", timeout=5000)
                    try:
                        self.page.locator("mat-dialog-container mat-form-field").first.wait_for(state="visible",
                                                                                                timeout=5000)
                    except:
                        pass
                    self.page.wait_for_timeout(500)
                    modal_campos = self.page.locator("mat-dialog-container mat-form-field").count()
                    log(f"   📊 Segunda tentativa: {modal_campos} campos")

                # ===== 3. SELECIONA UNIDADE (APENAS SE NÃO ESTIVER JÁ SELECIONADA) ===== ↓
                if unidade_ja_selecionada:
                    log(f"   ⏭️ Unidade já selecionada - pulando seleção de unidade")
                else:
                    log(f"   🖱️ Abrindo lista de unidades...")

                    # Tenta múltiplas formas de abrir o dropdown de unidades
                    dropdown_aberto = False

                    # Tentativa 1: Clica no mat-form-field
                    campo_unidade = self.page.locator("mat-dialog-container mat-form-field").first
                    if campo_unidade.is_visible(timeout=2000):
                        campo_unidade.click()
                        self.page.wait_for_timeout(1000)

                        # Verifica se abriu
                        if self.page.locator("mat-option").count() > 0:
                            dropdown_aberto = True
                            log(f"   ✅ Dropdown aberto (clique no form-field)")

                    # Tentativa 2: Clica no mat-select dentro do form-field
                    if not dropdown_aberto:
                        log(f"   🔄 Tentando mat-select...")
                        mat_select = self.page.locator("mat-dialog-container mat-select").first
                        if mat_select.is_visible(timeout=2000):
                            mat_select.click()
                            self.page.wait_for_timeout(1000)
                            if self.page.locator("mat-option").count() > 0:
                                dropdown_aberto = True
                                log(f"   ✅ Dropdown aberto (clique no mat-select)")

                    # Tentativa 3: Clica no input
                    if not dropdown_aberto:
                        log(f"   🔄 Tentando input...")
                        input_el = self.page.locator("mat-dialog-container mat-form-field input").first
                        if input_el.is_visible(timeout=2000):
                            input_el.click()
                            self.page.wait_for_timeout(1000)
                            if self.page.locator("mat-option").count() > 0:
                                dropdown_aberto = True
                                log(f"   ✅ Dropdown aberto (clique no input)")

                    # Tentativa 4: Clica na seta do dropdown
                    if not dropdown_aberto:
                        log(f"   🔄 Tentando seta do dropdown...")
                        seta = self.page.locator(
                            "mat-dialog-container mat-form-field .mat-select-arrow, mat-dialog-container mat-form-field mat-icon").first
                        if seta.is_visible(timeout=2000):
                            seta.click()
                            self.page.wait_for_timeout(1000)
                            if self.page.locator("mat-option").count() > 0:
                                dropdown_aberto = True
                                log(f"   ✅ Dropdown aberto (clique na seta)")

                    # Aguarda um pouco mais para garantir que carregou
                    self.page.wait_for_timeout(500)

                    # Seleciona a unidade na lista
                    log(f"   🔍 Procurando '{unidade_selecionada}' na lista...")

                    opcoes = self.page.locator("mat-option").all()
                    log(f"   📋 {len(opcoes)} opções no dropdown")

                    unidade_encontrada = False
                    unidade_sel_norm = unidade_selecionada.upper().strip()

                    for opcao in opcoes:
                        try:
                            texto = opcao.inner_text().strip()
                            texto_norm = texto.upper().strip()

                            if unidade_sel_norm in texto_norm or texto_norm in unidade_sel_norm:
                                log(f"   ✅ Selecionando: '{texto}'")
                                opcao.click()
                                unidade_encontrada = True
                                break
                        except:
                            continue

                    if not unidade_encontrada:
                        # Fuzzy nas opções visíveis
                        best_opt = None
                        best_opt_score = 0

                        for opcao in opcoes:
                            try:
                                texto = opcao.inner_text().strip()
                                score = fuzz.WRatio(unidade_sel_norm, texto.upper()) / 100.0
                                if score > best_opt_score:
                                    best_opt_score = score
                                    best_opt = opcao
                            except:
                                continue

                        if best_opt and best_opt_score >= 0.50:
                            texto = best_opt.inner_text().strip()
                            log(f"   ✅ Fuzzy no dropdown ({best_opt_score:.2f}): '{texto}'")
                            best_opt.click()
                            unidade_encontrada = True
                        else:
                            log(f"   ⚠️ Unidade não encontrada no dropdown - continuando sem filtro de unidade")
                            # Fecha o dropdown pressionando Escape
                            self.page.keyboard.press("Escape")
                            self.page.wait_for_timeout(500)

                    self.page.wait_for_timeout(1000)

                # ===== 4. SELECIONA STATUS "EM ABERTO" (SEMPRE) ===== ↓
                log("   🎯 Selecionando Status 'Em Aberto'...")

                # Clica no campo Status (segundo mat-form-field)
                campos = self.page.locator("mat-dialog-container mat-form-field").all()
                log(f"   📊 Encontrados {len(campos)} campos no modal")

                campo_status = None
                if len(campos) >= 2:
                    campo_status = campos[1]
                elif len(campos) == 1:
                    # Se só tem 1 campo, pode ser que a unidade já estava selecionada
                    # e só sobrou o campo de Status
                    campo_status = campos[0]
                    log(f"   📍 Usando único campo disponível como Status")

                if campo_status:
                    campo_status.click()
                    self.page.wait_for_timeout(1500)

                    # Procura "Em Aberto" nas opções
                    opcoes_status = self.page.locator("mat-option").all()
                    log(f"   📊 Encontradas {len(opcoes_status)} opções de status")

                    for opcao in opcoes_status:
                        try:
                            texto = opcao.inner_text().strip()
                            if "Aberto" in texto or "aberto" in texto:
                                log(f"   ✅ Status: '{texto}'")
                                opcao.click()
                                break
                        except:
                            continue

                    self.page.wait_for_timeout(500)
                else:
                    log(f"   ⚠️ Campo Status não encontrado")

                # ===== 5. CLICA EM FILTRAR ===== ↓
                log("   🚀 Aplicando filtros...")

                btn_filtrar = self.page.get_by_role("button", name=re.compile(r"Filtrar", re.I))
                if btn_filtrar.is_visible(timeout=3000):
                    btn_filtrar.click()
                else:
                    # Fallback: segundo botão no rodapé do modal
                    self.page.locator("mat-dialog-container button").nth(1).click()

                # Aguarda modal fechar
                try:
                    self.page.locator("mat-dialog-container").wait_for(state="hidden", timeout=10000)
                except:
                    # Tenta fechar clicando no X
                    log(f"   🔄 Fechando modal manualmente...")
                    self.page.locator("mat-dialog-container button mat-icon:has-text('close')").click()

                self.page.wait_for_load_state("networkidle")
                self.page.wait_for_timeout(2000)

                log("   ✅ Filtros aplicados!")
                # ===== FIM FILTROS ===== ↑

            except Exception as e:
                log(f"   ❌ Erro: {e}")
                raise

    def _fazer_logout_cipa_completo(self):
        """Faz logout completo do CIPA"""
        try:
            # 1. Tenta sair do condomínio primeiro (se estiver dentro)
            self._voltar_selecao_condo_cipa()

            # 2. Clica no menu do usuário (canto inferior esquerdo)
            menu_usuario = self.page.locator("div[class*='cursor-pointer']").filter(
                has_text=re.compile(r"@", re.I)
            ).first

            if menu_usuario.is_visible(timeout=3000):
                menu_usuario.click()
                self.page.wait_for_timeout(1000)

                # 3. Clica em "Sair"
                btn_sair = self.page.get_by_role("menuitem", name="Sair").first
                if btn_sair.is_visible(timeout=2000):
                    btn_sair.click()
                    self.page.wait_for_load_state("networkidle", timeout=10000)
                    log(f"   ✅ Logout completo realizado")
                    return

            # Fallback: limpa cookies
            log(f"   🛡️ Fallback: limpando cookies")
            self.page.context.clear_cookies()
            self.page.goto("https://cipafacil.digital/sign-in")
            self.page.wait_for_load_state("networkidle")

        except Exception as e:
            log(f"   ⚠️ Erro no logout: {e}")
            # Força limpeza
            self.page.context.clear_cookies()

    def _voltar_selecao_condo_cipa(self):
        """Volta para área de seleção de condomínios no CIPA"""
        try:
            # Procura menu do condomínio (canto superior direito)
            menu_condo = self.page.locator("div[class*='cursor-pointer']").filter(
                has_text=re.compile(r"Edifício|Condomínio|Residencial", re.I)
            ).first

            if menu_condo.is_visible(timeout=3000):
                log(f"   🏢 Clicando no menu do condomínio...")
                menu_condo.click()
                self.page.wait_for_timeout(1000)

                # Clica em "Sair" (do condomínio, não do sistema)
                btn_sair = self.page.get_by_role("menuitem", name="Sair").first
                if btn_sair.is_visible(timeout=2000):
                    btn_sair.click()
                    self.page.wait_for_timeout(2000)
                    log(f"   ✅ Voltou para seleção de condomínios")
                    return True

            # Fallback: navega direto
            log(f"   🔄 Fallback: navegando para lista de condomínios")
            self.page.goto("https://cipafacil.digital/condominiums")
            self.page.wait_for_load_state("networkidle")
            return True

        except Exception as e:
            log(f"   ⚠️ Erro ao voltar para seleção: {e}")
            return False


# ===================== CONFIGURAÇÃO DE FLUXOS =====================

# ----- LOGINS -----
LOGINS = {
    # ===== SUPERLÓGICA =====
    "superlogica": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # Verifica se já está logado (se já estiver, pula login)
        {"scope": "login", "action": "check_already_logged_superlogica"},

        # Passo 1: Preenche email
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='email']"},
         "value": "${USUARIO}"},
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "#salvar"}},

        # ← VERIFICA ERRO (Para se encontrar)
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "#divMsgErroArea"},
         "timeout": 2000,
         "stop_if_error": True},  # Para imediatamente se erro

        # Passo 2: Preenche senha (só chega aqui se passo 1 OK)
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#senha"},
         "value": "${SENHA}"},
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "#salvar"}},

        # ← VERIFICA ERRO novamente
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "#divMsgErroArea"},
         "timeout": 2000,
         "stop_if_error": True},
    ],

    # ===== iCondo =====
    "icondo": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#user_login"},
         "value": "${USUARIO}"},
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#user_pass"},
         "value": "${SENHA}"},
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "#loginform [type='submit'], button[type='submit']"}},

        # ← VERIFICA ERRO (Para se encontrar)
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "#login_error"},
         "timeout": 2000,
         "stop_if_error": True},
    ],

    # ===== WEBWARE (Conac, Lowndes, Acir) =====

    "webware": [
        # 1. Acesso
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # 2. Usuário (Tenta 'mem' OU 'usuario' OU 'login')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css",
                     "selector": "input[name='mem'], input[name='usuario'], input[name='login'], input[id='usuario']"},
         "value": "${USUARIO}"},

        # 3. Senha (Tenta 'pass' OU 'senha')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='pass'], input[name='senha'], input[id='senha']"},
         "value": "${SENHA}"},

        # 4. Botão Entrar (Tenta input Submit OU botão genérico)
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "input[type='Submit'], button[type='submit'], input[value='Entrar']"}},

        # ===== ADICIONA PAUSA ANTES DE VERIFICAR (NOVO!) ===== ↓
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css", "selector": "body"},  # Aguarda body (sempre existe)
         "timeout": 2000},  # 2 segundos de pausa
        # ===== FIM ===== ↑

        # 5. Verificação de erro
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "p.mensagem-erro, #dvContainer, .alert-danger"},
         "timeout": 30000,
         "stop_if_error": True}
    ],

    # ===== LiveFacilities (ADMLOC, CEGIL, EDEL, GERPEL, RASA, UNA) =====
    "livefacilities": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # Preenche login
        {"scope": "login", "action": "fill",
         "locator": {"by": "role", "role": "textbox", "name": "login"},
         "value": "${USUARIO}"},

        # Preenche senha
        {"scope": "login", "action": "fill",
         "locator": {"by": "role", "role": "textbox", "name": "senha"},
         "value": "${SENHA}"},

        # Clica Entrar
        {"scope": "login", "action": "click",
         "locator": {"by": "role", "role": "button", "name": "Entrar"}},

        # Verifica erro
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "#ucLoginSistema_lbErroEntrar"},
         "timeout": 3000,
         "stop_if_error": True},

        # === NOVO PASSO: Esperar a Home carregar ===
        # Aumentamos o timeout para 45 segundos (45000ms) para tolerar lentidão
        {"scope": "login", "action": "esperar_url",
         "pattern": "Home.aspx",
         "timeout": 45000}
    ],
    # ===== PROTEL =====
    "protel": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # Usuário
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#usuario_login"},
         "value": "${USUARIO}"},

        # Senha
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#usuario_senha"},
         "value": "${SENHA}"},

        # Botão Entrar
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "#botao_login"}},

        # Aguarda carregar a página após login (não verifica erro aqui, pois #conteudo sempre existe)
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css", "selector": "body"},
         "timeout": 3000}
    ],

    # ===== FERNANDO E FERNANDES =====
    "fernandoefernandes": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # Usuário (id='edit_user')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#edit_user"},
         "value": "${USUARIO}"},

        # Senha (id='edit_pw')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "#edit_pw"},
         "value": "${SENHA}"},

        # Botão Entrar (texto 'E N T R A R')
        {"scope": "login", "action": "click",
         "locator": {"by": "xpath", "selector": "//button[contains(text(), 'E N T R A R')]"}},

        # Verifica erro (XPath original do seu código)
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "form.login_form div:last-child"},
         "timeout": 3000,
         "stop_if_error": True}
    ],

    # ===== IMODATA (BLINDADO) =====
    "imodata": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # 1. (CRÍTICO) Verifica se caiu na tela de "Bem-vindo de volta"
        # O vídeo mostra botões: "Novo Login" e "Continuar com..."
        # Se esse botão existir, CLICA NELE para liberar o formulário.
        {
            "scope": "login",
            "action": "try_click",
            "locator": {
                "by": "xpath",
                "selector": "//input[@value='Novo Login'] | //button[contains(text(), 'Novo Login')] | //a[contains(text(), 'Novo Login')]"
            },
            "timeout": 5000
        },

        # 2. (SEGURANÇA) Espera o formulário animar/aparecer
        # Só tenta clicar em "Condomínio" se ele estiver visível
        {
            "scope": "login",
            "action": "esperar_selector",
            "locator": {"by": "xpath", "selector": "//*[text()='Condomínio']"},
            "timeout": 5000
        },

        # 3. Seleciona "Condomínio"
        {"scope": "login", "action": "click",
         "locator": {"by": "xpath", "selector": "//*[text()='Condomínio']"}},

        # 4. Preenche Login
        {"scope": "login", "action": "fill",
         "locator": {"by": "role", "role": "textbox", "name": "Seu login..."},
         "value": "${USUARIO}"},

        # 5. Clica Verificar
        {"scope": "login", "action": "click",
         "locator": {"by": "role", "role": "button", "name": "Verificar Usuário"}},

        # 6. Preenche Senha (com espera)
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "role", "role": "textbox", "name": "Sua senha..."},
         "timeout": 5000},

        {"scope": "login", "action": "fill",
         "locator": {"by": "role", "role": "textbox", "name": "Sua senha..."},
         "value": "${SENHA}"},

        # 7. Botão Acessar
        {"scope": "login", "action": "click",
         "locator": {"by": "role", "role": "button", "name": "Acessar"}},

        # 8. Verifica erro
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "xpath",
                     "selector": "//div[contains(@style, 'color:red')] | //div[contains(text(), 'não conferem')]"},
         "timeout": 3000,
         "stop_if_error": True}
    ],

    # ===== CIPA LOGIN =====
    "cipa": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # === VERIFICAÇÃO DE SESSÃO (NOVO!) ===
        {"scope": "login", "action": "verificar_sessao_cipa",
         "usuario": "${USUARIO}",
         "condominio": "${CONDOMINIO}",
         "timeout": 10000},

        # Aguarda página carregar
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css", "selector": "input#email, input[type='email'], condominium-list"},
         "timeout": 15000},

        # Preenche e-mail
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input#email"},
         "value": "${USUARIO}"},

        # Clica no botão para exibir campo senha
        {"scope": "login", "action": "try_click",
         "locator": {"by": "css", "selector": "#salvar, button[type='submit']"},
         "timeout": 3000},

        # Aguarda campo senha
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css", "selector": "input#password, input#senha, input[type='password']"},
         "timeout": 10000},

        # Preenche senha
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input#password, input#senha, input[type='password']"},
         "value": "${SENHA}"},

        # Clica Entrar
        {"scope": "login", "action": "click",
         "locator": {"by": "role", "role": "button", "name": "Entrar"},
         "timeout": 10000},

        # Verifica erro
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "css", "selector": "div.cipafacil-alert-message, #mat-error-2, .mat-error"},
         "timeout": 5000,
         "stop_if_error": True},

        # Aguarda qualquer indicador de área logada
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css",
                     "selector": "condominium-list, div.filter-article, a[href*='boleto'], a:has-text('Dashboard')"},
         "timeout": 20000},

        # Navega explicitamente para área de condomínios (caso tenha ido direto pro dashboard)
        {"scope": "login", "action": "goto", "value": "https://cipafacil.digital/condominios"},
    ],

    # ===== ADMINISTRADORA NACIONAL =====
    "nacional": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},

        # Preenche login (campo name='login')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='login']"},
         "value": "${USUARIO}"},

        # Preenche senha (campo name='senha')
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='senha']"},
         "value": "${SENHA}"},

        # Clica no botão Entrar (input submit)
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "input[type='submit']"}},

        # Aguarda a página após login
        {"scope": "login", "action": "esperar_selector",
         "locator": {"by": "css", "selector": "body"},
         "timeout": 5000},

        # Verifica erro de login ou condomínio desabilitado
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "xpath",
                     "selector": "//*[contains(text(), 'desabilitado') or contains(text(), 'Lamento') or contains(@class, 'erro') or contains(@class, 'mensagem-erro')]"},
         "timeout": 3000,
         "stop_if_error": True},
    ],

    # ===== ADMINISTRADORA NACIONAL =====
    "nacional": [
        {"scope": "login", "action": "goto", "value": "${SITE}"},
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='login']"},
         "value": "${USUARIO}"},
        {"scope": "login", "action": "fill",
         "locator": {"by": "css", "selector": "input[name='senha']"},
         "value": "${SENHA}"},
        {"scope": "login", "action": "click",
         "locator": {"by": "css", "selector": "input[type='submit']"}},
        {"scope": "login", "action": "verificar_erro_login",
         "locator": {"by": "xpath", "selector": "//*[contains(text(), 'desabilitado') or contains(text(), 'Lamento')]"},
         "timeout": 3000,
         "stop_if_error": True},
    ],
}

# ----- FLUXOS POR ADMINISTRADORA -----
ADMIN_FLOWS = {
    "superlogica": [
        # 1. Seleção do Condomínio
        {
            "action": "fuzzy_lista_fechada",
            "value": "${CONDOMINIO}",
            "container": {"by": "css", "selector": "ul#lista"},
            "items_css": "a",
            "toggle": {"by": "css", "selector": ".seta-menu"},
            "cutoff": 0.80,
        },

        # 2. ESPERA ROBUSTA: Aguarda o corpo da página carregar (aumentado para 30s)
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "#servicos-e-informacoes, .em-abertas"},
            "timeout": 30000
        },

        # 3. TENTA EXPANDIR: O robô agora tenta clicar no "Mais Itens" antes de checar erro
        {
            "action": "expandir_lista",
            "locator": {"by": "css", "selector": "a[id^='btn_mais_itens_cobrancas'], a:has-text('Mais itens')"},
            "timeout": 5000,  # Tempo para o botão aparecer
            "max_clicks": 500,
            "settle_ms": 800
        },

        # 4. SÓ AGORA CHECA SE ESTÁ VAZIO:
        # Se após tentar expandir a div de "nenhuma cobrança" ainda estiver lá, aí sim paramos.
        {
            "action": "parar_se_visivel",
            "locator": {"by": "css", "selector": "div[id^='div_nenhuma_cobranca'].div-nenhuma-cobranca"},
            "timeout": 2000,
            "message": "Não há boletos nesta unidade.",
        },

        # 5. Outras verificações (ABRJ, etc)
        {
            "action": "parar_se_visivel",
            "locator": {"by": "xpath", "selector": "//*[contains(text(), 'boleto indisponível')]"},
            "timeout": 1000,
            "message": "Boleto indisponível (vencido)",
        },

        # 6. Loop Final
        {
            "action": "loop_apartamentos_fuzzy_best",
            "alvo_apartamento": "${APARTAMENTO}",
            "container": {"by": "css", "selector": "div.em-abertas"},
            "items_css": "a.bloco-grid-cobrancas",
            "apto_parts": [".unidade .numero", ".unidade .complemento"],
            "click_target": ".unidade",
            "visualizar_spec": {"by": "role", "role": "button", "name": "Visualizar"},
            "download_priority": [
                {"by": "role", "role": "link", "name": "Pagar"},
                {"by": "role", "role": "link", "name": "Imprimir"},
                {"by": "role", "role": "button", "name": "Imprimir"},  # Fallback adicional
                {"by": "css", "selector": "a[id^='btn-imprimir']"}  # Fallback técnico
            ],
            "imprimir_spec": {"by": "role", "role": "link", "name": "Imprimir"},
            "identificador": "${IDENTIFICADOR}",
        }
    ],

    "icondo": [
        # Passo 1: Clicar no botão "segunda via"
        {
            "action": "click",
            "locator": {
                "by": "css",
                "selector": "a[href*='segunda-via-boleto']"
            },
            "timeout": 10000
        },
        # Passo 2: Verificar se tem mensagem de erro (sem boletos)
        {
            "action": "parar_se_visivel",
            "locator": {
                "by": "css",
                "selector": ".alert-danger"
            },
            "timeout": 2000,
            "message": "Não há boletos disponíveis (alert-danger visível)"
        },
        # Passo 3: Baixar todos os links "Recibo No."
        {
            "action": "download_recibos_icondo",
            "partial_text": "Recibo No.",
            "identificador": "${IDENTIFICADOR}",
            "max_boletos": 50,
            "timeout": 120000
        }
    ],

    # ===== WEBWARE (Conac, Lowndes, Acir) =====
    "webware": [
        # 1. Tenta selecionar condomínio + apartamento (OPCIONAL)
        {
            "action": "try_fuzzy_lista_fechada",
            "value": "${CONDOMINIO}",
            "alvo_apartamento": "${APARTAMENTO}",
            "container": {"by": "css", "selector": "body"},
            "items_css": ".card",
            "text_selector": ".txt-titulo",
            "click_selector": ".bt",
            "cutoff": 0.70,
            "timeout": 3000
        },
        # 2. Clica em "2ª Via de Boleto" (card NOVO ou link ANTIGO)
        {
            "action": "try_click_multi",
            "locators": [
                # Seletor 1: Pela imagem específica que você capturou (Super estável)
                {"by": "css", "selector": "img[src*='ic-boleto']"},
                # Seletor 2: Pela classe do parágrafo dentro do span
                {"by": "css", "selector": "span:has(p.nome)"},
                # Seletor 3: Fallback por container de estatística
                {"by": "css", "selector": ".statistic__item"}
            ],
            # "locators": [
            #     {"by": "css", "selector": "div.statistic__item:has(h2:has-text('2ª via de boleto')), a:has-text('VIA DE BOLETO')"},  # Card novo
            #     {"by": "css", "selector": "h2.number:has-text('2ª via')"},  # Fallback - clica direto no h2
            #     {"by": "role", "role": "link", "name": "VIA DE BOLETO", "exact": False}  # Link antigo
            # ],
            "timeout": 10000
        },

        # 3. Aguarda iframe carregar
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": ".prestacao-interativa"},
            "timeout": 10000
        },

        # 4. TENTA fechar alerta (OPCIONAL - nem sempre aparece)
        {
            "action": "try_click",
            "locator": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                "selector": ".popup-alerta-botoes"
            },
            "timeout": 3000
        },

        # 5. TENTA abrir dropdown de apartamentos (OPCIONAL - só se tiver combo)
        {
            "action": "try_click",
            "locator": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                "selector": ".v-input__control"
            },
            "timeout": 3000
        },

        # 6. TENTA selecionar apartamento (OPCIONAL - só se tiver combo)
        {
            "action": "try_fuzzy_lista_fechada",
            "value": "${APARTAMENTO}",
            "container": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                "selector": ".v-menu__content"
            },
            "items_css": ".v-list-item__content",
            "cutoff": 0.80,
            "timeout": 3000
        },

        # 7. TENTA clicar em Consultar (OPCIONAL - só se tiver combo)
        {
            "action": "try_click",
            "locator": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                "selector": "#container-btn-consulta"
            },
            "timeout": 3000
        },

        # 7.5. Aguarda Vue.js processar (espera loading sumir)
        {
            "action": "esperar_selector",
            "locator": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                "selector": "body"
            },
            "timeout": 3000
        },

        # ======================================================================
        # 🚀 ADAPTAÇÃO AQUI: DETECTOR DE OBJETO DE ERRO (ACIR/CONAC)
        # ======================================================================
        {
            "action": "parar_se_visivel",
            "locator": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "css",
                # Mira no ID específico do erro e nos alertas do Vuetify
                "selector": "#mensagem-indisponivel, .v-alert__text, .mensagem-erro"
            },
            "timeout": 10000,  # Tempo maior para o Vue.js processar a consulta (aumentado de 7s para 10s)
            "message": "Mensagem de indisponibilidade de boletos detectada"
        },
        # ======================================================================

        # 9. Loop para baixar boletos (TENTA ambos os seletores!)
        {
            "action": "loop_apartamentos_dinamico",  # ← NOVA AÇÃO ESPECIAL
            "item_spec_links": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "xpath",
                "selector": "//a[contains(@href,'/relatorioboleto/BuscarBoletoPorRecibo')]"
            },
            "stop_flow_if_fail": True,
            "item_spec_buttons": {
                "frame": {"selector": ".prestacao-interativa"},
                "by": "role",
                "role": "button",
                "name": "GERAR BOLETO"
            },
            "identificador": "${IDENTIFICADOR}"
        }
    ],
    # ===== LiveFacilities =====
    "livefacilities": [
        # 1. Menu MEUS DADOS
        {
            "action": "click",
            "locator": {"by": "role", "role": "link", "name": " MEUS DADOS - +"},
            "timeout": 10000
        },

        # 2. Submenu MINHAS UNIDADES
        {
            "action": "click",
            "locator": {"by": "role", "role": "link", "name": "MINHAS UNIDADES"},
            "timeout": 10000
        },

        # 3. Fuzzy match + seleção da unidade
        {
            "action": "try_fuzzy_livefacilities_unidade",
            "value": "${CONDOMINIO}",
            "apartamento": "${APARTAMENTO}",
            "timeout": 5000
        },

        # 4. Navega para 2ª VIA DE BOLETO via JavaScript
        # O link existe no DOM mesmo com o menu recolhido (não precisa abrir o menu).
        # Pega o data-url do elemento e navega direto, bypassa menus, overlays e visibilidade.
        {
            "action": "click",
            "locator": {"by": "css", "selector": "a[data-url*='boletoLista']"},
            "timeout": 15000,
            "js": True
        },

        # 6. Aguarda página carregar
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "body"},
            "timeout": 5000
        },

        # 7. Download do(s) boleto(s) - AGORA CORRIGIDO!
        {
            "action": "download_livefacilities_boletos",
            "identificador": "${IDENTIFICADOR}",
            "timeout": 30000
        }
    ],
    # ===== PROTEL =====
    "protel": [
        # 1. Clica no botão "Segunda Via" (se necessário navegação)
        {
            "action": "try_click",  # Tenta clicar, se já estiver na tela, segue
            "locator": {"by": "css", "selector": "#segunda_via, a:has-text('2ª Via')"},
            "timeout": 5000
        },

        # 2. Aguarda a página carregar (para aparecer boletos OU mensagem de erro)
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "#conteudo"},
            "timeout": 3000
        },

        # 2.5. Verifica se NÃO existe ícone de PDF (sem boletos disponíveis)
        {
            "action": "parar_se_nao_visivel",
            "locator": {"by": "css", "selector": "img[src*='icone_pdf.gif']"},
            "timeout": 2000,
            "stop_flow_if_not_found": True
        },

        # 3. Loop de Download (Reusando lógica de links do Webware)
        # O código antigo buscava imagens: //a//img[contains(@src,'/images/icone_pdf.gif')]
        # Então o link é o elemento 'a' que contém essa imagem.
        # Se não encontrar links, o loop vai falhar naturalmente e capturar a mensagem de erro
        {
            "action": "loop_apartamentos_dinamico",
            "identificador": "${IDENTIFICADOR}",
            "stop_flow_if_fail": True,

            # Cenário 1: Links diretos (PDFs)
            "item_spec_links": {
                "by": "xpath",
                "selector": "//a[descendant::img[contains(@src, 'icone_pdf.gif')]]"
            },

            # Cenário 2: Botões (Deixe um seletor impossível ou genérico se não tiver botões)
            "item_spec_buttons": {
                "by": "css", "selector": "botao_inexistente_fallback"
            }
        }
    ],
    # ===== FERNANDO E FERNANDES =====
    "fernandoefernandes": [
        # 1. Menu "Boletos"
        {
            "action": "click",
            "locator": {"by": "role", "role": "link", "name": "Boletos"},
            "timeout": 5000
        },

        # 2. Submenu "2ª Via de Boleto"
        {
            "action": "click",
            "locator": {"by": "role", "role": "link", "name": "2ª Via de Boleto"},
            "timeout": 5000
        },

        # 3. (NOVO) Se aparecer o aviso, PARA TUDO e lança ERRO
        {
            "action": "parar_se_visivel",
            "locator": {"by": "css", "selector": "p.panel_int"},  # Seletor do aviso "Não há informação..."
            "message": "Aviso do site: Não há boletos disponíveis.",  # Mensagem que vai para o Excel
            "timeout": 5000
        },

        # 4. Loop de Download (Só roda se o passo 3 não encontrar nada)
        {
            "action": "loop_apartamentos_dinamico",
            "identificador": "${IDENTIFICADOR}",
            "stop_flow_if_fail": True,
            "item_spec_buttons": {
                "by": "css",
                "selector": ".fa-download, .fa.fa-download"
            },
            "item_spec_links": {"by": "css", "selector": "a.botao-download-fallback"}
        }
    ],

    # ===== IMODATA (CORREÇÃO NUCLEAR POPUP) =====
    "imodata": [
        # 1. Espera inicial
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "body"},
            "timeout": 5000
        },

        # 2. (NOVO) Remove o popup na força bruta
        # Em vez de tentar achar o botão X, deletamos a div inteira
        {
            "action": "remover_elemento",
            "selector": "#modalAvisoCPF"  # O ID exato que deu erro no seu log
        },

        # 3. Clica em "2ª Via" (Agora o caminho estará livre)
        {
            "action": "click",
            "locator": {"by": "role", "role": "link", "name": "2a.Via", "exact": False},
            "timeout": 10000
        },

        # 4. Aguarda o Frame Principal carregar
        {
            "action": "esperar_selector",
            "locator": {"frame": {"name": "MyFrame"}, "by": "css", "selector": "body"},
            "timeout": 10000
        },

        # 5. Clica no link "Clique aqui para gerar" (Dentro do Frame)
        {
            "action": "try_click",
            "locator": {
                "frame": {"name": "MyFrame"},
                "by": "xpath",
                "selector": "//a[contains(text(), 'Clique aqui para gerar')]"
            },
            "timeout": 5000,
            "stop_flow_if_fail": True
        },

        # 6. Espera a DIV do boleto aparecer visualmente
        {
            "action": "esperar_selector",
            "locator": {
                "frame": {"name": "MyFrame"},
                "by": "css",
                "selector": "div.boleto"
            },
            "timeout": 15000
        },

        # 7. Loop de Popup (Salva o PDF)
        {
            "action": "loop_apartamentos_popup",
            "identificador": "${IDENTIFICADOR}",
            "locator": {
                "frame": {"name": "MyFrame"},
                "by": "css",
                "selector": "div.boleto"
            }
        }
    ],

    # ===== ADMINISTRADORA NACIONAL =====
    "nacional": [
        # NOTA: Fluxo genérico - ajuste conforme a estrutura real do site

        # 1. Aguarda a página logada carregar
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "body"},
            "timeout": 5000
        },

        # 2. Navega para a área de boletos/segunda via
        # (Ajuste o seletor conforme o menu real do site)
        {
            "action": "try_click",
            "locator": {"by": "css",
                        "selector": "a:has-text('2ª Via'), a:has-text('Boleto'), a:has-text('Segunda Via')"},
            "timeout": 5000
        },

        # 3. Aguarda carregar a área de boletos
        {
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "body"},
            "timeout": 5000
        },

        # 4. Download de boletos
        # (Ajuste conforme o sistema real - pode ser um link, botão, etc)
        {
            "action": "download_generico",
            "identificador": "${IDENTIFICADOR}",
            "timeout": 30000
        }
    ],

    "cipa": [
        # PASSO 1: Esperar o card do condomínio aparecer (scope=condo para poder pular)
        {
            "scope": "condo",
            "action": "esperar_selector",
            "locator": {"by": "css", "selector": "div.filter-article"},
            # "state": "visible",
            "timeout": 20000
        },

        # PASSO 2: Selecionar condomínio (Fuzzy)
        {
            "scope": "condo",
            "action": "try_fuzzy_lista_fechada",
            "value": "${CONDOMINIO}",
            "container": {"by": "css", "selector": "condominium-list"},
            "items_css": "div.filter-article",
            "text_selector": "div.text-md.font-semibold.leading-tight",
            "cutoff": 0.85,
            "timeout": 10000
        },

        # PASSO 3: Esperar a troca de página para o Dashboard
        {
            "scope": "condo",
            "action": "esperar_url",
            "pattern": "/dashboard",
            "timeout": 15000
        },

        # PASSO 4: Ir em Unidades para capturar a lista
        {
            "scope": "condo",
            "action": "capturar_unidades_cipa"
        },

        # PASSO 5: Clicar no link de Boletos
        {
            "scope": "condo",
            "action": "click",
            "locator": {"by": "css", "selector": "a[href*='/boletos']"},
            "timeout": 10000
        },

        # PASSO 6: Rodar a função de filtro com fuzzy match (scope=flow - sempre executa)
        {
            "action": "filtro_cipa",
            "apartamento": "${APARTAMENTO}"
        },

        # PASSO 7: Verificar se deu "Vazio"
        {
            "action": "parar_se_visivel",
            "locator": {"by": "css", "selector": "img[src*='boleto-empty.svg']"},
            "timeout": 5000,
            "message": "Nenhum boleto encontrado"
        },

        # PASSO 8: Download
        {
            "action": "loop_cipa_download",
            "identificador": "${IDENTIFICADOR}",
            "locator": {"by": "css", "selector": "img[src*='download']"},
            "timeout": 120000
        }
    ],

}

# ----- MAPEAMENTO: ADMINISTRADORA -> (LOGIN_TYPE, ADMIN_FLOW_TYPE) -----
FLOW_CONFIG = {
    # ===== SUPERLÓGICA PURA (login superlogica + flow superlogica) ===== ↓
    "ABRJ ADMINISTRADORA DE BENS": ("superlogica", "superlogica"),
    "PACIFICA ADMINSTRADORA DE IMOVEIS": ("superlogica", "superlogica"),
    "PROMENADE": ("superlogica", "superlogica"),
    "LUMARJ ADM E LOCAÇÃO": ("superlogica", "superlogica"),

    # ===== Superlógica (migrados do iCondo) ===== ↓
    "ADIPLANTEC": ("superlogica", "superlogica"),
    "HILMAR EMPREENDIMENTOS IMOBILIARIOS": ("superlogica", "superlogica"),
    "ESTASA": ("superlogica", "superlogica"),
    "A.G. RIO IMÓVEIS": ("superlogica", "superlogica"),
    "ADMINISTRADORA REALCAD": ("superlogica", "superlogica"),
    "CENTRIMÓVEIS": ("superlogica", "superlogica"),
    "CINOCRED IMÓVEIS": ("superlogica", "superlogica"),
    "CNI-AJF COSTA NETO CONDOMÍNIOS E LO": ("superlogica", "superlogica"),
    "IRIGON": ("superlogica", "superlogica"),
    "J.B.ANDRADE IMÓVEIS": ("superlogica", "superlogica"),
    "JC CONSULTORIA": ("superlogica", "superlogica"),
    "LUMEN ADMINISTRADORA": ("superlogica", "superlogica"),
    "ML ADMINISTRAÇÃO DE IMÓVEIS": ("superlogica", "superlogica"),
    "NOTA 10 X": ("superlogica", "superlogica"),
    "PERSONAL PRIME ADMINISTRADORA": ("superlogica", "superlogica"),
    "PRECISÃO ADMINISTRADORA DE IMÓVEIS": ("superlogica", "superlogica"),
    "PRED RIO ADMINISTRAÇÃO IMOBILIÁRIA": ("superlogica", "superlogica"),
    "PROHOME ADM E CONSULTORIA IMÓVEIS": ("superlogica", "superlogica"),
    "PROTEST": ("superlogica", "superlogica"),
    "QUALITY HOUSE": ("superlogica", "superlogica"),
    "ZIRTAEB": ("superlogica", "superlogica"),

    # ===== WEBWARE =====
    "CONAC ADMINISTRADORA": ("webware", "webware"),
    # "LOWNDES": ("webware", "webware"),
    "ACIR": ("webware", "webware"),

    # ===== LIVEFACITIES =====
    "BCF": ("livefacilities", "livefacilities"),
    "HFLEX": ("livefacilities", "livefacilities"),
    "VÓRTEX": ("livefacilities", "livefacilities"),
    "BAP": ("livefacilities", "livefacilities"),

    # ===== ADMINISTRADORA NACIONAL =====
    "ADMINISTRADORA NACIONAL": ("nacional", "nacional"),
    "Administradora Nacional": ("nacional", "nacional"),
    "ADM NACIONAL": ("nacional", "nacional"),

    # ===== PROTEL =====
    "PROTEL": ("protel", "protel"),

    # ===== FERNANDO E FERNANDES (migrado para fffacilities.com.br = LiveFacilities) =====
    "FERNANDO E FERNANDES": ("livefacilities", "livefacilities"),

    # ===== IMODATA =====
    "IMODATA": ("imodata", "imodata"),

    # ===== CIPA =====
    "CIPA": ("cipa", "cipa"),

    # ===== ADMINISTRADORA NACIONAL =====
    "ADMINISTRADORA NACIONAL": ("nacional", "nacional"),
}


def get_flow(admin_name: str):
    """
    Retorna o fluxo completo (login + steps) para uma administradora.
    Se não houver configuração específica, usa o padrão Superlógica.

    Returns:
        tuple: (fluxo_completo, login_type)
    """
    config = FLOW_CONFIG.get(admin_name, ("superlogica", "superlogica"))
    login_type, admin_flow_type = config

    login_steps = LOGINS.get(login_type, [])
    admin_steps = ADMIN_FLOWS.get(admin_flow_type, [])

    return (login_steps + admin_steps, login_type)


def carregar_lista_multipla():
    """Lê LISTAMULTIPLAS do .env e devolve mapas: admin->familia e familia->set(admins)."""
    raw = os.getenv("LISTAMULTIPLAS", "[]")
    try:
        arr = json.loads(raw)
    except Exception:
        arr = []
    admin_to_family = {}
    family_to_admins = {}
    for it in arr:
        fam = (it.get("Site") or "").strip()
        adm = (it.get("Administradora") or "").strip()
        if not fam or not adm:
            continue
        admin_to_family[adm] = fam
        family_to_admins.setdefault(fam, set()).add(adm)
    return admin_to_family, family_to_admins


# ===================== MAIN =====================
def main(filtro_adms: list | None = None,
         tabela_injetada=None,
         pasta_download: str | None = None,
         headless: bool = False,
         progress_cb=None):
    """
    Parâmetros extras para integração com a GUI:
      tabela_injetada  — resultado já consultado; se None faz a query interna.
      pasta_download   — substitui DOWNLOADS_DIR; se None usa o padrão do módulo.
      headless         — True para rodar sem janela visível.
      progress_cb      — callable(idx, total, cod, admin) para atualizar a GUI.
    """
    global DOWNLOADS_DIR

    from auxiliares import utils as aux
    load_dotenv()

    if pasta_download:
        DOWNLOADS_DIR = Path(pasta_download)

    preparar_perfil()


    # ===== CARREGA CACHE DE CÓDIGOS DE BARRAS EXISTENTES ===== ↓
    carregar_codigos_barras_existentes()
    # ===== FIM ===== ↑

    admin_to_family, _ = carregar_lista_multipla()

    lista_adms = ['ABRJ ADMINISTRADORA DE BENS'
                  ,'HILMAR EMPREENDIMENTOS IMOBILIARIOS'
                  ,'PACIFICA ADMINSTRADORA DE IMOVEIS'
                  ,'PROMENADE'
                  ,'A.G. RIO IMÓVEIS'
                  ,'ADIPLANTEC'
                  ,'ADMINISTRADORA REALCAD'
                  ,'CENTRIMÓVEIS'
                  ,'CINOCRED IMÓVEIS'
                  ,'CNI-AJF COSTA NETO CONDOMÍNIOS E LO'
                  ,'ESTASA'
                  ,'IRIGON'
                  ,'J.B.ANDRADE IMÓVEIS'
                  ,'JC CONSULTORIA'
                  ,'LUMARJ ADM E LOCAÇÃO'
                  ,'LUMEN ADMINISTRADORA'
                  ,'ML ADMINISTRAÇÃO DE IMÓVEIS'
                  ,'NOTA 10 X'
                  ,'PERSONAL PRIME ADMINISTRADORA'
                  ,'PRECISÃO ADMINISTRADORA DE IMÓVEIS'
                  ,'PRED RIO ADMINISTRAÇÃO IMOBILIÁRIA'
                  ,'PROHOME ADM E CONSULTORIA IMÓVEIS'
                  ,'PROTEST'
                  ,'QUALITY HOUSE'
                  ,'ZIRTAEB'
                  ,'CONAC ADMINISTRADORA'
                  ,'LOWNDES'
                  ,'ACIR'
                  ,'BCF'
                  ,'HFLEX'
                  ,'VÓRTEX'
                  ,'BAP'
                  ,'PROTEL'
                  ,'ADMINISTRADORA NACIONAL'
                  ,'FERNANDO E FERNANDES'
                  ,'IMODATA'
                  ,'CIPA']

    if filtro_adms:
        filtro_upper = [a.upper() for a in filtro_adms]
        lista_adms = [a for a in lista_adms if a.upper() in filtro_upper]
        if not lista_adms:
            log(f"[ERRO] Nenhuma administradora encontrada para o filtro: {filtro_adms}")
            return

    adms_formatados = ", ".join([f"'{adm}'" for adm in lista_adms])

    # SQL pedido
    SQL = f"""SELECT LEFT(TRIM(Tabela.CODIGO),4) AS Codigo, 
                     TRIM(Tabela.LogAdm) AS LogAdm, 
                     TRIM(Tabela.SenAdm) AS SenAdm, 
                     TRIM(Tabela.NomeAdm) AS NomeAdm, 
                     TRIM(Tabela.NomeCond) AS Condomínio, 
                     TRIM(Tabela.Unidade) AS Unidade, 
                     0 AS LoginMultiplo, 
                     '', 
                       '', 
                     '', 
                     '', 
                     ''
                    , REPLACE(REPLACE(REPLACE(TRIM(Prop.CGCCPF), '.', ''), '-', ''), '/', '') AS CPF

              FROM Inquilin_New AS Tabela
              LEFT JOIN Prop
                ON LEFT(TRIM(Tabela.CODIGO),4) = Prop.CODIGO
              WHERE Tabela.Situa NOT IN ('E', 'F', 'K', 'V')                 
                    AND TRIM(Tabela.NomeAdm) IN ({adms_formatados})
              ORDER BY Tabela.NomeAdm, Tabela.LogAdm, Tabela.NomeCond, Tabela.Unidade DESC;"""

    if tabela_injetada is not None:
        # Modo GUI: tabela já veio consultada pelo extracao.py
        tabela = tabela_injetada
    else:
        # Modo standalone: faz a própria query
        banco = aux.Banco(os.path.join(aux.caminhoprojeto(), 'Scai.WMB'))
        tabela = banco.consultar(SQL)

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            user_data_dir=str(USER_DATA_DIR),
            headless=headless,
            accept_downloads=True,
            args=["--no-first-run", "--no-default-browser-check", "--disable-extensions"],
        )

        # Isso sobrescreve a função de imprimir com uma função vazia em todas as abas e popups
        ctx.add_init_script("window.print = function(){};")

        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        # executor = ExecutorFluxoPW(page)
        # Passa o console global para o executor
        executor = ExecutorFluxoPW(page, console_rich=console)

        prev_creds = None
        prev_family = None
        prev_host = None
        prev_condo_base = None
        prev_condo_site = None
        prev_admin = None

        # Converte tabela para lista para permitir look-ahead
        linhas = list(tabela)
        total_linhas = len(linhas)

        # ===== NOVO: CONFIGURAÇÃO DA BARRA DE PROGRESSO FIXA =====
        global _progress_logger
        _progress_logger = FixedProgressLogger(
            total=total_linhas,
            description="[green]Processando condomínios..."
        )
        _progress_logger.start()

        try:
            for idx_linha, linha in enumerate(linhas):
                context = _make_context_from_row(linha)
                adm_raw = (context.get("ADMINISTRADORA") or "").strip()

                same_condo = False

                mensagem_reuso_condo = None  # ← INICIALIZA

                # ==============================================================================
                # 🛑 VALIDAÇÃO PRÉVIA: USUÁRIO E SENHA
                # ==============================================================================
                usr_check = (context.get("USUARIO") or "").strip()
                pwd_check = (context.get("SENHA") or "").strip()

                if not usr_check or not pwd_check:
                    # 1. Monta a mensagem detalhada com Condomínio e Apartamento
                    cond_info = context.get("CONDOMINIO", "???")
                    apto_info = context.get("APARTAMENTO", "???")
                    msg_erro = f"Erro: Usuário e/ou senha não preenchido | Cond.: {cond_info} - Apto.: {apto_info}"

                    codigo_check = context.get("IDENTIFICADOR", "???")
                    log(f"❌ [{codigo_check}] {msg_erro} | Pulando...")

                    # 2. Grava na coluna RESPOSTA
                    _gravar_resposta(linha, msg_erro)

                    # 3. Limpa explicitamente a coluna CHECKERRO para vir vazia (index 9)
                    try:
                        idx_err = _ensure_env_index("CHECKERRO")
                        if idx_err < len(linha):
                            linha[idx_err] = ""
                    except:
                        pass

                    # 4. Grava NOMEFUNCAO e PROBLEMALOGIN (Lógica mantida)
                    try:
                        config_check = FLOW_CONFIG.get(adm_raw, ("superlogica", "superlogica"))
                        idx_nm = _ensure_env_index("NOMEFUNCAO")
                        linha[idx_nm] = config_check[1]

                        idx_pl = _ensure_env_index("PROBLEMALOGIN")
                        linha[idx_pl] = "True"
                    except:
                        pass

                    _progress_logger.advance()  # Avança mesmo pulando
                    if progress_cb:
                        progress_cb(idx_linha, total_linhas,
                                    context.get("IDENTIFICADOR", ""), adm_raw)
                    continue  # ⛔ PULA para a próxima linha da tabela
                # ==============================================================================

                # Busca o fluxo pela função get_flow (retorna também o login_type = família)
                passos_base, family_from_config = get_flow(adm_raw)

                # Família do site: PRIORIDADE para family_from_config (vem do FLOW_CONFIG)
                site_family = family_from_config or ""

                # Fallback: LISTAMULTIPLAS
                if not site_family:
                    site_family = (admin_to_family.get(adm_raw, "") or "").strip()

                # Fallback final: tenta deduzir pelo host
                if not site_family:
                    site_url = (context.get("SITE") or "").strip()
                    host_deste = _host_from_url(site_url)
                    if "superlogica" in host_deste:
                        site_family = "superlogica"
                    elif "livefacilities" in host_deste or "bap.com.br" in host_deste or "servicos.condominio" in host_deste:
                        site_family = "livefacilities"

                if not passos_base:
                    msg = f"Sem fluxo configurado para administradora: {adm_raw}"
                    log(f"[ERRO] {msg}")
                    _gravar_resposta(linha, msg)
                    _progress_logger.advance()
                    if progress_cb:
                        progress_cb(idx_linha, total_linhas,
                                    context.get("IDENTIFICADOR", ""), adm_raw)
                    continue

                if not context.get("SITE"):
                    msg = f"Site não encontrado para administradora: {adm_raw}"
                    log(f"[ERRO] {msg}")
                    _gravar_resposta(linha, msg)
                    _progress_logger.advance()
                    if progress_cb:
                        progress_cb(idx_linha, total_linhas,
                                    context.get("IDENTIFICADOR", ""), adm_raw)
                    continue

                # Deep copy + resolve placeholders
                passos_resolvidos = json.loads(json.dumps(passos_base))

                def _resolve_obj(o):
                    if isinstance(o, dict):
                        return {k: _resolve_obj(_resolve_placeholders(v, context)) for k, v in o.items()}
                    elif isinstance(o, list):
                        return [_resolve_obj(x) for x in o]
                    elif isinstance(o, str):
                        return _resolve_placeholders(o, context)
                    return o

                passos_resolvidos = _resolve_obj(passos_resolvidos)

                # ---- Decisão de login / logout ----
                curr = creds_norm(context)
                target_site = (context.get("SITE") or "").strip()
                target_host = _host_from_url(target_site)

                on_login_page = tela_login_visivel(page, timeout_ms=600)
                family_now = (site_family or "superlogica").lower()
                ja_logado = esta_logado(page, family_now, timeout_ms=900)

                same_creds = (prev_creds is not None and curr == prev_creds)
                host_trocou = (prev_host is not None and target_host and prev_host != target_host)
                family_trocou = (prev_family is not None and family_now != prev_family)

                # DEBUG: Log do estado atual
                log(f"🔍 [DEBUG] family_now={family_now}, prev_creds={prev_creds is not None}, on_login_page={on_login_page}, ja_logado={ja_logado}")

                # ===== LOOK-AHEAD: VERIFICA SE O PRÓXIMO REGISTRO TEM MESMO USUÁRIO ===== ↓
                proximo_tem_mesmo_usuario = False
                proximo_tem_mesmo_host = False
                proximo_tem_mesma_family = False

                if idx_linha + 1 < total_linhas:
                    proxima_linha = linhas[idx_linha + 1]
                    proximo_context = _make_context_from_row(proxima_linha)
                    proximo_creds = creds_norm(proximo_context)
                    proximo_site = (proximo_context.get("SITE") or "").strip()
                    proximo_host = _host_from_url(proximo_site)

                    # Detecta família do próximo
                    proximo_adm = (proximo_context.get("ADMINISTRADORA") or "").strip()
                    proximo_family = (admin_to_family.get(proximo_adm, "") or "").strip()
                    if not proximo_family:
                        if "superlogica" in proximo_host:
                            proximo_family = "superlogica"
                        elif "livefacilities" in proximo_host or "bap.com.br" in proximo_host or "servicos.condominio" in proximo_host:
                            proximo_family = "livefacilities"
                    proximo_family = (proximo_family or "superlogica").lower()

                    proximo_tem_mesmo_usuario = (curr == proximo_creds)
                    proximo_tem_mesmo_host = (target_host == proximo_host) if (target_host and proximo_host) else False
                    proximo_tem_mesma_family = (family_now == proximo_family)
                # ===== FIM LOOK-AHEAD ===== ↑

                # ===== DETECTA ABOUT:BLANK ===== ↓
                url_atual = page.url.lower()
                esta_em_blank = url_atual in ("about:blank", "")
                # ===== FIM ===== ↑

                # ===== LÓGICA ESPECIAL PARA CIPA ===== ↓
                if "cipa" in family_now:  # "cipa" ou "cipafacil"
                    condo_alvo = context.get("CONDOMINIO", "")
                    condo_alvo_norm = _norm(condo_alvo)

                    # Compara condomínio atual vs alvo
                    same_condo_cipa = False
                    if prev_condo_base is not None and prev_admin == adm_raw:
                        if _norm(prev_condo_base) == condo_alvo_norm:
                            same_condo_cipa = True
                        else:
                            # Fuzzy
                            prev_norm = _norm(prev_condo_site) if prev_condo_site else ""
                            w_score = fuzz.WRatio(prev_norm, condo_alvo_norm) / 100.0
                            p_score = fuzz.partial_ratio(prev_norm, condo_alvo_norm) / 100.0
                            if max(w_score, p_score) >= 0.80:
                                same_condo_cipa = True

                    # ===== DECISÃO CIPA ===== ↓
                    # Caso 0: prev_creds é None (primeira execução ou após reset)
                    # Verifica se já está logado com o usuário correto usando verificar_sessao_cipa
                    if prev_creds is None:
                        # Não sabemos o estado anterior - deixa o login_steps decidir
                        # O passo verificar_sessao_cipa vai detectar se já está logado
                        log(f"ℹ️  [CIPA] Primeira execução ou após reset - verificando sessão via login_steps")
                        need_login = True  # Vai executar login_steps que inclui verificar_sessao_cipa

                    elif same_creds and same_condo_cipa:
                        # Caso 1: Usuário E condomínio iguais → NÃO FAZ NADA
                        log(f"♻️  [CIPA] Usuário e condomínio iguais - pulando logout/navegação")
                        need_login = False
                        same_condo = True  # pula passos scope=condo na filtragem abaixo

                    elif same_creds and not same_condo_cipa:
                        # Caso 2: Usuário igual, condomínio diferente → NAVEGA DIRETO PARA SELEÇÃO
                        log(f"♻️  [CIPA] Mesmo usuário, condomínio diferente - navegando para seleção de condomínios")

                        if not on_login_page and not esta_em_blank:
                            try:
                                # NAVEGAÇÃO DIRETA em vez de clicar no menu "Sair"
                                # (o botão "Sair" do menu faz logout completo!)
                                log(f"   🏢 Navegando para área de condomínios...")
                                _navegar_e_esperar(page, "https://cipafacil.digital/condominios")
                                log(f"   ✅ Na área de seleção de condomínios")

                            except Exception as e:
                                log(f"   ⚠️ Erro ao navegar: {e}")

                        need_login = False  # NÃO precisa fazer login!

                    else:
                        # Caso 3: Usuário diferente → LOGOUT COMPLETO
                        need_login = True

                        if not on_login_page and not esta_em_blank:
                            log(f"🔄 [LOGOUT] Usuário diferente - fazendo logout completo...")
                            try:
                                efetuar_logout(page, target_site)
                            except Exception as e:
                                log(f"⚠️ [LOGOUT] Erro: {e}")
                    # ===== FIM DECISÃO CIPA ===== ↑

                # ===== LÓGICA PADRÃO PARA OUTRAS PLATAFORMAS ===== ↓
                else:
                    need_login = False
                    skip_logout = False

                    # ===== LIVEFACILITIES: LÓGICA ESPECIAL ===== ↓
                    if family_now == "livefacilities":
                        if prev_creds is None:
                            # Primeira execução - força login limpo
                            log(f"🔄 [LIVEFACILITIES] Primeira execução - verificando estado...")
                            if not on_login_page and not esta_em_blank:
                                log(f"   → Forçando logout para estado limpo...")
                                need_login = True
                                skip_logout = False  # FORÇA logout
                            else:
                                log(f"   → Já está em tela de login ou blank, não precisa logout")
                                need_login = True
                                skip_logout = True
                        elif same_creds and not host_trocou and not family_trocou:
                            # Mesmas credenciais - REUSA sessão (não faz logout nem login)
                            # Ignora ja_logado pois pode ser falso positivo após reload
                            log(f"♻️  [LIVEFACILITIES] Mesmas credenciais - reusando sessão")
                            need_login = False
                            skip_logout = True
                        else:
                            # Credenciais diferentes - precisa logout e login
                            need_login = True
                            skip_logout = False
                    # ===== FIM LIVEFACILITIES ===== ↑

                    # ===== OUTRAS PLATAFORMAS ===== ↓
                    else:
                        if host_trocou or family_trocou or (not same_creds) or on_login_page or (not ja_logado):
                            need_login = True
                    # ===== FIM ===== ↑

                    # ===== OTIMIZAÇÃO: NÃO FAZ LOGOUT NEM LOGIN SE PRÓXIMO FOR IGUAL ===== ↓
                    # Se o PRÓXIMO registro tem mesmo usuário/host/family, assume sessão válida e reusa
                    # (evita logout E login desnecessários quando processa vários registros do mesmo usuário)
                    if need_login and proximo_tem_mesmo_usuario and proximo_tem_mesmo_host and proximo_tem_mesma_family:
                        # Próximo é igual → assume que está logado e reusa a sessão
                        # MAS só se JÁ TIVER FEITO LOGIN ANTES (prev_creds não é None)
                        # E só se realmente estiver logado no site certo (ja_logado=True)
                        # E só se a família/host não trocou (browser pode estar no site errado)
                        if not on_login_page and prev_creds is not None and ja_logado and not family_trocou and not host_trocou:
                            # Não está na tela de login, já logou antes, e está no site certo → REUSA a sessão
                            need_login = False  # NÃO faz login (evita recarregar página e matar sessão)
                            skip_logout = True  # NÃO faz logout
                        elif on_login_page or prev_creds is None:
                            # Está na página de login OU é o primeiro registro → precisa fazer login
                            need_login = True
                            skip_logout = True  # Pula logout (será feito no início se for LiveFacilities)
                    # ===== FIM OTIMIZAÇÃO ===== ↑

                    # Força logout se necessário

                    if need_login and not on_login_page and not esta_em_blank and not skip_logout:
                        log(f"🔄 [LOGOUT] Fazendo logout antes de novo login...")
                        try:
                            efetuar_logout(page, target_site)
                        except Exception as e:
                            log(f"⚠️ [LOGOUT] Erro: {e}")
                    elif need_login and on_login_page:
                        log(f"ℹ️  [LOGIN] Já está na tela de login - pulando logout")
                    elif need_login and esta_em_blank:
                        log(f"ℹ️  [INÍCIO] Browser em about:blank - pulando logout")
                    elif not need_login:
                        log(f"♻️  [REUSO] Mantendo login existente - pulando login/logout")
                # ===== FIM LÓGICA PADRÃO ===== ↑

                # ====================================================================
                # 👇👇 INSERIR ISTO AQUI 👇👇
                # CORREÇÃO: Garante que estamos na URL certa antes de tentar logar
                # (Resolve o problema de abrir o browser travado na tela de logout antiga)
                # ====================================================================
                if need_login and "livefacilities" in (site_family or "").lower():
                    log(f"   🛡️ [STARTUP] Forçando limpeza de sessão (garantia de Logoff)...")
                    try:
                        # 1. Limpa Cookies (sempre funciona)
                        page.context.clear_cookies()

                        # 2. Tenta limpar Storage (só se a página permitir)
                        try:
                            page.evaluate("() => { localStorage.clear(); sessionStorage.clear(); }")
                        except Exception:
                            pass  # Ignora se a página não permite (about:blank, etc)

                        # 3. Força navegação para a URL de login limpa
                        target_url = (context.get("SITE") or "").strip()
                        if target_url:
                            log(f"       ↪️ Recarregando página de login: {target_url}")
                            page.goto(target_url)
                            page.wait_for_load_state("networkidle")
                            # Pequena pausa para garantir que o formulário renderize
                            page.wait_for_timeout(1500)

                    except Exception as e:
                        log(f"       ⚠️ Erro na limpeza preventiva: {e}")
                # ====================================================================

                # Filtra por scope
                passos_a_executar = []
                for st in passos_resolvidos:
                    # scope = st.get("scope", "always")
                    # if scope == "login":
                    #     if need_login:
                    #         passos_a_executar.append(st)
                    # else:
                    #     passos_a_executar.append(st)
                    # Pega o escopo. Se não existir, o padrão é "always" (sempre executa)
                    scope = st.get("scope", "always")

                    # 1. Regra de Login (já existia, mantemos igual)
                    if scope == "login" and not need_login:
                        continue

                    # 2. Regra de Condomínio (Nova e Segura)
                    # Só vai entrar aqui se você ESPECIFICAR no JSON que o passo é de condomínio
                    if scope == "condo" and same_condo:
                        print(f"   ⚡ [OTIMIZAÇÃO] Pulando passo '{st.get('action')}' pois o condomínio já é o correto.")
                        continue

                    passos_a_executar.append(st)

                # ===== OTIMIZAÇÃO: REUSO DE CONDOMÍNIO (COM FUZZY!) ===== ↓
                condo_alvo = context.get("CONDOMINIO", "")
                condo_alvo_norm = _norm(condo_alvo)

                # Verifica se pode reusar condomínio já selecionado
                same_condo = False
                reuso_metodo = None  # ← Captura método usado
                reuso_score = 0.0  # ← Captura score
                reuso_w_score = 0.0  # ← WRatio
                reuso_p_score = 0.0  # ← partial_ratio

                # ===== CAPTURA FLOW_TYPE ANTES (NECESSÁRIO PARA DECISÃO DE REUSO) ===== ↓
                config = FLOW_CONFIG.get(adm_raw, ("superlogica", "superlogica"))
                login_type, admin_flow_type = config
                # ===== FIM ===== ↑

                # ADICIONADO: Verifica se NÃO é LiveFacilities E se o FLOW é Superlógica
                # LiveFacilities precisa selecionar a unidade SEMPRE, não existe "reuso de condomínio"
                # iCondo puro também não reusa, MAS iCondo com flow Superlógica (Estasa, Hilmar, etc) SIM!
                eh_livefacilities = "livefacilities" in family_now
                eh_flow_superlogica = "superlogica" in admin_flow_type.lower()

                if (prev_condo_base is not None and
                        prev_condo_site is not None and
                        prev_admin == adm_raw and
                        not need_login and
                        not eh_livefacilities and
                        eh_flow_superlogica):  # ← NOVA CONDIÇÃO!

                    # 1. Tenta comparação direta (rápido!)
                    if _norm(prev_condo_base) == condo_alvo_norm:
                        same_condo = True
                        reuso_metodo = "Comparação direta"
                        reuso_score = 1.0
                        reuso_w_score = 1.0
                        reuso_p_score = 1.0
                        log(f"♻️  [OTIMIZAÇÃO] Comparação direta: '{condo_alvo}' == '{prev_condo_base}'")

                    # 2. Tenta fuzzy match (robusto!)
                    else:
                        prev_norm = _norm(prev_condo_site)
                        w_score = fuzz.WRatio(prev_norm, condo_alvo_norm) / 100.0
                        p_score = fuzz.partial_ratio(prev_norm, condo_alvo_norm) / 100.0
                        fuzzy_score = max(w_score, p_score)

                        if fuzzy_score >= 0.80:  # Mesmo cutoff do fuzzy_lista_fechada
                            same_condo = True
                            reuso_score = fuzzy_score
                            reuso_w_score = w_score
                            reuso_p_score = p_score

                            # Define método usado (mesma lógica do fuzzy_lista_fechada)
                            if w_score >= 0.80:
                                reuso_metodo = "WRatio"
                            else:
                                reuso_metodo = "max(WRatio, partial)"

                            log(f"♻️  [OTIMIZAÇÃO] Fuzzy match: '{condo_alvo}' ~= '{prev_condo_site}' (score={fuzzy_score:.2f})")

                if same_condo:
                    # ===== MENSAGEM ESPECÍFICA PARA iCondo+Superlógica ===== ↓
                    if "icondo" in login_type.lower() and eh_flow_superlogica:
                        log(f"   → Pulando seleção/expansão de condomínio (iCondo + flow Superlógica)")
                    else:
                        log(f"   → Pulando seleção/expansão de condomínio")
                    # ===== FIM ===== ↑

                    # ===== ADICIONA MENSAGEM FORMATADA AO MATCH_LOG ===== ↓
                    lines = []
                    lines.append(f"[CONDOMÍNIO REUSADO] ♻️  (score={reuso_score:.2f}):")
                    lines.append(f"                     📋 Buscado na BASE:   '{condo_alvo}'")
                    lines.append(f"                     🌐 Já selecionado:    '{prev_condo_site}'")
                    lines.append(
                        f"                     📊 WRatio: {reuso_w_score:.2f} | partial_ratio: {reuso_p_score:.2f}")
                    lines.append(f"                     🎯 Método usado: {reuso_metodo}")
                    lines.append(f"                     ⚡ Pulou seleção/expansão (otimização)")

                    # Adiciona info do tipo de login/flow se for híbrido
                    if "icondo" in login_type.lower() and eh_flow_superlogica:
                        lines.append(f"                     🔧 Login: {login_type} | Flow: {admin_flow_type}")

                    # Guarda para adicionar DEPOIS na coluna RESPOSTA
                    mensagem_reuso_condo = "\n\n" + "\n".join(lines)
                    # ===== FIM ===== ↑

                    acoes_para_pular = {
                        'fuzzy_lista_fechada',  # Seleção do condomínio
                        'expandir_lista',  # "Mais itens"
                        'esperar_selector'  # Aguardar área de cobranças
                    }

                    passos_otimizados = []
                    for passo in passos_a_executar:
                        action = passo.get("action")

                        # Mantém verificações de erro (sempre executar)
                        if action == "parar_se_visivel":
                            passos_otimizados.append(passo)
                        # Pula seleção/expansão de condomínio
                        elif action not in acoes_para_pular:
                            passos_otimizados.append(passo)

                    passos_a_executar = passos_otimizados

                # ===== FIM OTIMIZAÇÃO ===== ↑

                # ===== INICIA PROCESSAMENTO ===== ↓
                codigo = context.get("IDENTIFICADOR", "????")
                cond = context.get("CONDOMINIO", "")
                apto = context.get("APARTAMENTO", "")
                usuario = context.get("USUARIO", "???")

                # NOTA: config/login_type/admin_flow_type já foram capturados acima (linha ~4120)
                # para serem usados na lógica de reuso de condomínio

                log(f"\n{'=' * 70}")
                log(f"🔹 PROCESSANDO [{codigo}] {cond} - {apto}")
                log(f"   👤 Usuário: {usuario}")
                log(f"{'=' * 70}")

                # Executa
                executor.reset_status()
                executor.codigo_cliente = codigo  # ← Define código

                # ===== PRESERVA condo_selecionado_site (SE pulou fuzzy) ===== ↓
                if same_condo:
                    executor.condo_selecionado_site = prev_condo_site
                # ===== FIM ===== ↑

                executor.executar(passos_a_executar, context)
                st = executor.get_status()

                # ===== GRAVA NOVAS COLUNAS NO DATAFRAME ===== ↓

                # ===== COLUNA NOMEFUNCAO (nome do fluxo) ===== ↓
                try:
                    idx_nome = _ensure_env_index("NOMEFUNCAO")

                    # Se erro de login: mostra login_type
                    # Senão: mostra admin_flow_type
                    if st.get("login_ok") is False:
                        nome_fluxo = login_type  # Ex: "icondo"
                    else:
                        nome_fluxo = admin_flow_type  # Ex: "superlogica"

                    if idx_nome < len(linha):
                        linha[idx_nome] = nome_fluxo
                    else:
                        while len(linha) <= idx_nome:
                            linha.append("")
                        linha[idx_nome] = nome_fluxo

                except Exception as e:
                    log(f"⚠️ Erro ao gravar NOMEFUNCAO: {e}")

                # Coluna MATCH_LOG
                try:
                    idx_match = _ensure_env_index("CHECKERRO")
                    if idx_match < len(linha):
                        linha[idx_match] = st.get("match_log", "")
                    else:
                        # Expande lista se necessário
                        while len(linha) <= idx_match:
                            linha.append("")
                        linha[idx_match] = st.get("match_log", "")
                except Exception as e:
                    log(f"⚠️ Erro ao gravar MATCH_LOG: {e}")

                # Coluna PROBLEMALOGIN
                try:
                    idx_login = _ensure_env_index("PROBLEMALOGIN")
                    valor_login = st.get("login_ok")

                    # ===== LÓGICA INVERTIDA (coluna é PROBLEMA, variável é OK) ===== ↓
                    if valor_login is True:
                        texto_login = "False"  # Login OK → NÃO tem problema
                    elif valor_login is False:
                        texto_login = "True"  # Login erro → SIM tem problema
                    elif not need_login:
                        # Reusou login (não precisou logar) → considera como OK
                        texto_login = "False"
                    else:
                        texto_login = ""  # Outros casos (não deveria acontecer)
                    # ===== FIM ===== ↑

                    if idx_login < len(linha):
                        linha[idx_login] = texto_login
                    else:
                        while len(linha) <= idx_login:
                            linha.append("")
                        linha[idx_login] = texto_login
                except Exception as e:
                    log(f"⚠️ Erro ao gravar PROBLEMALOGIN: {e}")

                # Coluna ERROS
                try:
                    idx_erros = _ensure_env_index("CHECKERRO")
                    if idx_erros < len(linha):
                        linha[idx_erros] = st.get("erros", "")
                    else:
                        while len(linha) <= idx_erros:
                            linha.append("")
                        linha[idx_erros] = st.get("erros", "")
                except Exception as e:
                    log(f"⚠️ Erro ao gravar ERROS: {e}")
                # ===== FIM ===== ↑

                # Atualiza baseline pós-execução
                # ===== OTIMIZAÇÃO: ATUALIZA BASELINE MESMO EM ERRO DE NEGÓCIO ===== ↓
                # Se próximo é igual, assume que está logado (evita chamada desnecessária a esta_logado)
                atualiza_baseline = False

                if proximo_tem_mesmo_usuario and proximo_tem_mesmo_host and proximo_tem_mesma_family:
                    # Próximo é igual → atualiza baseline mesmo que esta_logado() retorne False
                    atualiza_baseline = True
                    # log(f"   ✅ [BASELINE] Atualizando (próximo é igual)")
                else:
                    # Próximo é diferente → verifica se está logado
                    if esta_logado(page, family_now, timeout_ms=900):
                        atualiza_baseline = True
                        # log(f"   ✅ [BASELINE] Atualizando (verificou sessão)")
                    # else:
                    # log(f"   ⚠️  [BASELINE] NÃO atualizando (sessão inválida)")

                if atualiza_baseline:
                    prev_creds = curr
                    prev_family = family_now
                    prev_host = target_host
                    # ===== ATUALIZA BASELINE DE CONDOMÍNIO ===== ↓
                    # Atualiza se OK OU se condomínio foi selecionado (mesmo com erro no apto)
                    condo_foi_selecionado = st.get("condo_site") is not None
                    if st["ok"] or condo_foi_selecionado:
                        prev_condo_base = condo_alvo  # Nome da base
                        prev_condo_site = st.get("condo_site")  # Nome do site (fuzzy retornou)
                        prev_admin = adm_raw

                        if not st["ok"] and condo_foi_selecionado:
                            log(f"   ⚠️  [BASELINE] Condomínio salvo apesar do erro (erro foi no apartamento)")

                    # ===== FIM ===== ↑
                # ===== FIM OTIMIZAÇÃO ===== ↑

                if st["ok"]:
                    qtd = st.get("baixados", 0)
                    qtd_dup = st.get("duplicados", 0)
                    ok_msg = f"{qtd} boleto{'s' if qtd > 1 else ''} baixado{'s' if qtd > 1 else ''}"
                    if qtd_dup > 0:
                        ok_msg += f" ({qtd_dup} duplicado{'s' if qtd_dup > 1 else ''} descartado{'s' if qtd_dup > 1 else ''})"
                    log(f"\n ✅ [{codigo}] {ok_msg} | Cond.: {cond} - Apto.: {apto}")
                    log(f"{'=' * 70}\n")
                    resposta_final = f"✅ {ok_msg} | Cond.: {cond} - Apto.: {apto}"
                else:
                    err_msg = f"ERRO: {st['resposta']}"
                    log(f"\n❌ [{codigo}] {err_msg} | {cond} - {apto}")
                    log(f"{'=' * 70}\n")
                    resposta_final = f"❌ {err_msg} | Cond.: {cond} - Apto.: {apto}"

                # ===== ADICIONA FUZZY STATS + MENSAGEM DE REUSO (ORDEM CORRETA) ===== ↓
                # 1. Adiciona mensagem de reuso PRIMEIRO (se houver)
                #    Isso coloca informação do condomínio antes do apartamento
                if mensagem_reuso_condo:
                    resposta_final += mensagem_reuso_condo

                # 2. Adiciona match_log DEPOIS
                #    - Se reusou: contém só apartamento
                #    - Se não reusou: contém condomínio + apartamento
                match_log = st.get("match_log", "")
                if match_log:
                    resposta_final += "\n\n" + match_log

                _gravar_resposta(linha, resposta_final)
                # ===== FIM ===== ↑

                # ====================================================================
                # RESETA BASELINE APENAS EM CASO DE ERRO DE LOGIN/SESSÃO
                # Erros de negócio (sem boletos, boleto vencido) NÃO invalidam a sessão
                # ====================================================================

                if not st["ok"]:
                    # Verifica se é erro de login ou sessão inválida
                    login_falhou = st.get("login_ok") is False

                    # ===== OTIMIZAÇÃO: NÃO VERIFICA SESSÃO SE PRÓXIMO FOR IGUAL ===== ↓
                    # Se o próximo registro tem mesmo usuário/host/family, NÃO precisa verificar sessão
                    # (evita reset desnecessário da baseline)
                    if proximo_tem_mesmo_usuario and proximo_tem_mesmo_host and proximo_tem_mesma_family:
                        # Próximo é igual → assume que sessão está válida
                        sessao_invalida = False
                        # log(f"ℹ️  [OTIMIZAÇÃO] Próximo registro é igual - assumindo sessão válida")
                    else:
                        # Próximo é diferente ou não existe → verifica sessão
                        # EXCETO para CIPA: erros de negócio (unidade não encontrada, etc) não invalidam sessão
                        if "cipa" in family_now.lower():
                            # CIPA: só considera sessão inválida se login_falhou
                            sessao_invalida = login_falhou
                        else:
                            sessao_invalida = not esta_logado(page, family_now, timeout_ms=900)
                    # ===== FIM OTIMIZAÇÃO ===== ↑

                    if login_falhou or sessao_invalida:
                        log(f"🔄 [ERRO DE LOGIN/SESSÃO] Resetando baseline para forçar login na próxima linha.")

                        # Limpa variáveis de memória para forçar novo login
                        prev_creds = None
                        prev_family = None
                        prev_host = None
                        prev_condo_base = None
                        prev_condo_site = None
                        prev_admin = None
                    else:
                        # Erro de negócio (sem boletos, boleto vencido, etc.)
                        # Sessão ainda válida - mantém baseline para reaproveitar login
                        log(f"ℹ️  [ERRO DE NEGÓCIO] Sessão válida - mantendo login para próxima linha.")
                # ====================================================================

                # ===== AVANÇA BARRA DE PROGRESSO =====
                _progress_logger.advance()
                if progress_cb:
                    progress_cb(idx_linha, total_linhas,
                                context.get("IDENTIFICADOR", ""), adm_raw)
                # ===== FIM =====

        finally:
            # Para o logger ao terminar
            if _progress_logger:
                _progress_logger.stop()
                _progress_logger = None

        ctx.close()

        # ==============================================================================
        # 🕒 TIMESTAMP ÚNICO PARA AMBOS OS ARQUIVOS
        # ==============================================================================
        timestamp_comum = datetime.now().strftime("%Y%m%d_%H%M%S")
        PASTA_LOGS = "logs"  # Define a pasta de destino

        # Garante que a pasta existe (embora na sua imagem já exista)
        if not os.path.exists(PASTA_LOGS):
            os.makedirs(PASTA_LOGS)
        # ==============================================================================

        # ===== SALVA DATAFRAME DE BOLETOS ===== ↓
        if executor.boletos_processados:
            log("\n" + "=" * 60)
            log("SALVANDO DATAFRAME DE BOLETOS...")
            log("=" * 60)

            df_boletos = pd.DataFrame(executor.boletos_processados)

            # Ordena por vencimento
            df_boletos['Vencimento_dt'] = pd.to_datetime(df_boletos['Vencimento'], format="%d/%m/%Y")
            df_boletos = df_boletos.sort_values('Vencimento_dt')
            df_boletos = df_boletos.drop('Vencimento_dt', axis=1)

            df_boletos_excel = df_boletos.drop(['Pasta_Mes', 'Caminho_Completo'], axis=1)

            # --- ALTERAÇÃO: Caminho aponta para pasta logs ---
            nome_arq_boletos = f"Boletos_Processados_{timestamp_comum}.xlsx"
            arquivo_boletos = os.path.join(PASTA_LOGS, nome_arq_boletos)

            # Salva Excel
            df_boletos_excel.to_excel(arquivo_boletos, index=False, engine='openpyxl')

            # ===== FORMATA EXCEL ===== ↓
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Alignment

                wb = load_workbook(arquivo_boletos)
                ws = wb.active

                # Wrap text em todas as colunas
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Ajusta largura das colunas
                ws.column_dimensions['A'].width = 12  # Cliente
                ws.column_dimensions['B'].width = 50  # Cod_Barras
                ws.column_dimensions['C'].width = 12  # Tipo_Cod_B
                ws.column_dimensions['D'].width = 25  # Nome_Arquivo
                ws.column_dimensions['E'].width = 50  # Linha_Digitavel
                ws.column_dimensions['F'].width = 15  # Valor
                ws.column_dimensions['G'].width = 15  # Vencimento
                # ws.column_dimensions['H'].width = 20  # Pasta_Mes
                # ws.column_dimensions['I'].width = 60  # Caminho_Completo

                wb.save(arquivo_boletos)
                log("✅ Formatação Excel aplicada!")

            except Exception as e:
                log(f"⚠️ Erro ao formatar Excel: {e}")
            # ===== FIM FORMATAÇÃO ===== ↑

            log(f"\n✅ {len(df_boletos)} boletos processados salvos em: {arquivo_boletos}")

            # ===== RESUMO POR MÊS ===== ↓
            log(f"\n📁 RESUMO POR MÊS:")
            resumo = df_boletos.groupby('Pasta_Mes').agg({
                'Cliente': 'count',
                'Valor': 'sum'
            }).rename(columns={'Cliente': 'Qtd_Boletos', 'Valor': 'Total_R$'})

            # Formata coluna de valor
            resumo['Total_R$'] = resumo['Total_R$'].apply(lambda x: f"R$ {x:,.2f}")

            log(resumo.to_string())
            log("=" * 60 + "\n")
            # ===== FIM RESUMO ===== ↑
        else:
            log("\n⚠️ Nenhum boleto foi processado com sucesso.\n")
        # ===== FIM ===== ↑

        # ===== SALVA BOLETOS DUPLICADOS (SE HOUVER) ===== ↓
        if BOLETOS_DUPLICADOS:
            log("\n" + "=" * 60)
            log("⚠️ BOLETOS DUPLICADOS DETECTADOS")
            log("=" * 60)

            df_duplicados = pd.DataFrame(BOLETOS_DUPLICADOS)

            nome_arq_duplicados = f"Boletos_Duplicados_{timestamp_comum}.xlsx"
            arquivo_duplicados = os.path.join(PASTA_LOGS, nome_arq_duplicados)

            df_duplicados.to_excel(arquivo_duplicados, index=False, engine='openpyxl')

            log(f"\n⚠️ {len(BOLETOS_DUPLICADOS)} boletos duplicados encontrados e descartados!")
            log(f"📋 Lista salva em: {arquivo_duplicados}")
            log("=" * 60 + "\n")
        # ===== FIM DUPLICADOS ===== ↑

        log("\n" + "=" * 60)
        log("SALVANDO RESULTADOS...")
        log("=" * 60)

        colunas = [
            "IDENTIFICADOR", "USUARIO", "SENHA", "ADMINISTRADORA",
            "CONDOMINIO", "APARTAMENTO", "LOGINMULTIPLO", "RESPOSTA",
            "CHECKARQUIVO", "CHECKERRO", "NOMEFUNCAO", "PROBLEMALOGIN"
        ]

        # Remove a coluna CPF (índice 12) antes de criar o DataFrame
        tabela_sem_cpf = [linha[:12] for linha in tabela]
        df = pd.DataFrame(tabela_sem_cpf, columns=colunas)

        # --- ALTERAÇÃO: Caminho aponta para pasta logs ---
        nome_arq_resultado = f"Resultado_Boletos_{timestamp_comum}.xlsx"
        arquivo = os.path.join(PASTA_LOGS, nome_arq_resultado)

        df.to_excel(arquivo, index=False, engine='openpyxl')

        # ===== CONFIGURA QUEBRA DE LINHA NO EXCEL ===== ↓
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        wb = load_workbook(arquivo)
        ws = wb.active

        # Pega índices do .env
        try:
            idx_match = _ensure_env_index("NOMEFUNCAO")  # 10
            idx_erros = _ensure_env_index("CHECKERRO")  # 9
            idx_resp = _ensure_env_index("RESPOSTA")  # 7

            # Converte para letras de coluna Excel
            col_match = _idx_to_excel_col(idx_match)  # 'K'
            col_erros = _idx_to_excel_col(idx_erros)  # 'J'
            col_resp = _idx_to_excel_col(idx_resp)  # 'H'

            # Configura wrap text nas colunas de texto longo
            for row in ws.iter_rows(min_row=2):  # Pula cabeçalho
                # NOMEFUNCAO
                cell_match = row[idx_match]
                if cell_match.value:
                    cell_match.alignment = Alignment(wrap_text=True, vertical='top')

                # CHECKERRO
                cell_erros = row[idx_erros]
                if cell_erros.value:
                    cell_erros.alignment = Alignment(wrap_text=True, vertical='top')

                # RESPOSTA
                cell_resp = row[idx_resp]
                if cell_resp.value:
                    cell_resp.alignment = Alignment(wrap_text=True, vertical='top')

            # Ajusta largura das colunas (baseado nos índices do .env)
            ws.column_dimensions[col_match].width = 80  # NOMEFUNCAO
            ws.column_dimensions[col_erros].width = 60  # CHECKERRO
            ws.column_dimensions[col_resp].width = 50  # RESPOSTA

            # Ajusta altura padrão das linhas (auto-ajuste)
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = None  # Auto

            wb.save(arquivo)
            log("✅ Formatação Excel aplicada com sucesso!")

        except Exception as e:
            log(f"⚠️ Erro ao formatar Excel (não crítico): {e}")
            # Salva mesmo com erro de formatação
            try:
                wb.save(arquivo)
            except:
                pass
        # ===== FIM ===== ↑

        log(f"\n✅ {len(tabela)} registros salvos em: {arquivo}")
        log("=" * 60 + "\n")
        # ===== FIM ===== ↑

        # ===== ALERTA: ADMINISTRADORAS COM ITENS COM PROBLEMA ===== ↓
        try:
            from collections import defaultdict

            idx_adm   = ADMINISTRADORA   # 3
            idx_resp  = RESPOSTA         # 7
            idx_id    = 0                # IDENTIFICADOR
            idx_condo = 4                # CONDOMINIO
            idx_apto  = 5                # APARTAMENTO
            idx_login = PROBLEMALOGIN    # 11

            # Agrupa linhas por administradora
            grupos_adm = defaultdict(list)
            for linha in tabela:
                adm = (linha[idx_adm] if idx_adm < len(linha) else "") or ""
                if adm:
                    grupos_adm[adm].append(linha)

            adms_com_problema = []

            for adm, linhas in grupos_adm.items():
                if not linhas:
                    continue

                itens_com_erro = []

                for l in linhas:
                    resp       = (l[idx_resp]  if idx_resp  < len(l) else "") or ""
                    login_prob = (l[idx_login] if idx_login < len(l) else "") or ""
                    codigo     = (l[idx_id]    if idx_id    < len(l) else "") or ""
                    condo      = (l[idx_condo] if idx_condo < len(l) else "") or ""
                    apto       = (l[idx_apto]  if idx_apto  < len(l) else "") or ""

                    tem_erro = resp.startswith("❌") or (resp and not resp.startswith("✅"))

                    if tem_erro:
                        # Extrai motivo curto da RESPOSTA
                        motivo = resp
                        if motivo.startswith("❌"):
                            motivo = motivo[1:].strip()
                        if motivo.upper().startswith("ERRO:"):
                            motivo = motivo[5:].strip()
                        # Remove sufixo "| Cond.: ..." que é redundante
                        if " | Cond.:" in motivo:
                            motivo = motivo.split(" | Cond.:")[0].strip()
                        # Só primeira linha (ignora match_log abaixo)
                        motivo = motivo.split("\n")[0].strip()
                        # Trunca se muito longo
                        if len(motivo) > 90:
                            motivo = motivo[:87] + "..."

                        itens_com_erro.append({
                            "codigo"     : codigo,
                            "condo"      : condo,
                            "apto"       : apto,
                            "motivo"     : motivo,
                            "login_prob" : login_prob == "True",
                        })

                # Reporta se houver ao menos 1 erro que não seja só de login/senha
                if itens_com_erro:
                    todos_login = all(i["login_prob"] for i in itens_com_erro)
                    if not todos_login:
                        adms_com_problema.append({
                            "adm"        : adm,
                            "total"      : len(linhas),
                            "com_erro"   : len(itens_com_erro),
                            "todos_erro" : len(itens_com_erro) == len(linhas),
                            "itens"      : itens_com_erro,
                        })

            if adms_com_problema:
                log("\n" + "=" * 60)
                log("🔴 ATENÇÃO: ADMINISTRADORAS COM ITENS COM PROBLEMA")
                log("=" * 60)
                for entry in adms_com_problema:
                    n_erro = entry["com_erro"]
                    n_total = entry["total"]
                    icone = "🔴" if entry["todos_erro"] else "🟡"
                    log(f"\n  {icone} {entry['adm']}  ({n_erro} de {n_total} item{'s' if n_total > 1 else ''} com problema)")
                    for item in entry["itens"]:
                        flag_login = "  ⚠️ [LOGIN]" if item["login_prob"] else ""
                        desc = f"{item['condo']} - {item['apto']}".strip(" -")
                        log(f"     • [{item['codigo']}] {desc}{flag_login}")
                        if item["motivo"]:
                            log(f"          ↳ {item['motivo']}")
                log("\n" + "=" * 60 + "\n")

        except Exception as e:
            log(f"⚠️ Erro ao gerar relatório de administradoras com problema: {e}")
        # ===== FIM ALERTA ADMINISTRADORAS ===== ↑


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Extrator Unificado")
    parser.add_argument("administradoras", nargs="*", metavar="ADM",
                        help="Administradoras a processar (sem informar, processa todas)")
    args = parser.parse_args()
    aux.prevent_sleep()
    main(filtro_adms=args.administradoras or None)
    aux.allow_sleep()