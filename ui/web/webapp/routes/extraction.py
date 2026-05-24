"""
extraction.py — Rotas para disparar e monitorar extrações.
Usa SSE (Server-Sent Events) para progresso em tempo real.
"""

import asyncio
import datetime
import json
import os
import secrets
import sys
import threading
import time
from pathlib import Path
from types import SimpleNamespace

from fastapi import APIRouter, Depends, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

# Adiciona raiz do projeto ao path para importar módulos existentes
_project_root = str(Path(__file__).resolve().parent.parent.parent)
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

from ..models import get_db, Execucao, Usuario, Empresa, FonteDado
from ..auth import get_usuario_logado, get_admin, contar_extracoes_hoje, gerar_csrf_token, verificar_csrf_token

# ── Definição de templates para upload Excel ─────────────────────────────────
# Cada tipo/subtipo define as colunas esperadas e exemplos
TEMPLATES_UPLOAD = {
    "Prefeitura__IPTU": {
        "colunas": ["Codigo", "IPTU"],
        "exemplos": [["1001", "0012345678"], ["1002", "0087654321"]],
        "descricao": "Preencha com o código do imóvel e o número do IPTU.",
    },
    "Prefeitura__IPTU_Faltante": {
        "colunas": ["Codigo", "IPTU", "Nr Cod IPTU"],
        "exemplos": [["1001", "0012345678", "1"], ["1002", "0087654321", "3"]],
        "descricao": "Preencha com código, IPTU e número do código IPTU (posição 1-11).",
    },
    "Prefeitura__Nada_Consta": {
        "colunas": ["Codigo", "IPTU"],
        "exemplos": [["1001", "0012345678"], ["1002", "0087654321"]],
        "descricao": "Preencha com o código do imóvel e o número do IPTU.",
    },
    "Prefeitura__Certidao_Negativa": {
        "colunas": ["Codigo", "IPTU"],
        "exemplos": [["1001", "0012345678"], ["1002", "0087654321"]],
        "descricao": "Preencha com o código do imóvel e o número do IPTU.",
    },
    "Bombeiros": {
        "colunas": ["Codigo", "CBM", "IPTU", "Cidade"],
        "exemplos": [["1001", "0012345678", "0098765432", "Rio de Janeiro"]],
        "descricao": "Preencha com código do imóvel, inscrição CBMERJ, IPTU e cidade.",
    },
    "Condomínios": {
        "colunas": ["Codigo", "LogAdm", "SenAdm", "NomeAdm", "Condominio", "Unidade", "LoginMultiplo"],
        "exemplos": [["1001", "usuario", "senha123", "Adm XYZ", "Cond. Flores", "101", "False"]],
        "descricao": "Preencha com os dados de login da administradora de condomínios.",
    },
}

router = APIRouter()
from pathlib import Path as _Path
templates = Jinja2Templates(directory=str(_Path(__file__).resolve().parent.parent / "templates"))

# ── Store de progresso em memória ─────────────────────────────────────────────
# job_id → {status, total, current, mensagem, log[], resultados[]}
_jobs: dict = {}
_jobs_lock = threading.Lock()


def _criar_job(job_id: str, tipo: str, total: int, usuario_id: int = 0):
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "rodando",
            "tipo": tipo,
            "total": total,
            "current": 0,
            "mensagem": "Iniciando...",
            "log": [],
            "erros": [],          # Lista de erros estruturados
            "resultados": [],
            "inicio": time.time(),
            "usuario_id": usuario_id,
        }


def _atualizar_job(job_id: str, **kwargs):
    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id].update(kwargs)


def _log_job(job_id: str, msg: str, nivel: str = "INFO"):
    """Adiciona linha de log com timestamp e nível."""
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{ts}] {nivel:<5} | {msg}"
    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id]["log"].append(linha)


def _erro_job(job_id: str, codigo: str, contexto: str, erro: str):
    """Registra um erro estruturado no job."""
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    with _jobs_lock:
        if job_id in _jobs:
            _jobs[job_id]["erros"].append({
                "codigo": codigo,
                "contexto": contexto,
                "erro": erro,
                "hora": ts,
            })
    _log_job(job_id, f"{contexto} - Cód. {codigo} - {erro}", nivel="ERRO")


def _get_job(job_id: str) -> dict:
    with _jobs_lock:
        return dict(_jobs.get(job_id, {}))


# ── Dashboard principal ───────────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    execucoes_recentes = (
        db.query(Execucao)
        .filter(Execucao.usuario_id == usuario.id)
        .order_by(Execucao.iniciado_em.desc())
        .limit(20)
        .all()
    )
    csrf_token = gerar_csrf_token()
    response = templates.TemplateResponse(
        request, "dashboard.html",
        {
            "usuario": usuario,
            "uso_hoje": uso_hoje,
            "execucoes": execucoes_recentes,
            "csrf_token": csrf_token,
        },
    )
    response.set_cookie(
        key="csrf_token", value=csrf_token,
        httponly=True, samesite="strict", max_age=3600,
    )
    return response


# ── APIs para fontes dinâmicas ───────────────────────────────────────────────

@router.get("/api/empresas-disponiveis")
async def empresas_disponiveis(
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Retorna empresas ativas para popular dropdown (admin global)."""
    if not usuario.is_admin:
        # Usuário não-admin: retorna apenas a empresa dele (se tiver)
        if usuario.empresa_id:
            emp = db.query(Empresa).get(usuario.empresa_id)
            if emp:
                return [{"id": emp.id, "nome": emp.nome}]
        return []

    empresas = db.query(Empresa).filter(
        Empresa.ativo == True
    ).order_by(Empresa.nome).all()
    return [{"id": e.id, "nome": e.nome} for e in empresas]


@router.get("/api/fontes-disponiveis")
async def fontes_disponiveis(
    empresa_id: int = 0,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Retorna fontes de dados ativas para uma empresa.
    Admin global passa empresa_id; usuário vinculado usa a empresa dele.
    """
    # Determinar empresa
    eid = empresa_id
    if not usuario.is_admin:
        eid = usuario.empresa_id or 0
    if not eid:
        return []

    fontes = db.query(FonteDado).filter(
        FonteDado.empresa_id == eid,
        FonteDado.ativo == True,
    ).order_by(FonteDado.tipo_extracao).all()

    return [
        {
            "id": f.id,
            "tipo_extracao": f.tipo_extracao,
            "nome_fonte": f.nome_fonte,
            "colunas": json.loads(f.colunas_esperadas) if f.colunas_esperadas else [],
        }
        for f in fontes
    ]


# ── Disparar extração ────────────────────────────────────────────────────────

@router.post("/api/extrair")
async def disparar_extracao(
    request: Request,
    tipo: str = Form(...),
    subtipo: str = Form(""),
    tipopagamento: int = Form(2),
    somente_valores: bool = Form(False),
    codigos_barra: bool = Form(False),
    somente_faltantes: bool = Form(True),
    inicio_mes: bool = Form(False),
    data_mes: str = Form(""),
    num_workers: int = Form(1),
    delay_iptu: int = Form(5),
    arquivo_banco: UploadFile = File(None),
    caminho_download: str = Form(""),
    csrf_token: str = Form(""),
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    # Verificar CSRF
    cookie_csrf = request.cookies.get("csrf_token", "")
    if not verificar_csrf_token(csrf_token, cookie_csrf):
        return JSONResponse(status_code=403, content={"erro": "Requisição inválida (CSRF)."})

    # Verificar permissão de upload de banco
    if arquivo_banco and arquivo_banco.filename:
        if not usuario.is_admin and not usuario.pode_upload_banco:
            return JSONResponse(
                status_code=403,
                content={"erro": "Você não tem permissão para enviar arquivo de banco. Use o upload de planilha."},
            )

    # Verificar limite diário
    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    if uso_hoje >= usuario.limite_diario:
        return JSONResponse(
            status_code=429,
            content={"erro": f"Limite diário atingido ({usuario.limite_diario} extrações)."},
        )

    job_id = secrets.token_urlsafe(24)

    # Registrar execução no banco
    execucao = Execucao(
        usuario_id=usuario.id,
        tipo=tipo,
        subtipo=subtipo,
        status="rodando",
    )
    db.add(execucao)
    db.commit()
    exec_id = execucao.id

    # Sanitizar caminhos contra path traversal
    def _sanitizar_caminho(caminho: str, tipo_caminho: str) -> str:
        if not caminho:
            return ""
        caminho_real = os.path.realpath(caminho)
        # Bloquear path traversal (deve ser absoluto e não conter ..)
        if ".." in caminho or not os.path.isabs(caminho_real):
            raise HTTPException(status_code=400, detail=f"Caminho inválido para {tipo_caminho}.")
        return caminho_real

    # Processar upload do arquivo de banco
    caminho_banco = ""
    if arquivo_banco and arquivo_banco.filename:
        import tempfile
        upload_dir = os.path.join(tempfile.gettempdir(), "extrator_uploads", str(usuario.id))
        os.makedirs(upload_dir, exist_ok=True)
        # Sanitizar nome do arquivo
        nome_seguro = os.path.basename(arquivo_banco.filename)
        caminho_banco = os.path.join(upload_dir, nome_seguro)
        conteudo = await arquivo_banco.read()
        with open(caminho_banco, "wb") as f:
            f.write(conteudo)

    caminho_download = _sanitizar_caminho(caminho_download, "download") if caminho_download else ""

    # Disparar em thread separada
    params = {
        "tipo": tipo,
        "subtipo": subtipo,
        "tipopagamento": tipopagamento,
        "somente_valores": somente_valores,
        "codigos_barra": codigos_barra,
        "somente_faltantes": somente_faltantes,
        "inicio_mes": inicio_mes,
        "data_mes": data_mes,
        "num_workers": num_workers,
        "delay_iptu": delay_iptu,
        "caminho_banco": caminho_banco,
        "caminho_download": caminho_download,
        "usuario_id": usuario.id,
    }

    t = threading.Thread(
        target=_executar_extracao,
        args=(job_id, exec_id, params),
        daemon=True,
        name=f"job-{job_id}",
    )
    t.start()

    return JSONResponse({"job_id": job_id, "exec_id": exec_id})


# ── Download de template Excel ───────────────────────────────────────────────

@router.get("/api/template/{tipo_extracao}")
async def download_template(
    tipo_extracao: str,
    empresa_id: int = 0,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Gera e retorna um Excel template com as colunas da FonteDado configurada.
    Prioridade: FonteDado da empresa → fallback TEMPLATES_UPLOAD (legado).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile

    # Determinar empresa
    eid = empresa_id if usuario.is_admin and empresa_id else usuario.empresa_id
    colunas = None
    descricao = ""
    nome_fonte = tipo_extracao

    # Tentar buscar da FonteDado da empresa
    if eid:
        fonte = db.query(FonteDado).filter(
            FonteDado.empresa_id == eid,
            FonteDado.tipo_extracao == tipo_extracao,
            FonteDado.ativo == True,
        ).first()
        if fonte:
            try:
                colunas = json.loads(fonte.colunas_esperadas) if fonte.colunas_esperadas else None
            except (json.JSONDecodeError, TypeError):
                colunas = None
            nome_fonte = fonte.nome_fonte
            descricao = f"Fonte: {fonte.nome_fonte} — Tipo: {fonte.tipo_extracao}"

    # Fallback: dict legado TEMPLATES_UPLOAD
    if not colunas:
        # Compatibilidade com formato antigo "Prefeitura" + subtipo
        chave = tipo_extracao
        template_info = TEMPLATES_UPLOAD.get(chave)
        if not template_info:
            raise HTTPException(status_code=400, detail=f"Template não disponível para: {tipo_extracao}")
        colunas = template_info["colunas"]
        descricao = template_info.get("descricao", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"

    # Estilos
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A2E44", end_color="1A2E44", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    example_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    example_font = Font(name="Calibri", color="888888", italic=True, size=10)

    # Header
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Exemplos legados (só se veio do TEMPLATES_UPLOAD)
    exemplos = template_info.get("exemplos", []) if 'template_info' in locals() else []
    for row_idx, exemplo in enumerate(exemplos, 2):
        for col_idx, valor in enumerate(exemplo, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = example_font
            cell.fill = example_fill
            cell.border = thin_border

    # Ajustar largura das colunas
    for col_idx, col_name in enumerate(colunas, 1):
        largura = max(len(str(col_name)) + 4, 15)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura

    # Aba de instruções
    ws_inst = wb.create_sheet("Instruções")
    ws_inst.cell(row=1, column=1, value="Instruções de Preenchimento").font = Font(bold=True, size=14)
    ws_inst.cell(row=3, column=1, value=f"Tipo de Extração: {tipo_extracao}")
    ws_inst.cell(row=4, column=1, value=descricao)
    ws_inst.cell(row=6, column=1, value="Regras:").font = Font(bold=True)
    ws_inst.cell(row=7, column=1, value="1. Não altere os nomes das colunas na aba 'Dados'.")
    ws_inst.cell(row=8, column=1, value="2. Apague as linhas de exemplo antes de preencher seus dados.")
    ws_inst.cell(row=9, column=1, value="3. Preencha uma linha por registro.")
    ws_inst.cell(row=10, column=1, value="4. Salve o arquivo e faça upload no sistema.")
    ws_inst.column_dimensions["A"].width = 70

    # Salvar em memória
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    nome_arquivo = f"template_{tipo_extracao.lower().replace(' ', '_').replace('/', '_')}.xlsx"

    with open(tmp.name, "rb") as f:
        conteudo = f.read()
    os.unlink(tmp.name)

    from starlette.responses import Response
    return Response(
        content=conteudo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


# ── Upload de dados (CSV do helper ou Excel da web) ─────────────────────────

@router.post("/api/upload-dados")
async def upload_dados(
    request: Request,
    tipo: str = Form(...),
    subtipo: str = Form(""),
    tipopagamento: int = Form(2),
    somente_valores: bool = Form(False),
    codigos_barra: bool = Form(False),
    somente_faltantes: bool = Form(True),
    inicio_mes: bool = Form(False),
    data_mes: str = Form(""),
    num_workers: int = Form(1),
    delay_iptu: int = Form(5),
    arquivo_dados: UploadFile = File(...),
    csrf_token: str = Form(""),
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """
    Recebe dados via upload (CSV ou Excel) e dispara a extração.
    Substitui a necessidade de ter o banco .WMB no servidor.
    """
    # Verificar CSRF (pular se veio do helper desktop via API)
    cookie_csrf = request.cookies.get("csrf_token", "")
    if cookie_csrf and csrf_token:
        if not verificar_csrf_token(csrf_token, cookie_csrf):
            return JSONResponse(status_code=403, content={"erro": "Requisição inválida (CSRF)."})

    # Verificar limite diário
    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    if uso_hoje >= usuario.limite_diario:
        return JSONResponse(
            status_code=429,
            content={"erro": f"Limite diário atingido ({usuario.limite_diario} extrações)."},
        )

    # Ler arquivo enviado
    nome_arquivo = arquivo_dados.filename or ""
    conteudo_bytes = await arquivo_dados.read()

    if not conteudo_bytes:
        return JSONResponse(status_code=400, content={"erro": "Arquivo vazio."})

    # Parsear conforme extensão
    try:
        if nome_arquivo.lower().endswith((".xlsx", ".xls")):
            linhas = _parsear_excel(conteudo_bytes)
        elif nome_arquivo.lower().endswith(".csv"):
            linhas = _parsear_csv(conteudo_bytes)
        else:
            # Tentar CSV como fallback
            linhas = _parsear_csv(conteudo_bytes)
    except Exception as e:
        return JSONResponse(
            status_code=400,
            content={"erro": f"Erro ao ler arquivo: {str(e)[:200]}"},
        )

    if not linhas:
        return JSONResponse(status_code=400, content={"erro": "Arquivo sem dados (apenas cabeçalho ou vazio)."})

    # Criar job e execução
    job_id = secrets.token_urlsafe(24)

    execucao = Execucao(
        usuario_id=usuario.id,
        tipo=tipo,
        subtipo=subtipo,
        status="rodando",
    )
    db.add(execucao)
    db.commit()
    exec_id = execucao.id

    # Preparar params
    caminho_download = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "..", "Downloads"
    )
    os.makedirs(caminho_download, exist_ok=True)

    params = {
        "tipo": tipo,
        "subtipo": subtipo,
        "tipopagamento": tipopagamento,
        "somente_valores": somente_valores,
        "codigos_barra": codigos_barra,
        "somente_faltantes": somente_faltantes,
        "inicio_mes": inicio_mes,
        "data_mes": data_mes,
        "num_workers": num_workers,
        "delay_iptu": delay_iptu,
        "caminho_download": os.path.realpath(caminho_download),
        "usuario_id": usuario.id,
        "dados_upload": linhas,  # Dados já parseados
    }

    t = threading.Thread(
        target=_executar_extracao_upload,
        args=(job_id, exec_id, params),
        daemon=True,
        name=f"job-upload-{job_id}",
    )
    t.start()

    return JSONResponse({
        "job_id": job_id,
        "exec_id": exec_id,
        "registros": len(linhas),
        "mensagem": f"{len(linhas)} registros carregados. Extração iniciada.",
    })


def _parsear_excel(conteudo_bytes: bytes) -> list:
    """Lê Excel e retorna lista de tuplas (sem cabeçalho)."""
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), read_only=True, data_only=True)
    # Usar primeira aba (ou aba "Dados" se existir)
    ws = wb["Dados"] if "Dados" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Primeira linha é cabeçalho, pular
    dados = []
    for row in rows[1:]:
        # Ignorar linhas completamente vazias
        if any(cell is not None and str(cell).strip() for cell in row):
            dados.append(tuple(str(cell).strip() if cell is not None else "" for cell in row))

    return dados


def _parsear_csv(conteudo_bytes: bytes) -> list:
    """Lê CSV e retorna lista de tuplas (sem cabeçalho)."""
    import csv as csv_mod
    import io

    # Tentar detectar encoding
    try:
        texto = conteudo_bytes.decode("utf-8")
    except UnicodeDecodeError:
        texto = conteudo_bytes.decode("latin-1")

    reader = csv_mod.reader(io.StringIO(texto))
    rows = list(reader)

    if not rows:
        return []

    # Primeira linha é cabeçalho, pular
    dados = []
    for row in rows[1:]:
        if any(cell.strip() for cell in row):
            dados.append(tuple(cell.strip() for cell in row))

    return dados


def _executar_extracao_upload(job_id: str, exec_id: int, params: dict):
    """Executa extração usando dados já carregados do upload (sem precisar do banco Access)."""
    from ..models import SessionLocal, Execucao as ExecModel

    tipo = params["tipo"]
    db = SessionLocal()
    linhas = params.get("dados_upload", [])

    try:
        total = len(linhas)
        _criar_job(job_id, tipo, total, usuario_id=params.get("usuario_id", 0))
        _log_job(job_id, f"Iniciando extração via upload: {tipo}", nivel="INICIO")
        _log_job(job_id, f"{total} registros carregados do arquivo enviado")

        caminho_download = params.get("caminho_download", "")
        subtipo = params.get("subtipo", "")

        # Determinar subpasta de download
        match tipo:
            case "Prefeitura":
                subpasta = {"IPTU": "IPTU", "Nada Consta": "Nada Consta",
                            "Certidão Negativa": "Certidão Enfitêutica"}.get(subtipo, "IPTU")
            case "Bombeiros":
                subpasta = "Bombeiros"
            case "Condomínios":
                subpasta = "Condomínios"
            case _:
                _log_job(job_id, f"Tipo desconhecido: {tipo}", nivel="ERRO")
                _atualizar_job(job_id, status="erro", mensagem=f"Tipo desconhecido: {tipo}")
                _finalizar_execucao(db, exec_id, "erro", 0, 0, 0, job_id=job_id)
                return

        pasta_final = os.path.join(caminho_download, subpasta)
        os.makedirs(pasta_final, exist_ok=True)

        _atualizar_job(job_id, total=total, mensagem=f"Processando {total} registros...")

        # Executar extração conforme tipo
        ok_count = 0
        err_count = 0

        if tipo == "Condomínios":
            ok_count, err_count = _extrair_condominios_api(
                job_id, linhas, pasta_final, params
            )
        elif tipo == "Prefeitura":
            ok_count, err_count = _extrair_prefeitura(
                job_id, linhas, pasta_final, params
            )
        elif tipo == "Bombeiros":
            ok_count, err_count = _extrair_bombeiros(
                job_id, linhas, pasta_final, params
            )

        _log_job(job_id, f"Concluído: {ok_count} OK, {err_count} erros de {total} total", nivel="FIM")
        _atualizar_job(
            job_id,
            status="concluido",
            mensagem=f"Concluído: {ok_count} OK, {err_count} erros de {total} total",
        )
        _finalizar_execucao(db, exec_id, "concluido", total, ok_count, err_count, job_id=job_id)

    except Exception as e:
        import traceback
        _log_job(job_id, f"ERRO INTERNO: {traceback.format_exc()}", nivel="ERRO")
        _atualizar_job(job_id, status="erro", mensagem="Erro interno na extração. Consulte o log.")
        _finalizar_execucao(db, exec_id, "erro", 0, 0, 0, job_id=job_id)
    finally:
        db.close()


def _executar_extracao(job_id: str, exec_id: int, params: dict):
    """Executa a extração em background. Atualiza o job store e o banco."""
    from ..models import SessionLocal, Execucao

    tipo = params["tipo"]
    db = SessionLocal()

    try:
        _criar_job(job_id, tipo, 0, usuario_id=params.get("usuario_id", 0))
        _log_job(job_id, f"Iniciando extração: {tipo}", nivel="INICIO")

        from auxiliares import utils as aux
        from auxiliares import sensiveis as senha

        # Determinar caminho do banco
        caminho_banco = params.get("caminho_banco", "")
        if not caminho_banco:
            caminho_banco = os.path.join(aux.caminhoprojeto(), "Scai.WMB")

        if not os.path.isfile(caminho_banco):
            _log_job(job_id, f"Banco não encontrado: {caminho_banco}", nivel="ERRO")
            _atualizar_job(job_id, status="erro", mensagem=f"Banco não encontrado: {caminho_banco}")
            _finalizar_execucao(db, exec_id, "erro", 0, 0, 0, job_id=job_id)
            return

        # Determinar caminho de download
        caminho_download = params.get("caminho_download", "")
        if not caminho_download:
            caminho_download = os.path.join(aux.caminhoprojeto(), "Downloads")

        _log_job(job_id, "Conectando ao banco de dados...")

        bd = aux.Banco(caminho_banco)

        # Determinar SQL e executar consulta
        match tipo:
            case "Prefeitura":
                subtipo = params.get("subtipo", "IPTU")
                if params.get("somente_faltantes") and subtipo == "IPTU":
                    sql = senha.SQLIPTUFALTANTE
                else:
                    sql = senha.SQLIPTUCOMPLETO
                subpasta = {"IPTU": "IPTU", "Nada Consta": "Nada Consta",
                            "Certidão Negativa": "Certidão Enfitêutica"}.get(subtipo, "IPTU")

            case "Bombeiros":
                sql = senha.SQLCBM
                subpasta = "Bombeiros"

            case "Condomínios":
                data_mes = params.get("data_mes", "")
                if params.get("inicio_mes"):
                    bd.executarsql("DELETE * FROM BoletosCondominios")
                sql = aux.retornarlistaboletos(anomes=data_mes)
                subpasta = "Condomínios"

            case _:
                _log_job(job_id, f"Tipo desconhecido: {tipo}", nivel="ERRO")
                _atualizar_job(job_id, status="erro", mensagem=f"Tipo desconhecido: {tipo}")
                _finalizar_execucao(db, exec_id, "erro", 0, 0, 0, job_id=job_id)
                return

        resultado = bd.consultar(sql)
        bd.fecharbanco()

        if not resultado:
            _log_job(job_id, "Nenhum registro encontrado.")
            _atualizar_job(job_id, status="concluido", mensagem="Nenhum registro encontrado.")
            _finalizar_execucao(db, exec_id, "concluido", 0, 0, 0, job_id=job_id)
            return

        total = len(resultado)
        pasta_final = os.path.join(caminho_download, subpasta)
        os.makedirs(pasta_final, exist_ok=True)

        _atualizar_job(job_id, total=total, mensagem=f"Processando {total} registros...")
        _log_job(job_id, f"{total} registros carregados do banco")

        # ── Executar extração conforme tipo ───────────────────────────
        ok_count = 0
        err_count = 0

        if tipo == "Condomínios":
            ok_count, err_count = _extrair_condominios_api(
                job_id, resultado, pasta_final, params
            )
        elif tipo == "Prefeitura":
            ok_count, err_count = _extrair_prefeitura(
                job_id, resultado, pasta_final, params
            )
        elif tipo == "Bombeiros":
            ok_count, err_count = _extrair_bombeiros(
                job_id, resultado, pasta_final, params
            )

        _log_job(job_id, f"Concluído: {ok_count} OK, {err_count} erros de {total} total", nivel="FIM")
        _atualizar_job(
            job_id,
            status="concluido",
            mensagem=f"Concluído: {ok_count} OK, {err_count} erros de {total} total",
        )
        _finalizar_execucao(db, exec_id, "concluido", total, ok_count, err_count, job_id=job_id)

    except Exception as e:
        import traceback
        _log_job(job_id, f"ERRO INTERNO: {traceback.format_exc()}", nivel="ERRO")
        _atualizar_job(job_id, status="erro", mensagem="Erro interno na extração. Consulte o log para detalhes.")
        _finalizar_execucao(db, exec_id, "erro", 0, 0, 0, job_id=job_id)
    finally:
        db.close()


def _finalizar_execucao(db, exec_id, status, total, ok, erros, job_id: str = ""):
    """Atualiza o registro de execução no banco, incluindo log e erros."""
    try:
        exec_obj = db.query(Execucao).get(exec_id)
        if exec_obj:
            exec_obj.status = status
            exec_obj.total_registros = total
            exec_obj.registros_ok = ok
            exec_obj.registros_erro = erros
            exec_obj.finalizado_em = datetime.datetime.utcnow()
            if exec_obj.iniciado_em:
                exec_obj.duracao_seg = (
                    exec_obj.finalizado_em - exec_obj.iniciado_em
                ).total_seconds()

            # Persistir log e erros do job em memória para o banco
            if job_id:
                job = _get_job(job_id)
                if job:
                    # Log completo como texto
                    exec_obj.log_texto = "\n".join(job.get("log", []))
                    # Erros estruturados como JSON
                    exec_obj.erros_detalhados = json.dumps(
                        job.get("erros", []), ensure_ascii=False
                    )

            db.commit()
    except Exception:
        pass


# ── Extrações por tipo ────────────────────────────────────────────────────────

def _extrair_condominios_api(job_id, linhas, pasta_download, params):
    """Extração de condomínios via API headless (extracao_api.py)."""
    from extratores.Condominios import extracao_api

    extracao_api.set_downloads_dir(pasta_download)

    # Progress callback para atualizar o job
    def progress_cb(idx, total, cod, admin):
        _atualizar_job(
            job_id,
            current=idx + 1,
            mensagem=f"[{idx+1}/{total}] {admin} - Cód. {cod}",
        )
        _log_job(job_id, f"{admin} - Cód. {cod}", nivel="OK")

    try:
        resultados = extracao_api.processar(
            linhas,
            progress_cb=progress_cb,
        )

        ok = sum(1 for r in resultados if r["status"] == "ok")
        dup = sum(1 for r in resultados if r["status"] == "duplicado")
        err = len(resultados) - ok - dup

        # Registrar erros estruturados
        for r in resultados:
            if r["status"] not in ("ok", "duplicado"):
                _erro_job(job_id, str(r.get("codigo", "")),
                          r.get("administradora", "Condomínios"),
                          r.get("erro", r["status"]))

        _atualizar_job(job_id, resultados=resultados)
        _log_job(job_id, f"Condomínios: OK={ok}, Duplicados={dup}, Erros={err}")

        return ok + dup, err

    except Exception as e:
        _log_job(job_id, f"Erro fatal em condomínios: {e}", nivel="ERRO")
        return 0, len(linhas)


def _extrair_prefeitura(job_id, linhas, pasta_download, params):
    """Extração de IPTU / Nada Consta / Certidão via Playwright híbrido."""
    from playwright.sync_api import sync_playwright
    from extratores.Prefeitura import Biptu_hibrido as Bh

    subtipo = params.get("subtipo", "IPTU")
    visual = SimpleNamespace(
        somentevalores=SimpleNamespace(get=lambda: params.get("somente_valores", False)),
        codigosdebarra=SimpleNamespace(get=lambda: params.get("codigos_barra", False)),
        tipopagamento=SimpleNamespace(get=lambda: params.get("tipopagamento", 2)),
        faltantes=SimpleNamespace(get=lambda: params.get("somente_faltantes", True)),
        tiposervico=SimpleNamespace(get=lambda: subtipo),
    )

    visual.mudartexto = lambda *a, **kw: None
    visual.configurarbarra = lambda *a, **kw: None

    objeto = SimpleNamespace(pastadownload=pasta_download, visual=visual)
    ok_count = 0
    err_count = 0
    total = len(linhas)

    with sync_playwright() as p:
        profile_name = os.getenv("NOMEPROFILEIPTU", "iptu_profile")
        profile_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    "..", "Profile", profile_name)
        headless = subtipo != "Nada Consta"

        context = p.chromium.launch_persistent_context(
            user_data_dir=profile_path,
            headless=headless,
            accept_downloads=True,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"],
        )
        page = context.pages[0] if context.pages else context.new_page()

        try:
            for i, linha in enumerate(linhas):
                cod = str(linha[Bh.Codigo])
                _atualizar_job(
                    job_id, current=i + 1,
                    mensagem=f"[{i+1}/{total}] {subtipo} - Cód. {cod}",
                )

                try:
                    match subtipo:
                        case "IPTU":
                            dados, df = Bh.extrairIPTU_hibrido(page, objeto, linha)
                        case "Nada Consta":
                            dados, df = Bh.extrairnadaconsta_hibrido(page, objeto, linha)
                        case "Certidão Negativa":
                            dados, df = Bh.extraircertidaonegativa_requests(objeto, linha)
                        case _:
                            dados = [cod, "", "", "Subtipo desconhecido"]

                    status_item = dados[-1] if dados else "ERRO"
                    if "ERRO" in str(status_item).upper():
                        err_count += 1
                        _erro_job(job_id, cod, subtipo, str(status_item))
                    else:
                        ok_count += 1
                        _log_job(job_id, f"{subtipo} - Cód. {cod} - OK", nivel="OK")

                except Exception as e:
                    _erro_job(job_id, cod, subtipo, str(e))
                    err_count += 1

                # Delay entre requisições (IPTU)
                if subtipo == "IPTU" and params.get("delay_iptu", 0) > 0:
                    time.sleep(params["delay_iptu"])

        finally:
            context.close()

    return ok_count, err_count


def _extrair_bombeiros(job_id, linhas, pasta_download, params):
    """Extração de Bombeiros via Playwright híbrido."""
    from playwright.sync_api import sync_playwright
    from extratores.Prefeitura import Biptu_hibrido as Bh

    visual = SimpleNamespace(
        tipopagamento=SimpleNamespace(get=lambda: params.get("tipopagamento", 2)),
        somentevalores=SimpleNamespace(get=lambda: False),
        codigosdebarra=SimpleNamespace(get=lambda: False),
        faltantes=SimpleNamespace(get=lambda: False),
        tiposervico=SimpleNamespace(get=lambda: ""),
    )
    visual.mudartexto = lambda *a, **kw: None
    visual.configurarbarra = lambda *a, **kw: None

    objeto = SimpleNamespace(pastadownload=pasta_download, visual=visual)
    ok_count = 0
    err_count = 0
    total = len(linhas)
    num_workers = max(1, min(params.get("num_workers", 1), 8))

    import queue as queue_mod

    work_queue = queue_mod.Queue()
    for i, linha in enumerate(linhas):
        work_queue.put((i, linha))

    results_lock = threading.Lock()

    def _worker(worker_id):
        nonlocal ok_count, err_count
        profile_base = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "..", "Profile",
            os.getenv("NOMEPROFILECBM", "cbm_profile"),
        )
        profile_path = f"{profile_base}_{worker_id}"

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=profile_path,
                headless=True,
                accept_downloads=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
            page = context.pages[0] if context.pages else context.new_page()
            try:
                while True:
                    try:
                        i, linha = work_queue.get_nowait()
                    except queue_mod.Empty:
                        break

                    cod = str(linha[Bh.Codigo])
                    _atualizar_job(
                        job_id, current=i + 1,
                        mensagem=f"[{i+1}/{total}] Bombeiros - Cód. {cod}",
                    )

                    try:
                        dados, df = Bh.extrairbombeiros_hibrido(page, objeto, linha)
                        status_item = dados[3] if len(dados) > 3 else "ERRO"
                        with results_lock:
                            if "ERRO" in str(status_item).upper():
                                err_count += 1
                                _erro_job(job_id, cod, "Bombeiros", str(status_item))
                            else:
                                ok_count += 1
                                _log_job(job_id, f"Bombeiros - Cód. {cod} - OK", nivel="OK")
                    except Exception as e:
                        _erro_job(job_id, cod, "Bombeiros", str(e))
                        with results_lock:
                            err_count += 1

                    work_queue.task_done()
            finally:
                context.close()

    if num_workers == 1:
        _worker(0)
    else:
        threads = []
        for wid in range(num_workers):
            t = threading.Thread(target=_worker, args=(wid,), daemon=True)
            threads.append(t)
        for t in threads:
            t.start()
        for t in threads:
            t.join()

    return ok_count, err_count


# ── SSE: Progresso em tempo real ──────────────────────────────────────────────

@router.get("/api/progresso/{job_id}")
async def stream_progresso(job_id: str, usuario: Usuario = Depends(get_usuario_logado)):
    """Server-Sent Events para progresso de uma extração."""

    # Verificar ownership do job
    job_check = _get_job(job_id)
    if job_check and job_check.get("usuario_id") and job_check["usuario_id"] != usuario.id and not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    async def event_generator():
        while True:
            job = _get_job(job_id)
            if not job:
                yield f"data: {json.dumps({'erro': 'Job não encontrado'})}\n\n"
                break

            elapsed = time.time() - job.get("inicio", time.time())
            current = job.get("current", 0)
            total = job.get("total", 0)

            # Calcular ETA
            eta = 0
            if current > 0 and total > 0 and current < total:
                media = elapsed / current
                eta = media * (total - current)

            payload = {
                "status": job.get("status", ""),
                "total": total,
                "current": current,
                "mensagem": job.get("mensagem", ""),
                "elapsed": round(elapsed, 1),
                "eta": round(eta, 1),
                "log_count": len(job.get("log", [])),
            }
            yield f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"

            if job.get("status") in ("concluido", "erro"):
                break

            await asyncio.sleep(1)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/api/progresso/{job_id}/log")
async def get_job_log(job_id: str, usuario: Usuario = Depends(get_usuario_logado)):
    """Retorna o log completo de um job."""
    job = _get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado")
    # Verificar ownership
    if job.get("usuario_id") and job["usuario_id"] != usuario.id and not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso negado.")
    return {"log": job.get("log", []), "resultados": job.get("resultados", [])}


# ── Histórico ─────────────────────────────────────────────────────────────────

@router.get("/historico", response_class=HTMLResponse)
async def historico(
    request: Request,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    execucoes = (
        db.query(Execucao)
        .filter(Execucao.usuario_id == usuario.id)
        .order_by(Execucao.iniciado_em.desc())
        .limit(100)
        .all()
    )
    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    return templates.TemplateResponse(
        request, "historico.html",
        {"usuario": usuario, "execucoes": execucoes, "uso_hoje": uso_hoje},
    )


# ── Detalhes e Log de uma execução ──────────────────────────────────────────

@router.get("/execucao/{exec_id}", response_class=HTMLResponse)
async def detalhe_execucao(
    request: Request,
    exec_id: int,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Página de detalhes de uma execução com log completo."""
    execucao = db.query(Execucao).get(exec_id)
    if not execucao:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    # Ownership: só o dono ou admin pode ver
    if execucao.usuario_id != usuario.id and not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    # Parsear erros do JSON
    erros = []
    try:
        erros = json.loads(execucao.erros_detalhados or "[]")
    except (json.JSONDecodeError, TypeError):
        pass

    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    return templates.TemplateResponse(
        request, "detalhe_execucao.html",
        {
            "usuario": usuario,
            "uso_hoje": uso_hoje,
            "exec": execucao,
            "erros": erros,
            "log_linhas": (execucao.log_texto or "").split("\n") if execucao.log_texto else [],
        },
    )


@router.get("/execucao/{exec_id}/download")
async def download_log(
    exec_id: int,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Download do log como arquivo .txt."""
    execucao = db.query(Execucao).get(exec_id)
    if not execucao:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    if execucao.usuario_id != usuario.id and not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    log_texto = execucao.log_texto or "(Log vazio)"

    # Adicionar cabeçalho informativo
    header = (
        f"ExtratorUnificado - Log de Execução #{exec_id}\n"
        f"Tipo: {execucao.tipo} {('/ ' + execucao.subtipo) if execucao.subtipo else ''}\n"
        f"Status: {execucao.status}\n"
        f"Iniciado: {execucao.iniciado_em}\n"
        f"Finalizado: {execucao.finalizado_em or 'N/A'}\n"
        f"Registros: {execucao.registros_ok} OK / {execucao.registros_erro} erros / {execucao.total_registros} total\n"
        f"{'=' * 60}\n\n"
    )

    conteudo = header + log_texto

    # Se há erros, adicionar seção separada
    if execucao.erros_detalhados and execucao.erros_detalhados != "[]":
        try:
            erros = json.loads(execucao.erros_detalhados)
            if erros:
                conteudo += f"\n\n{'=' * 60}\nERROS DETALHADOS ({len(erros)} erros)\n{'=' * 60}\n"
                for i, e in enumerate(erros, 1):
                    conteudo += f"\n{i}. [{e.get('hora', '')}] {e.get('contexto', '')} - Cód. {e.get('codigo', '')} - {e.get('erro', '')}"
        except (json.JSONDecodeError, TypeError):
            pass

    nome_arquivo = f"log_execucao_{exec_id}_{execucao.tipo.lower().replace(' ', '_')}.txt"

    from starlette.responses import Response
    return Response(
        content=conteudo,
        media_type="text/plain; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
        },
    )


@router.get("/execucao/{exec_id}/erros")
async def download_erros_json(
    exec_id: int,
    usuario: Usuario = Depends(get_usuario_logado),
    db: Session = Depends(get_db),
):
    """Download dos erros como arquivo JSON estruturado."""
    execucao = db.query(Execucao).get(exec_id)
    if not execucao:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    if execucao.usuario_id != usuario.id and not usuario.is_admin:
        raise HTTPException(status_code=403, detail="Acesso negado.")

    try:
        erros = json.loads(execucao.erros_detalhados or "[]")
    except (json.JSONDecodeError, TypeError):
        erros = []

    nome_arquivo = f"erros_execucao_{exec_id}.json"

    from starlette.responses import Response
    return Response(
        content=json.dumps(erros, ensure_ascii=False, indent=2),
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
        },
    )


# ── Admin: Painel de Logs ───────────────────────────────────────────────────

@router.get("/admin/logs", response_class=HTMLResponse)
async def admin_logs(
    request: Request,
    status_filtro: str = "",
    tipo_filtro: str = "",
    usuario_filtro: str = "",
    pagina: int = 1,
    usuario: Usuario = Depends(get_admin),
    db: Session = Depends(get_db),
):
    """Painel admin: todas as execuções com filtros."""
    from sqlalchemy import desc, func

    query = db.query(Execucao).join(Usuario)

    # Aplicar filtros
    if status_filtro:
        query = query.filter(Execucao.status == status_filtro)
    if tipo_filtro:
        query = query.filter(Execucao.tipo == tipo_filtro)
    if usuario_filtro:
        query = query.filter(Usuario.nome.ilike(f"%{usuario_filtro}%"))

    total = query.count()
    por_pagina = 25
    total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
    pagina = max(1, min(pagina, total_paginas))

    execucoes = (
        query.order_by(desc(Execucao.iniciado_em))
        .offset((pagina - 1) * por_pagina)
        .limit(por_pagina)
        .all()
    )

    # Estatísticas rápidas
    stats = {
        "total": db.query(func.count(Execucao.id)).scalar() or 0,
        "erros_hoje": db.query(func.count(Execucao.id)).filter(
            Execucao.status == "erro",
            Execucao.iniciado_em >= datetime.datetime.combine(
                datetime.date.today(), datetime.time.min
            ),
        ).scalar() or 0,
        "rodando": db.query(func.count(Execucao.id)).filter(
            Execucao.status == "rodando"
        ).scalar() or 0,
    }

    # Tipos distintos pra filtro
    tipos = [r[0] for r in db.query(Execucao.tipo).distinct().all()]

    uso_hoje = contar_extracoes_hoje(db, usuario.id)
    return templates.TemplateResponse(
        request, "admin_logs.html",
        {
            "usuario": usuario,
            "uso_hoje": uso_hoje,
            "execucoes": execucoes,
            "stats": stats,
            "tipos": tipos,
            "filtros": {
                "status": status_filtro,
                "tipo": tipo_filtro,
                "usuario": usuario_filtro,
            },
            "pagina": pagina,
            "total_paginas": total_paginas,
            "total": total,
            "csrf_token": gerar_csrf_token(),
        },
    )
    response.set_cookie(
        key="csrf_token", value=response.context["csrf_token"],
        httponly=True, samesite="strict", max_age=3600,
    )
    return response


@router.post("/admin/execucao/{exec_id}/cancelar")
async def cancelar_execucao(
    request: Request,
    exec_id: int,
    csrf_token: str = Form(""),
    db: Session = Depends(get_db),
    _admin: Usuario = Depends(get_admin),
):
    """Marca uma execução travada como 'erro' (só admin)."""
    cookie_csrf = request.cookies.get("csrf_token", "")
    if not verificar_csrf_token(csrf_token, cookie_csrf):
        return RedirectResponse(url="/admin/logs", status_code=303)

    execucao = db.query(Execucao).get(exec_id)
    if execucao and execucao.status == "rodando":
        execucao.status = "erro"
        execucao.mensagem = "Cancelado manualmente pelo administrador."
        execucao.finalizado_em = datetime.datetime.utcnow()
        if execucao.iniciado_em:
            execucao.duracao_seg = (execucao.finalizado_em - execucao.iniciado_em).total_seconds()
        db.commit()

    # Volta para a página de origem (dashboard ou admin/logs)
    referer = request.headers.get("referer", "")
    redirect_url = referer if referer else "/admin/logs"
    return RedirectResponse(url=redirect_url, status_code=303)
