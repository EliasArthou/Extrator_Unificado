"""
Extrator IPTU em Lote: busca inscrições no banco e extrai boletos
usando Playwright + AntiCaptcha reCAPTCHA v2.

Uso:
    python testar_iptu_lote.py                                  # banco padrão (Scai.WMB)
    python testar_iptu_lote.py -b "C:\\caminho\\Scai.WMB"       # banco específico
    python testar_iptu_lote.py -c 0954                          # a partir do cliente 0954
    python testar_iptu_lote.py -n 10                            # só os 10 primeiros
    python testar_iptu_lote.py -w 4                             # 4 workers em paralelo
    python testar_iptu_lote.py -t 1                             # tipo pagamento: 1=cota única
    python testar_iptu_lote.py --faltante                       # só IPTUs não extraídos ainda
    python testar_iptu_lote.py --visible                        # modo visível
    python testar_iptu_lote.py --dry-run                        # lista sem extrair
"""

import os
import sys
import time
import queue
import threading
import argparse
import pandas as pd
from datetime import datetime
from types import SimpleNamespace
from dotenv import load_dotenv

load_dotenv()

# Suprime warnings de conexão (anticaptcha SDK)
import logging
import urllib3
import io
logging.getLogger("urllib3").setLevel(logging.CRITICAL)
logging.getLogger("requests").setLevel(logging.CRITICAL)
urllib3.disable_warnings()


class _FilteredStream:
    """Filtro de stdout/stderr que suprime linhas com palavras-chave indesejadas."""
    BLOCKED = ("ConnectionResetError", "Connection aborted", "RemoteDisconnected")

    def __init__(self, original):
        self._original = original

    def write(self, text):
        if any(kw in text for kw in self.BLOCKED):
            return len(text)
        return self._original.write(text)

    def flush(self):
        return self._original.flush()

    def __getattr__(self, name):
        return getattr(self._original, name)


from rich.console import Console
from rich.live import Live
from rich.panel import Panel
from rich.progress import (
    Progress, ProgressColumn, SpinnerColumn, TextColumn, BarColumn, Task,
)
from rich.text import Text

import auxiliares as aux
import sensiveis as senha
import Biptu_hibrido as Bh

console = Console(force_terminal=True, color_system="truecolor")


class _MockVar:
    """Simula variável tkinter (objeto.visual.xxx.get()) para uso sem GUI."""
    def __init__(self, val):
        self._val = val
    def get(self):
        return self._val


class HMSElapsedColumn(ProgressColumn):
    def render(self, task: Task) -> Text:
        elapsed = task.finished_time if task.finished else task.elapsed
        if elapsed is None:
            return Text("00:00:00", style="cyan")
        h, resto = divmod(int(elapsed), 3600)
        m, s = divmod(resto, 60)
        return Text(f"{h:02d}:{m:02d}:{s:02d}", style="cyan")


class HMSRemainingColumn(ProgressColumn):
    def render(self, task: Task) -> Text:
        remaining = task.time_remaining
        if remaining is None:
            return Text("--:--:--", style="cyan")
        h, resto = divmod(int(remaining), 3600)
        m, s = divmod(resto, 60)
        return Text(f"{h:02d}:{m:02d}:{s:02d}", style="cyan")


class _WorkerConsole:
    """
    Wrapper do Console com buffer por thread.
    Cada worker acumula mensagens durante o processamento e as escreve
    todas de uma vez ao final do cliente (escrita sequencial com lock).
    """
    def __init__(self, base: Console, write_lock: threading.Lock):
        self._base = base
        self._write_lock = write_lock
        self._local = threading.local()

    def _buffer(self) -> list:
        if not hasattr(self._local, 'buf'):
            self._local.buf = []
        return self._local.buf

    def print(self, *args, **kwargs):
        name = threading.current_thread().name
        if name.startswith("worker-"):
            self._buffer().append((args, kwargs))
        else:
            self._base.print(*args, **kwargs)

    def flush(self, header: str = "", footer: str = ""):
        """Escreve buffer acumulado no console, de forma sequencial.

        header: linha(s) impressas antes do buffer (ex: separador + cliente).
        footer: linha(s) impressas depois do buffer (ex: status final OK/SKIP/ERRO).
        Tudo dentro do mesmo lock para evitar interleaving entre workers.
        """
        buf = self._buffer()
        if not buf and not header and not footer:
            return
        wid = threading.current_thread().name.split("-")[-1]
        with self._write_lock:
            if header:
                self._base.print(header)
            for args, kwargs in buf:
                if args:
                    args = (f"[dim]\\[W{wid}][/] {args[0]}",) + args[1:]
                self._base.print(*args, **kwargs)
            if footer:
                self._base.print(footer)
        buf.clear()

    def __getattr__(self, name):
        return getattr(self._base, name)


def _worker(worker_id: int, work_queue: queue.Queue, objeto,
            headless: bool, caminho_perfil_base: str,
            progress, task_id, lock: threading.Lock,
            log_resultados: list, counters: dict,
            worker_console: _WorkerConsole):
    """
    Worker paralelo: cada thread tem seu próprio browser.
    Mensagens bufferizadas e impressas agrupadas por cliente.
    """
    from playwright.sync_api import sync_playwright

    perfil_worker = f"{caminho_perfil_base}_{worker_id}"

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=perfil_worker,
            headless=headless,
            accept_downloads=True,
            args=["--disable-blink-features=AutomationControlled", "--start-maximized"]
        )
        page = context.pages[0] if context.pages else context.new_page()

        try:
            while True:
                try:
                    i, linha, total = work_queue.get_nowait()
                except queue.Empty:
                    break

                cod = str(linha[Bh.Codigo])
                nriptu = str(linha[Bh.NrIPTU]).strip()

                def _montar_header(n_processado: int) -> str:
                    return (
                        f"\n[bold]{'='*60}[/]\n"
                        f"[bold cyan][W{worker_id}][/] [{n_processado}/{total}] "
                        f"Cliente: {cod} | IPTU: {nriptu}"
                        f"\n[bold]{'='*60}[/]"
                    )

                # Skip se arquivo já existe
                caminho_existente = os.path.join(objeto.pastadownload, f"{cod}_{nriptu}.pdf")
                if os.path.isfile(caminho_existente):
                    status = 'SKIP (arquivo existente)'
                    # Atômico: incrementa contador, monta header, faz flush — mesmo lock
                    # pra garantir que a numeração apareça em ordem cronológica de término.
                    with lock:
                        counters['processed'] += 1
                        counters['skip'] += 1
                        log_resultados.append({'Cod Cliente': cod, 'IPTU': nriptu, 'Status': status})
                        progress.advance(task_id)
                        worker_console.flush(
                            header=_montar_header(counters['processed']),
                            footer=f"  [yellow]-> SKIP[/] {status}",
                        )
                    work_queue.task_done()
                    continue

                try:
                    dados, df = Bh.extrairIPTU_hibrido(page, objeto, linha)
                    if dados and isinstance(dados[0], list):
                        status = dados[0][7] if len(dados[0]) > 7 else 'Ok'
                    else:
                        status = dados[7] if dados and len(dados) > 7 else 'Ok'
                except Exception as e:
                    status = f"ERRO: {e}"

                # Define linha de status (footer) p/ feedback em tempo real
                if 'Ok' in status or 'OK' in status:
                    status_footer = f"  [green]-> OK[/]"
                elif 'SKIP' in status or 'Já extraído' in status:
                    status_footer = f"  [yellow]-> SKIP[/] {status}"
                elif 'PAGO' in status or 'Quitad' in status:
                    status_footer = f"  [cyan]-> PAGO[/] {status}"
                else:
                    status_footer = f"  [red]-> ERRO[/] {str(status)[:160]}"

                # Atômico: incrementa processed + contadores, faz flush — tudo sob mesmo lock
                with lock:
                    counters['processed'] += 1
                    if 'Ok' in status or 'OK' in status:
                        counters['ok'] += 1
                    elif 'SKIP' in status or 'Já extraído' in status:
                        counters['skip'] += 1
                    elif 'PAGO' in status or 'Quitad' in status:
                        counters['pago'] = counters.get('pago', 0) + 1
                    else:
                        counters['erros'] += 1

                    log_resultados.append({
                        'Cod Cliente': cod,
                        'IPTU': nriptu,
                        'Status': status,
                    })

                    worker_console.flush(
                        header=_montar_header(counters['processed']),
                        footer=status_footer,
                    )

                    progress.advance(task_id)

                work_queue.task_done()

        finally:
            context.close()


def main():
    sys.stdout = _FilteredStream(sys.stdout)
    sys.stderr = _FilteredStream(sys.stderr)

    parser = argparse.ArgumentParser(description="Extrator IPTU em Lote")
    parser.add_argument("-b", "--banco", default="",
                        help="Caminho do banco de dados (.WMB ou .xlsx)")
    parser.add_argument("-c", "--cliente", default="0",
                        help="Código do cliente inicial (0 = todos)")
    parser.add_argument("-n", "--max", type=int, default=0,
                        help="Máximo de registros a processar (0 = todos)")
    parser.add_argument("-p", "--pasta", default="Downloads/IPTU",
                        help="Pasta para salvar os PDFs")
    parser.add_argument("-w", "--workers", type=int, default=1,
                        help="Número de workers em paralelo (padrão: 1)")
    parser.add_argument("-t", "--tipo", default="1",
                        choices=["1", "2", "3", "4"],
                        help="Tipo de pagamento: 1=Cota única, 2=1ª parcela, "
                             "3=2ª parcela, 4=3ª parcela (padrão: 1)")
    parser.add_argument("--faltante", action="store_true",
                        help="Usar SQL de IPTUs faltantes (não extraídos ainda)")
    parser.add_argument("--visible", action="store_true",
                        help="Abrir janela do navegador (padrão: headless)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Só lista as inscrições, sem extrair")
    args = parser.parse_args()

    num_workers = max(1, args.workers)

    # ── 1. Banco ────────────────────────────────────────────
    caminho_banco = args.banco
    if not caminho_banco:
        caminho_projeto = aux.caminhoprojeto()
        candidato = os.path.join(caminho_projeto, 'Scai.WMB')
        if os.path.isfile(candidato):
            caminho_banco = candidato
        else:
            console.print("[red][ERRO][/] Banco não encontrado. Use -b para especificar o caminho.")
            console.print(f"  Procurado em: {candidato}")
            sys.exit(1)

    console.print(f"[cyan][BD][/] Abrindo banco: {caminho_banco}")
    bd = aux.Banco(caminho_banco)

    # ── 2. Consulta SQL ─────────────────────────────────────
    sql = senha.SQLIPTUFALTANTE if args.faltante else senha.SQLIPTUCOMPLETO
    tipo_sql = "faltantes" if args.faltante else "completo"

    cliente_corte = str(args.cliente).zfill(4)
    if cliente_corte != '0000':
        sql = sql.rstrip(';') + f" WHERE Codigo >= '{cliente_corte}' ORDER BY Codigo;"

    console.print(f"[cyan][BD][/] Executando consulta IPTU ({tipo_sql})...")
    resultado = bd.consultar(sql)

    if not resultado:
        console.print("[yellow][BD][/] Nenhum registro encontrado.")
        sys.exit(0)

    resultado.sort(key=lambda x: x[0])

    total = len(resultado)
    if args.max > 0:
        resultado = resultado[:args.max]

    console.print(f"[cyan][BD][/] {total} inscrição(ões) encontrada(s)"
                  + (f", processando {len(resultado)}" if args.max > 0 else ""))

    # ── 3. Lista inscrições ──────────────────────────────────
    console.print(f"\n{'#':>4}  {'Código':<8}  {'IPTU':<12}")
    console.print(f"{'─'*4}  {'─'*8}  {'─'*12}")
    for i, linha in enumerate(resultado):
        cod = str(linha[Bh.Codigo])
        nriptu = str(linha[Bh.NrIPTU]).strip()
        console.print(f"{i+1:>4}  {cod:<8}  {nriptu:<12}")

    if args.dry_run:
        console.print(f"\n[yellow][DRY-RUN][/] Nenhuma extração realizada.")
        return

    # ── 4. Prepara execução ──────────────────────────────────
    os.makedirs(args.pasta, exist_ok=True)

    caminho_perfil_base = os.path.join(
        os.path.dirname(__file__), "Profile",
        os.getenv('NOMEPROFILEIPTU', 'iptu_profile')
    )

    headless = not args.visible
    console.print(f"\n[bold cyan][PW][/] Iniciando {num_workers} worker(s) "
                  f"({'visível' if args.visible else 'headless'}) | "
                  f"Tipo pagamento: {args.tipo}...")

    # Objeto mock (substitui o objeto GUI da aplicação)
    objeto = SimpleNamespace(
        pastadownload=args.pasta,
        visual=SimpleNamespace(
            somentevalores=_MockVar(False),       # False = gera boleto PDF
            codigosdebarra=_MockVar(False),       # False = não salva no banco
            tipopagamento=_MockVar(args.tipo),    # tipo de pagamento
        )
    )

    aux.prevent_sleep()
    tempo_inicio = time.time()
    log_resultados = []
    counters = {'ok': 0, 'skip': 0, 'pago': 0, 'erros': 0, 'processed': 0}
    lock = threading.Lock()

    # Fila de trabalho
    work_queue = queue.Queue()
    for i, linha in enumerate(resultado):
        work_queue.put((i, linha, len(resultado)))

    # Worker console
    _write_lock = threading.Lock()
    worker_console = _WorkerConsole(console, _write_lock)
    _console_original_bh = Bh.console
    Bh.console = worker_console

    # ── Barra de progresso ───────────────────────────────────
    progress = Progress(
        SpinnerColumn(),
        TextColumn(
            f"[bold blue]{num_workers} worker(s)[/]" if num_workers > 1
            else "[bold blue]IPTU[/]",
            justify="left"
        ),
        BarColumn(bar_width=40),
        TextColumn("{task.completed:>4}/{task.total:<4}", style="bold cyan"),
        TextColumn("│ Restante:", style="dim"),
        HMSRemainingColumn(),
        TextColumn("│ Decorrido:", style="dim"),
        HMSElapsedColumn(),
        expand=False,
    )
    task_id = progress.add_task("", total=len(resultado))

    def _build_panel():
        return Panel(
            progress,
            title="[bold green]Progresso IPTU",
            border_style="green",
            padding=(0, 1),
        )

    live = Live(_build_panel(), console=console, refresh_per_second=1,
                transient=False, screen=False, vertical_overflow="visible")
    live.start()

    try:
        if num_workers == 1:
            # ── Modo sequencial ──────────────────────────────
            from playwright.sync_api import sync_playwright

            with sync_playwright() as p:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=caminho_perfil_base,
                    headless=headless,
                    accept_downloads=True,
                    args=["--disable-blink-features=AutomationControlled",
                          "--start-maximized"]
                )
                page = context.pages[0] if context.pages else context.new_page()

                try:
                    while not work_queue.empty():
                        try:
                            i, linha, total_itens = work_queue.get_nowait()
                        except queue.Empty:
                            break

                        cod = str(linha[Bh.Codigo])
                        nriptu = str(linha[Bh.NrIPTU]).strip()

                        console.print(
                            f"\n[bold]{'='*60}[/]\n"
                            f"[bold cyan][{i+1}/{total_itens}][/] "
                            f"Cliente: {cod} | IPTU: {nriptu}"
                            f"\n[bold]{'='*60}[/]"
                        )

                        # Skip se arquivo já existe
                        caminho_existente = os.path.join(objeto.pastadownload, f"{cod}_{nriptu}.pdf")
                        if os.path.isfile(caminho_existente):
                            status = 'SKIP (arquivo existente)'
                            console.print(f"[yellow][SKIP][/] Já existe: {os.path.basename(caminho_existente)}")
                            counters['skip'] += 1
                            log_resultados.append({'Cod Cliente': cod, 'IPTU': nriptu, 'Status': status})
                            progress.advance(task_id)
                            work_queue.task_done()
                            continue

                        try:
                            dados, df = Bh.extrairIPTU_hibrido(page, objeto, linha)
                            if dados and isinstance(dados[0], list):
                                status = dados[0][7] if len(dados[0]) > 7 else 'Ok'
                            else:
                                status = dados[7] if dados and len(dados) > 7 else 'Ok'
                        except Exception as e:
                            status = f"ERRO: {e}"

                        if 'Ok' in status or 'OK' in status:
                            counters['ok'] += 1
                        elif 'SKIP' in status or 'Já extraído' in status:
                            counters['skip'] += 1
                        elif 'PAGO' in status or 'Quitad' in status:
                            counters['pago'] += 1
                        else:
                            counters['erros'] += 1

                        log_resultados.append({
                            'Cod Cliente': cod,
                            'IPTU': nriptu,
                            'Status': status,
                        })

                        progress.advance(task_id)
                        work_queue.task_done()

                finally:
                    context.close()

        else:
            # ── Modo paralelo ────────────────────────────────
            threads = []
            for wid in range(num_workers):
                t = threading.Thread(
                    target=_worker,
                    args=(wid, work_queue, objeto, headless,
                          caminho_perfil_base, progress, task_id,
                          lock, log_resultados, counters, worker_console),
                    daemon=True,
                    name=f"worker-{wid}",
                )
                threads.append(t)

            for t in threads:
                t.start()
            for t in threads:
                t.join()

    finally:
        live.stop()
        Bh.console = _console_original_bh
        aux.allow_sleep()

    # ── 5. Log Excel ─────────────────────────────────────────
    tempo_total = time.time() - tempo_inicio
    h, resto = divmod(tempo_total, 3600)
    m, s = divmod(resto, 60)
    tempo_fmt = f"{int(h):02d}:{int(m):02d}:{int(s):02d}"

    if log_resultados:
        timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        pasta_log = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logs")
        os.makedirs(pasta_log, exist_ok=True)
        caminho_log = os.path.join(pasta_log, f"Log_IPTU_{timestamp}.xlsx")

        df_log = pd.DataFrame(log_resultados)

        try:
            df_log.to_excel(caminho_log, index=False, engine='openpyxl')

            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = load_workbook(caminho_log)
            ws = wb.active

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                      fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                status_cell = row[2]  # coluna C = Status
                if status_cell.value:
                    v = str(status_cell.value)
                    if 'Ok' in v or 'OK' in v:
                        status_cell.font = Font(color='008000', bold=True)
                    elif 'SKIP' in v or 'Já extraído' in v:
                        status_cell.font = Font(color='0070C0')
                    else:
                        status_cell.font = Font(color='FF0000', bold=True)

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 35

            wb.save(caminho_log)
        except Exception as e:
            console.print(f"[yellow][AVISO][/] Erro ao formatar Excel: {e}")

        console.print(f"\n[green][LOG][/] Resultado salvo em: {caminho_log}")


    # ── 6. Resumo ─────────────────────────────────────────────
    console.print(f"\n[bold]{'='*60}[/]")
    console.print(f"[bold green]RESUMO[/]")
    console.print(f"[bold]{'='*60}[/]")
    console.print(f"  Total processados: {len(resultado)}")
    console.print(f"  Workers:           {num_workers}")
    console.print(f"  Tipo pagamento:    {args.tipo}")
    console.print(f"  [green]OK:                {counters['ok']}[/]")
    console.print(f"  [cyan]Skip (existente):  {counters['skip']}[/]")
    console.print(f"  [cyan]Pago (quitado):    {counters.get('pago', 0)}[/]")
    console.print(f"  [red]Erros:             {counters['erros']}[/]")
    console.print(f"  Tempo total:       {tempo_fmt}")
    console.print(f"  Pasta:             {os.path.abspath(args.pasta)}")

    if counters['erros'] > 0:
        console.print(f"\n[bold red]Registros com erro:[/]")
        for reg in log_resultados:
            v = reg['Status']
            if ('Ok' not in v and 'OK' not in v and 'SKIP' not in v
                    and 'PAGO' not in v and 'Quitad' not in v):
                console.print(f"  {reg['Cod Cliente']} | {reg['IPTU']} | {v}")


if __name__ == "__main__":
    main()
